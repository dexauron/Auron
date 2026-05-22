#!/usr/bin/env python3
"""WAY MARKET v8 — Full Build (10 sheets)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import date

today = date.today()
first_month = today.replace(day=1)

NAVY="FF111827"; INDIGO="FF4F46E5"
GREEN="FF10B981"; GREEN_L="FFD1FAE5"
BLUE="FF3B82F6";  BLUE_L="FFEFF6FF"
RED="FFEF4444";   RED_L="FFFEE2E2"
AMBER="FFF59E0B"; AMBER_L="FFFEF3C7"
PURPLE="FF8B5CF6";PURP_L="FFEDE9FE"
TEAL="FF14B8A6";  TEAL_L="FFCCFBF1"
GRAY="FF6B7280";  LGRAY="FFF7F8FA"
BORDER="FFE5E7EB";WHITE="FFFFFFFF"
INP="FFEFF6FF";   INP_BD="FF3B82F6"
DISABLED="FFF3F4F6"
MONEY="#,##0;[Red]-#,##0"; DATE_F="DD.MM.YYYY"

def F(c): return PatternFill("solid",start_color=c,fgColor=c)
def fnt(sz=10,bold=False,col="FF000000",it=False):
    return Font(name="Calibri",size=sz,bold=bold,color=col,italic=it)
def brd(c=BORDER,s="thin"):
    sd=Side(style=s,color=c); return Border(left=sd,right=sd,top=sd,bottom=sd)
def CA(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def LA(): return Alignment(horizontal="left",vertical="center",wrap_text=True)
def RA(): return Alignment(horizontal="right",vertical="center")

def banner(ws,txt,merge,bg=NAVY,sz=14):
    ws.merge_cells(merge)
    c=ws[merge.split(":")[0]]; c.value=txt
    c.font=fnt(sz,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=CA()
    r=int(''.join(x for x in merge.split(":")[0] if x.isdigit()))
    ws.row_dimensions[r].height=38

def sec(ws,row,txt,ncols,bg):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row,1); c.value=txt
    c.font=fnt(11,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=LA()
    ws.row_dimensions[row].height=26

def hrow(ws,row,headers,bg=NAVY,h=24):
    for ci,h_ in enumerate(headers,1):
        c=ws.cell(row,ci); c.value=h_
        c.font=fnt(9,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=CA(); c.border=brd()
    ws.row_dimensions[row].height=h

def cw(ws,widths):
    for col,w in widths.items(): ws.column_dimensions[col].width=w

def lbl(ws,row,c1,c2,text,bg=LGRAY):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=text
    c.font=fnt(10); c.fill=F(bg); c.border=brd(); c.alignment=LA()
    ws.row_dimensions[row].height=26

def inp(ws,row,c1,c2,money=True):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1)
    c.font=fnt(12,True,INDIGO); c.fill=F(INP); c.border=brd(INP_BD,"medium"); c.alignment=CA()
    if money: c.number_format=MONEY
    return c

def calc(ws,row,c1,c2,formula,money=True,col=GREEN,bg=GREEN_L):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=formula
    c.font=fnt(11,True,col); c.fill=F(bg); c.border=brd()
    c.alignment=RA() if money else CA()
    if money: c.number_format=MONEY
    return c

def hint(ws,row,c1,c2,text):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=text
    c.font=fnt(9,it=True,col=GRAY); c.fill=F(LGRAY); c.border=brd(); c.alignment=LA()

wb=Workbook()

# ════════════════════════════════════════════════
# НАСТРОЙКИ
# ════════════════════════════════════════════════
ws=wb.active; ws.title="НАСТРОЙКИ"; ws.sheet_view.showGridLines=False
banner(ws,"⚙  НАСТРОЙКИ МАГАЗИНА — все параметры в одном месте","A1:H1",INDIGO)
ws.merge_cells("A2:H2"); ws.cell(2,1).value="Листы ВВОД_КАССА и ВВОД_РАСХОДЫ подстраиваются автоматически."
ws.cell(2,1).font=fnt(10,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY)
ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22

sec(ws,4,"  📋 ПАРАМЕТРЫ МАГАЗИНА",8,INDIGO)
params=[("Название магазина","WAY MARKET №2","text"),("Дата начала учёта","01.01.2026","date"),
        ("Доля в фонд рентабельности (%)","25","num"),("Лимит на закуп (%)","75","num"),
        ("Начальный долг поставщикам (₽)",500000,"money"),("Округление сумм","До рубля","round"),
        ("Период сравнения","Прошлый месяц","period")]
for i,(label,val,t) in enumerate(params,5):
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
    c=ws.cell(i,1); c.value=label; c.font=fnt(10); c.fill=F(LGRAY); c.border=brd(); c.alignment=LA()
    ws.merge_cells(start_row=i,start_column=5,end_row=i,end_column=8)
    c=ws.cell(i,5); c.value=val; c.font=fnt(11,True,INDIGO); c.fill=F(INP)
    c.border=brd(INP_BD,"medium"); c.alignment=CA()
    if t=="money": c.number_format=MONEY
    if t=="date": c.number_format=DATE_F
    if t=="round":
        dv=DataValidation(type="list",formula1='"До рубля,До копейки,До 100 руб,До 1000 руб"')
        ws.add_data_validation(dv); dv.add(c)
    if t=="period":
        dv=DataValidation(type="list",formula1='"Прошлый месяц,Прошлый квартал,Прошлый год"')
        ws.add_data_validation(dv); dv.add(c)
    ws.row_dimensions[i].height=24

sec(ws,13,"  ☀ РЕЖИМ РАБОТЫ — активные смены",8,AMBER)
ws.merge_cells("A14:D14"); ws.cell(14,1).value="Количество смен"
ws.cell(14,1).font=fnt(10); ws.cell(14,1).fill=F(LGRAY); ws.cell(14,1).border=brd(); ws.cell(14,1).alignment=LA()
ws.merge_cells("E14:H14"); ws.cell(14,5).value=2
ws.cell(14,5).font=fnt(11,True,AMBER); ws.cell(14,5).fill=F(INP)
ws.cell(14,5).border=brd(INP_BD,"medium"); ws.cell(14,5).alignment=CA()
dv_cnt=DataValidation(type="list",formula1='"1,2,3"'); ws.add_data_validation(dv_cnt); dv_cnt.add(ws.cell(14,5))
dv_oo=DataValidation(type="list",formula1='"Вкл,Выкл"'); ws.add_data_validation(dv_oo)
for i,(sn,sv) in enumerate([("Смена ДЕНЬ","Вкл"),("Смена ВЕЧЕР","Вкл"),("Смена НОЧЬ","Выкл")],15):
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
    c=ws.cell(i,1); c.value=sn; c.font=fnt(10); c.fill=F(LGRAY); c.border=brd(); c.alignment=LA()
    ws.merge_cells(start_row=i,start_column=5,end_row=i,end_column=8)
    c=ws.cell(i,5); c.value=sv; c.font=fnt(11,True); c.fill=F(INP)
    c.border=brd(INP_BD,"medium"); c.alignment=CA(); dv_oo.add(c); ws.row_dimensions[i].height=22

sec(ws,19,"  💳 Z-ОТЧЁТ (источники выручки)",8,BLUE)
for i,(p,v) in enumerate([("Эквайринг","Вкл"),("Перевод","Вкл"),("Онлайн торговля","Выкл"),
                           ("Иман (хозяин)","Вкл"),("Выплата с кассы","Вкл")],20):
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
    c=ws.cell(i,1); c.value=p; c.font=fnt(10); c.fill=F(LGRAY); c.border=brd(); c.alignment=LA()
    ws.merge_cells(start_row=i,start_column=5,end_row=i,end_column=8)
    c=ws.cell(i,5); c.value=v; c.font=fnt(11,True); c.fill=F(INP)
    c.border=brd(INP_BD,"medium"); c.alignment=CA(); dv_oo.add(c); ws.row_dimensions[i].height=22

sec(ws,26,"  ✅ КОНТРОЛЬ КАССЫ",8,GREEN)
for i,(p,v) in enumerate([("Сверка по наличке","Вкл"),("Сверка по эквайрингу","Выкл"),("Сверка по переводу","Выкл")],27):
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
    c=ws.cell(i,1); c.value=p; c.font=fnt(10); c.fill=F(LGRAY); c.border=brd(); c.alignment=LA()
    ws.merge_cells(start_row=i,start_column=5,end_row=i,end_column=8)
    c=ws.cell(i,5); c.value=v; c.font=fnt(11,True); c.fill=F(INP)
    c.border=brd(INP_BD,"medium"); c.alignment=CA(); dv_oo.add(c); ws.row_dimensions[i].height=22

sec(ws,31,"  📋 ИНВЕНТАРЬ И ДОПОЛНИТЕЛЬНЫЙ УЧЁТ",8,PURPLE)
for i,(p,v) in enumerate([("Списание товара","Вкл"),("Возврат поставщику","Вкл"),("Касса утром/вечером (остатки)","Вкл")],32):
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
    c=ws.cell(i,1); c.value=p; c.font=fnt(10); c.fill=F(LGRAY); c.border=brd(); c.alignment=LA()
    ws.merge_cells(start_row=i,start_column=5,end_row=i,end_column=8)
    c=ws.cell(i,5); c.value=v; c.font=fnt(11,True); c.fill=F(INP)
    c.border=brd(INP_BD,"medium"); c.alignment=CA(); dv_oo.add(c); ws.row_dimensions[i].height=22

sec(ws,36,"  🔔 ПОРОГИ УВЕДОМЛЕНИЙ",8,RED)
for i,(p,v,t) in enumerate([("Расхождение кассы больше (₽)",5000,"money"),
                              ("Общий долг больше (₽)",1000000,"money"),("Просрочка больше (дней)",7,"num")],37):
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
    c=ws.cell(i,1); c.value=p; c.font=fnt(10); c.fill=F(LGRAY); c.border=brd(); c.alignment=LA()
    ws.merge_cells(start_row=i,start_column=5,end_row=i,end_column=8)
    c=ws.cell(i,5); c.value=v; c.font=fnt(11,True,RED); c.fill=F(INP)
    c.border=brd(INP_BD,"medium"); c.alignment=CA()
    if t=="money": c.number_format=MONEY; ws.row_dimensions[i].height=22

for addr in ["E15","E16","E17","E20","E21","E22","E23","E24","E27","E28","E29","E32","E33","E34"]:
    ws.conditional_formatting.add(addr,FormulaRule(formula=[f'{addr}="Вкл"'],fill=F(GREEN_L),font=fnt(11,True,GREEN)))
    ws.conditional_formatting.add(addr,FormulaRule(formula=[f'{addr}="Выкл"'],fill=F(RED_L),font=fnt(11,True,RED)))

sec(ws,41,"  📂 СПРАВОЧНИКИ (выпадающие списки)",8,INDIGO)
# Кассиры col A
ws.cell(43,1).value="КАССИРЫ"; ws.cell(43,1).font=fnt(10,True,"FFFFFFFF")
ws.cell(43,1).fill=F(PURPLE); ws.cell(43,1).alignment=CA(); ws.cell(43,1).border=brd()
for i,n in enumerate(["Иванов А.","Петров П.","Сидоров С.","Козлов К."],44):
    c=ws.cell(i,1); c.value=n; c.font=fnt(10); c.fill=F(PURP_L if i%2==0 else WHITE); c.border=brd(); c.alignment=LA()
for i in range(48,80):
    c=ws.cell(i,1); c.font=fnt(10); c.fill=F(PURP_L if i%2==0 else WHITE); c.border=brd(); c.alignment=LA()
# Категории col C
ws.cell(43,3).value="КАТЕГОРИИ"; ws.cell(43,3).font=fnt(10,True,"FFFFFFFF")
ws.cell(43,3).fill=F(AMBER); ws.cell(43,3).alignment=CA(); ws.cell(43,3).border=brd()
cats=["Закуп товара","ГСМ","Расходный материал","Зарплата","Аренда","Коммунальные",
      "Налог","Прочие расходы","Списание","Возврат","Маркетинг","Охрана"]
for i,n in enumerate(cats,44):
    c=ws.cell(i,3); c.value=n; c.font=fnt(10); c.fill=F(AMBER_L if i%2==0 else WHITE); c.border=brd(); c.alignment=LA()
for i in range(56,80):
    c=ws.cell(i,3); c.font=fnt(10); c.fill=F(AMBER_L if i%2==0 else WHITE); c.border=brd(); c.alignment=LA()
# Способы оплаты col E
ws.cell(43,5).value="СПОСОБЫ ОПЛАТЫ"; ws.cell(43,5).font=fnt(10,True,"FFFFFFFF")
ws.cell(43,5).fill=F(TEAL); ws.cell(43,5).alignment=CA(); ws.cell(43,5).border=brd()
for i,n in enumerate(["Наличка","Эквайринг","Перевод","Онлайн","Иман","Долг"],44):
    c=ws.cell(i,5); c.value=n; c.font=fnt(10); c.fill=F(TEAL_L if i%2==0 else WHITE); c.border=brd(); c.alignment=LA()
for i in range(50,80):
    c=ws.cell(i,5); c.font=fnt(10); c.fill=F(TEAL_L if i%2==0 else WHITE); c.border=brd(); c.alignment=LA()
# Типы операций col G
ws.cell(43,7).value="ТИПЫ ОПЕРАЦИЙ"; ws.cell(43,7).font=fnt(10,True,"FFFFFFFF")
ws.cell(43,7).fill=F(BLUE); ws.cell(43,7).alignment=CA(); ws.cell(43,7).border=brd()
for i,n in enumerate(["Доход","Расход","Долг","Оплата долга","Расхождение","Иман","Списание","Возврат","Касса"],44):
    c=ws.cell(i,7); c.value=n; c.font=fnt(10); c.fill=F(BLUE_L if i%2==0 else WHITE); c.border=brd(); c.alignment=LA()
for i in range(53,80):
    c=ws.cell(i,7); c.font=fnt(10); c.fill=F(BLUE_L if i%2==0 else WHITE); c.border=brd(); c.alignment=LA()

# ПОСТОЯННЫЕ РАСХОДЫ + умная таблица
sec(ws,82,"  💰 ПОСТОЯННЫЕ РАСХОДЫ (умная таблица tblПостоянные)",8,TEAL)
hrow(ws,83,["Месяц","Зарплата","Аренда","Коммунальные","Налог","Маркетинг","Охрана","ИТОГО"],TEAL,24)
months_ru=["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
for mi,mon in enumerate(months_ru):
    r=84+mi
    for ci,v in enumerate([mon,540000,366000,90000,90000,30000,25000],1):
        c=ws.cell(r,ci); c.value=v; c.font=fnt(10,bold=(ci==1))
        c.fill=F(INP if ci>=2 else (LGRAY if mi%2==0 else WHITE))
        c.border=brd(); c.alignment=LA() if ci==1 else RA()
        if ci>=2: c.number_format=MONEY
    ws.cell(r,8).value=f"=SUM(B{r}:G{r})"; ws.cell(r,8).font=fnt(10,True)
    ws.cell(r,8).fill=F(GREEN_L); ws.cell(r,8).border=brd(); ws.cell(r,8).alignment=RA()
    ws.cell(r,8).number_format=MONEY; ws.row_dimensions[r].height=22
tbl_c=Table(displayName="tblПостоянные",ref="A83:H95")
tbl_c.tableStyleInfo=TableStyleInfo(name="TableStyleLight2",showRowStripes=True); ws.add_table(tbl_c)

# СПРАВОЧНИК ТП
sec(ws,97,"  🤝 СПРАВОЧНИК ПОСТАВЩИКОВ / ТП (до 1000 строк)",8,RED)
hrow(ws,98,["№","Название / ФИО","Телефон"],RED,24)
for i,(n,nm,tel) in enumerate([(1,"Молочка-Сервис","+7-900-001"),(2,"Мясо-Премиум","+7-900-002"),
                                (3,"Хлеб-Опт","+7-900-003"),(4,"Бакалея-Юг","+7-900-004"),(5,"Овощи-Сад","+7-900-005")],99):
    ws.cell(i,1).value=n; ws.cell(i,2).value=nm; ws.cell(i,3).value=tel
    for ci in range(1,4):
        c=ws.cell(i,ci); c.font=fnt(10,bold=(ci==2))
        c.fill=F(RED_L if i%2==0 else WHITE); c.border=brd(); c.alignment=CA() if ci==1 else LA()
for i in range(104,1099):
    for ci in range(1,4):
        c=ws.cell(i,ci); c.font=fnt(10); c.fill=F(RED_L if i%2==0 else WHITE); c.border=brd(); c.alignment=CA() if ci==1 else LA()
cw(ws,{"A":18,"B":18,"C":18,"D":14,"E":15,"F":15,"G":18,"H":14})
ws.freeze_panes="A3"; ws.sheet_properties.tabColor="FF374151"
print("✓ НАСТРОЙКИ")

# ════════════════════════════════════════════════
# БАЗА_ДДС (умная таблица tblБаза)
# ════════════════════════════════════════════════
ws=wb.create_sheet("БАЗА_ДДС"); ws.sheet_view.showGridLines=False
banner(ws,"🗄  БАЗА ДДС — ТРАНЗАКЦИОННЫЙ ЛОГ (умная таблица tblБаза)","A1:H1",NAVY)
ws.merge_cells("A2:H2")
ws.cell(2,1).value="Умная таблица tblБаза — фильтрация по любому столбцу. НЕ удаляйте строки с данными!"
ws.cell(2,1).font=fnt(10,it=True,col=BLUE); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22
headers=["Дата","Смена","Кассир","Тип операции","Категория","Способ оплаты","Сумма","Комментарий"]
hrow(ws,3,headers,NAVY,32)
for r in range(4,3004):
    alt=r%2==0
    for ci in range(1,9):
        c=ws.cell(r,ci); c.border=brd(); c.fill=F(LGRAY if alt else WHITE); c.font=fnt(10)
        c.alignment=CA() if ci in [2,4,6] else LA() if ci in [3,5,8] else RA() if ci==7 else CA()
    ws.cell(r,1).number_format=DATE_F; ws.cell(r,7).number_format=MONEY; ws.row_dimensions[r].height=20
# Условное форматирование типов
for tipo,fill_,font_ in [("Доход",BLUE_L,BLUE),("Расход",RED_L,RED),("Долг",RED_L,RED),
                          ("Оплата долга",GREEN_L,GREEN),("Расхождение",AMBER_L,AMBER),
                          ("Иман",PURP_L,PURPLE),("Списание",RED_L,RED),("Возврат",GREEN_L,GREEN),("Касса",TEAL_L,TEAL)]:
    ws.conditional_formatting.add("D4:D3003",FormulaRule(formula=[f'$D4="{tipo}"'],fill=F(fill_),font=fnt(10,True,font_)))
# Data Validation
dv_tp=DataValidation(type="list",formula1="=НАСТРОЙКИ!$G$44:$G$52"); ws.add_data_validation(dv_tp); dv_tp.add("D4:D3003")
dv_pay=DataValidation(type="list",formula1="=НАСТРОЙКИ!$E$44:$E$49"); ws.add_data_validation(dv_pay); dv_pay.add("F4:F3003")
dv_cat=DataValidation(type="list",formula1="=НАСТРОЙКИ!$C$44:$C$55"); ws.add_data_validation(dv_cat); dv_cat.add("E4:E3003")
dv_ksr=DataValidation(type="list",formula1="=НАСТРОЙКИ!$A$44:$A$79"); ws.add_data_validation(dv_ksr); dv_ksr.add("C4:C3003")
dv_sm=DataValidation(type="list",formula1='"День,Вечер,Ночь,—"'); ws.add_data_validation(dv_sm); dv_sm.add("B4:B3003")
# Умная таблица
tbl_b=Table(displayName="tblБаза",ref="A3:H3003")
tbl_b.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showFirstColumn=False)
ws.add_table(tbl_b)
cw(ws,{"A":12,"B":10,"C":18,"D":16,"E":18,"F":14,"G":14,"H":30})
ws.freeze_panes="A4"; ws.sheet_properties.tabColor="FF6B7280"
print("✓ БАЗА_ДДС (tblБаза)")

# ════════════════════════════════════════════════
# ВВОД_КАССА
# ════════════════════════════════════════════════
ws=wb.create_sheet("ВВОД_КАССА"); ws.sheet_view.showGridLines=False
banner(ws,"💰  ВВОД ДАННЫХ КАССЫ — Z-отчёты и факт","A1:M1",BLUE)
ws.merge_cells("A2:M2")
ws.cell(2,1).value="1. Заполните дату (кнопки СЕГОДНЯ/ВЧЕРА)  2. Выберите кассира  3. Введите Z-отчёты смен  4. Нажмите СОХРАНИТЬ КАССУ"
ws.cell(2,1).font=fnt(10,it=True,col=BLUE); ws.cell(2,1).fill=F(BLUE_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22

# Дата
sec(ws,4,"  📅 ДАТА СМЕНЫ",13,INDIGO)
for col_,lbl_,w_ in [(1,"День",4),(4,"Месяц",1),(7,"Год",1),(10,"Итоговая дата:",1)]:
    ws.cell(5,col_).value=lbl_; ws.cell(5,col_).font=fnt(10); ws.cell(5,col_).fill=F(LGRAY)
    ws.cell(5,col_).border=brd(); ws.cell(5,col_).alignment=LA()
c=inp(ws,5,2,3,money=False)
dv_dd=DataValidation(type="whole",operator="between",formula1=1,formula2=31,allow_blank=True)
ws.add_data_validation(dv_dd); dv_dd.add(c)
c=inp(ws,5,5,6,money=False)
dv_mm=DataValidation(type="list",formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"')
ws.add_data_validation(dv_mm); dv_mm.add(c)
c=inp(ws,5,8,9,money=False)
dv_yy=DataValidation(type="whole",operator="between",formula1=2020,formula2=2099,allow_blank=True)
ws.add_data_validation(dv_yy); dv_yy.add(c)
ws.merge_cells("K5:M5")
c=ws.cell(5,11)
c.value='=IFERROR(DATE(H5,MATCH(E5,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0),C5),"")'
c.font=fnt(12,True,GREEN); c.fill=F(GREEN_L); c.border=brd(); c.alignment=CA(); c.number_format='[$-419]dddd, d mmmm yyyy'
ws.row_dimensions[5].height=28
ws.merge_cells("A6:M6"); ws.cell(6,1).value="💡 Используйте кнопки СЕГОДНЯ и ВЧЕРА (установите макрос: Alt+F8 → УстановитьВсеКнопки → Run)"
ws.cell(6,1).font=fnt(9,it=True,col=GRAY); ws.cell(6,1).fill=F(LGRAY); ws.cell(6,1).alignment=CA(); ws.row_dimensions[6].height=20

# Кассир
sec(ws,8,"  👤 КАССИР И АКТИВНЫЕ СМЕНЫ",13,PURPLE)
lbl(ws,9,1,2,"Кассир")
c=inp(ws,9,3,5,money=False)
dv_k=DataValidation(type="list",formula1="=НАСТРОЙКИ!$A$44:$A$79"); ws.add_data_validation(dv_k); dv_k.add(c)
ws.cell(9,6).value="Смены:"; ws.cell(9,6).font=fnt(10); ws.cell(9,6).alignment=LA()
for si,(sn,addr,col) in enumerate([("ДЕНЬ","НАСТРОЙКИ!$E$15",AMBER),("ВЕЧЕР","НАСТРОЙКИ!$E$16",PURPLE),("НОЧЬ","НАСТРОЙКИ!$E$17",INDIGO)]):
    cl=7+si*2
    ws.merge_cells(start_row=9,start_column=cl,end_row=9,end_column=cl+1)
    c=ws.cell(9,cl); c.value=f'=IF({addr}="Вкл","{sn} ✓","{sn} —")'
    c.font=fnt(10,True); c.fill=F(LGRAY); c.border=brd(); c.alignment=CA()
    ws.conditional_formatting.add(f"{get_column_letter(cl)}9:{get_column_letter(cl+1)}9",
        FormulaRule(formula=[f'{addr}="Вкл"'],fill=F(GREEN_L),font=fnt(10,True,GREEN)))

def shift_block(ws,row_start,label_color,shift_label,addrs_Z,addr_fnal,addr_fecv,row_rasx):
    sec(ws,row_start,f"  {shift_label}",13,label_color)
    z_fields=[("Наличная торговля (Z из кассы)",None),("Эквайринг (Z)",    "НАСТРОЙКИ!$E$20"),
              ("Перевод (Z)",                    "НАСТРОЙКИ!$E$21"),("Онлайн (Z)","НАСТРОЙКИ!$E$22"),
              ("Иман (хозяин)",                  "НАСТРОЙКИ!$E$23"),("Выплата с кассы","НАСТРОЙКИ!$E$24")]
    for ri,(label,sw) in enumerate(z_fields,row_start+1):
        lbl(ws,ri,1,4,label)
        inp(ws,ri,5,7)
        if sw:
            hint(ws,ri,8,13,f'=IF({sw}="Вкл","✓ Активно — введите сумму","⛔ Отключено в НАСТРОЙКИ")')
            ws.conditional_formatting.add(f"E{ri}:G{ri}",FormulaRule(formula=[f'{sw}="Выкл"'],fill=F(DISABLED),font=fnt(10,col=GRAY,it=True)))
        else:
            hint(ws,ri,8,13,"Основная выручка наличными по Z-отчёту")
    r_fnal=row_start+8; r_fecv=row_start+9
    lbl(ws,r_fnal,1,4,"Наличка по факту (пересчёт)")
    inp(ws,r_fnal,5,7)
    hint(ws,r_fnal,8,13,'=IF(НАСТРОЙКИ!$E$27="Вкл","Введите фактическую сумму нала","⛔ Сверка наличных отключена")')
    ws.conditional_formatting.add(f"E{r_fnal}:G{r_fnal}",FormulaRule(formula=['НАСТРОЙКИ!$E$27="Выкл"'],fill=F(DISABLED),font=fnt(10,col=GRAY,it=True)))
    lbl(ws,r_fecv,1,4,"Эквайринг по факту")
    inp(ws,r_fecv,5,7)
    hint(ws,r_fecv,8,13,'=IF(НАСТРОЙКИ!$E$28="Вкл","Введите факт эквайринга","⛔ Сверка экв. отключена")')
    ws.conditional_formatting.add(f"E{r_fecv}:G{r_fecv}",FormulaRule(formula=['НАСТРОЙКИ!$E$28="Выкл"'],fill=F(DISABLED),font=fnt(10,col=GRAY,it=True)))
    r_rx=row_start+11
    lbl(ws,r_rx,1,4,"Расхождение нал (авто)")
    z0,z5,f_=addrs_Z[0],addrs_Z[5],addr_fnal
    calc(ws,r_rx,5,7,f'=IF(НАСТРОЙКИ!$E$27="Вкл",IFERROR({z0}-{z5}-{f_},0),"—")',col=AMBER,bg=AMBER_L)
    hint(ws,r_rx,8,13,"Z Нал − Выплата − Факт нал (авто)")

# Три смены
shift_block(ws,11,"FF"+AMBER[2:],"☀ СМЕНА ДЕНЬ",
    ["E12","E13","E14","E15","E16","E17"],"E19","E20",22)
shift_block(ws,24,"FF"+PURPLE[2:],"🌙 СМЕНА ВЕЧЕР",
    ["E25","E26","E27","E28","E29","E30"],"E32","E33",35)
shift_block(ws,37,"FF"+INDIGO[2:],"🌑 СМЕНА НОЧЬ",
    ["E38","E39","E40","E41","E42","E43"],"E45","E46",48)

# Остатки кассы
sec(ws,50,"  💵 ОСТАТКИ В КАССЕ",13,TEAL)
lbl(ws,51,1,4,"Касса утром (остаток на начало дня)"); inp(ws,51,5,7)
hint(ws,51,8,13,'=IF(НАСТРОЙКИ!$E$34="Вкл","Введите остаток нала утром","⛔ Учёт остатков отключён")')
ws.conditional_formatting.add("E51:G51",FormulaRule(formula=['НАСТРОЙКИ!$E$34="Выкл"'],fill=F(DISABLED),font=fnt(10,col=GRAY,it=True)))
lbl(ws,52,1,4,"Касса вечером (остаток на конец дня)"); inp(ws,52,5,7)
hint(ws,52,8,13,'=IF(НАСТРОЙКИ!$E$34="Вкл","Введите остаток нала вечером","⛔ Учёт остатков отключён")')
ws.conditional_formatting.add("E52:G52",FormulaRule(formula=['НАСТРОЙКИ!$E$34="Выкл"'],fill=F(DISABLED),font=fnt(10,col=GRAY,it=True)))

# Итоги
sec(ws,54,"  📈 ИТОГИ СМЕНЫ (авто-расчёт)",13,NAVY)
day_f='=E12+IF(НАСТРОЙКИ!$E$20="Вкл",E13,0)+IF(НАСТРОЙКИ!$E$21="Вкл",E14,0)+IF(НАСТРОЙКИ!$E$22="Вкл",E15,0)+IF(НАСТРОЙКИ!$E$23="Вкл",E16,0)-IF(НАСТРОЙКИ!$E$24="Вкл",E17,0)'
eve_f='=E25+IF(НАСТРОЙКИ!$E$20="Вкл",E26,0)+IF(НАСТРОЙКИ!$E$21="Вкл",E27,0)+IF(НАСТРОЙКИ!$E$22="Вкл",E28,0)+IF(НАСТРОЙКИ!$E$23="Вкл",E29,0)-IF(НАСТРОЙКИ!$E$24="Вкл",E30,0)'
nig_f='=E38+IF(НАСТРОЙКИ!$E$20="Вкл",E39,0)+IF(НАСТРОЙКИ!$E$21="Вкл",E40,0)+IF(НАСТРОЙКИ!$E$22="Вкл",E41,0)+IF(НАСТРОЙКИ!$E$23="Вкл",E42,0)-IF(НАСТРОЙКИ!$E$24="Вкл",E43,0)'
lbl(ws,55,1,4,"Выручка ДЕНЬ"); c=calc(ws,55,5,7,day_f); c.font=fnt(12,True,AMBER); c.fill=F(AMBER_L)
lbl(ws,56,1,4,"Выручка ВЕЧЕР"); c=calc(ws,56,5,7,eve_f); c.font=fnt(12,True,PURPLE); c.fill=F(PURP_L)
lbl(ws,57,1,4,"Выручка НОЧЬ"); c=calc(ws,57,5,7,nig_f); c.font=fnt(12,True,INDIGO); c.fill=F(BLUE_L)
lbl(ws,58,1,4,"ИТОГО ВЫРУЧКА ЗА ДЕНЬ"); c=calc(ws,58,5,7,"=E55+E56+E57"); c.font=fnt(14,True,NAVY); ws.row_dimensions[58].height=32
lbl(ws,59,1,4,f"Лимит на закуп ({chr(37)} из НАСТРОЙКИ)"); calc(ws,59,5,7,"=ROUND(E58*НАСТРОЙКИ!$E$8/100,0)")
lbl(ws,60,1,4,"Фонд рентабельности"); calc(ws,60,5,7,"=ROUND(E58*НАСТРОЙКИ!$E$7/100,0)")

# Кнопки-заглушка
sec(ws,62,"  ✋ УПРАВЛЕНИЕ (кнопки после установки макроса: Alt+F8 → УстановитьВсеКнопки)",13,NAVY)
ws.merge_cells("A63:M63")
ws.cell(63,1).value="📅 СЕГОДНЯ | 📅 ВЧЕРА | 💾 СОХРАНИТЬ КАССУ | 🧹 ОЧИСТИТЬ | 🔄 ОБНОВИТЬ ФОРМУ"
ws.cell(63,1).font=fnt(10,it=True,col=GRAY); ws.cell(63,1).fill=F(LGRAY); ws.cell(63,1).alignment=CA(); ws.row_dimensions[63].height=28

cw(ws,{"A":5,"B":5,"C":7,"D":16,"E":12,"F":5,"G":5,"H":6,"I":6,"J":5,"K":14,"L":5,"M":12,"N":14,"O":14})
ws.freeze_panes="A4"; ws.sheet_properties.tabColor="FF"+BLUE[2:]
print("✓ ВВОД_КАССА")

# ════════════════════════════════════════════════
# ВВОД_РАСХОДЫ
# ════════════════════════════════════════════════
ws=wb.create_sheet("ВВОД_РАСХОДЫ"); ws.sheet_view.showGridLines=False
banner(ws,"💸  ВВОД РАСХОДОВ ДНЯ — закуп, долги, операционные","A1:M1",GREEN)
ws.merge_cells("A2:M2")
ws.cell(2,1).value="Заполняйте в течение дня по мере появления данных. Кнопка СОХРАНИТЬ РАСХОДЫ — записывает всё в БАЗА_ДДС."
ws.cell(2,1).font=fnt(10,it=True,col=GREEN); ws.cell(2,1).fill=F(GREEN_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22

# Дата
sec(ws,4,"  📅 ДАТА РАСХОДОВ",13,INDIGO)
for col_,lbl_ in [(1,"День"),(4,"Месяц"),(7,"Год"),(10,"Итоговая дата:")]:
    ws.cell(5,col_).value=lbl_; ws.cell(5,col_).font=fnt(10); ws.cell(5,col_).fill=F(LGRAY)
    ws.cell(5,col_).border=brd(); ws.cell(5,col_).alignment=LA()
c=inp(ws,5,2,3,money=False); dv_dd2=DataValidation(type="whole",operator="between",formula1=1,formula2=31,allow_blank=True)
ws.add_data_validation(dv_dd2); dv_dd2.add(c)
c=inp(ws,5,5,6,money=False); dv_mm2=DataValidation(type="list",formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"')
ws.add_data_validation(dv_mm2); dv_mm2.add(c)
c=inp(ws,5,8,9,money=False); dv_yy2=DataValidation(type="whole",operator="between",formula1=2020,formula2=2099,allow_blank=True)
ws.add_data_validation(dv_yy2); dv_yy2.add(c)
ws.merge_cells("K5:M5"); c=ws.cell(5,11)
c.value='=IFERROR(DATE(H5,MATCH(E5,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0),C5),"")'
c.font=fnt(12,True,GREEN); c.fill=F(GREEN_L); c.border=brd(); c.alignment=CA(); c.number_format='[$-419]dddd, d mmmm yyyy'
ws.row_dimensions[5].height=28
ws.merge_cells("A6:M6"); ws.cell(6,1).value="💡 Кнопки СЕГОДНЯ/ВЧЕРА работают после установки макроса (Alt+F8 → УстановитьВсеКнопки)"
ws.cell(6,1).font=fnt(9,it=True,col=GRAY); ws.cell(6,1).fill=F(LGRAY); ws.cell(6,1).alignment=CA(); ws.row_dimensions[6].height=20

# Ответственный
sec(ws,8,"  👤 ОТВЕТСТВЕННЫЙ ЗА РАСХОДЫ",13,TEAL)
lbl(ws,9,1,2,"Ответственный"); c=inp(ws,9,3,5,money=False)
dv_k2=DataValidation(type="list",formula1="=НАСТРОЙКИ!$A$44:$A$79"); ws.add_data_validation(dv_k2); dv_k2.add(c)
hint(ws,9,6,13,"Кто проводит закуп и расходы сегодня")

# ЗАКУП ТОВАРА
sec(ws,11,"  🛒 ЗАКУП ТОВАРА",13,AMBER)
purch=[("Закуп нал из кассы","Платим сразу при получении товара"),
       ("Закуп нал из офиса","Принёс хозяин или директор"),
       ("Закуп товара В ДОЛГ","Взяли товар — заплатим позже"),
       ("Выплата долга поставщикам","Гасим старый долг (любой ТП)")]
for ri,(label,h_) in enumerate(purch,12):
    lbl(ws,ri,1,4,label); inp(ws,ri,5,7); hint(ws,ri,8,13,h_)

# ОПЕРАЦИОННЫЕ
sec(ws,17,"  ⚙ ОПЕРАЦИОННЫЕ РАСХОДЫ",13,TEAL)
ops_=[("ГСМ (топливо, доставка)","Бензин, газ, грузоперевозки"),
      ("Расходный материал","Лента, пакеты, скотч, перчатки"),
      ("Прочие расходы","Всё остальное что не вошло выше")]
for ri,(label,h_) in enumerate(ops_,18):
    lbl(ws,ri,1,4,label); inp(ws,ri,5,7); hint(ws,ri,8,13,h_)

# ИНВЕНТАРЬ
sec(ws,22,"  📋 ИНВЕНТАРЬ",13,PURPLE)
lbl(ws,23,1,4,"Списание товара (истёк срок/поломка)")
inp(ws,23,5,7); hint(ws,23,8,13,'=IF(НАСТРОЙКИ!$E$32="Вкл","Введите сумму списания","⛔ Отключено в НАСТРОЙКИ")')
ws.conditional_formatting.add("E23:G23",FormulaRule(formula=['НАСТРОЙКИ!$E$32="Выкл"'],fill=F(DISABLED),font=fnt(10,col=GRAY,it=True)))
lbl(ws,24,1,4,"Возврат поставщику")
inp(ws,24,5,7); hint(ws,24,8,13,'=IF(НАСТРОЙКИ!$E$33="Вкл","Уменьшает долг поставщику","⛔ Отключено в НАСТРОЙКИ")')
ws.conditional_formatting.add("E24:G24",FormulaRule(formula=['НАСТРОЙКИ!$E$33="Выкл"'],fill=F(DISABLED),font=fnt(10,col=GRAY,it=True)))

# Итоги расходов
sec(ws,26,"  📊 ИТОГИ РАСХОДОВ ДНЯ (авто-расчёт)",13,NAVY)
lbl(ws,27,1,4,"Закуп товара всего"); calc(ws,27,5,7,"=E12+E13+E14",col=AMBER,bg=AMBER_L)
lbl(ws,28,1,4,"Выплата долга"); calc(ws,28,5,7,"=E15",col=RED,bg=RED_L)
lbl(ws,29,1,4,"Операционные расходы"); calc(ws,29,5,7,"=E18+E19+E20",col=TEAL,bg=TEAL_L)
lbl(ws,30,1,4,"ИТОГО РАСХОДОВ ЗА ДЕНЬ"); c=calc(ws,30,5,7,"=E27+E28+E29",col=RED,bg=RED_L)
c.font=fnt(14,True,RED); ws.row_dimensions[30].height=32
lbl(ws,31,1,4,"Лимит на закуп (из НАСТРОЙКИ)")
calc(ws,31,5,7,'=IFERROR(TEXT(НАСТРОЙКИ!$E$8,"0")&"% от выручки кассы","Настройте лимит")',money=False,col=GRAY,bg=LGRAY)

# Кнопки-заглушка
sec(ws,33,"  ✋ УПРАВЛЕНИЕ",13,NAVY)
ws.merge_cells("A34:M34"); ws.cell(34,1).value="📅 СЕГОДНЯ | 📅 ВЧЕРА | 💾 СОХРАНИТЬ РАСХОДЫ | 🧹 ОЧИСТИТЬ"
ws.cell(34,1).font=fnt(10,it=True,col=GRAY); ws.cell(34,1).fill=F(LGRAY); ws.cell(34,1).alignment=CA(); ws.row_dimensions[34].height=28

cw(ws,{"A":5,"B":5,"C":7,"D":16,"E":12,"F":5,"G":5,"H":6,"I":6,"J":5,"K":14,"L":5,"M":12,"N":14,"O":14})
ws.freeze_panes="A4"; ws.sheet_properties.tabColor="FF"+GREEN[2:]
print("✓ ВВОД_РАСХОДЫ")

# ════════════════════════════════════════════════
# ЗАПИСЬ_НА_ВЫПЛАТУ (умная таблица tblВыплаты)
# ════════════════════════════════════════════════
ws=wb.create_sheet("ЗАПИСЬ_НА_ВЫПЛАТУ"); ws.sheet_view.showGridLines=False
banner(ws,"📋  ЗАПИСЬ НА ВЫПЛАТУ ПОСТАВЩИКАМ (умная таблица tblВыплаты)","A1:K1",PURPLE)
ws.merge_cells("A2:K2"); ws.cell(2,1).value="Запишите: Дата → Поставщик → Сумма → Статус. Календарь выплат обновится автоматически."
ws.cell(2,1).font=fnt(10,it=True,col=PURPLE); ws.cell(2,1).fill=F(PURP_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=24
hdrs=["№","Дата выплаты","Поставщик (ТП)","Сумма (₽)","Статус","Накладная №","Способ оплаты","Комментарий","","","Idx"]
hrow(ws,3,hdrs,PURPLE,32)
ws.column_dimensions["K"].hidden=True
dv_tp2=DataValidation(type="list",formula1="=НАСТРОЙКИ!$B$99:$B$1098"); ws.add_data_validation(dv_tp2); dv_tp2.add("C4:C503")
dv_st=DataValidation(type="list",formula1='"Запланировано,Выплачено,Просрочено,Отменено"'); ws.add_data_validation(dv_st); dv_st.add("E4:E503")
dv_sp=DataValidation(type="list",formula1='"Наличка,Эквайринг,Перевод"',allow_blank=True); ws.add_data_validation(dv_sp); dv_sp.add("G4:G503")
for r in range(4,504):
    ws.cell(r,1).value=f'=IF(C{r}="","",ROW()-3)'
    for ci in range(1,9):
        c=ws.cell(r,ci); c.border=brd(); c.fill=F(LGRAY if r%2==0 else WHITE); c.font=fnt(10,bold=(ci==3))
        c.alignment=CA() if ci in [1,2,5,7] else LA() if ci in [3,6,8] else RA() if ci==4 else CA()
    ws.cell(r,2).number_format=DATE_F; ws.cell(r,4).number_format=MONEY
    ws.cell(r,11).value=f'=IF($B{r}="",99999,COUNTIFS($B$4:$B{r},$B{r}))'
    ws.cell(r,11).font=fnt(8,col=GRAY); ws.row_dimensions[r].height=22
for tipo,f_,fn_ in [("Запланировано",BLUE_L,BLUE),("Выплачено",GREEN_L,GREEN),("Просрочено",RED_L,RED),("Отменено",LGRAY,GRAY)]:
    ws.conditional_formatting.add("E4:E503",FormulaRule(formula=[f'$E4="{tipo}"'],fill=F(f_),font=fnt(10,True,fn_)))
# Умная таблица
tbl_v=Table(displayName="tblВыплаты",ref="A3:K503")
tbl_v.tableStyleInfo=TableStyleInfo(name="TableStyleMedium5",showRowStripes=True); ws.add_table(tbl_v)
cw(ws,{"A":6,"B":14,"C":26,"D":15,"E":16,"F":14,"G":14,"H":26})
ws.freeze_panes="A4"; ws.sheet_properties.tabColor="FF"+PURPLE[2:]
print("✓ ЗАПИСЬ_НА_ВЫПЛАТУ (tblВыплаты)")

# ════════════════════════════════════════════════
# КАЛЕНДАРЬ_ВЫПЛАТ
# ════════════════════════════════════════════════
ws=wb.create_sheet("КАЛЕНДАРЬ_ВЫПЛАТ"); ws.sheet_view.showGridLines=False
banner(ws,"📅  КАЛЕНДАРЬ ВЫПЛАТ ПОСТАВЩИКАМ","A1:N1",PURPLE)
ws.merge_cells("A2:N2"); ws.cell(2,1).value="Выберите месяц и год — видите все запланированные выплаты по датам"
ws.cell(2,1).font=fnt(10,it=True,col=PURPLE); ws.cell(2,1).fill=F(PURP_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22
ws.cell(4,1).value="Месяц:"; ws.cell(4,1).font=fnt(11,True); ws.cell(4,1).alignment=RA()
ws.merge_cells("B4:D4"); ws.cell(4,2).value="Май"
ws.cell(4,2).font=fnt(14,True,INDIGO); ws.cell(4,2).fill=F(INP); ws.cell(4,2).border=brd(INP_BD,"medium"); ws.cell(4,2).alignment=CA()
dv_mon=DataValidation(type="list",formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"')
ws.add_data_validation(dv_mon); dv_mon.add("B4")
ws.cell(4,6).value="Год:"; ws.cell(4,6).font=fnt(11,True); ws.cell(4,6).alignment=RA()
ws.merge_cells("G4:H4"); ws.cell(4,7).value=2026
ws.cell(4,7).font=fnt(14,True,INDIGO); ws.cell(4,7).fill=F(INP); ws.cell(4,7).border=brd(INP_BD,"medium"); ws.cell(4,7).alignment=CA()
ws.row_dimensions[4].height=36
ws.cell(4,16).value='=MATCH(B4,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0)'
ws.cell(5,16).value='=DATE(G4,P4,1)'; ws.cell(5,16).number_format=DATE_F
ws.cell(6,16).value='=EOMONTH(P5,0)'; ws.cell(6,16).number_format=DATE_F
ws.cell(7,16).value='=WEEKDAY(P5,2)'
ws.column_dimensions['P'].hidden=True
# KPI сводка
sec(ws,6,"  📊 СВОДКА ПО ВЫБРАННОМУ МЕСЯЦУ",14,INDIGO)
def kpi_cal(ws,row,c1,c2,lbl_,f_,bg):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=lbl_; c.font=fnt(10,True,"FFBBBBBB"); c.fill=F("FF1F2937"); c.alignment=CA()
    ws.merge_cells(start_row=row+1,start_column=c1,end_row=row+1,end_column=c2)
    c=ws.cell(row+1,c1); c.value=f_; c.font=fnt(18,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=CA(); c.number_format=MONEY
    ws.row_dimensions[row].height=20; ws.row_dimensions[row+1].height=40
kpi_cal(ws,7,1,3,"💰 К ВЫПЛАТЕ ВСЕГО",
    '=IFERROR(SUMIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!D:D,ЗАПИСЬ_НА_ВЫПЛАТУ!B:B,">="&P5,ЗАПИСЬ_НА_ВЫПЛАТУ!B:B,"<="&P6),0)',PURPLE)
kpi_cal(ws,7,5,7,"✅ ВЫПЛАЧЕНО",
    '=IFERROR(SUMIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!D:D,ЗАПИСЬ_НА_ВЫПЛАТУ!B:B,">="&P5,ЗАПИСЬ_НА_ВЫПЛАТУ!B:B,"<="&P6,ЗАПИСЬ_НА_ВЫПЛАТУ!E:E,"Выплачено"),0)',GREEN)
kpi_cal(ws,7,9,11,"📋 ЗАПЛАНИРОВАНО",
    '=IFERROR(SUMIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!D:D,ЗАПИСЬ_НА_ВЫПЛАТУ!B:B,">="&P5,ЗАПИСЬ_НА_ВЫПЛАТУ!B:B,"<="&P6,ЗАПИСЬ_НА_ВЫПЛАТУ!E:E,"Запланировано"),0)',AMBER)
kpi_cal(ws,7,12,14,"⚠ ПРОСРОЧЕНО",
    '=IFERROR(SUMIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!D:D,ЗАПИСЬ_НА_ВЫПЛАТУ!B:B,">="&P5,ЗАПИСЬ_НА_ВЫПЛАТУ!B:B,"<="&P6,ЗАПИСЬ_НА_ВЫПЛАТУ!E:E,"Просрочено"),0)',RED)
# Дни недели
ws.row_dimensions[10].height=28
for i,d in enumerate(['ПН','ВТ','СР','ЧТ','ПТ','СБ','ВС']):
    cs=1+i*2
    ws.merge_cells(start_row=10,start_column=cs,end_row=10,end_column=cs+1)
    c=ws.cell(10,cs); c.value=d; c.font=fnt(10,True,"FFFFFFFF")
    c.fill=F(RED if i>=5 else NAVY); c.alignment=CA(); c.border=brd()
# Сетка
CELL_ROWS=4; SR=11
for week in range(6):
    br=SR+week*CELL_ROWS
    for dp in range(7):
        cs=1+dp*2; ce=cs+1; di=week*7+dp
        cl=get_column_letter(cs); check=f"${cl}${br}"
        dr=f'DATE($G$4,$P$4,1)+{di}-($P$7-1)'
        ws.cell(br,cs).value=f'=IFERROR(IF(AND({dr}>=$P$5,{dr}<=$P$6),DAY({dr}),""),"")'
        ws.cell(br,cs).font=fnt(13,True,NAVY); ws.cell(br,cs).alignment=LA()
        ws.cell(br,ce).value=f'=IFERROR(IF({check}="","",SUMIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!$D:$D,ЗАПИСЬ_НА_ВЫПЛАТУ!$B:$B,{dr})),"")'
        ws.cell(br,ce).font=fnt(11,True,RED); ws.cell(br,ce).alignment=RA(); ws.cell(br,ce).number_format='#,##0;;;'
        for tp_i in range(2):
            rt=br+1+tp_i; n=tp_i+1
            ws.cell(rt,cs).value=f'=IFERROR(IF({check}="","",INDEX(ЗАПИСЬ_НА_ВЫПЛАТУ!$C:$C,SUMPRODUCT((ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503={dr})*(ЗАПИСЬ_НА_ВЫПЛАТУ!$K$4:$K$503={n})*ROW(ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503)))),"")'
            ws.cell(rt,cs).font=fnt(8,col=BLUE); ws.cell(rt,cs).alignment=LA()
            ws.cell(rt,ce).value=f'=IFERROR(IF({check}="","",SUMPRODUCT((ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503={dr})*(ЗАПИСЬ_НА_ВЫПЛАТУ!$K$4:$K$503={n})*ЗАПИСЬ_НА_ВЫПЛАТУ!$D$4:$D$503)),"")'
            ws.cell(rt,ce).font=fnt(8,True,RED); ws.cell(rt,ce).alignment=RA(); ws.cell(rt,ce).number_format='#,##0;;;'
        ws.cell(br+3,cs).value=f'=IFERROR(IF(OR({check}="",COUNTIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!$B:$B,{dr})<=2),"","+ ещё "&(COUNTIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!$B:$B,{dr})-2)),"")'
        ws.cell(br+3,cs).font=fnt(8,it=True,col=GRAY); ws.cell(br+3,cs).alignment=CA()
        ws.merge_cells(start_row=br+3,start_column=cs,end_row=br+3,end_column=ce)
        bg_c="FFFFF5F5" if dp>=5 else "FFFAFAFA"
        for ri in range(br,br+4):
            for ci_ in [cs,ce]:
                ws.cell(ri,ci_).fill=F(bg_c); ws.cell(ri,ci_).border=brd("FF999999" if ri==br else BORDER)
        for ri in [br,br+1,br+2,br+3]:
            ws.row_dimensions[ri].height=22 if ri==br else 16
ws.sheet_properties.tabColor="FF"+PURPLE[2:]
for col_i in range(1,15): ws.column_dimensions[get_column_letter(col_i)].width=13
print("✓ КАЛЕНДАРЬ_ВЫПЛАТ")

# ════════════════════════════════════════════════
# ДАШБОРД — формульные KPI + общий фильтр дат
# ════════════════════════════════════════════════
ws=wb.create_sheet("ДАШБОРД"); ws.sheet_view.showGridLines=False
banner(ws,"📊  ДАШБОРД — ПОЛНАЯ АНАЛИТИКА","A1:L1",NAVY,16)
ws.merge_cells("A2:L2")
ws.cell(2,1).value='=НАСТРОЙКИ!E5&"  |  Фильтр дат: меняйте ячейки B4 и E4 — все KPI обновятся автоматически"'
ws.cell(2,1).font=fnt(11,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=24

# ФИЛЬТР ДАТ (строка 4) — ключевые ячейки B4 и E4
ws.cell(4,1).value="🗓 ОТ:"; ws.cell(4,1).font=fnt(12,True,INDIGO); ws.cell(4,1).alignment=RA(); ws.cell(4,1).fill=F(BLUE_L); ws.cell(4,1).border=brd()
ws.cell(4,2).value=first_month; ws.cell(4,2).number_format=DATE_F  # START DATE — $B$4
ws.cell(4,2).font=fnt(14,True,INDIGO); ws.cell(4,2).fill=F(INP); ws.cell(4,2).border=brd(INP_BD,"medium"); ws.cell(4,2).alignment=CA()
ws.cell(4,3).value="ДО:"; ws.cell(4,3).font=fnt(12,True,INDIGO); ws.cell(4,3).alignment=CA(); ws.cell(4,3).fill=F(BLUE_L); ws.cell(4,3).border=brd()
ws.cell(4,4).value=today; ws.cell(4,4).number_format=DATE_F  # END DATE — $D$4
ws.cell(4,4).font=fnt(14,True,INDIGO); ws.cell(4,4).fill=F(INP); ws.cell(4,4).border=brd(INP_BD,"medium"); ws.cell(4,4).alignment=CA()
ws.merge_cells("E4:G4")
ws.cell(4,5).value='="Период: "&TEXT(B4,"DD.MM.YYYY")&" — "&TEXT(D4,"DD.MM.YYYY")&"  |  "&(D4-B4+1)&" дн."'
ws.cell(4,5).font=fnt(10,it=True,col=GRAY); ws.cell(4,5).fill=F(LGRAY); ws.cell(4,5).alignment=CA(); ws.cell(4,5).border=brd()
ws.merge_cells("H4:L4")
ws.cell(4,8).value="◀ Кнопки фильтра появятся после: Alt+F8 → УстановитьВсеКнопки → Run"
ws.cell(4,8).font=fnt(9,it=True,col=GRAY); ws.cell(4,8).fill=F(LGRAY); ws.cell(4,8).alignment=CA(); ws.cell(4,8).border=brd()
ws.row_dimensions[4].height=32

# START/END helpers — all formulas reference $B$4 and $D$4
S="$B$4"; E_="$D$4"
def sf(tipo=None,cat=None,pay=None,extra_tipo=None):
    """SUMIFS formula for dashboard KPIs"""
    conds=f"БАЗА_ДДС!A:A,\">=\"&{S},БАЗА_ДДС!A:A,\"<=\"&{E_}"
    if tipo: conds+=f",БАЗА_ДДС!D:D,\"{tipo}\""
    if extra_tipo: conds+=f',БАЗА_ДДС!D:D,"{extra_tipo}"'  # wrong, need different approach
    if cat: conds+=f",БАЗА_ДДС!E:E,\"{cat}\""
    if pay: conds+=f",БАЗА_ДДС!F:F,\"{pay}\""
    return f"=IFERROR(SUMIFS(БАЗА_ДДС!G:G,{conds}),0)"

def kpi_block(ws,row,c1,ncols,label,val_formula,bg,val_col="FFFFFFFF",sz=20):
    c2=c1+ncols-1
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=label; c.font=fnt(9,True,"FFAAAAAA"); c.fill=F("FF1E293B"); c.alignment=CA()
    ws.merge_cells(start_row=row+1,start_column=c1,end_row=row+1,end_column=c2)
    c=ws.cell(row+1,c1); c.value=val_formula; c.font=fnt(sz,True,val_col)
    c.fill=F(bg); c.alignment=CA(); c.number_format=MONEY
    ws.row_dimensions[row].height=22; ws.row_dimensions[row+1].height=44

# ── БЛОК 1: ВЫРУЧКА ──
sec(ws,6,"  💚  ВЫРУЧКА",12,GREEN)
rev_total=sf("Доход")
kpi_block(ws,7,1,3,"Общая выручка",rev_total,GREEN)
kpi_block(ws,7,4,3,"Ср. выручка в день",f"=IFERROR({rev_total[1:]}/MAX(1,{E_}-{S}+1),0)",GREEN)
kpi_block(ws,7,7,3,"Ср. выручка в смену",
    f'=IFERROR({rev_total[1:]}/MAX(1,COUNTIFS(БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Доход")),0)',GREEN)
kpi_block(ws,7,10,3,"Лучший день",
    f'=IFERROR(MAXIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Доход"),0)','FF047857')
ws.row_dimensions[9].height=10

# ── БЛОК 2: КОНТРОЛЬ КАССЫ ──
sec(ws,10,"  💙  КОНТРОЛЬ КАССЫ",12,BLUE)
rev_by_z=f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Доход",БАЗА_ДДС!H:H,"по Z"),0)'
vypls=f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!H:H,"Выплата с кассы"),0)'
raskhozh=f'=ABS(IFERROR(SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расхождение"),0))'
kpi_block(ws,11,1,3,"Выручка по Z-отчётам",rev_by_z,BLUE)
kpi_block(ws,11,4,3,"Выплаты из кассы",vypls,BLUE)
kpi_block(ws,11,7,3,"Расхождения кассы",raskhozh,"FFDC2626","FFFFFFFF")
kpi_block(ws,11,10,3,"Кол-во расхождений",
    f'=IFERROR(COUNTIFS(БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расхождение"),0)',BLUE,sz=24)
ws.cell(12,10).number_format="0 шт."; ws.row_dimensions[13].height=10

# ── БЛОК 3: ДОЛГИ ПОСТАВЩИКАМ ──
sec(ws,14,"  🔴  ДОЛГИ ПОСТАВЩИКАМ",12,RED)
vzyal_v_dolg=sf("Долг"); oplatil_dolg=sf("Оплата долга")
tek_dolg=f'=НАСТРОЙКИ!$E$9+{vzyal_v_dolg[1:]}-{oplatil_dolg[1:]}'
kpi_block(ws,15,1,3,"Текущий долг (всего)",tek_dolg,RED)
kpi_block(ws,15,4,3,"Взято в долг (период)",vzyal_v_dolg,RED)
kpi_block(ws,15,7,3,"Выплачено долга (период)",oplatil_dolg,"FF166534","FFFFFFFF")
kpi_block(ws,15,10,3,"План выплат (запланировано)",
    '=IFERROR(SUMIF(ЗАПИСЬ_НА_ВЫПЛАТУ!E:E,"Запланировано",ЗАПИСЬ_НА_ВЫПЛАТУ!D:D),0)','FF7C3AED',"FFFFFFFF")
ws.row_dimensions[17].height=10

# ── БЛОК 4: ПРИБЫЛЬ И РЕНТАБЕЛЬНОСТЬ ──
sec(ws,18,"  🟢  ПРИБЫЛЬ И РЕНТАБЕЛЬНОСТЬ",12,PURPLE)
zakup_all=sf("Расход","Закуп товара")
post_r=f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Зарплата")+SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Аренда")+SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Коммунальные")+SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Налог"),0)'
perm_r=f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"ГСМ")+SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Расходный материал")+SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Прочие расходы"),0)'
kpi_block(ws,19,1,3,"Закуп товара",zakup_all,PURPLE)
kpi_block(ws,19,4,3,"Постоянные расходы",post_r,PURPLE)
kpi_block(ws,19,7,3,"Переменные расходы",perm_r,PURPLE)
chist=f'={rev_total[1:]}-{zakup_all[1:]}-{post_r[1:]}-{perm_r[1:]}'
kpi_block(ws,19,10,3,"ЧИСТАЯ ПРИБЫЛЬ",f"=IFERROR({chist[1:]},0)","FF064E3B","FFFFFFFF",sz=18)
ws.row_dimensions[21].height=10

# ── ДЕТАЛИЗАЦИЯ РАСХОДОВ (строки 22-36) ──
sec(ws,22,"  📊  ДЕТАЛИЗАЦИЯ РАСХОДОВ ПО СТАТЬЯМ",12,NAVY)
ws.cell(23,1).value="Статья расхода"; ws.cell(23,1).font=fnt(9,True,"FFFFFFFF"); ws.cell(23,1).fill=F(NAVY); ws.cell(23,1).alignment=CA(); ws.cell(23,1).border=brd()
ws.merge_cells("B23:C23"); ws.cell(23,2).value="Сумма (₽)"; ws.cell(23,2).font=fnt(9,True,"FFFFFFFF"); ws.cell(23,2).fill=F(NAVY); ws.cell(23,2).alignment=CA(); ws.cell(23,2).border=brd()
ws.cell(23,4).value="% выручки"; ws.cell(23,4).font=fnt(9,True,"FFFFFFFF"); ws.cell(23,4).fill=F(NAVY); ws.cell(23,4).alignment=CA(); ws.cell(23,4).border=brd()
ws.merge_cells("E23:L23"); ws.cell(23,5).value="Визуальный индикатор"; ws.cell(23,5).font=fnt(9,True,"FFFFFFFF"); ws.cell(23,5).fill=F(NAVY); ws.cell(23,5).alignment=CA(); ws.cell(23,5).border=brd()
ws.row_dimensions[23].height=22

exp_cats=[
    ("Закуп товара","Расход","Закуп товара",RED),
    ("Зарплата","Расход","Зарплата",PURPLE),
    ("Аренда","Расход","Аренда",PURPLE),
    ("Коммунальные + Налог","Расход",None,PURPLE),
    ("ГСМ","Расход","ГСМ",TEAL),
    ("Расходный материал","Расход","Расходный материал",TEAL),
    ("Прочие расходы","Расход","Прочие расходы",TEAL),
    ("Маркетинг / Охрана","Расход",None,AMBER),
    ("Списание товара","Списание",None,GRAY),
    ("Возврат поставщику","Возврат",None,GREEN),
]
# Формулы для строк 24-33
row_sums={}
for ri,(name,tipo,cat,col) in enumerate(exp_cats,24):
    alt=ri%2==0
    ws.cell(ri,1).value=name; ws.cell(ri,1).font=fnt(10,bold=True); ws.cell(ri,1).fill=F(LGRAY if alt else WHITE)
    ws.cell(ri,1).border=brd(); ws.cell(ri,1).alignment=LA()
    # Сумма
    if name=="Коммунальные + Налог":
        f_=f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Коммунальные")+SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Налог"),0)'
    elif name=="Маркетинг / Охрана":
        f_=f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Маркетинг")+SUMIFS(БАЗА_ДДС!G:G,БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"Расход",БАЗА_ДДС!E:E,"Охрана"),0)'
    else:
        base=f'БАЗА_ДДС!A:A,">="&{S},БАЗА_ДДС!A:A,"<="&{E_},БАЗА_ДДС!D:D,"{tipo}"'
        if cat: base+=f',БАЗА_ДДС!E:E,"{cat}"'
        f_=f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,{base}),0)'
    ws.merge_cells(start_row=ri,start_column=2,end_row=ri,end_column=3)
    c=ws.cell(ri,2); c.value=f_; c.font=fnt(11,True,col); c.fill=F(LGRAY if alt else WHITE)
    c.border=brd(); c.alignment=RA(); c.number_format=MONEY; row_sums[ri]=f_
    # % выручки
    ws.cell(ri,4).value=f'=IFERROR(B{ri}/{rev_total[1:]},0)'; ws.cell(ri,4).font=fnt(10,col=GRAY)
    ws.cell(ri,4).fill=F(LGRAY if alt else WHITE); ws.cell(ri,4).border=brd(); ws.cell(ri,4).alignment=CA()
    ws.cell(ri,4).number_format="0.0%"
    # Бар-индикатор
    ws.merge_cells(start_row=ri,start_column=5,end_row=ri,end_column=12)
    c=ws.cell(ri,5)
    c.value=f'=IFERROR(IF(B{ri}=0,"",REPT("█",MAX(1,INT(B{ri}/MAX($B$24:$B$33)*18)))),"")'
    c.font=Font(name="Calibri",size=10,color=col); c.fill=F(LGRAY if alt else WHITE)
    c.border=brd(); c.alignment=LA()
    ws.row_dimensions[ri].height=24

# Итого расходов
ws.cell(34,1).value="ИТОГО РАСХОДОВ"; ws.cell(34,1).font=fnt(11,True,RED); ws.cell(34,1).fill=F(RED_L); ws.cell(34,1).border=brd(); ws.cell(34,1).alignment=LA()
ws.merge_cells("B34:C34"); c=ws.cell(34,2); c.value="=SUM(B24:B33)"; c.font=fnt(13,True,RED); c.fill=F(RED_L); c.border=brd(); c.alignment=RA(); c.number_format=MONEY
ws.cell(34,4).value=f'=IFERROR(B34/{rev_total[1:]},0)'; ws.cell(34,4).font=fnt(10,True,RED)
ws.cell(34,4).fill=F(RED_L); ws.cell(34,4).border=brd(); ws.cell(34,4).alignment=CA(); ws.cell(34,4).number_format="0.0%"
ws.merge_cells("E34:L34"); c=ws.cell(34,5); c.value=f'=IFERROR("ПРИБЫЛЬ: "&TEXT({chist[1:]},"#,##0 ₽")&"  |  Рентабельность: "&TEXT(IFERROR({chist[1:]}/{rev_total[1:]},0),"0.0%"),"")'
c.font=fnt(11,True,GREEN); c.fill=F(GREEN_L); c.border=brd(); c.alignment=LA(); ws.row_dimensions[34].height=28

# Подсказка
ws.row_dimensions[35].height=10
ws.merge_cells("A36:L36")
ws.cell(36,1).value="💡 Изменяйте даты в ячейках B4 (ОТ) и D4 (ДО) — все показатели обновятся. После установки макроса — используйте кнопки быстрого выбора периода."
ws.cell(36,1).font=fnt(9,it=True,col=GRAY); ws.cell(36,1).fill=F(AMBER_L); ws.cell(36,1).alignment=LA(); ws.row_dimensions[36].height=28

cw(ws,{"A":20,"B":14,"C":10,"D":10,"E":8,"F":8,"G":8,"H":8,"I":8,"J":8,"K":8,"L":8,"M":12,"N":12,"O":12,"P":12,"Q":12})
ws.freeze_panes="A5"; ws.sheet_properties.tabColor="FF"+NAVY[2:]
print("✓ ДАШБОРД (формульные KPI + общий фильтр B4:D4)")

# ════════════════════════════════════════════════
# ОТЧЁТ_РУКОВОДИТЕЛЮ
# ════════════════════════════════════════════════
ws=wb.create_sheet("ОТЧЁТ_РУКОВОДИТЕЛЮ"); ws.sheet_view.showGridLines=False
banner(ws,"📑  УПРАВЛЕНЧЕСКИЙ ОТЧЁТ ДЛЯ РУКОВОДИТЕЛЯ","A1:E1",INDIGO)
ws.merge_cells("A2:E2"); ws.cell(2,1).value="Выберите месяц — все данные обновятся автоматически"
ws.cell(2,1).font=fnt(10,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=24
ws.cell(4,1).value="Месяц:"; ws.cell(4,1).font=fnt(12,True); ws.cell(4,1).alignment=RA()
ws.merge_cells("B4:C4"); ws.cell(4,2).value="Май"
ws.cell(4,2).font=fnt(14,True,INDIGO); ws.cell(4,2).fill=F(INP); ws.cell(4,2).border=brd(INP_BD,"medium"); ws.cell(4,2).alignment=CA()
dv_mr=DataValidation(type="list",formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"')
ws.add_data_validation(dv_mr); dv_mr.add("B4")
ws.cell(4,4).value="Год:"; ws.cell(4,4).font=fnt(12,True); ws.cell(4,4).alignment=RA()
ws.cell(4,5).value=2026; ws.cell(4,5).font=fnt(14,True,INDIGO); ws.cell(4,5).fill=F(INP)
ws.cell(4,5).border=brd(INP_BD,"medium"); ws.cell(4,5).alignment=CA(); ws.row_dimensions[4].height=32
ws.merge_cells("A6:E6"); ws.cell(6,1).value='=НАСТРОЙКИ!E5&" — УПРАВЛЕНЧЕСКИЙ ОТЧЁТ"'
ws.cell(6,1).font=fnt(14,True,NAVY); ws.cell(6,1).alignment=CA(); ws.row_dimensions[6].height=30
ws.merge_cells("A7:E7"); ws.cell(7,1).value='="Период: "&B4&" "&E4&" г."'
ws.cell(7,1).font=fnt(11,it=True,col=GRAY); ws.cell(7,1).alignment=CA(); ws.row_dimensions[7].height=22
hrow(ws,9,["СТАТЬЯ","Текущий мес.","Прошлый мес.","Изменение","% от выручки"],INDIGO,28)

def sec_r(ws,row,txt,bg):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
    c=ws.cell(row,1); c.value=txt; c.font=fnt(11,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=LA(); ws.row_dimensions[row].height=24

MKEYS='{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"}'
def cur_f(tipo=None,cat=None,pay=None):
    base=f'БАЗА_ДДС!A:A,">="&DATE(E4,MATCH(B4,{MKEYS},0),1),БАЗА_ДДС!A:A,"<="&EOMONTH(DATE(E4,MATCH(B4,{MKEYS},0),1),0)'
    if tipo: base+=f',БАЗА_ДДС!D:D,"{tipo}"'
    if cat: base+=f',БАЗА_ДДС!E:E,"{cat}"'
    if pay: base+=f',БАЗА_ДДС!F:F,"{pay}"'
    return f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,{base}),0)'
def prev_f(tipo=None,cat=None,pay=None):
    base=f'БАЗА_ДДС!A:A,">="&EOMONTH(DATE(E4,MATCH(B4,{MKEYS},0),1),-2)+1,БАЗА_ДДС!A:A,"<"&DATE(E4,MATCH(B4,{MKEYS},0),1)'
    if tipo: base+=f',БАЗА_ДДС!D:D,"{tipo}"'
    if cat: base+=f',БАЗА_ДДС!E:E,"{cat}"'
    if pay: base+=f',БАЗА_ДДС!F:F,"{pay}"'
    return f'=IFERROR(SUMIFS(БАЗА_ДДС!G:G,{base}),0)'
def cur_pay(pay): return cur_f(tipo="Доход",pay=pay)
def prev_pay(pay): return prev_f(tipo="Доход",pay=pay)

def rep_row(ws,row,label,v_f,p_f,is_total=False,col=NAVY):
    alt=row%2==0
    bg=LGRAY if is_total else (LGRAY if alt else WHITE)
    ws.cell(row,1).value=label; ws.cell(row,1).font=fnt(10,bold=is_total); ws.cell(row,1).fill=F(bg); ws.cell(row,1).border=brd(); ws.cell(row,1).alignment=LA()
    ws.cell(row,2).value=v_f; ws.cell(row,2).font=fnt(10,bold=is_total,col=col if is_total else "FF000000")
    ws.cell(row,2).fill=F(bg); ws.cell(row,2).border=brd(); ws.cell(row,2).alignment=RA(); ws.cell(row,2).number_format=MONEY
    ws.cell(row,3).value=p_f; ws.cell(row,3).font=fnt(10,col=GRAY); ws.cell(row,3).fill=F(bg); ws.cell(row,3).border=brd(); ws.cell(row,3).alignment=RA(); ws.cell(row,3).number_format=MONEY
    ws.cell(row,4).value=f'=IFERROR(IF(C{row}=0,"—",(B{row}-C{row})/C{row}),"—")'
    ws.cell(row,4).font=fnt(10,True,GREEN); ws.cell(row,4).fill=F(bg); ws.cell(row,4).border=brd(); ws.cell(row,4).alignment=CA(); ws.cell(row,4).number_format="+0.0%;-0.0%;—"
    ws.cell(row,5).value=f'=IFERROR(B{row}/B16,"—")'; ws.cell(row,5).font=fnt(10,col=GRAY)
    ws.cell(row,5).fill=F(bg); ws.cell(row,5).border=brd(); ws.cell(row,5).alignment=CA(); ws.cell(row,5).number_format="0.0%"
    ws.row_dimensions[row].height=22

sec_r(ws,10,"  💚  ДОХОДЫ",GREEN)
rep_row(ws,11,"Наличная торговля",cur_pay("Наличка"),prev_pay("Наличка"))
rep_row(ws,12,"Эквайринг",cur_pay("Эквайринг"),prev_pay("Эквайринг"))
rep_row(ws,13,"Перевод",cur_pay("Перевод"),prev_pay("Перевод"))
rep_row(ws,14,"Онлайн",cur_pay("Онлайн"),prev_pay("Онлайн"))
rep_row(ws,15,"Иман",cur_f("Иман"),prev_f("Иман"))
rep_row(ws,16,"ИТОГО ВЫРУЧКА","=SUM(B11:B15)","=SUM(C11:C15)",True,GREEN)
sec_r(ws,18,"  🔴  РАСХОДЫ",RED)
for ri,(label,cat) in enumerate([("Закуп товара","Закуп товара"),("Зарплата","Зарплата"),("Аренда","Аренда"),
                                  ("Коммунальные","Коммунальные"),("Налог","Налог"),("ГСМ","ГСМ"),
                                  ("Расходный материал","Расходный материал"),
                                  ("Списание товара","Списание"),("Возврат поставщику","Возврат"),
                                  ("Прочие расходы","Прочие расходы")],19):
    tipo_="Расход" if cat not in ["Списание","Возврат"] else cat
    cat__=None if cat in ["Списание","Возврат"] else cat
    rep_row(ws,ri,label,cur_f(tipo_,cat__),prev_f(tipo_,cat__))
rep_row(ws,29,"ИТОГО РАСХОДЫ","=SUM(B19:B28)","=SUM(C19:C28)",True,RED)
sec_r(ws,31,"  🔵  ДОЛГИ ПОСТАВЩИКАМ",INDIGO)
rep_row(ws,32,"Взято в долг (период)",cur_f("Долг"),prev_f("Долг"))
rep_row(ws,33,"Выплачено долга (период)",cur_f("Оплата долга"),prev_f("Оплата долга"))
rep_row(ws,34,"Долг на конец месяца","=НАСТРОЙКИ!E9+B32-B33","=НАСТРОЙКИ!E9+C32-C33",True,INDIGO)
sec_r(ws,36,"  🟢  ПРИБЫЛЬ",GREEN)
rep_row(ws,37,"ЧИСТАЯ ПРИБЫЛЬ","=B16-B29","=C16-C29",True,GREEN)
ws.merge_cells("A39:E39")
ws.cell(39,1).value='="Рентабельность: "&IFERROR(TEXT(B37/B16,"0.0%"),"—")&"  |  Прибыль на день: "&IFERROR(TEXT(B37/DAY(EOMONTH(DATE(E4,MATCH(B4,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0),1),0)),"#,##0 ₽"),"—")'
ws.cell(39,1).font=fnt(12,True,GREEN); ws.cell(39,1).fill=F(GREEN_L); ws.cell(39,1).alignment=CA(); ws.row_dimensions[39].height=30
cw(ws,{"A":30,"B":18,"C":18,"D":16,"E":18})
ws.freeze_panes="A10"; ws.sheet_properties.tabColor="FF"+INDIGO[2:]
print("✓ ОТЧЁТ_РУКОВОДИТЕЛЮ")

# ════════════════════════════════════════════════
# ИНСТРУКЦИЯ
# ════════════════════════════════════════════════
ws=wb.create_sheet("ИНСТРУКЦИЯ"); ws.sheet_view.showGridLines=False
ws.merge_cells("A1:H1"); c=ws["A1"]
c.value="📘  WAY MARKET v8 — РУКОВОДСТВО ПОЛЬЗОВАТЕЛЯ"; c.font=fnt(16,True,"FFFFFFFF"); c.fill=F(NAVY); c.alignment=CA(); ws.row_dimensions[1].height=40
ws.merge_cells("A2:H2"); ws.cell(2,1).value="v8: ВВОД_КАССА + ВВОД_РАСХОДЫ (разделены). Все таблицы умные. Дашборд с общим фильтром дат."
ws.cell(2,1).font=fnt(11,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=28

content=[
    ("sec","🚀 ПЕРВЫЙ ЗАПУСК (один раз)",GREEN),
    ("step","Шаг 1.","Файл → Сохранить как → WAY_MARKET_v8.xlsm (обязательно .xlsm для макросов)"),
    ("step","Шаг 2.","Включите макросы (жёлтая полоса → «Включить содержимое»)"),
    ("step","Шаг 3.","Откройте редактор VBA: Alt + F11"),
    ("step","Шаг 4.","File → Import File → выберите Модуль_WM8.bas"),
    ("step","Шаг 5.","Закройте редактор. Alt+F8 → УстановитьВсеКнопки → Run"),
    ("step","✅","Шаблон готов. Кнопки появятся на всех листах. Дашборд настроен на текущий месяц."),
    ("space","",""),
    ("sec","📝 ЕЖЕДНЕВНАЯ РАБОТА",PURPLE),
    ("step","Утром →","ВВОД_КАССА: нажмите ВЧЕРА → впишите Z-отчёты смен → кнопка СОХРАНИТЬ КАССУ"),
    ("step","Днём →","ВВОД_РАСХОДЫ: добавляйте закупы и расходы по мере появления → СОХРАНИТЬ РАСХОДЫ"),
    ("step","При выплате ТП →","ЗАПИСЬ_НА_ВЫПЛАТУ: дата + поставщик + сумма + статус"),
    ("step","Анализ →","ДАШБОРД: меняйте даты в B4 и D4, или нажимайте кнопки периода"),
    ("step","Отчёт →","ОТЧЁТ_РУКОВОДИТЕЛЮ: выберите месяц — всё считается автоматически"),
    ("space","",""),
    ("sec","🗄 УМНЫЕ ТАБЛИЦЫ (Excel Tables)",TEAL),
    ("step","tblБаза","БАЗА_ДДС: вся история транзакций. Автофильтр по любому столбцу."),
    ("step","tblВыплаты","ЗАПИСЬ_НА_ВЫПЛАТУ: история выплат поставщикам."),
    ("step","tblПостоянные","НАСТРОЙКИ: постоянные расходы по месяцам."),
    ("step","Как использовать","Кликните на таблицу → вкладка «Конструктор» → фильтр по нужному столбцу."),
    ("space","",""),
    ("sec","📊 ДАШБОРД — ОБЩИЙ ФИЛЬТР ДАТ",INDIGO),
    ("step","Ячейка B4","ОТ (начало периода) — введите любую дату"),
    ("step","Ячейка D4","ДО (конец периода) — введите любую дату"),
    ("step","Кнопки периода","После установки макроса: Этот месяц / Прошлый месяц / С начала года / Всё время"),
    ("step","Все KPI","Обновляются одновременно при изменении B4 или D4"),
    ("space","",""),
    ("sec","⚙ НАСТРОЙКИ",AMBER),
    ("step","Смены","Раздел 2: Вкл/Выкл для Дня, Вечера, Ночи"),
    ("step","Z-отчёт","Раздел 3: какие источники выручки использовать"),
    ("step","Пороги","Раздел 6: когда показывать предупреждения"),
    ("step","Кассиры","Справочник: колонка A от строки 44"),
    ("step","Поставщики","Справочник ТП: от строки 99 (до 1000 шт.)"),
    ("space","",""),
    ("sec","🛡 ВАЖНЫЕ ПРАВИЛА",RED),
    ("step","1.","НЕ удаляйте данные из БАЗА_ДДС — это основа всех отчётов"),
    ("step","2.","НЕ редактируйте зелёные ячейки (формулы-расчёты)"),
    ("step","3.","Скрытая колонка K в ЗАПИСЬ_НА_ВЫПЛАТУ — служебная, не трогать"),
    ("step","4.","Файл обязательно в формате .xlsm (иначе макросы не работают)"),
    ("step","5.","Перед крупными изменениями — резервная копия"),
]
r=3
for tp,lb,val in content:
    if tp=="space": ws.row_dimensions[r].height=8; r+=1; continue
    if tp=="sec":
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
        c=ws.cell(r,1); c.value=lb; c.font=fnt(12,True,"FFFFFFFF")
        c.fill=F(val); c.alignment=LA(); ws.row_dimensions[r].height=28
    else:
        ws.cell(r,1).value=lb; ws.cell(r,1).font=fnt(10,True,BLUE); ws.cell(r,1).alignment=LA()
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
        c=ws.cell(r,2); c.value=val; c.font=fnt(10); c.alignment=LA(); ws.row_dimensions[r].height=22
    r+=1
cw(ws,{"A":14,"B":22,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16})
ws.sheet_properties.tabColor="FF"+AMBER[2:]
print(f"✓ ИНСТРУКЦИЯ ({r} строк)")

# ════════════════════════════════════════════════
# КАК_СДЕЛАТЬ_ДАШБОРД
# ════════════════════════════════════════════════
ws=wb.create_sheet("КАК_СДЕЛАТЬ_ДАШБОРД"); ws.sheet_view.showGridLines=False
ws.merge_cells("A1:H1"); c=ws["A1"]
c.value="📊  КАК НАСТРОИТЬ ДАШБОРД — ПОШАГОВАЯ ИНСТРУКЦИЯ"; c.font=fnt(16,True,"FFFFFFFF"); c.fill=F(NAVY); c.alignment=CA(); ws.row_dimensions[1].height=40
ws.merge_cells("A2:H2"); ws.cell(2,1).value="Дашборд ГОТОВ: 16 KPI + детализация расходов. Для интерактивности — установите макрос и настройте обновление."
ws.cell(2,1).font=fnt(11,it=True,col=GREEN); ws.cell(2,1).fill=F(GREEN_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=28

dash_content=[
    ("sec","📌 ЧТО УЖЕ ГОТОВО НА ЛИСТЕ ДАШБОРД",GREEN),
    ("step","✅ 4 блока KPI","ВЫРУЧКА / КОНТРОЛЬ КАССЫ / ДОЛГИ / ПРИБЫЛЬ — 16 показателей"),
    ("step","✅ Общий фильтр","Ячейки B4 (ОТ) и D4 (ДО) — меняете одно число, обновляется всё"),
    ("step","✅ Детализация","10 статей расходов с визуальным баром и % от выручки"),
    ("step","✅ Кнопки VBA","После установки макроса: Этот месяц / Прошлый / С начала года / Всё время"),
    ("space","",""),
    ("sec","🚀 БЫСТРЫЙ СТАРТ (5 минут)",INDIGO),
    ("step","1.","Установите макрос: Alt+F11 → File → Import → Модуль_WM8.bas"),
    ("step","2.","Alt+F8 → УстановитьВсеКнопки → Run"),
    ("step","3.","Перейдите на ДАШБОРД — появятся кнопки периода"),
    ("step","4.","Введите тестовые данные через ВВОД_КАССА — нажмите СОХРАНИТЬ КАССУ"),
    ("step","5.","Вернитесь на ДАШБОРД — данные обновятся автоматически"),
    ("space","",""),
    ("sec","📊 ДОБАВИТЬ СВОДНЫЕ ТАБЛИЦЫ (опционально, 30 мин)",BLUE),
    ("step","Зачем?","Для графиков и дополнительных срезов данных"),
    ("step","Шаг 1.","Лист ДАШБОРД → свободная область → Вставка → Сводная таблица"),
    ("step","Шаг 2.","Источник: tblБаза (умная таблица БАЗА_ДДС)"),
    ("step","Шаг 3.","Добавьте поля: Дата в СТРОКИ, Сумма в ЗНАЧЕНИЯ, Тип операции в ФИЛЬТР"),
    ("step","Шаг 4.","Вставить срез (Slicer) → Дата → подключить ко всем сводным"),
    ("step","Важно!","MAXIFS требует Excel 2019+ или LibreOffice 6.0+"),
    ("space","",""),
    ("sec","📐 КАРТА KPI — ЯЧЕЙКИ ДАШБОРДА",NAVY),
    ("hdr","Ячейка | KPI | Формула",""),
    ("kpi","B8","Общая выручка — SUMIFS по типу Доход"),
    ("kpi","E8","Ср. выручка в день — Общая / кол-во дней"),
    ("kpi","H8","Ср. выручка в смену — Общая / кол-во смен"),
    ("kpi","K8","Лучший день — MAXIFS по типу Доход"),
    ("kpi","B12","Выручка по Z — только с пометкой 'по Z'"),
    ("kpi","E12","Выплаты из кассы — по комментарию"),
    ("kpi","H12","Расхождения — тип Расхождение"),
    ("kpi","K12","Кол-во расхождений — COUNTIFS"),
    ("kpi","B16","Текущий долг — начальный + Долг - Оплата долга"),
    ("kpi","E16","Взято в долг — тип Долг"),
    ("kpi","H16","Выплачено — тип Оплата долга"),
    ("kpi","K16","План выплат — из ЗАПИСЬ_НА_ВЫПЛАТУ"),
    ("kpi","B20","Закуп товара — категория Закуп товара"),
    ("kpi","E20","Постоянные расходы — Зарплата+Аренда+Коммуналка+Налог"),
    ("kpi","H20","Переменные расходы — ГСМ+Материал+Прочие"),
    ("kpi","K20","Чистая прибыль — Выручка - все расходы"),
    ("space","",""),
    ("sec","💡 СОВЕТЫ",AMBER),
    ("step","Авто-обновление","Правый клик на сводной → Параметры → «Обновить при открытии»"),
    ("step","Цветовой код","🟢 Зелёный=хорошо | 🔴 Красный=расход | 🟡 Жёлтый=внимание | 🔵 Синий=нейтраль"),
    ("step","Печать отчёта","ОТЧЁТ_РУКОВОДИТЕЛЮ → Ctrl+P → готово к печати"),
]
r=4
for tp,lb,val in dash_content:
    if tp=="space": ws.row_dimensions[r].height=8; r+=1; continue
    if tp=="sec":
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
        c=ws.cell(r,1); c.value=lb; c.font=fnt(12,True,"FFFFFFFF"); c.fill=F(val); c.alignment=LA(); ws.row_dimensions[r].height=28
    elif tp in ("step","kpi"):
        ws.cell(r,1).value=lb; ws.cell(r,1).font=fnt(10,True,BLUE if tp=="step" else GREEN); ws.cell(r,1).alignment=LA()
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
        c=ws.cell(r,2); c.value=val; c.font=fnt(10); c.alignment=LA(); ws.row_dimensions[r].height=22
    elif tp=="hdr":
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
        c=ws.cell(r,1); c.value=lb; c.font=fnt(10,True,GRAY); c.fill=F(LGRAY); c.alignment=CA(); ws.row_dimensions[r].height=22
    r+=1
cw(ws,{"A":14,"B":22,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16})
ws.sheet_properties.tabColor="FF"+NAVY[2:]
print(f"✓ КАК_СДЕЛАТЬ_ДАШБОРД ({r} строк)")

# ════════════════════════════════════════════════
# Порядок листов
# ════════════════════════════════════════════════
order=["ИНСТРУКЦИЯ","КАК_СДЕЛАТЬ_ДАШБОРД","ВВОД_КАССА","ВВОД_РАСХОДЫ",
       "ЗАПИСЬ_НА_ВЫПЛАТУ","КАЛЕНДАРЬ_ВЫПЛАТ","БАЗА_ДДС","ДАШБОРД","ОТЧЁТ_РУКОВОДИТЕЛЮ","НАСТРОЙКИ"]
for idx,name in enumerate(order):
    if name in wb.sheetnames:
        cur=wb.sheetnames.index(name)
        if cur!=idx: wb.move_sheet(name,offset=idx-cur)

out_path="/home/user/Auron/WAY_MARKET_v8.xlsx"
wb.save(out_path)
print(f"\n✅ Файл сохранён: {out_path}")
print(f"   Листов: {len(wb.sheetnames)}")
print(f"   Листы: {', '.join(wb.sheetnames)}")
