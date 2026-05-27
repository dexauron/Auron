#!/usr/bin/env python3
"""
build_wm_macros.py — ФИНАНСОВЫЙ КОНТРОЛЬ cash book builder
Block 1 of 5: Foundation + БАЗА_ДДС

Outputs:
  ФИНАНСОВЫЙ_КОНТРОЛЬ.xlsx       — full workbook (rename to .xlsm after VBA import)
  ФИНАНСОВЫЙ_КОНТРОЛЬ_VBA.bas    — VBA code to import (created in Block 5)
"""

import os
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel

random.seed(42)

# ═══════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════

YEAR = 2025
SHOP = "ФИНАНСОВЫЙ КОНТРОЛЬ"
FONT = "Calibri"

TEAL   = "FF0B4F54"
TEAL_M = "FF0D9488"
TEAL_L = "FFE6FFFA"
GOLD   = "FFB45309"
GREEN  = "FF059669"
RED    = "FFDC2626"
AMBER  = "FFD97706"
PURPLE = "FF7C3AED"
NAVY   = "FF111827"
WHITE  = "FFFFFFFF"
GRAY_L = "FFF9FAFB"
GRAY_M = "FFE5E7EB"
GRAY_D = "FF6B7280"
BLUE   = "FF1D4ED8"

FMT_RUB  = '#,##0\\ ₽'
FMT_DATE = 'DD.MM.YYYY'

CASHIERS     = ["Айгуль", "Зарина", "Данияр"]
SHIFTS       = ["Утро", "Вечер", "Ночная"]
SHIFT_FACTOR = {"Утро": 0.30, "Вечер": 0.50, "Ночная": 0.20}
PAY_METHODS  = ["Наличные", "Карта", "Перевод"]
SUPPLIERS    = ["ТД Метро", "Лента", "Вкусвилл", "Магнит", "Х5 Ритейл", "Юнилевер"]
CATS_EXPENSE = ["ЗП", "Аренда", "Налоги", "Интернет", "Закуп товара",
                "Оплата ТП", "Коммуналка", "Реклама", "Другое"]
TYPES_ALL    = ["Приход", "Расход", "Долг"]
MONTHS_RU    = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

MONTH_FACTOR = {
    1: 1.15, 2: 1.05, 3: 1.10, 4: 0.90, 5: 0.88, 6: 0.85,
    7: 0.90, 8: 0.92, 9: 0.95, 10: 1.00, 11: 1.05, 12: 1.25,
}

TYPE_COLORS = {"Приход": GREEN, "Расход": RED, "Долг": AMBER}

# ═══════════════════════════════════════════════════════════════
#  STYLE FACTORIES
# ═══════════════════════════════════════════════════════════════

def mkfill(color):
    return PatternFill("solid", fgColor=color)


def mkfont(color=NAVY, size=10, bold=False):
    return Font(name=FONT, size=size, bold=bold, color=color)


def mkalign(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def mkborder(style="thin", color="FFD1D5DB"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ═══════════════════════════════════════════════════════════════
#  SHEET HELPERS
# ═══════════════════════════════════════════════════════════════

def set_widths(ws, pairs):
    for col, w in pairs:
        ws.column_dimensions[col].width = w


def sheet_title(ws, text, subtitle="", ncols=9):
    cl = get_column_letter(ncols)
    ws.merge_cells(f"A1:{cl}1")
    c = ws.cell(1, 1, text)
    c.fill = mkfill(TEAL)
    c.font = mkfont(WHITE, 16, True)
    c.alignment = mkalign("left", "center")
    ws.row_dimensions[1].height = 38
    if subtitle:
        ws.merge_cells(f"A2:{cl}2")
        c = ws.cell(2, 1, subtitle)
        c.fill = mkfill(TEAL_M)
        c.font = mkfont(WHITE, 10)
        c.alignment = mkalign("left", "center")
        ws.row_dimensions[2].height = 20


def sec_hdr(ws, row, text, size=11, bg=TEAL, fg=WHITE,
            ncols=9, height=22, bold=True):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.fill = mkfill(bg)
    c.font = mkfont(fg, size, bold)
    c.alignment = mkalign("left", "center")
    ws.row_dimensions[row].height = height


def tbl_hdr(ws, row, headers, bg=TEAL_M, fg=WHITE, height=20, start_col=1):
    for i, h in enumerate(headers, start_col):
        c = ws.cell(row, i, h)
        c.fill = mkfill(bg)
        c.font = mkfont(fg, 10, True)
        c.alignment = mkalign("center", "center")
        c.border = mkborder()
    ws.row_dimensions[row].height = height


def d_cell(ws, row, col, v, alt=False, halign="left", fmt=None,
           color=None, bold=False):
    c = ws.cell(row, col, v)
    c.fill = mkfill(GRAY_L if alt else WHITE)
    c.font = mkfont(color or NAVY, 10, bold)
    c.alignment = mkalign(halign, "center")
    c.border = mkborder()
    if fmt:
        c.number_format = fmt
    return c


# ═══════════════════════════════════════════════════════════════
#  DEMO DATA GENERATION
# ═══════════════════════════════════════════════════════════════

def rr(lo, hi, step=100):
    """Random rounded number between lo and hi."""
    lo_s = max(1, int(lo / step))
    hi_s = max(1, int(hi / step))
    if lo_s > hi_s:
        hi_s = lo_s
    return random.randint(lo_s, hi_s) * step


def last_day(m, y=YEAR):
    if m == 2:
        return 28
    if m in (4, 6, 9, 11):
        return 30
    return 31


def gen_baza():
    """Generate full-year demo data for БАЗА_ДДС.
    Returns list of tuples: (date, shift, cashier, type, cat, method, amount, disc, comment)
    """
    rows = []

    # ─── ПРИХОД (revenue) ───
    d = date(YEAR, 1, 1)
    while d <= date(YEAR, 12, 31):
        mf = MONTH_FACTOR[d.month]
        for shift in SHIFTS:
            cashier = random.choice(CASHIERS)
            shift_base = rr(70000 * mf, 140000 * mf, 1000)
            sf = SHIFT_FACTOR.get(shift, 0.33)
            total = int(shift_base * sf)

            cp = random.uniform(0.34, 0.44)
            kp = random.uniform(0.46, 0.56)
            tp = max(0.0, 1.0 - cp - kp)

            for method, pct in [("Наличные", cp), ("Карта", kp), ("Перевод", tp)]:
                amt = round(total * pct / 100) * 100
                disc = 0
                if method == "Наличные" and random.random() < 0.12:
                    disc = random.choice([-200, -100, -50, 50, 100, 200])
                rows.append((d, shift, cashier, "Приход", "Продажи",
                             method, amt, disc, ""))
        d += timedelta(1)

    # ─── FIXED MONTHLY EXPENSES ───
    for m in range(1, 13):
        ld = last_day(m)
        # ЗП (advance + main)
        for day, amt, cmt in [
            (random.randint(10, 12), rr(40000, 50000), "Аванс"),
            (random.randint(25, ld), rr(95000, 110000), "Зарплата"),
        ]:
            rows.append((date(YEAR, m, day), "Утро", random.choice(CASHIERS),
                         "Расход", "ЗП", "Наличные", amt, 0, cmt))

        rows.append((date(YEAR, m, random.randint(1, 3)), "Утро",
                     random.choice(CASHIERS), "Расход", "Аренда", "Перевод",
                     80000, 0, f"Аренда {m:02d}.{YEAR}"))

        rows.append((date(YEAR, m, 10), "Утро", random.choice(CASHIERS),
                     "Расход", "Интернет", "Перевод", 3500, 0, "Провайдер"))

        rows.append((date(YEAR, m, 15), "Утро", random.choice(CASHIERS),
                     "Расход", "Коммуналка", "Перевод",
                     rr(14000, 22000, 500), 0, ""))

        if m in (1, 4, 7, 10):
            rows.append((date(YEAR, m, 20), "Утро", random.choice(CASHIERS),
                         "Расход", "Налоги", "Перевод",
                         rr(35000, 55000), 0, "Квартальный налог"))

        rows.append((date(YEAR, m, random.randint(5, 15)), "Утро",
                     random.choice(CASHIERS), "Расход", "Реклама", "Перевод",
                     rr(5000, 15000), 0, ""))

    # ─── CASH GOODS PURCHASE (main COGS) ───
    d = date(YEAR, 1, 1)
    while d <= date(YEAR, 12, 31):
        # 1-2 purchases per day to cover realistic COGS
        for _ in range(random.randint(1, 2)):
            if random.random() < 0.75:
                mf = MONTH_FACTOR[d.month]
                rows.append((d, random.choice(SHIFTS), random.choice(CASHIERS),
                             "Расход", "Закуп товара",
                             random.choice(["Наличные", "Карта", "Перевод"]),
                             rr(35000 * mf, 90000 * mf), 0,
                             random.choice(SUPPLIERS)))
        d += timedelta(1)

    # ─── CREDIT PURCHASE (ДОЛГ) ───
    for m in range(1, 13):
        ld = last_day(m)
        mf = MONTH_FACTOR[m]
        for _ in range(random.randint(2, 4)):
            sup = random.choice(SUPPLIERS)
            rows.append((date(YEAR, m, random.randint(1, ld)), "Утро",
                         random.choice(CASHIERS), "Долг", "Закуп товара", "Перевод",
                         rr(60000 * mf, 150000 * mf), 0, sup))

    # ─── DEBT PAYMENTS ───
    for m in range(1, 13):
        ld = last_day(m)
        for _ in range(random.randint(1, 2)):
            rows.append((date(YEAR, m, random.randint(10, ld)), "Утро",
                         random.choice(CASHIERS), "Расход", "Оплата ТП", "Перевод",
                         rr(30000, 70000), 0, random.choice(SUPPLIERS)))

    # ─── MISC ───
    for m in range(1, 13):
        ld = last_day(m)
        for _ in range(random.randint(2, 5)):
            rows.append((date(YEAR, m, random.randint(1, ld)),
                         random.choice(SHIFTS), random.choice(CASHIERS),
                         "Расход", "Другое",
                         random.choice(["Наличные", "Карта"]),
                         rr(300, 5000, 100), 0, ""))

    rows.sort(key=lambda r: (r[0], r[1], r[3]))
    return rows


def gen_vyplaty():
    """Generate ЗАПИСЬ_ВЫПЛАТ demo data."""
    rows = []
    # Use a fixed cutoff tied to YEAR so statuses don't flip with today's date.
    # Payments up to and including Oct 2025 are mostly paid; Nov-Dec are planned.
    paid_cutoff = date(YEAR, 10, 31)
    idx = 1
    for m in range(1, 13):
        ld = last_day(m)
        for sup in random.sample(SUPPLIERS, random.randint(2, 4)):
            pd = date(YEAR, m, random.randint(10, min(25, ld)))
            amt = rr(30000, 80000)
            if pd <= paid_cutoff and random.random() > 0.15:
                status = "Оплачено"
            else:
                status = "Запланировано"
            fd = None
            if status == "Оплачено":
                fd = date(YEAR, m, min(pd.day + random.randint(0, 5), ld))
            rows.append({
                "idx": idx, "plan_date": pd, "supplier": sup,
                "amount": amt, "status": status,
                "invoice": f"НК-{YEAR}-{idx:04d}",
                "method": random.choice(["Перевод", "Наличные"]),
                "fact_date": fd, "comment": "",
            })
            idx += 1
    rows.sort(key=lambda r: r["plan_date"])
    return rows


# ═══════════════════════════════════════════════════════════════
#  BLOCK 1: БАЗА_ДДС
# ═══════════════════════════════════════════════════════════════

BAZA_HEADERS = ["Дата", "Смена", "Кассир", "Тип", "Категория",
                "Способ оплаты", "Сумма", "Расхождение", "Комментарий"]
BAZA_FMTS = [FMT_DATE, None, None, None, None, None, FMT_RUB, FMT_RUB, None]
BAZA_ALNS = ["center", "center", "center", "center", "left",
             "center", "right", "right", "left"]


def build_baza(ws, rows):
    # Title
    sheet_title(ws, "  БАЗА ДДС",
                "  Единая база всех финансовых операций магазина")

    # Stats row 3
    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:D3")
    c = ws.cell(3, 1, f"Магазин: {SHOP}   |   Год: {YEAR}")
    c.font = mkfont(GRAY_D, 9)
    c.alignment = mkalign("left", "center")

    ws.merge_cells("E3:I3")
    c = ws.cell(3, 5)
    c.value = ('=COUNTA(tblБаза[Дата])&" записей  |  Обновлено: "'
               '&TEXT(MAX(tblБаза[Дата]),"DD.MM.YYYY")')
    c.font = mkfont(GRAY_D, 9)
    c.alignment = mkalign("right", "center")

    ws.row_dimensions[4].height = 5  # spacer

    # Header row 5
    tbl_hdr(ws, 5, BAZA_HEADERS)

    # Data rows from row 6
    for i, r in enumerate(rows):
        rn = 6 + i
        alt = (i % 2 == 1)
        d, shift, cashier, typ, cat, method, amt, disc, cmt = r

        vals = [d, shift, cashier, typ, cat, method, amt,
                disc if disc != 0 else None, cmt]

        for col, (v, fmt, aln) in enumerate(zip(vals, BAZA_FMTS, BAZA_ALNS), 1):
            c = d_cell(ws, rn, col, v, alt, aln, fmt)
            if col == 4:  # Тип
                c.font = mkfont(TYPE_COLORS.get(v, NAVY), 10, True)

        ws.row_dimensions[rn].height = 17

    # Named table — pre-sized to 3000 rows
    last_row = max(5 + len(rows), 3005)
    tbl = Table(displayName="tblБаза", ref=f"A5:I{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(tbl)

    ws.freeze_panes = "A6"
    set_widths(ws, [("A", 13), ("B", 10), ("C", 12), ("D", 12),
                    ("E", 18), ("F", 16), ("G", 14), ("H", 14), ("I", 25)])


# ═══════════════════════════════════════════════════════════════
#  BLOCK 2: ВВОД_КАССА + ВВОД_РАСХОДЫ
# ═══════════════════════════════════════════════════════════════

def _form_label(ws, row, col, text, ncols=1, bg=TEAL_L, fg=NAVY, size=10):
    """Label cell with light teal bg, bold text."""
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + ncols - 1)
    c = ws.cell(row, col, text)
    c.fill = mkfill(bg)
    c.font = mkfont(fg, size, True)
    c.alignment = mkalign("left", "center")
    return c


def _form_input(ws, row, col, value=None, ncols=1, fmt=None, halign="left"):
    """Editable input cell with strong teal underline."""
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + ncols - 1)
    c = ws.cell(row, col, value)
    c.fill = mkfill(WHITE)
    c.font = mkfont(NAVY, 11, True)
    c.alignment = mkalign(halign, "center")
    c.border = Border(
        left=Side(style="thin", color="FFD1D5DB"),
        right=Side(style="thin", color="FFD1D5DB"),
        top=Side(style="thin", color="FFD1D5DB"),
        bottom=Side(style="medium", color="FF0B4F54"),
    )
    if fmt:
        c.number_format = fmt
    return c


def _form_btn(ws, row, col, text, bg=GREEN, ncols=1, nrows=1, size=12):
    """Button-styled cell."""
    if ncols > 1 or nrows > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row + nrows - 1, end_column=col + ncols - 1)
    c = ws.cell(row, col, text)
    c.fill = mkfill(bg)
    c.font = mkfont(WHITE, size, True)
    c.alignment = mkalign("center", "center")
    c.border = Border(
        left=Side(style="thin", color=bg),
        right=Side(style="thin", color=bg),
        top=Side(style="thin", color=bg),
        bottom=Side(style="medium", color=NAVY),
    )
    return c


def _add_dv(ws, formula, cell_ref, show_error=True):
    """Attach data validation dropdown to a cell."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    if show_error:
        dv.error = "Выберите значение из списка"
        dv.errorTitle = "Неверное значение"
        dv.showErrorMessage = True
    else:
        dv.showErrorMessage = False  # allow free-type for autocomplete fields
    ws.add_data_validation(dv)
    dv.add(cell_ref)


def build_vvod_kassa(ws):
    """ВВОД_КАССА — Form for daily cash register entry. 16 rows × 7 cols."""
    ws.sheet_view.showGridLines = False

    # Column widths (7 cols)
    set_widths(ws, [("A", 18), ("B", 14), ("C", 14), ("D", 14),
                    ("E", 14), ("F", 14), ("G", 14)])

    # Title block
    sheet_title(ws, "  ВВОД КАССЫ",
                "  Заполните данные смены и нажмите СОХРАНИТЬ", ncols=7)

    # Row 3: Дата | Смена | Кассир
    ws.row_dimensions[3].height = 30
    _form_label(ws, 3, 1, "  Дата:")
    _form_input(ws, 3, 2, value=None, fmt=FMT_DATE, halign="center")
    _form_label(ws, 3, 3, "  Смена:")
    _form_input(ws, 3, 4, value=None, halign="center")
    _form_label(ws, 3, 5, "  Кассир:")
    _form_input(ws, 3, 6, value=None, ncols=2, halign="center")

    # Data validations for D3 (shift), F3 (cashier) — reference Настройки справочники
    _add_dv(ws, "'Настройки'!$H$22:$H$45", "D3")
    _add_dv(ws, "'Настройки'!$B$22:$B$45", "F3", show_error=False)  # autocomplete

    # Row 4: TODAY button at E4 (replaces mockup text)
    ws.row_dimensions[4].height = 26
    _form_btn(ws, 4, 5, "📅 СЕГОДНЯ", bg=BLUE, ncols=2, size=11)

    ws.row_dimensions[5].height = 8  # spacer

    # Row 6: section header
    sec_hdr(ws, 6, "  Z-ОТЧЁТ vs ФАКТ", size=11, bg=TEAL_M, ncols=7, height=24)

    # Row 7: column headers (only cols A-D, no padding)
    ws.row_dimensions[7].height = 24
    headers = ["  Способ оплаты", "Z-отчёт", "Факт", "Расхождение"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(7, i, h)
        c.fill = mkfill(TEAL)
        c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center" if i > 1 else "left", "center")
        c.border = mkborder()

    # Rows 8-10: data rows for each payment method
    for i, method in enumerate(PAY_METHODS):
        rn = 8 + i
        ws.row_dimensions[rn].height = 28
        c = ws.cell(rn, 1, f"  {method}")
        c.fill = mkfill(GRAY_L if i % 2 == 0 else WHITE)
        c.font = mkfont(NAVY, 11, True)
        c.alignment = mkalign("left", "center")
        c.border = mkborder()
        _form_input(ws, rn, 2, value=0, fmt='#,##0', halign="right")
        _form_input(ws, rn, 3, value=0, fmt='#,##0', halign="right")
        c = ws.cell(rn, 4, f"=C{rn}-B{rn}")
        c.fill = mkfill(GRAY_L if i % 2 == 0 else WHITE)
        c.font = mkfont(NAVY, 11, True)
        c.alignment = mkalign("right", "center")
        c.border = mkborder()
        c.number_format = '#,##0;[Red]-#,##0'

    # Row 11: ИТОГО
    ws.row_dimensions[11].height = 30
    c = ws.cell(11, 1, "  ИТОГО")
    c.fill = mkfill(TEAL_M)
    c.font = mkfont(WHITE, 12, True)
    c.alignment = mkalign("left", "center")
    c.border = mkborder()
    for col, formula in [(2, "=SUM(B8:B10)"), (3, "=SUM(C8:C10)"), (4, "=SUM(D8:D10)")]:
        c = ws.cell(11, col, formula)
        c.fill = mkfill(TEAL_M)
        c.font = mkfont(WHITE, 12, True)
        c.alignment = mkalign("right", "center")
        c.border = mkborder()
        c.number_format = '#,##0' if col != 4 else '#,##0;[Red]-#,##0'

    # Row 12: spacer
    ws.row_dimensions[12].height = 8

    # Row 13: Выручка за смену (big highlight)
    ws.row_dimensions[13].height = 36
    _form_label(ws, 13, 1, "  Выручка за смену:", ncols=2, bg=GRAY_M, size=11)
    c = ws.cell(13, 3, "=C11")
    c.fill = mkfill(GREEN)
    c.font = mkfont(WHITE, 16, True)
    c.alignment = mkalign("center", "center")
    c.number_format = FMT_RUB
    c.border = mkborder()
    _form_label(ws, 13, 4, "  Комментарий:", bg=GRAY_M, size=11)
    _form_input(ws, 13, 5, value="", ncols=3)

    # Row 14: Выплата из кассы
    ws.row_dimensions[14].height = 30
    _form_label(ws, 14, 1, "  Выплата из кассы:", ncols=3, bg=AMBER, fg=WHITE, size=11)
    _form_input(ws, 14, 4, value=0, ncols=4, fmt=FMT_RUB, halign="right")

    # Rows 15-16: spacer
    ws.row_dimensions[15].height = 12
    ws.row_dimensions[16].height = 12

    # Row 17: SAVE button (large, merged across A:G)
    ws.row_dimensions[17].height = 44
    _form_btn(ws, 17, 1, "💾  СОХРАНИТЬ КАССУ", bg=GREEN, ncols=7, size=14)


def build_vvod_rashody(ws):
    """ВВОД_РАСХОДЫ — Form for daily expenses + Закуп в долг section."""
    ws.sheet_view.showGridLines = False

    set_widths(ws, [("A", 18), ("B", 16), ("C", 16), ("D", 16)])

    sheet_title(ws, "  ВВОД РАСХОДОВ",
                "  Заполните операцию и нажмите СОХРАНИТЬ", ncols=4)

    # Row 3: Дата | TODAY
    ws.row_dimensions[3].height = 30
    _form_label(ws, 3, 1, "  Дата:")
    _form_input(ws, 3, 2, value=None, fmt=FMT_DATE, halign="center")
    _form_btn(ws, 3, 3, "📅 СЕГОДНЯ", bg=BLUE, ncols=2, size=11)

    ws.row_dimensions[4].height = 8  # spacer

    # Row 5: section header — Обычный расход
    sec_hdr(ws, 5, "  РАСХОД", size=11, bg=RED, ncols=4, height=24)

    # Row 6: Категория
    ws.row_dimensions[6].height = 30
    _form_label(ws, 6, 1, "  Категория:")
    _form_input(ws, 6, 2, ncols=3, halign="center")
    _add_dv(ws, "'Настройки'!$C$22:$C$45", "B6", show_error=False)  # autocomplete

    # Row 7: Способ оплаты
    ws.row_dimensions[7].height = 30
    _form_label(ws, 7, 1, "  Способ оплаты:")
    _form_input(ws, 7, 2, ncols=3, halign="center")
    _add_dv(ws, "'Настройки'!$D$22:$D$45", "B7")

    # Row 8: Сумма
    ws.row_dimensions[8].height = 30
    _form_label(ws, 8, 1, "  Сумма:")
    _form_input(ws, 8, 2, value=0, ncols=3, fmt=FMT_RUB, halign="right")

    # Row 9: Комментарий
    ws.row_dimensions[9].height = 30
    _form_label(ws, 9, 1, "  Комментарий:")
    _form_input(ws, 9, 2, value="", ncols=3, halign="left")

    # Row 10: spacer
    ws.row_dimensions[10].height = 12

    # Row 11: section header — Закуп в долг
    sec_hdr(ws, 11, "  ЗАКУП В ДОЛГ  (заполнять ТОЛЬКО если закуп товара в долг)",
            size=11, bg=AMBER, ncols=4, height=24)

    # Row 12: Поставщик
    ws.row_dimensions[12].height = 30
    _form_label(ws, 12, 1, "  Поставщик:")
    _form_input(ws, 12, 2, ncols=3, halign="center")
    _add_dv(ws, "'Настройки'!$G$22:$G$45", "B12", show_error=False)  # autocomplete

    # Row 13: Сумма долга
    ws.row_dimensions[13].height = 30
    _form_label(ws, 13, 1, "  Сумма долга:")
    _form_input(ws, 13, 2, value=0, ncols=3, fmt=FMT_RUB, halign="right")

    # Row 14: spacer
    ws.row_dimensions[14].height = 14

    # Row 15: hint
    ws.merge_cells("A15:D15")
    c = ws.cell(15, 1,
                "  Можно заполнить ТОЛЬКО один раздел: либо РАСХОД, либо ЗАКУП В ДОЛГ")
    c.fill = mkfill(GRAY_L)
    c.font = mkfont(GRAY_D, 9, False)
    c.alignment = mkalign("center", "center")
    ws.row_dimensions[15].height = 18

    # Row 16: SAVE button
    ws.row_dimensions[16].height = 44
    _form_btn(ws, 16, 1, "💾  СОХРАНИТЬ", bg=GREEN, ncols=4, size=14)


# ═══════════════════════════════════════════════════════════════
#  BLOCK 3: ЗАПИСЬ_ВЫПЛАТ + НАСТРОЙКИ
# ═══════════════════════════════════════════════════════════════

ZV_HEADERS = ["№", "Дата плановой оплаты", "Поставщик (ТП)", "Сумма",
              "Статус", "Накладная", "Способ оплаты",
              "Дата фактической оплаты", "Примечание"]
ZV_LAST_ROW = 506  # pre-allocate space


def build_zapis_vyplat(ws, rows):
    """ЗАПИСЬ_ВЫПЛАТ — Supplier payments tracker."""
    ws.sheet_view.showGridLines = False

    set_widths(ws, [("A", 6), ("B", 18), ("C", 18), ("D", 14),
                    ("E", 15), ("F", 14), ("G", 14), ("H", 18), ("I", 22)])

    sheet_title(ws, "  ЗАПИСЬ ВЫПЛАТ",
                "  Платежи поставщикам — план и факт", ncols=9)

    # Rows 3-4: KPI summary (4 tiles)
    tiles = [
        ("Запланировано", GREEN,
         f'=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
         f'tblВыплаты[Дата плановой оплаты],">="&TODAY())'),
        ("Просрочено", RED,
         f'=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
         f'tblВыплаты[Дата плановой оплаты],"<"&TODAY())'),
        ("Оплачено в году", BLUE,
         f'=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Оплачено")'),
        ("Долг общий (БАЗА)", AMBER,
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
         f'-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП")'),
    ]

    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 32
    for i, (lbl, color, formula) in enumerate(tiles):
        c1 = i * 2 + 1
        c2 = c1 + 1
        # Label row 3
        ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)
        c = ws.cell(3, c1, lbl)
        c.fill = mkfill(color)
        c.font = mkfont(WHITE, 9, True)
        c.alignment = mkalign("center", "center")
        # Value row 4
        ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
        c = ws.cell(4, c1, formula)
        c.fill = mkfill(WHITE)
        c.font = mkfont(color, 14, True)
        c.alignment = mkalign("center", "center")
        c.number_format = FMT_RUB
        c.border = Border(
            left=Side(style="medium", color=color),
            right=Side(style="medium", color=color),
            bottom=Side(style="medium", color=color),
        )

    ws.row_dimensions[5].height = 8  # spacer

    # Row 6: Table header
    tbl_hdr(ws, 6, ZV_HEADERS, bg=PURPLE, height=36)
    # Wrap headers
    for col in range(1, 10):
        ws.cell(6, col).alignment = mkalign("center", "center", wrap=True)

    # Data rows from row 7
    PLAN_DATE_COL = 2
    STATUS_COL = 5
    FACT_DATE_COL = 8
    for i, r in enumerate(rows):
        rn = 7 + i
        alt = (i % 2 == 1)
        vals = [
            r["idx"],
            r["plan_date"],
            r["supplier"],
            r["amount"],
            r["status"],
            r["invoice"],
            r["method"],
            r["fact_date"],
            r["comment"],
        ]
        fmts = [None, FMT_DATE, None, FMT_RUB, None, None, None, FMT_DATE, None]
        alns = ["center", "center", "left", "right", "center",
                "center", "center", "center", "left"]
        for col, (v, fmt, aln) in enumerate(zip(vals, fmts, alns), 1):
            c = d_cell(ws, rn, col, v, alt, aln, fmt)
            # Color-code status
            if col == STATUS_COL:
                if v == "Оплачено":
                    c.font = mkfont(GREEN, 10, True)
                elif v == "Запланировано":
                    c.font = mkfont(BLUE, 10, True)
        ws.row_dimensions[rn].height = 18

    # Named table tblВыплаты
    tbl = Table(displayName="tblВыплаты", ref=f"A6:I{ZV_LAST_ROW}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showRowStripes=False, showColumnStripes=False,
    )
    ws.add_table(tbl)

    # Data validations for columns E (Статус) and G (Способ)
    dv_st = DataValidation(type="list", formula1='"Запланировано,Оплачено"',
                            allow_blank=True)
    ws.add_data_validation(dv_st)
    dv_st.add(f"E7:E{ZV_LAST_ROW}")

    dv_m = DataValidation(type="list", formula1="'Настройки'!$D$22:$D$45",
                           allow_blank=True)
    ws.add_data_validation(dv_m)
    dv_m.add(f"G7:G{ZV_LAST_ROW}")

    dv_sup = DataValidation(type="list",
                            formula1="'Настройки'!$G$22:$G$45",
                            allow_blank=True)
    ws.add_data_validation(dv_sup)
    dv_sup.add(f"C7:C{ZV_LAST_ROW}")

    ws.freeze_panes = "A7"


def build_nastroyki(ws):
    """НАСТРОЙКИ — Parameters, lookups, plan-vs-fact table."""
    ws.sheet_view.showGridLines = False

    set_widths(ws, [("A", 4), ("B", 22), ("C", 16), ("D", 16), ("E", 16),
                    ("F", 16), ("G", 16), ("H", 16)])

    sheet_title(ws, "  НАСТРОЙКИ",
                "  Параметры, справочники, план постоянных расходов", ncols=8)

    # ── Р1: Параметры (rows 4-9) ──
    sec_hdr(ws, 4, "  Р1 · ПАРАМЕТРЫ", bg=TEAL, ncols=8, height=24)
    params = [
        ("Магазин:", SHOP),
        ("Год для дашборда:", YEAR),
        ("Год для Р8 (план vs факт):", YEAR),
        ("Начальный остаток кассы:", 0),
        ("Валюта:", "RUB"),
    ]
    for i, (lbl, val) in enumerate(params):
        rn = 5 + i
        ws.row_dimensions[rn].height = 22
        _form_label(ws, rn, 2, "  " + lbl)
        c = _form_input(ws, rn, 3, value=val, ncols=2, halign="left")
        if isinstance(val, (int, float)) and lbl.startswith("Начал"):
            c.number_format = FMT_RUB
    # Спейсер
    ws.row_dimensions[10].height = 10

    # ── Р4: Кассы / даты (row 11-13) ──
    sec_hdr(ws, 11, "  Р4 · ПЕРИОД РАБОТЫ", bg=TEAL, ncols=8, height=24)
    ws.row_dimensions[12].height = 22
    _form_label(ws, 12, 2, "  Дата открытия магазина:")
    _form_input(ws, 12, 3, value=date(YEAR, 1, 1), ncols=2,
                fmt=FMT_DATE, halign="center")
    ws.row_dimensions[13].height = 10

    # ── Р6: Пороги (row 14-19) ──
    sec_hdr(ws, 14, "  Р6 · ПОРОГИ (для сигналов)", bg=TEAL, ncols=8, height=24)
    thresholds = [
        ("Долг — внимание (₽):", 200000),
        ("Долг — критично (₽):", 500000),
        ("Дней до оплаты — внимание:", 7),
        ("Расхождение кассы — критично (₽):", 1000),
    ]
    for i, (lbl, val) in enumerate(thresholds):
        rn = 15 + i
        ws.row_dimensions[rn].height = 22
        _form_label(ws, rn, 2, "  " + lbl)
        c = _form_input(ws, rn, 3, value=val, ncols=2,
                        fmt=FMT_RUB if "₽" in lbl else "0",
                        halign="right")
    ws.row_dimensions[19].height = 10

    # ── Р7: Справочники (rows 20-45) ──
    sec_hdr(ws, 20, "  Р7 · СПРАВОЧНИКИ", bg=TEAL, ncols=8, height=24)

    # Sub-headers row 21
    sub_headers = ["", "Кассиры", "Категории расход", "Способы оплаты",
                   "Типы операций", "Месяцы", "Поставщики", "Смены"]
    ws.row_dimensions[21].height = 24
    for i, h in enumerate(sub_headers, 1):
        c = ws.cell(21, i, h)
        c.fill = mkfill(TEAL_M)
        c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center", "center")
        c.border = mkborder()

    # Lookup data — 24 rows (22-45)
    columns_data = {
        2: list(CASHIERS),
        3: list(CATS_EXPENSE),
        4: PAY_METHODS,
        5: TYPES_ALL,
        6: MONTHS_RU,
        7: SUPPLIERS,
        8: list(SHIFTS),
    }
    for rn in range(22, 46):
        ws.row_dimensions[rn].height = 18
        idx = rn - 22
        for col, lst in columns_data.items():
            v = lst[idx] if idx < len(lst) else None
            d_cell(ws, rn, col, v,
                   alt=(idx % 2 == 1), halign="left")
        # Col 1 padding only (col 8 now has Смены data)
        c = ws.cell(rn, 1)
        c.fill = mkfill(GRAY_L if idx % 2 == 1 else WHITE)
        c.border = mkborder()

    ws.row_dimensions[46].height = 10

    # ── Р8: План-Факт постоянных расходов (rows 47-62) ──
    sec_hdr(ws, 47, "  Р8 · ПОСТОЯННЫЕ РАСХОДЫ — ПЛАН vs ФАКТ", bg=TEAL,
            ncols=8, height=24)

    # Header row 48
    ws.row_dimensions[48].height = 32
    p8_headers = ["", "Месяц", "ЗП план", "Аренда план", "Налоги план",
                  "Интернет план", "Коммуналка план", "ИТОГО план"]
    for i, h in enumerate(p8_headers, 1):
        c = ws.cell(48, i, h)
        c.fill = mkfill(TEAL_M)
        c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center", "center", wrap=True)
        c.border = mkborder()

    plan_defaults = {
        "ЗП": 145000, "Аренда": 80000, "Налоги": 0,
        "Интернет": 3500, "Коммуналка": 18000,
    }
    quarterly_tax = 45000
    for i, mname in enumerate(MONTHS_RU):
        rn = 49 + i
        m = i + 1
        ws.row_dimensions[rn].height = 20
        alt = (i % 2 == 1)
        d_cell(ws, rn, 1, m, alt, "center")
        d_cell(ws, rn, 2, mname, alt, "left")
        d_cell(ws, rn, 3, plan_defaults["ЗП"], alt, "right", FMT_RUB)
        d_cell(ws, rn, 4, plan_defaults["Аренда"], alt, "right", FMT_RUB)
        d_cell(ws, rn, 5, quarterly_tax if m in (1, 4, 7, 10) else 0,
               alt, "right", FMT_RUB)
        d_cell(ws, rn, 6, plan_defaults["Интернет"], alt, "right", FMT_RUB)
        d_cell(ws, rn, 7, plan_defaults["Коммуналка"], alt, "right", FMT_RUB)
        # ИТОГО plan
        c = d_cell(ws, rn, 8, f"=SUM(C{rn}:G{rn})", alt, "right", FMT_RUB)
        c.font = mkfont(NAVY, 10, True)

    # Fact sub-header (row 61)
    ws.row_dimensions[61].height = 12
    sec_hdr(ws, 62, "  Р8 · ФАКТ (из БАЗА_ДДС за год $C$7)", bg=TEAL_M,
            ncols=8, height=22)
    ws.row_dimensions[63].height = 32
    p8_fact_headers = ["", "Месяц", "ЗП факт", "Аренда факт", "Налоги факт",
                       "Интернет факт", "Коммуналка факт", "ИТОГО факт"]
    for i, h in enumerate(p8_fact_headers, 1):
        c = ws.cell(63, i, h)
        c.fill = mkfill(GRAY_M)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("center", "center", wrap=True)
        c.border = mkborder()

    cats_fact = ["ЗП", "Аренда", "Налоги", "Интернет", "Коммуналка"]
    for i, mname in enumerate(MONTHS_RU):
        rn = 64 + i
        m = i + 1
        ws.row_dimensions[rn].height = 20
        alt = (i % 2 == 1)
        d_cell(ws, rn, 1, m, alt, "center")
        d_cell(ws, rn, 2, mname, alt, "left")
        for ci, cat in enumerate(cats_fact):
            col = 3 + ci
            formula = (
                f'=SUMPRODUCT('
                f'(YEAR(tblБаза[Дата])=$C$7)*'
                f'(MONTH(tblБаза[Дата])={m})*'
                f'(tblБаза[Тип]="Расход")*'
                f'(tblБаза[Категория]="{cat}")*'
                f'tblБаза[Сумма])'
            )
            c = d_cell(ws, rn, col, formula, alt, "right", FMT_RUB)
        # ИТОГО факт
        c = d_cell(ws, rn, 8, f"=SUM(C{rn}:G{rn})", alt, "right", FMT_RUB)
        c.font = mkfont(NAVY, 10, True)

    ws.row_dimensions[76].height = 10

    # ── Р9: Поставщики (rows 77+) ──
    sec_hdr(ws, 77, "  Р9 · ПОСТАВЩИКИ (ТП)", bg=TEAL, ncols=8, height=24)
    ws.row_dimensions[78].height = 24
    p9_hdr = ["", "№", "Название", "Контакт", "Условия оплаты", "", "", ""]
    for i, h in enumerate(p9_hdr, 1):
        c = ws.cell(78, i, h)
        c.fill = mkfill(TEAL_M)
        c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center", "center")
        c.border = mkborder()

    for i, sup in enumerate(SUPPLIERS):
        rn = 79 + i
        ws.row_dimensions[rn].height = 20
        alt = (i % 2 == 1)
        d_cell(ws, rn, 1, "", alt)
        d_cell(ws, rn, 2, i + 1, alt, "center")
        d_cell(ws, rn, 3, sup, alt, "left")
        d_cell(ws, rn, 4, f"+7 (495) {random.randint(100, 999)}-{random.randint(10, 99)}-{random.randint(10, 99)}",
               alt, "left")
        d_cell(ws, rn, 5, random.choice(["30 дней", "14 дней", "Предоплата", "По факту"]),
               alt, "left")
        for col in (6, 7, 8):
            d_cell(ws, rn, col, "", alt)

    # ── Автокомплит: вспомогательные колонки I, J, K (скрытые) ──
    # VBA записывает сюда отфильтрованные списки и указывает на них DV
    ac_lists = {
        9:  list(CASHIERS),      # I — Кассиры
        10: list(CATS_EXPENSE),  # J — Категории расходов
        11: list(SUPPLIERS),     # K — Поставщики
    }
    for col, items in ac_lists.items():
        for row_i, item in enumerate(items):
            ws.cell(row=1 + row_i, column=col, value=item)
        # Pad remaining rows with empty so VBA can safely clear a fixed range
        for row_i in range(len(items), 15):
            ws.cell(row=1 + row_i, column=col, value=None)
    # Hide these helper columns from view
    for col_letter in ("I", "J", "K"):
        ws.column_dimensions[col_letter].hidden = True


# ═══════════════════════════════════════════════════════════════
#  BLOCK 4a: ПУЛЬТ
# ═══════════════════════════════════════════════════════════════

def _kpi_tile(ws, row, col, ncols, label, formula, color,
              fmt=FMT_RUB, lbl_size=9, val_size=14):
    """Render a 2-row KPI tile (label + value)."""
    # Label
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + ncols - 1)
    c = ws.cell(row, col, label)
    c.fill = mkfill(color)
    c.font = mkfont(WHITE, lbl_size, True)
    c.alignment = mkalign("center", "center")

    # Value
    ws.merge_cells(start_row=row + 1, start_column=col,
                   end_row=row + 1, end_column=col + ncols - 1)
    c = ws.cell(row + 1, col, formula)
    c.fill = mkfill(WHITE)
    c.font = mkfont(color, val_size, True)
    c.alignment = mkalign("center", "center")
    c.number_format = fmt
    c.border = Border(
        left=Side(style="medium", color=color),
        right=Side(style="medium", color=color),
        bottom=Side(style="medium", color=color),
    )


def build_pult(ws):
    """ПУЛЬТ — Quick overview with KPI tiles."""
    ws.sheet_view.showGridLines = False
    set_widths(ws, [("A", 12), ("B", 12), ("C", 12), ("D", 12),
                    ("E", 12), ("F", 12)])

    sheet_title(ws, "  ПУЛЬТ", "  Быстрый обзор работы магазина", ncols=6)

    # Row 3: period selectors
    ws.row_dimensions[3].height = 30
    _form_label(ws, 3, 1, "  Год:")
    _form_input(ws, 3, 2, value=YEAR, halign="center")
    _form_label(ws, 3, 3, "  Период:")
    _form_input(ws, 3, 4, value="Весь год", ncols=2, halign="center")
    _add_dv(ws, '"' + ",".join(["Весь год"] + MONTHS_RU) + '"', "D3")

    # Hidden helper cells row 4 (period start/end derived from B3, D3)
    # B3 = year, D3 = month name or "Весь год"
    # Hidden row 4 with helpers
    ws.row_dimensions[4].height = 1
    ws.cell(4, 1, '=IF(D3="Весь год",DATE(B3,1,1),DATE(B3,MATCH(D3,'
                  + '{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";'
                  + '"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0),1))')
    ws.cell(4, 2, '=IF(D3="Весь год",DATE(B3,12,31),EOMONTH(A4,0))')
    # Hide row visually
    for col in range(1, 7):
        c = ws.cell(4, col)
        c.font = mkfont(WHITE, 1)
        c.number_format = FMT_DATE

    ws.row_dimensions[5].height = 8

    # ── СЕГОДНЯ section (rows 6-9) ──
    sec_hdr(ws, 6, "  СЕГОДНЯ", bg=NAVY, ncols=6, height=22)
    today_tiles = [
        ("Выручка",
         '=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",tblБаза[Дата],TODAY())',
         GREEN),
        ("Расход",
         '=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",tblБаза[Дата],TODAY())',
         RED),
        ("Прибыль",
         '=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",tblБаза[Дата],TODAY())'
         '-SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",tblБаза[Дата],TODAY())',
         BLUE),
        ("Расхождение",
         '=SUMIFS(tblБаза[Расхождение],tblБаза[Тип],"Приход",tblБаза[Дата],TODAY())',
         AMBER),
    ]
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 32
    # 4 tiles in cols A-D, E-F left blank
    for i, (lbl, f, color) in enumerate(today_tiles):
        _kpi_tile(ws, 7, i + 1, 1, lbl, f, color, val_size=10)

    ws.row_dimensions[9].height = 10

    # ── ЗА ПЕРИОД (rows 10-13) ──
    sec_hdr(ws, 10, "  ЗА ПЕРИОД", bg=TEAL, ncols=6, height=22)
    P = '$A$4'  # period start
    Q = '$B$4'  # period end
    # Tiles are at row 11 (label) / row 12 (value), cols 1..6
    period_tiles = [
        ("Выручка",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         GREEN),
        ("Расход",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         RED),
        ("Прибыль", '=A12-B12', BLUE),
        ("Смен закрыто",
         f'=ROUND(COUNTIFS(tblБаза[Тип],"Приход",tblБаза[Дата],">="&{P},'
         f'tblБаза[Дата],"<="&{Q})/3,0)', PURPLE),
        ("Сред. выручка",
         '=IFERROR(A12/D12,0)', TEAL_M),
        ("Долг новый",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         AMBER),
    ]
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 32
    for i, (lbl, f, color) in enumerate(period_tiles):
        fmt = "0" if lbl == "Смен закрыто" else FMT_RUB
        _kpi_tile(ws, 11, i + 1, 1, lbl, f, color, fmt=fmt, val_size=10)

    ws.row_dimensions[13].height = 10

    # ── СИГНАЛЫ ТП (rows 14-17) ──
    sec_hdr(ws, 14, "  СИГНАЛЫ ПОСТАВЩИКАМ", bg=AMBER, ncols=6, height=22)
    tp_tiles = [
        ("Запланировано",
         f'=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
         f'tblВыплаты[Дата плановой оплаты],">="&TODAY(),'
         f'tblВыплаты[Дата плановой оплаты],"<="&{Q})', GREEN),
        ("Просрочено",
         f'=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
         f'tblВыплаты[Дата плановой оплаты],"<"&TODAY())', RED),
        ("Оплачено в периоде",
         f'=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Оплачено",'
         f'tblВыплаты[Дата фактической оплаты],">="&{P},'
         f'tblВыплаты[Дата фактической оплаты],"<="&{Q})', BLUE),
        ("Долг общий (сейчас)",
         '=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
         '-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП")', AMBER),
    ]
    ws.row_dimensions[15].height = 18
    ws.row_dimensions[16].height = 32
    for i, (lbl, f, color) in enumerate(tp_tiles):
        _kpi_tile(ws, 15, i + 1, 1, lbl, f, color, val_size=10)

    ws.row_dimensions[17].height = 10

    # Footer note
    ws.merge_cells("A18:F18")
    c = ws.cell(18, 1, "  Подсказка: при первом открытии нажмите F9 для пересчёта")
    c.font = mkfont(GRAY_D, 9)
    c.alignment = mkalign("left", "center")


# ═══════════════════════════════════════════════════════════════
#  BLOCK 4b: КАЛЕНДАРЬ_ВЫПЛАТ
# ═══════════════════════════════════════════════════════════════

def build_calendar(ws):
    """КАЛЕНДАРЬ_ВЫПЛАТ — Monthly calendar grid 7×6 with payment summaries."""
    ws.sheet_view.showGridLines = False

    # 9 cols total: A-G = calendar, H = spacer, I-J = side panel
    set_widths(ws, [("A", 13), ("B", 13), ("C", 13), ("D", 13),
                    ("E", 13), ("F", 13), ("G", 13), ("H", 2),
                    ("I", 18), ("J", 14)])

    sheet_title(ws, "  КАЛЕНДАРЬ ВЫПЛАТ",
                "  Платежи поставщикам по дням месяца", ncols=10)

    # Row 3: Month / Year selectors
    ws.row_dimensions[3].height = 32
    _form_label(ws, 3, 1, "  Месяц:")
    _form_input(ws, 3, 2, value="Январь", halign="center")
    _add_dv(ws, '"' + ",".join(MONTHS_RU) + '"', "B3")
    _form_label(ws, 3, 3, "  Год:")
    _form_input(ws, 3, 4, value=YEAR, halign="center")
    # Today button hint
    _form_btn(ws, 3, 9, "📅 СЕГОДНЯ", bg=BLUE, ncols=2, size=11)

    # Helper cells row 4 (hidden small)
    ws.row_dimensions[4].height = 6
    # I4 = selected date (set by VBA click handler); default = first of month
    months_arr = '{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"}'
    ws.cell(4, 1, f'=DATE(D3,MATCH(B3,{months_arr},0),1)')
    ws.cell(4, 1).number_format = FMT_DATE
    ws.cell(4, 2, '=EOMONTH(A4,0)')
    ws.cell(4, 2).number_format = FMT_DATE
    ws.cell(4, 3, '=WEEKDAY(A4,2)')  # 1=Mon..7=Sun
    for col in range(1, 11):
        ws.cell(4, col).font = mkfont(WHITE, 1)

    ws.row_dimensions[5].height = 8

    # Row 6: weekday header
    weekdays = ["Понедельник", "Вторник", "Среда", "Четверг",
                "Пятница", "Суббота", "Воскресенье"]
    ws.row_dimensions[6].height = 26
    for i, wd in enumerate(weekdays, 1):
        c = ws.cell(6, i, wd)
        c.fill = mkfill(TEAL)
        c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center", "center")
        c.border = mkborder()

    # 6 weeks × 5 rows per day = rows 7..36
    # Each "week" occupies 5 rows: date, оплачено, план, просроч, итого
    sub_rows = ["date", "paid", "plan", "over", "total"]
    sub_labels = {"date": "", "paid": "✓ Оплачено",
                  "plan": "🟡 Планир.", "over": "🔴 Просроч.", "total": "Σ Итого"}
    sub_colors = {"date": NAVY, "paid": GREEN, "plan": AMBER,
                  "over": RED, "total": NAVY}

    start_row = 7
    for week in range(6):
        week_top = start_row + week * 5
        # heights
        ws.row_dimensions[week_top].height = 18      # date number
        for i in range(1, 5):
            ws.row_dimensions[week_top + i].height = 16

        for day_col in range(1, 8):
            # day position 0-based in 6×7 grid
            day_pos = week * 7 + (day_col - 1)
            # Day number formula
            # day_n = day_pos - (WEEKDAY-1) + 1 = day_pos - C4 + 2
            day_formula = (
                f'=IF(AND({day_pos}-$C$4+2>=1,'
                f'{day_pos}-$C$4+2<=DAY($B$4)),'
                f'{day_pos}-$C$4+2,"")'
            )
            c = ws.cell(week_top, day_col, day_formula)
            c.fill = mkfill(GRAY_L)
            c.font = mkfont(NAVY, 12, True)
            c.alignment = mkalign("center", "center")
            c.border = Border(
                left=Side(style="thin", color="FFD1D5DB"),
                right=Side(style="thin", color="FFD1D5DB"),
                top=Side(style="medium", color="FF9CA3AF"),
            )

            # Date object formula (for SUMIFS)
            # row "date" in helper: use IFERROR to handle empty cells
            # Build the actual date from the day number: DATE(year, month, day)
            date_ref_formula = (
                f'=IFERROR(DATE($D$3,MONTH($A$4),'
                f'{get_column_letter(day_col)}{week_top}),"")'
            )

            # Оплачено row
            r_paid = week_top + 1
            f_paid = (
                f'=IFERROR(IF({get_column_letter(day_col)}{week_top}="","",'
                f'SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Оплачено",'
                f'tblВыплаты[Дата фактической оплаты],'
                f'DATE($D$3,MONTH($A$4),{get_column_letter(day_col)}{week_top}))),"")'
            )
            c = ws.cell(r_paid, day_col, f_paid)
            c.fill = mkfill(WHITE)
            c.font = mkfont(GREEN, 9, True)
            c.alignment = mkalign("right", "center")
            c.border = Border(
                left=Side(style="thin", color="FFD1D5DB"),
                right=Side(style="thin", color="FFD1D5DB"),
            )
            c.number_format = '[=0]"";#,##0'

            # План row — ALL "Запланировано" on their plan_date (no TODAY filter)
            r_plan = week_top + 2
            f_plan = (
                f'=IFERROR(IF({get_column_letter(day_col)}{week_top}="","",'
                f'SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
                f'tblВыплаты[Дата плановой оплаты],'
                f'DATE($D$3,MONTH($A$4),{get_column_letter(day_col)}{week_top}))),"")'
            )
            c = ws.cell(r_plan, day_col, f_plan)
            c.fill = mkfill(WHITE)
            c.font = mkfont(AMBER, 9, True)
            c.alignment = mkalign("right", "center")
            c.border = Border(
                left=Side(style="thin", color="FFD1D5DB"),
                right=Side(style="thin", color="FFD1D5DB"),
            )
            c.number_format = '[=0]"";#,##0'

            # Просрочено row — planned AND past-due (plan_date < TODAY); visual only,
            # NOT added to total (it is a subset of the Plan row above)
            r_over = week_top + 3
            f_over = (
                f'=IFERROR(IF({get_column_letter(day_col)}{week_top}="","",'
                f'SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
                f'tblВыплаты[Дата плановой оплаты],'
                f'DATE($D$3,MONTH($A$4),{get_column_letter(day_col)}{week_top}),'
                f'tblВыплаты[Дата плановой оплаты],"<"&TODAY())),"")'
            )
            c = ws.cell(r_over, day_col, f_over)
            c.fill = mkfill(WHITE)
            c.font = mkfont(RED, 9, True)
            c.alignment = mkalign("right", "center")
            c.border = Border(
                left=Side(style="thin", color="FFD1D5DB"),
                right=Side(style="thin", color="FFD1D5DB"),
            )
            c.number_format = '[=0]"";#,##0'

            # ИТОГО row — paid + plan (over is already included in plan, don't double-count)
            r_total = week_top + 4
            cl = get_column_letter(day_col)
            f_total = f'=IFERROR(IF({cl}{week_top}="","",{cl}{r_paid}+{cl}{r_plan}),"")'
            c = ws.cell(r_total, day_col, f_total)
            c.fill = mkfill(GRAY_M)
            c.font = mkfont(NAVY, 10, True)
            c.alignment = mkalign("right", "center")
            c.border = Border(
                left=Side(style="thin", color="FFD1D5DB"),
                right=Side(style="thin", color="FFD1D5DB"),
                bottom=Side(style="medium", color="FF9CA3AF"),
            )
            c.number_format = '[=0]"—";#,##0'

    # Side panel (col I-J), rows 6-12: monthly summary
    ws.cell(6, 9, "СВОДКА ЗА МЕСЯЦ").fill = mkfill(TEAL)
    ws.cell(6, 9).font = mkfont(WHITE, 10, True)
    ws.cell(6, 9).alignment = mkalign("center", "center")
    ws.merge_cells("I6:J6")

    side_kpis = [
        ("Оплачено",
         '=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Оплачено",'
         'tblВыплаты[Дата фактической оплаты],">="&$A$4,'
         'tblВыплаты[Дата фактической оплаты],"<="&$B$4)', GREEN),
        ("Запланировано",
         '=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
         'tblВыплаты[Дата плановой оплаты],">="&$A$4,'
         'tblВыплаты[Дата плановой оплаты],"<="&$B$4)', AMBER),
        ("Просрочено",
         '=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
         'tblВыплаты[Дата плановой оплаты],">="&$A$4,'
         'tblВыплаты[Дата плановой оплаты],"<="&$B$4,'
         'tblВыплаты[Дата плановой оплаты],"<"&TODAY())', RED),
        ("Платежей",
         '=COUNTIFS(tblВыплаты[Дата плановой оплаты],">="&$A$4,'
         'tblВыплаты[Дата плановой оплаты],"<="&$B$4)', BLUE),
    ]
    for i, (lbl, formula, color) in enumerate(side_kpis):
        rn = 7 + i
        ws.row_dimensions[rn].height = 22
        c = ws.cell(rn, 9, lbl)
        c.fill = mkfill(GRAY_L)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("left", "center")
        c.border = mkborder()

        c = ws.cell(rn, 10, formula)
        c.fill = mkfill(WHITE)
        c.font = mkfont(color, 11, True)
        c.alignment = mkalign("right", "center")
        c.border = mkborder()
        c.number_format = "#,##0" if lbl != "Платежей" else "0"


# ═══════════════════════════════════════════════════════════════
#  BLOCK 4c: ДАШБОРД
# ═══════════════════════════════════════════════════════════════

def build_dashboard(ws):
    """ДАШБОРД — Full analytics dashboard."""
    ws.sheet_view.showGridLines = False
    set_widths(ws, [(get_column_letter(i), 12) for i in range(1, 13)])

    sheet_title(ws, "  ДАШБОРД", "  Полная аналитика магазина", ncols=12)

    # Row 3: Period selector
    ws.row_dimensions[3].height = 32
    _form_label(ws, 3, 1, "  Год от:")
    _form_input(ws, 3, 2, value=YEAR, halign="center")
    _form_label(ws, 3, 3, "  Месяц от:")
    _form_input(ws, 3, 4, value="Январь", halign="center")
    _add_dv(ws, '"' + ",".join(MONTHS_RU) + '"', "D3")
    _form_label(ws, 3, 5, "  Год до:")
    _form_input(ws, 3, 6, value=YEAR, halign="center")
    _form_label(ws, 3, 7, "  Месяц до:")
    _form_input(ws, 3, 8, value="Декабрь", halign="center")
    _add_dv(ws, '"' + ",".join(MONTHS_RU) + '"', "H3")
    _form_btn(ws, 3, 11, "🔄 ОБНОВИТЬ", bg=AMBER, ncols=2, size=11)

    # Hidden helpers row 4
    months_arr = '{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"}'
    ws.cell(4, 1, f'=DATE(B3,MATCH(D3,{months_arr},0),1)')  # period start
    ws.cell(4, 2, f'=EOMONTH(DATE(F3,MATCH(H3,{months_arr},0),1),0)')  # period end
    ws.cell(4, 1).number_format = FMT_DATE
    ws.cell(4, 2).number_format = FMT_DATE
    for col in range(1, 13):
        ws.cell(4, col).font = mkfont(WHITE, 1)
    ws.row_dimensions[4].height = 6

    P = '$A$4'
    Q = '$B$4'
    ws.row_dimensions[5].height = 8

    # ── BLOCK 1: ВЫРУЧКА (rows 6-8) ──
    sec_hdr(ws, 6, "  ВЫРУЧКА", bg=GREEN, ncols=12, height=22)
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 32
    rev_tiles = [
        ("Выручка всего",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
        ("Наличные",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
         f'tblБаза[Способ оплаты],"Наличные",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
        ("Карта",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
         f'tblБаза[Способ оплаты],"Карта",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
        ("Перевод",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
         f'tblБаза[Способ оплаты],"Перевод",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
    ]
    for i, (lbl, f) in enumerate(rev_tiles):
        _kpi_tile(ws, 7, i * 3 + 1, 3, lbl, f, GREEN, val_size=12)

    # ── BLOCK 2: РАСХОДЫ (rows 9-11) ──
    sec_hdr(ws, 9, "  РАСХОДЫ", bg=RED, ncols=12, height=22)
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 32
    exp_tiles = [
        ("Расход всего",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
        ("Закуп товара",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
         f'tblБаза[Категория],"Закуп товара",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
        ("ФОТ + Аренда",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
         f'tblБаза[Категория],"ЗП",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'
         f'+SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
         f'tblБаза[Категория],"Аренда",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
        ("Прочее",
         f'=A11-(D11+G11)'),
    ]
    # Note: tile values land in row 8, but row 11 is value row. Need offsets.
    for i, (lbl, f) in enumerate(exp_tiles):
        col = i * 3 + 1
        _kpi_tile(ws, 10, col, 3, lbl, f, RED, val_size=12)

    # ── BLOCK 3: ПРИБЫЛЬ + ДОЛГ (rows 12-14) ──
    sec_hdr(ws, 12, "  ПРИБЫЛЬ И ДОЛГИ", bg=BLUE, ncols=12, height=22)
    ws.row_dimensions[13].height = 18
    ws.row_dimensions[14].height = 32
    pl_tiles = [
        ("Прибыль (касса)",
         f'=A8-A11'),
        ("Рентабельность",
         f'=IFERROR((A8-A11)/A8,0)', "0.0%"),
        ("Долг — взято",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
        ("Долг общий (текущ.)",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
         f'-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП")'),
    ]
    for i, t in enumerate(pl_tiles):
        col = i * 3 + 1
        if len(t) == 3:
            lbl, f, fmt = t
        else:
            lbl, f = t
            fmt = FMT_RUB
        _kpi_tile(ws, 13, col, 3, lbl, f, BLUE, fmt=fmt, val_size=12)

    # ── BLOCK 4: ОПЕРАЦИИ (rows 15-17) ──
    sec_hdr(ws, 15, "  ОПЕРАЦИИ И КАССА", bg=PURPLE, ncols=12, height=22)
    ws.row_dimensions[16].height = 18
    ws.row_dimensions[17].height = 32
    op_tiles = [
        ("Смен закрыто",
         f'=ROUND(COUNTIFS(tblБаза[Тип],"Приход",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})/3,0)', "0"),
        ("Сред. выручка",
         f'=IFERROR(A8/A17,0)'),
        ("Расхождение касс",
         f'=SUMIFS(tblБаза[Расхождение],tblБаза[Тип],"Приход",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'),
        ("Транзакций",
         f'=COUNTIFS(tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})', "0"),
    ]
    for i, t in enumerate(op_tiles):
        col = i * 3 + 1
        if len(t) == 3:
            lbl, f, fmt = t
        else:
            lbl, f = t
            fmt = FMT_RUB
        _kpi_tile(ws, 16, col, 3, lbl, f, PURPLE, fmt=fmt, val_size=12)

    ws.row_dimensions[18].height = 10

    # ── ДЕТАЛИЗАЦИЯ РАСХОДОВ (rows 19-30) ──
    sec_hdr(ws, 19, "  ДЕТАЛИЗАЦИЯ РАСХОДОВ ПО КАТЕГОРИЯМ", bg=NAVY,
            ncols=12, height=22)
    ws.row_dimensions[20].height = 22
    detail_hdr = ["", "Категория", "", "", "Сумма", "", "", "Доля %", "", "", "", ""]
    for i, h in enumerate(detail_hdr, 1):
        c = ws.cell(20, i, h)
        c.fill = mkfill(TEAL_M)
        c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center", "center")
        c.border = mkborder()

    # Use CATS_EXPENSE in detail
    for i, cat in enumerate(CATS_EXPENSE):
        rn = 21 + i
        ws.row_dimensions[rn].height = 20
        alt = (i % 2 == 1)
        d_cell(ws, rn, 1, "", alt)
        c = ws.cell(rn, 2, cat)
        ws.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=4)
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("left", "center")
        c.border = mkborder()
        for col in (3, 4):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

        # Sum
        formula = (
            f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
            f'tblБаза[Категория],"{cat}",'
            f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'
        )
        ws.merge_cells(start_row=rn, start_column=5, end_row=rn, end_column=7)
        c = ws.cell(rn, 5, formula)
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("right", "center")
        c.border = mkborder()
        c.number_format = FMT_RUB
        for col in (6, 7):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

        # Доля
        ws.merge_cells(start_row=rn, start_column=8, end_row=rn, end_column=12)
        c = ws.cell(rn, 8, f'=IFERROR(E{rn}/$A$11,0)')
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10)
        c.alignment = mkalign("right", "center")
        c.border = mkborder()
        c.number_format = "0.0%"
        for col in range(9, 13):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

    next_row = 21 + len(CATS_EXPENSE) + 1
    ws.row_dimensions[next_row].height = 10
    next_row += 1

    # ── ВЫРУЧКА ПО ДНЯМ НЕДЕЛИ ──
    sec_hdr(ws, next_row, "  ВЫРУЧКА ПО ДНЯМ НЕДЕЛИ", bg=TEAL, ncols=12, height=22)
    next_row += 1
    weekdays = ["Понедельник", "Вторник", "Среда", "Четверг",
                "Пятница", "Суббота", "Воскресенье"]
    ws.row_dimensions[next_row].height = 22
    hdrs_wd = ["", "День недели", "", "", "Выручка", "", "", "Доля %", "", "", "", ""]
    for i, h in enumerate(hdrs_wd, 1):
        c = ws.cell(next_row, i, h)
        c.fill = mkfill(TEAL_M); c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center", "center"); c.border = mkborder()
    base_wd_row = next_row + 1
    for i, wd in enumerate(weekdays):
        rn = base_wd_row + i
        ws.row_dimensions[rn].height = 20
        alt = (i % 2 == 1)
        d_cell(ws, rn, 1, "", alt)
        ws.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=4)
        c = ws.cell(rn, 2, wd)
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("left", "center")
        c.border = mkborder()
        for col in (3, 4):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

        ws.merge_cells(start_row=rn, start_column=5, end_row=rn, end_column=7)
        formula = (
            f'=SUMPRODUCT((WEEKDAY(tblБаза[Дата],2)={i+1})*'
            f'(tblБаза[Тип]="Приход")*'
            f'(tblБаза[Дата]>={P})*(tblБаза[Дата]<={Q})*tblБаза[Сумма])'
        )
        c = ws.cell(rn, 5, formula)
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("right", "center")
        c.border = mkborder(); c.number_format = FMT_RUB
        for col in (6, 7):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

        ws.merge_cells(start_row=rn, start_column=8, end_row=rn, end_column=12)
        c = ws.cell(rn, 8, f'=IFERROR(E{rn}/$A$8,0)')
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10)
        c.alignment = mkalign("right", "center")
        c.border = mkborder(); c.number_format = "0.0%"
        for col in range(9, 13):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

    next_row = base_wd_row + 7 + 1
    ws.row_dimensions[next_row].height = 10
    next_row += 1

    # ── ВЫРУЧКА ПО СМЕНАМ ──
    sec_hdr(ws, next_row, "  ВЫРУЧКА ПО СМЕНАМ", bg=TEAL, ncols=12, height=22)
    next_row += 1
    ws.row_dimensions[next_row].height = 22
    for i, h in enumerate(["", "Смена", "", "", "Выручка", "", "", "Доля %", "", "", "", ""], 1):
        c = ws.cell(next_row, i, h)
        c.fill = mkfill(TEAL_M); c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center", "center"); c.border = mkborder()
    base_sh_row = next_row + 1
    for i, sh in enumerate(SHIFTS):
        rn = base_sh_row + i
        ws.row_dimensions[rn].height = 20
        alt = (i % 2 == 1)
        d_cell(ws, rn, 1, "", alt)
        ws.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=4)
        c = ws.cell(rn, 2, sh)
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("left", "center"); c.border = mkborder()
        for col in (3, 4):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

        ws.merge_cells(start_row=rn, start_column=5, end_row=rn, end_column=7)
        formula = (
            f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
            f'tblБаза[Смена],"{sh}",'
            f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'
        )
        c = ws.cell(rn, 5, formula)
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("right", "center")
        c.border = mkborder(); c.number_format = FMT_RUB
        for col in (6, 7):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

        ws.merge_cells(start_row=rn, start_column=8, end_row=rn, end_column=12)
        c = ws.cell(rn, 8, f'=IFERROR(E{rn}/$A$8,0)')
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10)
        c.alignment = mkalign("right", "center")
        c.border = mkborder(); c.number_format = "0.0%"
        for col in range(9, 13):
            ws.cell(rn, col).fill = mkfill(GRAY_L if alt else WHITE)
            ws.cell(rn, col).border = mkborder()

    # Footer hint
    footer_row = base_sh_row + len(SHIFTS) + 2
    ws.merge_cells(start_row=footer_row, start_column=1,
                   end_row=footer_row, end_column=12)
    c = ws.cell(footer_row, 1, "  Подсказка: после изменения периода нажмите F9 для пересчёта")
    c.font = mkfont(GRAY_D, 9)
    c.alignment = mkalign("left", "center")

    # ── РАСШИРЕННАЯ АНАЛИТИКА (добавлено) ──────────────────────
    pp_r = footer_row + 2
    ws.row_dimensions[footer_row + 1].height = 16
    ws.row_dimensions[pp_r].height = 0          # hidden helper row
    _pc1 = ws.cell(pp_r, 1, '=DATE(YEAR($A$4),MONTH($A$4)-1,1)')
    _pc1.number_format = FMT_DATE
    _pc2 = ws.cell(pp_r, 2, '=$A$4-1')
    _pc2.number_format = FMT_DATE
    for _ci in range(1, 13):
        ws.cell(pp_r, _ci).font = mkfont(WHITE, 1)

    P  = '$A$4'
    Q  = '$B$4'
    PP = f'$A${pp_r}'
    QP = f'$B${pp_r}'

    def _rev(p, q):
        return (f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
                f'tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')

    def _exp(p, q, cat=''):
        if cat:
            return (f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
                    f'tblБаза[Категория],"{cat}",'
                    f'tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')
        return (f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
                f'tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')

    def _cnt(p, q, typ=''):
        if typ:
            return (f'=COUNTIFS(tblБаза[Тип],"{typ}",'
                    f'tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')
        return (f'=COUNTIFS(tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')

    r = pp_r + 1
    sec_hdr(ws, r,
            "  РАСШИРЕННАЯ АНАЛИТИКА  —  сравнение с предыдущим месяцем",
            bg=NAVY, ncols=12, height=26)
    r += 1
    ws.row_dimensions[r].height = 4
    r += 1

    def _xt(lr, vr, col, lbl, cf, pf, fmt, bg, inv=False, has_trend=True):
        """Write one 3-col KPI tile: label + value + optional trend delta."""
        n = 3
        ws.merge_cells(start_row=lr, start_column=col,
                       end_row=lr, end_column=col + n - 1)
        c = ws.cell(lr, col, lbl)
        c.fill = mkfill(bg)
        c.font = mkfont(WHITE, 9)
        c.alignment = mkalign("center", "center")
        for ci in range(col + 1, col + n):
            ws.cell(lr, ci).fill = mkfill(bg)

        v_end = col + 1 if has_trend else col + 2
        ws.merge_cells(start_row=vr, start_column=col,
                       end_row=vr, end_column=v_end)
        c = ws.cell(vr, col, cf)
        c.fill = mkfill(WHITE)
        c.font = mkfont(NAVY, 13, True)
        c.alignment = mkalign("center", "center")
        c.number_format = fmt
        c.border = mkborder()
        for ci in range(col + 1, v_end + 1):
            ws.cell(vr, ci).fill = mkfill(WHITE)
            ws.cell(vr, ci).border = mkborder()

        if has_trend and pf is not None:
            tc = col + 2
            tf = f'={cf.lstrip("=")}-({pf.lstrip("=")})'
            c  = ws.cell(vr, tc, tf)
            c.fill  = mkfill("FFF0F4FF")
            c.font  = mkfont(NAVY, 9, True)
            c.alignment = mkalign("center", "center")
            c.border = mkborder()
            if "%" in str(fmt):
                tf_fmt = ('[Green]"▲ "0.0%;[Red]"▼ "0.0%;"-"' if not inv
                          else '[Red]"▲ "0.0%;[Green]"▼ "0.0%;"-"')
            elif '"x"' in str(fmt):
                tf_fmt = ('[Green]"▲ "0.00"x";[Red]"▼ "0.00"x";"-"' if not inv
                          else '[Red]"▲ "0.00"x";[Green]"▼ "0.00"x";"-"')
            elif str(fmt) == "0":
                tf_fmt = ('[Green]"▲ "0;[Red]"▼ "0;"-"' if not inv
                          else '[Red]"▲ "0;[Green]"▼ "0;"-"')
            else:
                tf_fmt = ('[Green]"▲ "#,##0;[Red]"▼ "#,##0;"-"' if not inv
                          else '[Red]"▲ "#,##0;[Green]"▼ "#,##0;"-"')
            c.number_format = tf_fmt
        elif has_trend:
            tc = col + 2
            c  = ws.cell(vr, tc, '"-"')
            c.fill  = mkfill("FFF0F4FF")
            c.font  = mkfont(GRAY_D, 9)
            c.alignment = mkalign("center", "center")
            c.border = mkborder()

    def _sec(title, bg, tiles):
        nonlocal r
        sec_hdr(ws, r, title, bg=bg, ncols=12, height=22)
        ws.row_dimensions[r + 1].height = 18
        ws.row_dimensions[r + 2].height = 36
        for i, tile in enumerate(tiles):
            col = i * 3 + 1
            lbl, cf, pf, fmt, inv = tile[:5]
            ht  = tile[5] if len(tile) > 5 else True
            _xt(r + 1, r + 2, col, lbl, cf, pf, fmt, bg, inv, ht)
        r += 3
        ws.row_dimensions[r].height = 5
        r += 1

    _sec("  ВЫРУЧКА", GREEN, [
        ("Общая выручка",
         _rev(P, Q), _rev(PP, QP), FMT_RUB, False),
        ("Среднее в день",
         f'=IFERROR(({_rev(P, Q)[1:]})/({Q}-{P}+1),0)',
         f'=IFERROR(({_rev(PP, QP)[1:]})/({QP}-{PP}+1),0)',
         FMT_RUB, False),
        ("Среднее за смену",
         f'=IFERROR(({_rev(P, Q)[1:]})/MAX(1,({_cnt(P, Q, "Приход")[1:]})/3),0)',
         f'=IFERROR(({_rev(PP, QP)[1:]})/MAX(1,({_cnt(PP, QP, "Приход")[1:]})/3),0)',
         FMT_RUB, False),
        ("Лучший день",
         f'=IFERROR(AGGREGATE(14,7,'
         f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",tblБаза[Дата],tblБаза[Дата])'
         f'/((tblБаза[Дата]>={P})*(tblБаза[Дата]<={Q})*(tblБаза[Тип]="Приход")),1),0)',
         f'=IFERROR(AGGREGATE(14,7,'
         f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",tblБаза[Дата],tblБаза[Дата])'
         f'/((tblБаза[Дата]>={PP})*(tblБаза[Дата]<={QP})*(tblБаза[Тип]="Приход")),1),0)',
         FMT_RUB, False),
    ])

    _sec("  КОНТРОЛЬ КАССЫ", BLUE, [
        ("Выплаты из кассы",
         _exp(P, Q, "Выплата"), _exp(PP, QP, "Выплата"), FMT_RUB, True),
        ("Расхождений сумма",
         f'=SUMIFS(tblБаза[Расхождение],tblБаза[Тип],"Приход",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         f'=SUMIFS(tblБаза[Расхождение],tblБаза[Тип],"Приход",'
         f'tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})',
         FMT_RUB, True),
        ("Кол-во расхождений",
         f'=COUNTIFS(tblБаза[Расхождение],"<>0",tblБаза[Тип],"Приход",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         f'=COUNTIFS(tblБаза[Расхождение],"<>0",tblБаза[Тип],"Приход",'
         f'tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})',
         "0", True),
        ("Остаток кассы",
         '=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход")'
         '-SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход")'
         '-SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
         '+SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП")',
         f'={_rev(P, Q)[1:]}-({_exp(P, Q)[1:]})',
         FMT_RUB, False),
    ])

    _sec("  ДОЛГИ И ОБЯЗАТЕЛЬСТВА", AMBER, [
        ("Текущий долг",
         '=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
         '-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП")',
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'
         f'-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         FMT_RUB, True),
        ("Взято в долг",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
         f'tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})',
         FMT_RUB, True),
        ("Выплачено долгов",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП",'
         f'tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})',
         FMT_RUB, False),
        ("К оплате (=долг)",
         '=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
         '-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП")',
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'
         f'-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         FMT_RUB, True),
    ])

    _sec("  ПРИБЫЛЬ", GREEN, [
        ("Закуп товара",
         _exp(P, Q, 'Закуп товара'), _exp(PP, QP, 'Закуп товара'),
         FMT_RUB, True),
        ("Все расходы",
         _exp(P, Q), _exp(PP, QP), FMT_RUB, True),
        ("Чистая прибыль",
         f'={_rev(P, Q)[1:]}-({_exp(P, Q)[1:]})',
         f'={_rev(PP, QP)[1:]}-({_exp(PP, QP)[1:]})',
         FMT_RUB, False),
        ("Рентабельность %",
         f'=IFERROR(({_rev(P, Q)[1:]}-{_exp(P, Q)[1:]})'
         f'/MAX(1,{_rev(P, Q)[1:]}),0)',
         f'=IFERROR(({_rev(PP, QP)[1:]}-{_exp(PP, QP)[1:]})'
         f'/MAX(1,{_rev(PP, QP)[1:]}),0)',
         "0.0%", False),
    ])

    _sec("  ЭФФЕКТИВНОСТЬ", PURPLE, [
        ("Маржа %",
         f'=IFERROR(({_rev(P, Q)[1:]}-{_exp(P, Q, "Закуп товара")[1:]})'
         f'/MAX(1,{_rev(P, Q)[1:]}),0)',
         f'=IFERROR(({_rev(PP, QP)[1:]}-{_exp(PP, QP, "Закуп товара")[1:]})'
         f'/MAX(1,{_rev(PP, QP)[1:]}),0)',
         "0.0%", False),
        ("Эффект. закупа (x)",
         f'=IFERROR({_rev(P, Q)[1:]}/MAX(1,{_exp(P, Q, "Закуп товара")[1:]}),0)',
         f'=IFERROR({_rev(PP, QP)[1:]}/MAX(1,{_exp(PP, QP, "Закуп товара")[1:]}),0)',
         '0.00"x"', False),
        ("Нагрузка долга %",
         f'=IFERROR((SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
         f'-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП"))'
         f'/MAX(1,{_rev(P, Q)[1:]}),0)',
         f'=IFERROR((SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
         f'-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП"))'
         f'/MAX(1,{_rev(PP, QP)[1:]}),0)',
         "0.0%", True),
        ("Ср. расход/день",
         f'=IFERROR({_exp(P, Q)[1:]}/({Q}-{P}+1),0)',
         f'=IFERROR({_exp(PP, QP)[1:]}/({QP}-{PP}+1),0)',
         FMT_RUB, True),
    ])

    _sec("  ОПЕРАЦИИ И ВЫПЛАТЫ", RED, [
        ("Просроч. выплаты",
         '=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
         'tblВыплаты[Дата плановой оплаты],"<"&TODAY())',
         None, FMT_RUB, True, True),
        ("Просроч. кол-во",
         '=COUNTIFS(tblВыплаты[Статус],"Запланировано",'
         'tblВыплаты[Дата плановой оплаты],"<"&TODAY())',
         None, "0", True, True),
        ("Выплачено %",
         '=IFERROR(SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Оплачено")'
         '/MAX(1,SUMIFS(tblВыплаты[Сумма],tblВыплаты[Сумма],">0")),0)',
         None, "0.0%", False, True),
        ("Закуп в долг %",
         f'=IFERROR(SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})'
         f'/MAX(1,{_exp(P, Q, "Закуп товара")[1:]}),0)',
         f'=IFERROR(SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
         f'tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})'
         f'/MAX(1,{_exp(PP, QP, "Закуп товара")[1:]}),0)',
         "0.0%", True),
    ])

    _sec("  СТАТИСТИКА ПЕРИОДА", TEAL, [
        ("Дней с данными",
         f'=IFERROR(SUMPRODUCT((tblБаза[Дата]>={P})'
         f'*(tblБаза[Дата]<={Q})*(tblБаза[Тип]="Приход"))/3,0)',
         None, "0", False, False),
        ("Всего операций",
         _cnt(P, Q), None, "0", False, False),
        ("Макс. выручка/день",
         f'=IFERROR(AGGREGATE(14,7,'
         f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",tblБаза[Дата],tblБаза[Дата])'
         f'/((tblБаза[Дата]>={P})*(tblБаза[Дата]<={Q})*(tblБаза[Тип]="Приход")),1),0)',
         f'=IFERROR(AGGREGATE(14,7,'
         f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",tblБаза[Дата],tblБаза[Дата])'
         f'/((tblБаза[Дата]>={PP})*(tblБаза[Дата]<={QP})*(tblБаза[Тип]="Приход")),1),0)',
         FMT_RUB, False),
        ("Мин. выручка/день",
         f'=IFERROR(AGGREGATE(15,7,'
         f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",tblБаза[Дата],tblБаза[Дата])'
         f'/((tblБаза[Дата]>={P})*(tblБаза[Дата]<={Q})*(tblБаза[Тип]="Приход")),1),0)',
         None, FMT_RUB, False, False),
    ])

    _sec("  ДОПОЛНИТЕЛЬНО", NAVY, [
        ("Хозяин",
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
         f'tblБаза[Комментарий],"Хозяин",'
         f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
         f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
         f'tblБаза[Комментарий],"Хозяин",'
         f'tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})',
         FMT_RUB, True),
        ("Списания+Возвраты",
         _exp(P, Q, 'Другое'), _exp(PP, QP, 'Другое'),
         FMT_RUB, True),
        ("Налоги",
         _exp(P, Q, 'Налоги'), None, FMT_RUB, True, False),
        ("Коммуналка+Интернет",
         f'={_exp(P, Q, "Коммуналка")[1:]}+{_exp(P, Q, "Интернет")[1:]}',
         None, FMT_RUB, True, False),
    ])


def build_dashboard_charts(ws):
    """Append 6 charts to Дашборд after extended analytics (~row 86)."""

    # ── HELPER TABLE: monthly aggregates ────────────────────────
    # Cols N-Q (14-17), rows 3-15  (piggyback on empty right side)
    HC = 14   # col N — month label
    HR = 15   # col O — revenue
    HE = 16   # col P — expenses
    HP = 17   # col Q — profit

    for col, txt in [(HC, "Месяц"), (HR, "Выручка"), (HE, "Расходы"), (HP, "Прибыль")]:
        c = ws.cell(3, col, txt)
        c.font  = mkfont(GRAY_D, 8)
        c.fill  = mkfill(GRAY_L)

    for m in range(1, 13):
        row = 3 + m   # rows 4-15
        c = ws.cell(row, HC, MONTHS_RU[m - 1])
        c.font = mkfont(NAVY, 8)

        rf = (f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
              f'tblБаза[Дата],">="&DATE($B$3,{m},1),'
              f'tblБаза[Дата],"<="&EOMONTH(DATE($B$3,{m},1),0))')
        ef = (f'=SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
              f'tblБаза[Дата],">="&DATE($B$3,{m},1),'
              f'tblБаза[Дата],"<="&EOMONTH(DATE($B$3,{m},1),0))')

        c = ws.cell(row, HR, rf);  c.number_format = FMT_RUB; c.font = mkfont(NAVY, 8)
        c = ws.cell(row, HE, ef);  c.number_format = FMT_RUB; c.font = mkfont(NAVY, 8)
        cl_r = get_column_letter(HR); cl_e = get_column_letter(HE)
        c = ws.cell(row, HP, f'={cl_r}{row}-{cl_e}{row}')
        c.number_format = FMT_RUB; c.font = mkfont(NAVY, 8)

    # ── HELPER TABLE: payment methods ───────────────────────────
    # Cols N-O (14-15), rows 17-20
    # D8=Наличные, G8=Карта, J8=Перевод (first cell of each merged KPI tile)
    for col, txt in [(14, "Способ"), (15, "Сумма")]:
        c = ws.cell(17, col, txt); c.font = mkfont(GRAY_D, 8); c.fill = mkfill(GRAY_L)
    for i, (lbl, src) in enumerate([("Наличные", "D8"), ("Карта", "G8"), ("Перевод", "J8")]):
        r = 18 + i
        ws.cell(r, 14, lbl).font = mkfont(NAVY, 8)
        c = ws.cell(r, 15, f'={src}'); c.number_format = FMT_RUB; c.font = mkfont(NAVY, 8)

    # ── SECTION HEADER ───────────────────────────────────────────
    # Extended analytics ends at row ~85 (r=86 after last _sec).
    # CSEC must be past the shifts section: shifts end at row ~(49 + extra_shifts)
    CSEC = 87 + (len(SHIFTS) - 2)
    ws.row_dimensions[CSEC].height = 8
    sec_hdr(ws, CSEC + 1, "  ДИАГРАММЫ — ВИЗУАЛЬНАЯ АНАЛИТИКА",
            bg=TEAL, ncols=12, height=26)
    ws.row_dimensions[CSEC + 2].height = 6
    C1 = CSEC + 3   # row 90 — first chart anchor row

    # Row numbers for existing dashboard data (computed from build_dashboard layout)
    # Expense categories: header row 20, data rows 21-29, col B(2), col E(5)
    # Weekday:            header row 33, data rows 34-40, col B(2), col E(5)
    # Shift:              header row 44, data rows 45-46, col B(2), col E(5)
    EXP_HDR = 20;  EXP_R1 = 21;  EXP_R2 = 29
    WD_HDR  = 33;  WD_R1  = 34;  WD_R2  = 40
    SH_HDR  = 44;  SH_R1  = 45;  SH_R2  = 45 + len(SHIFTS) - 1

    # ── CHART 1: Monthly trend — Line chart ──────────────────────
    ch1 = LineChart()
    ch1.title  = "Выручка / Расходы / Прибыль по месяцам"
    ch1.style  = 10
    ch1.height = 12;  ch1.width = 24
    ch1.y_axis.numFmt = '#,##0'
    ch1.x_axis.title  = "Месяц"

    cats1 = Reference(ws, min_col=HC, min_row=4, max_row=15)
    data1 = Reference(ws, min_col=HR, min_row=3, max_col=HP, max_row=15)
    ch1.add_data(data1, titles_from_data=True)
    ch1.set_categories(cats1)
    ws.add_chart(ch1, f"A{C1}")

    # ── CHART 2: Payment method — Pie chart ──────────────────────
    C2 = C1 + 19   # ~row 109
    ch2 = PieChart()
    ch2.title  = "Выручка по способу оплаты"
    ch2.style  = 10
    ch2.height = 12;  ch2.width = 12

    data2 = Reference(ws, min_col=15, min_row=18, max_row=20)
    cats2 = Reference(ws, min_col=14, min_row=18, max_row=20)
    ch2.add_data(data2)
    ch2.set_categories(cats2)
    ws.add_chart(ch2, f"A{C2}")

    # ── CHART 3: Expense by category — Horizontal bar chart ──────
    ch3 = BarChart()
    ch3.barDir  = "bar"
    ch3.title   = "Расходы по категориям"
    ch3.style   = 10
    ch3.height  = 12;  ch3.width = 12
    ch3.x_axis.title = "Сумма, ₽"

    data3 = Reference(ws, min_col=5, min_row=EXP_HDR, max_row=EXP_R2)
    cats3 = Reference(ws, min_col=2, min_row=EXP_R1,  max_row=EXP_R2)
    ch3.add_data(data3, titles_from_data=True)
    ch3.set_categories(cats3)
    ws.add_chart(ch3, f"G{C2}")

    # ── CHART 4: Revenue by weekday — Column chart ───────────────
    C3 = C2 + 19   # ~row 128
    ch4 = BarChart()
    ch4.barDir  = "col"
    ch4.title   = "Выручка по дням недели"
    ch4.style   = 10
    ch4.height  = 12;  ch4.width = 24
    ch4.y_axis.numFmt = '#,##0'
    ch4.x_axis.title  = "День"

    data4 = Reference(ws, min_col=5, min_row=WD_HDR, max_row=WD_R2)
    cats4 = Reference(ws, min_col=2, min_row=WD_R1,  max_row=WD_R2)
    ch4.add_data(data4, titles_from_data=True)
    ch4.set_categories(cats4)
    ws.add_chart(ch4, f"A{C3}")

    # ── CHART 5: Revenue by shift — Column chart ─────────────────
    C4 = C3 + 19   # ~row 147
    ch5 = BarChart()
    ch5.barDir  = "col"
    ch5.title   = "Выручка по сменам"
    ch5.style   = 10
    ch5.height  = 10;  ch5.width = 12

    data5 = Reference(ws, min_col=5, min_row=SH_HDR, max_row=SH_R2)
    cats5 = Reference(ws, min_col=2, min_row=SH_R1,  max_row=SH_R2)
    ch5.add_data(data5, titles_from_data=True)
    ch5.set_categories(cats5)
    ws.add_chart(ch5, f"A{C4}")

    # ── CHART 6: Expense share by category — Pie chart ───────────
    ch6 = PieChart()
    ch6.title  = "Доля расходов по категориям"
    ch6.style  = 10
    ch6.height = 10;  ch6.width = 12

    data6 = Reference(ws, min_col=5, min_row=EXP_R1, max_row=EXP_R2)
    cats6 = Reference(ws, min_col=2, min_row=EXP_R1, max_row=EXP_R2)
    ch6.add_data(data6)
    ch6.set_categories(cats6)
    ws.add_chart(ch6, f"G{C4}")


# ═══════════════════════════════════════════════════════════════
#  BLOCK 5: VBA EXPORT (.bas file)
# ═══════════════════════════════════════════════════════════════

VBA_CODE = r'''Attribute VB_Name = "FinKontrolMacros"
' ═══════════════════════════════════════════════════════════════
'  ФИНАНСОВЫЙ КОНТРОЛЬ — VBA Macros
'  Импорт: Alt+F11 → File → Import File → выбрать этот .bas
'  Назначить кнопкам: правый клик по кнопке → Назначить макрос
' ═══════════════════════════════════════════════════════════════

Option Explicit

' ── Константы — имена листов ──
Private Const SH_BAZA  As String = "БАЗА_ДДС"
Private Const SH_KASSA As String = "Ввод_Касса"
Private Const SH_RASH  As String = "Ввод_Расходы"
Private Const SH_VYPL  As String = "Запись_Выплат"
Private Const SH_CAL   As String = "Календарь_Выплат"
Private Const SH_DASH  As String = "Дашборд"
Private Const SH_PULT  As String = "Пульт"
Private Const SH_SETS  As String = "Настройки"
Private Const SH_SVOD  As String = "Сводные"

' ── Автокомплит: вспомогательные колонки в Настройки ──
' Col I (9) = фильтр Кассиры, Col J (10) = Категории, Col K (11) = Поставщики
Private Const AC_COL_KASSA As Integer = 9
Private Const AC_COL_CAT   As Integer = 10
Private Const AC_COL_SUP   As Integer = 11
Private Const AC_MAX_ROWS  As Integer = 15

' ═══════════════════════════════════════════════════════════════
'  SAVE KASSA — Сохранить кассу в БАЗА_ДДС
'  Назначить на кнопку A17:G17 на листе Ввод_Касса
' ═══════════════════════════════════════════════════════════════
Public Sub SaveKassa()
    Dim wsK As Worksheet, wsB As Worksheet
    Set wsK = ThisWorkbook.Worksheets(SH_KASSA)
    Set wsB = ThisWorkbook.Worksheets(SH_BAZA)

    ' Валидация
    Dim dtVal As Variant, shVal As String, cashVal As String
    dtVal   = wsK.Range("B3").Value
    shVal   = CStr(wsK.Range("D3").Value)
    cashVal = CStr(wsK.Range("F3").Value)

    If Not IsDate(dtVal) Then
        MsgBox "Введите дату смены (B3)", vbExclamation, "Касса"
        Exit Sub
    End If
    If Len(Trim(shVal)) = 0 Then
        MsgBox "Выберите смену (D3)", vbExclamation, "Касса"
        Exit Sub
    End If
    If Len(Trim(cashVal)) = 0 Then
        MsgBox "Выберите кассира (F3)", vbExclamation, "Касса"
        Exit Sub
    End If

    ' Добавляем строки через ListObject — таблица расширяется автоматически
    Dim tblB As ListObject
    Set tblB = wsB.ListObjects("tblБаза")

    ' 3 Приход-строки (Наличные, Карта, Перевод)
    Dim methods As Variant, i As Long, r As Long
    Dim factVal As Double, zVal As Double, discVal As Double
    methods = Array("Наличные", "Карта", "Перевод")
    For i = 0 To 2
        r = tblB.ListRows.Add.Range.Row
        zVal    = CDbl(Nz(wsK.Cells(8 + i, 2).Value))
        factVal = CDbl(Nz(wsK.Cells(8 + i, 3).Value))
        discVal = factVal - zVal

        wsB.Cells(r, 1).Value = CDate(dtVal)
        wsB.Cells(r, 2).Value = shVal
        wsB.Cells(r, 3).Value = cashVal
        wsB.Cells(r, 4).Value = "Приход"
        wsB.Cells(r, 5).Value = "Продажи"
        wsB.Cells(r, 6).Value = methods(i)
        wsB.Cells(r, 7).Value = factVal
        If discVal <> 0 Then wsB.Cells(r, 8).Value = discVal
        wsB.Cells(r, 9).Value = CStr(wsK.Range("E13").Value)
        wsB.Cells(r, 1).NumberFormat = "DD.MM.YYYY"
        wsB.Cells(r, 7).NumberFormat = "#,##0"
        wsB.Cells(r, 8).NumberFormat = "#,##0"
    Next i

    ' Выплата из кассы (D14) — если сумма > 0
    Dim vyplAmt As Double
    vyplAmt = CDbl(Nz(wsK.Range("D14").Value))
    If vyplAmt > 0 Then
        r = tblB.ListRows.Add.Range.Row
        wsB.Cells(r, 1).Value = CDate(dtVal)
        wsB.Cells(r, 2).Value = shVal
        wsB.Cells(r, 3).Value = cashVal
        wsB.Cells(r, 4).Value = "Расход"
        wsB.Cells(r, 5).Value = "Выплата"
        wsB.Cells(r, 6).Value = "Наличные"
        wsB.Cells(r, 7).Value = vyplAmt
        wsB.Cells(r, 8).Value = ""
        wsB.Cells(r, 9).Value = "Выплата из кассы"
        wsB.Cells(r, 1).NumberFormat = "DD.MM.YYYY"
        wsB.Cells(r, 7).NumberFormat = "#,##0"
    End If

    ' Очистить форму
    wsK.Range("B3").ClearContents
    wsK.Range("D3").ClearContents
    wsK.Range("F3").ClearContents
    For i = 8 To 10
        wsK.Cells(i, 2).Value = 0
        wsK.Cells(i, 3).Value = 0
    Next i
    wsK.Range("D14").Value = 0
    wsK.Range("E13").ClearContents

    MsgBox "Сохранено в БАЗА_ДДС за " & Format(dtVal, "DD.MM.YYYY") _
           & " (" & shVal & ", " & cashVal & ")", _
           vbInformation, "Касса сохранена"
End Sub


' ═══════════════════════════════════════════════════════════════
'  SAVE RASHOD — Сохранить расход или закуп в долг в БАЗА_ДДС
'  Назначить на кнопку A16:D16 на листе Ввод_Расходы
' ═══════════════════════════════════════════════════════════════
Public Sub SaveRashod()
    Dim wsR As Worksheet, wsB As Worksheet
    Set wsR = ThisWorkbook.Worksheets(SH_RASH)
    Set wsB = ThisWorkbook.Worksheets(SH_BAZA)

    Dim dtVal As Variant
    dtVal = wsR.Range("B3").Value
    If Not IsDate(dtVal) Then
        MsgBox "Введите дату (B3)", vbExclamation, "Расход"
        Exit Sub
    End If

    ' Проверяем, какой раздел заполнен
    Dim rashSum As Double, dolgSum As Double
    rashSum = CDbl(Nz(wsR.Range("B8").Value))
    dolgSum = CDbl(Nz(wsR.Range("B13").Value))

    If rashSum > 0 And dolgSum > 0 Then
        MsgBox "Заполните ТОЛЬКО один раздел: либо РАСХОД, либо ЗАКУП В ДОЛГ", _
               vbExclamation, "Расход"
        Exit Sub
    End If
    If rashSum <= 0 And dolgSum <= 0 Then
        MsgBox "Введите сумму в одном из разделов (B8 или B13)", _
               vbExclamation, "Расход"
        Exit Sub
    End If

    Dim tblB As ListObject
    Set tblB = wsB.ListObjects("tblБаза")
    Dim r As Long

    If rashSum > 0 Then
        ' Обычный расход
        Dim catVal As String, mthVal As String
        catVal = CStr(wsR.Range("B6").Value)
        mthVal = CStr(wsR.Range("B7").Value)
        If Len(Trim(catVal)) = 0 Then
            MsgBox "Выберите категорию (B6)", vbExclamation, "Расход"
            Exit Sub
        End If
        If Len(Trim(mthVal)) = 0 Then
            MsgBox "Выберите способ оплаты (B7)", vbExclamation, "Расход"
            Exit Sub
        End If

        r = tblB.ListRows.Add.Range.Row
        wsB.Cells(r, 1).Value = CDate(dtVal)
        wsB.Cells(r, 2).Value = ""
        wsB.Cells(r, 3).Value = ""
        wsB.Cells(r, 4).Value = "Расход"
        wsB.Cells(r, 5).Value = catVal
        wsB.Cells(r, 6).Value = mthVal
        wsB.Cells(r, 7).Value = rashSum
        wsB.Cells(r, 8).Value = ""
        wsB.Cells(r, 9).Value = CStr(wsR.Range("B9").Value)

        MsgBox "Расход сохранён: " & Format(rashSum, "#,##0") & " " & ChrW(8381) & " (" & catVal & ")", _
               vbInformation, "Расход"
    Else
        ' Закуп в долг
        Dim supVal As String
        supVal = CStr(wsR.Range("B12").Value)
        If Len(Trim(supVal)) = 0 Then
            MsgBox "Выберите поставщика (B12)", vbExclamation, "Долг"
            Exit Sub
        End If

        r = tblB.ListRows.Add.Range.Row
        wsB.Cells(r, 1).Value = CDate(dtVal)
        wsB.Cells(r, 2).Value = ""
        wsB.Cells(r, 3).Value = ""
        wsB.Cells(r, 4).Value = "Долг"
        wsB.Cells(r, 5).Value = "Закуп товара"
        wsB.Cells(r, 6).Value = "Перевод"
        wsB.Cells(r, 7).Value = dolgSum
        wsB.Cells(r, 8).Value = ""
        wsB.Cells(r, 9).Value = supVal

        MsgBox "Закуп в долг сохранён: " & Format(dolgSum, "#,##0") & " " & ChrW(8381) & _
               " (" & supVal & ")" & vbCrLf & vbCrLf & _
               "Не забудьте добавить запись в ЗАПИСЬ_ВЫПЛАТ для планирования оплаты!", _
               vbInformation, "Долг"
    End If

    ' Форматирование новой строки
    wsB.Cells(r, 1).NumberFormat = "DD.MM.YYYY"
    wsB.Cells(r, 7).NumberFormat = "#,##0"

    ' Очистить форму
    wsR.Range("B3").ClearContents
    wsR.Range("B6").ClearContents
    wsR.Range("B7").ClearContents
    wsR.Range("B8").Value = 0
    wsR.Range("B9").ClearContents
    wsR.Range("B12").ClearContents
    wsR.Range("B13").Value = 0
End Sub


' ═══════════════════════════════════════════════════════════════
'  INSERT TODAY — Кнопки СЕГОДНЯ
' ═══════════════════════════════════════════════════════════════
Public Sub InsertToday_Kassa()
    ThisWorkbook.Worksheets(SH_KASSA).Range("B3").Value = Date
    ThisWorkbook.Worksheets(SH_KASSA).Range("B3").NumberFormat = "DD.MM.YYYY"
End Sub

Public Sub InsertToday_Rashod()
    ThisWorkbook.Worksheets(SH_RASH).Range("B3").Value = Date
    ThisWorkbook.Worksheets(SH_RASH).Range("B3").NumberFormat = "DD.MM.YYYY"
End Sub

Public Sub InsertToday_Calendar()
    Dim months As Variant
    months = Array("Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", _
                   "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь")
    ThisWorkbook.Worksheets(SH_CAL).Range("B3").Value = months(Month(Date) - 1)
    ThisWorkbook.Worksheets(SH_CAL).Range("D3").Value = Year(Date)
End Sub


' ═══════════════════════════════════════════════════════════════
'  REFRESH DASHBOARD — Пересчёт + обновление всех таблиц
' ═══════════════════════════════════════════════════════════════
Public Sub RefreshDashboard()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationAutomatic
    Application.CalculateFull
    ThisWorkbook.RefreshAll
    Application.ScreenUpdating = True
    MsgBox "Все формулы и сводные пересчитаны.", vbInformation, "Обновление"
End Sub


' ═══════════════════════════════════════════════════════════════
'  АВТОКОМПЛИТ — умный поиск по подстроке
'  Как работает:
'    1) Пользователь начинает печатать в поле (F3 / B6 / B12)
'    2) Worksheet_Change вызывает AC_Kassa / AC_Category / AC_Supplier
'    3) VBA фильтрует список по введённому тексту (case-insensitive)
'    4) Записывает совпадения во вспомогательные колонки I/J/K листа Настройки
'    5) Обновляет выпадающий список поля — теперь видны только совпадения
'    6) Если совпадение ровно одно — вставляет полное значение автоматически
'
'  УСТАНОВКА (сделать 1 раз после импорта этого .bas файла):
'    Открыть Alt+F11, затем для каждого листа:
'
'  ─── Вставить в модуль листа "Ввод_Касса" ──────────────────
'  (правый клик на вкладку → "Просмотр кода", вставить ниже)
'
'  Private Sub Worksheet_Change(ByVal Target As Range)
'      If Target.Cells.Count > 1 Then Exit Sub
'      If Target.Address = "$F$3" Then
'          Call FinKontrolMacros.AC_Kassa(Target)
'      End If
'  End Sub
'
'  ─── Вставить в модуль листа "Ввод_Расходы" ────────────────
'
'  Private Sub Worksheet_Change(ByVal Target As Range)
'      If Target.Cells.Count > 1 Then Exit Sub
'      If Target.Address = "$B$6" Then
'          Call FinKontrolMacros.AC_Category(Target)
'      ElseIf Target.Address = "$B$12" Then
'          Call FinKontrolMacros.AC_Supplier(Target)
'      End If
'  End Sub
'
' ═══════════════════════════════════════════════════════════════

Private Sub AC_DoFilter(inputCell As Range, masterList As Variant, helperCol As Integer)
    Dim wsS As Worksheet
    Set wsS = ThisWorkbook.Worksheets(SH_SETS)

    Dim typed As String
    typed = LCase(Trim(CStr(inputCell.Value)))

    On Error GoTo errHandler
    Application.EnableEvents = False
    Application.ScreenUpdating = False

    ' Очистить вспомогательный столбец
    wsS.Range(wsS.Cells(1, helperCol), wsS.Cells(AC_MAX_ROWS, helperCol)).ClearContents

    Dim i As Long, j As Long
    j = 0

    ' Фильтр по подстроке (регистр игнорируется)
    For i = 0 To UBound(masterList)
        If Len(typed) = 0 Or InStr(1, LCase(CStr(masterList(i))), typed, vbTextCompare) > 0 Then
            wsS.Cells(j + 1, helperCol).Value = masterList(i)
            j = j + 1
        End If
    Next i

    ' Нет совпадений — показать весь список
    If j = 0 Then
        For i = 0 To UBound(masterList)
            wsS.Cells(i + 1, helperCol).Value = masterList(i)
        Next i
        j = UBound(masterList) + 1
    End If

    ' Адрес диапазона для DataValidation
    Dim colLetter As String
    colLetter = Chr(64 + helperCol)   ' 9→I, 10→J, 11→K
    Dim dvAddr As String
    dvAddr = "=" & wsS.Name & "!$" & colLetter & "$1:$" & colLetter & "$" & j

    ' Обновить DataValidation (без блокировки — пользователь может печатать свободно)
    With inputCell.Validation
        .Delete
        .Add Type:=xlValidateList, AlertStyle:=xlValidAlertInformation, _
             Operator:=xlBetween, Formula1:=dvAddr
        .IgnoreBlank = True
        .InCellDropdown = True
        .ShowInput = False
        .ShowError = False
    End With

    ' Автозаполнение при единственном совпадении
    If j = 1 And Len(typed) > 0 Then
        inputCell.Value = wsS.Cells(1, helperCol).Value
    End If

errHandler:
    Application.EnableEvents = True
    Application.ScreenUpdating = True
End Sub

' ── Публичные методы — вызываются из Worksheet_Change листов ──

Public Sub AC_Kassa(inputCell As Range)
    ' Читаем кассиров из Настройки Р7 (столбец B, строки 22+)
    Dim wsS As Worksheet
    Set wsS = ThisWorkbook.Worksheets(SH_SETS)
    Dim startRow As Long, n As Long, i As Long
    startRow = 22
    n = 0
    Do While Len(CStr(wsS.Cells(startRow + n, 2).Value)) > 0
        n = n + 1
    Loop
    If n = 0 Then Exit Sub
    Dim lst() As String
    ReDim lst(n - 1)
    For i = 0 To n - 1
        lst(i) = CStr(wsS.Cells(startRow + i, 2).Value)
    Next i
    Call AC_DoFilter(inputCell, lst, AC_COL_KASSA)
End Sub

Public Sub AC_Category(inputCell As Range)
    ' Читаем категории из Настройки Р7 (столбец C, строки 22+)
    Dim wsS As Worksheet
    Set wsS = ThisWorkbook.Worksheets(SH_SETS)
    Dim startRow As Long, n As Long, i As Long
    startRow = 22
    n = 0
    Do While Len(CStr(wsS.Cells(startRow + n, 3).Value)) > 0
        n = n + 1
    Loop
    If n = 0 Then Exit Sub
    Dim lst() As String
    ReDim lst(n - 1)
    For i = 0 To n - 1
        lst(i) = CStr(wsS.Cells(startRow + i, 3).Value)
    Next i
    Call AC_DoFilter(inputCell, lst, AC_COL_CAT)
End Sub

Public Sub AC_Supplier(inputCell As Range)
    ' Читаем поставщиков из Настройки P9 (столбец C, строки 79+)
    ' Добавляй новых поставщиков прямо в таблицу Настройки — список бесконечный
    Dim wsS As Worksheet
    Set wsS = ThisWorkbook.Worksheets(SH_SETS)
    Dim startRow As Long, n As Long, i As Long
    startRow = 79
    n = 0
    Do While Len(CStr(wsS.Cells(startRow + n, 3).Value)) > 0
        n = n + 1
    Loop
    If n = 0 Then Exit Sub
    Dim lst() As String
    ReDim lst(n - 1)
    For i = 0 To n - 1
        lst(i) = CStr(wsS.Cells(startRow + i, 3).Value)
    Next i
    Call AC_DoFilter(inputCell, lst, AC_COL_SUP)
End Sub


' ═══════════════════════════════════════════════════════════════
'  УТИЛИТЫ
' ═══════════════════════════════════════════════════════════════
Private Function Nz(v As Variant) As Variant
    If IsEmpty(v) Or IsNull(v) Or (VarType(v) = vbString And v = "") Then
        Nz = 0
    Else
        Nz = v
    End If
End Function


' ═══════════════════════════════════════════════════════════════
'  УСТАНОВКА КНОПОК — ЗАПУСТИТЬ ОДИН РАЗ
'  После импорта .bas:  Alt+F8 → SetupAll → Выполнить
'  Добавляет кнопки-фигуры с макросами + автокомплит.
' ═══════════════════════════════════════════════════════════════
Public Sub SetupAll()
    On Error GoTo setupErr
    Application.ScreenUpdating = False

    Dim wsK As Worksheet, wsR As Worksheet
    Dim wsC As Worksheet, wsD As Worksheet
    Set wsK = ThisWorkbook.Worksheets(SH_KASSA)
    Set wsR = ThisWorkbook.Worksheets(SH_RASH)
    Set wsC = ThisWorkbook.Worksheets(SH_CAL)
    Set wsD = ThisWorkbook.Worksheets(SH_DASH)

    ' ── Ввод_Касса ──────────────────────────────────────────────
    Call AddBtn(wsK, "A17:G17", "  СОХРАНИТЬ КАССУ", _
                "FinKontrolMacros.SaveKassa", RGB(5, 150, 105))
    Call AddBtn(wsK, "E4:F4", "  СЕГОДНЯ", _
                "FinKontrolMacros.InsertToday_Kassa", RGB(29, 78, 216))

    ' ── Ввод_Расходы ────────────────────────────────────────────
    Call AddBtn(wsR, "A16:D16", "  СОХРАНИТЬ", _
                "FinKontrolMacros.SaveRashod", RGB(5, 150, 105))
    Call AddBtn(wsR, "C3:D3", "  СЕГОДНЯ", _
                "FinKontrolMacros.InsertToday_Rashod", RGB(29, 78, 216))

    ' ── Календарь_Выплат ────────────────────────────────────────
    Call AddBtn(wsC, "I3:J3", "  СЕГОДНЯ", _
                "FinKontrolMacros.InsertToday_Calendar", RGB(29, 78, 216))

    ' ── Дашборд ─────────────────────────────────────────────────
    Call AddBtn(wsD, "K3:L3", "  ОБНОВИТЬ", _
                "FinKontrolMacros.RefreshDashboard", RGB(217, 119, 6))

    ' ── Сводные ─────────────────────────────────────────────────
    Call AddBtn(ThisWorkbook.Worksheets(SH_SVOD), "A3:L3", _
                "  СОЗДАТЬ СВОДНЫЕ ТАБЛИЦЫ", _
                "FinKontrolMacros.CreatePivotTables", RGB(14, 116, 144))

    ' ── Автокомплит (Worksheet_Change) ──────────────────────────
    Call TryInjectAutocomplete

    Application.ScreenUpdating = True
    MsgBox "Кнопки установлены!" & vbCrLf & _
           "Сохраните файл как .xlsm чтобы сохранить макросы.", _
           vbInformation, "ФИНАНСОВЫЙ КОНТРОЛЬ — Установка завершена"
    Exit Sub
setupErr:
    Application.ScreenUpdating = True
    MsgBox "Ошибка установки: " & Err.Description, vbCritical, "SetupAll"
End Sub


Private Sub AddBtn(ws As Worksheet, rngAddr As String, caption As String, _
                   macro As String, clr As Long)
    Dim rng As Range
    Set rng = ws.Range(rngAddr)

    ' Удалить предыдущую кнопку с тем же макросом
    Dim shp As Shape
    For Each shp In ws.Shapes
        If shp.OnAction = macro Then shp.Delete: Exit For
    Next shp

    ' Добавить прямоугольную фигуру с назначенным макросом
    Set shp = ws.Shapes.AddShape(msoShapeRoundedRectangle, _
        rng.Left + 1, rng.Top + 1, rng.Width - 2, rng.Height - 2)

    shp.OnAction = macro
    shp.Fill.ForeColor.RGB = clr
    shp.Fill.Solid
    shp.Line.Visible = msoFalse

    On Error Resume Next
    shp.TextFrame2.TextRange.Text = caption
    shp.TextFrame2.TextRange.Font.Bold = True
    shp.TextFrame2.TextRange.Font.Size = 12
    shp.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = RGB(255, 255, 255)
    shp.TextFrame2.VerticalAnchor = msoAnchorMiddle
    shp.TextFrame2.TextRange.ParagraphFormat.Alignment = 2  ' ppAlignCenter=2 (mso center); ppAlignCenter not in Excel VBA
    On Error GoTo 0
End Sub


Private Sub TryInjectAutocomplete()
    ' Требует: Файл → Параметры → Центр управления безопасностью →
    '   Параметры макросов → "Доверять доступу к объектной модели VBA"
    On Error GoTo noAccess

    Dim proj As Object
    Set proj = ThisWorkbook.VBProject

    Call InjectWSChange(ThisWorkbook.Worksheets(SH_KASSA), _
        "Private Sub Worksheet_Change(ByVal Target As Range)" & vbCrLf & _
        "    If Target.Cells.Count > 1 Then Exit Sub" & vbCrLf & _
        "    If Target.Address = ""$F$3"" Then" & vbCrLf & _
        "        Call FinKontrolMacros.AC_Kassa(Target)" & vbCrLf & _
        "    End If" & vbCrLf & _
        "End Sub")

    Call InjectWSChange(ThisWorkbook.Worksheets(SH_RASH), _
        "Private Sub Worksheet_Change(ByVal Target As Range)" & vbCrLf & _
        "    If Target.Cells.Count > 1 Then Exit Sub" & vbCrLf & _
        "    If Target.Address = ""$B$6"" Then" & vbCrLf & _
        "        Call FinKontrolMacros.AC_Category(Target)" & vbCrLf & _
        "    ElseIf Target.Address = ""$B$12"" Then" & vbCrLf & _
        "        Call FinKontrolMacros.AC_Supplier(Target)" & vbCrLf & _
        "    End If" & vbCrLf & _
        "End Sub")
    Exit Sub

noAccess:
    ' Кнопки работают и без этой настройки. Только автокомплит при наборе недоступен.
End Sub


Private Sub InjectWSChange(ws As Worksheet, code As String)
    Dim cm As Object
    Set cm = ThisWorkbook.VBProject.VBComponents(ws.CodeName).CodeModule
    If InStr(1, cm.Lines(1, cm.CountOfLines), "Worksheet_Change") = 0 Then
        cm.InsertLines cm.CountOfLines + 1, vbCrLf & code
    End If
End Sub


' ═══════════════════════════════════════════════════════════════
'  СВОДНЫЕ ТАБЛИЦЫ — реальные Excel PivotTables из tblБаза
'  Alt+F8 -> CreatePivotTables -> Выполнить  (или нажать кнопку на листе)
' ═══════════════════════════════════════════════════════════════
Public Sub CreatePivotTables()
    On Error GoTo pivErr
    Application.ScreenUpdating = False
    Application.DisplayAlerts = False

    Dim wbk As Workbook
    Set wbk = ThisWorkbook

    Dim wsSrc As Worksheet
    Set wsSrc = wbk.Worksheets(SH_BAZA)

    Dim wsPT As Worksheet
    Set wsPT = wbk.Worksheets(SH_SVOD)

    ' Очистить старые сводные таблицы
    Dim pt As PivotTable
    For Each pt In wsPT.PivotTables
        pt.TableRange2.Clear
    Next pt

    ' Источник — умная таблица tblБаза
    Dim tbl As ListObject
    Set tbl = wsSrc.ListObjects("tblБаза")
    Dim srcAddr As String
    srcAddr = "'" & wsSrc.Name & "'!" & tbl.Range.Address

    ' Создать общий PivotCache
    Dim pc As PivotCache
    Set pc = wbk.PivotCaches.Create(SourceType:=xlDatabase, SourceData:=srcAddr)

    ' ── ПТ1: Выручка по месяцам ─────────────────────────────────
    Dim pt1 As PivotTable
    Set pt1 = pc.CreatePivotTable( _
        TableDestination:=wsPT.Cells(10, 2), TableName:="PT_VyruchkaMesyac")
    With pt1
        With .PivotFields("Тип")
            .Orientation = xlPageField
            .CurrentPage = "Приход"
        End With
        With .PivotFields("Дата")
            .Orientation = xlRowField
            .Position = 1
        End With
        With .PivotFields("Сумма")
            .Orientation = xlDataField
            .Function = xlSum
            .NumberFormat = "#,##0"
            .Name = "Выручка (руб.)"
        End With
        .NullString = "0"
        .RowAxisLayout xlTabularRow
    End With
    On Error Resume Next
    pt1.PivotFields("Дата").AutoGroup
    On Error GoTo pivErr

    ' ── ПТ2: Расходы по категориям ──────────────────────────────
    Dim pt2 As PivotTable
    Set pt2 = pc.CreatePivotTable( _
        TableDestination:=wsPT.Cells(10, 9), TableName:="PT_RaskhodKat")
    With pt2
        With .PivotFields("Тип")
            .Orientation = xlPageField
            .CurrentPage = "Расход"
        End With
        With .PivotFields("Категория")
            .Orientation = xlRowField
            .Position = 1
        End With
        With .PivotFields("Сумма")
            .Orientation = xlDataField
            .Function = xlSum
            .NumberFormat = "#,##0"
            .Name = "Сумма расходов"
        End With
        .NullString = "0"
    End With

    ' ── ПТ3: Выручка кассиры x смены ────────────────────────────
    Dim pt3 As PivotTable
    Set pt3 = pc.CreatePivotTable( _
        TableDestination:=wsPT.Cells(30, 2), TableName:="PT_VyruchkaKassir")
    With pt3
        With .PivotFields("Тип")
            .Orientation = xlPageField
            .CurrentPage = "Приход"
        End With
        With .PivotFields("Кассир")
            .Orientation = xlRowField
            .Position = 1
        End With
        With .PivotFields("Смена")
            .Orientation = xlColumnField
            .Position = 1
        End With
        With .PivotFields("Сумма")
            .Orientation = xlDataField
            .Function = xlSum
            .NumberFormat = "#,##0"
            .Name = "Выручка"
        End With
        .NullString = "0"
    End With

    ' ── ПТ4: Долги по поставщикам ───────────────────────────────
    Dim pt4 As PivotTable
    Set pt4 = pc.CreatePivotTable( _
        TableDestination:=wsPT.Cells(30, 9), TableName:="PT_DolgiPost")
    With pt4
        With .PivotFields("Тип")
            .Orientation = xlPageField
            .CurrentPage = "Долг"
        End With
        With .PivotFields("Комментарий")
            .Orientation = xlRowField
            .Position = 1
        End With
        With .PivotFields("Сумма")
            .Orientation = xlDataField
            .Function = xlSum
            .NumberFormat = "#,##0"
            .Name = "Долг (руб.)"
        End With
        .NullString = "0"
    End With

    ' ── ПТ5: Итоговая сводная по всем типам ─────────────────────
    Dim pt5 As PivotTable
    Set pt5 = pc.CreatePivotTable( _
        TableDestination:=wsPT.Cells(50, 2), TableName:="PT_ObshchayaSvodnaya")
    With pt5
        With .PivotFields("Тип")
            .Orientation = xlRowField
            .Position = 1
        End With
        With .PivotFields("Категория")
            .Orientation = xlRowField
            .Position = 2
        End With
        With .PivotFields("Способ оплаты")
            .Orientation = xlColumnField
            .Position = 1
        End With
        With .PivotFields("Сумма")
            .Orientation = xlDataField
            .Function = xlSum
            .NumberFormat = "#,##0"
            .Name = "Итого"
        End With
        .NullString = "0"
    End With

    Application.DisplayAlerts = True
    Application.ScreenUpdating = True
    MsgBox "5 сводных таблиц созданы!" & vbCrLf & vbCrLf & _
           "ПТ1 - Выручка по месяцам" & vbCrLf & _
           "ПТ2 - Расходы по категориям" & vbCrLf & _
           "ПТ3 - Выручка: кассиры x смены" & vbCrLf & _
           "ПТ4 - Долги по поставщикам" & vbCrLf & _
           "ПТ5 - Итоговая сводная (тип x категория x метод оплаты)", _
           vbInformation, "ФИНАНСОВЫЙ КОНТРОЛЬ — Сводные таблицы"
    Exit Sub
pivErr:
    Application.DisplayAlerts = True
    Application.ScreenUpdating = True
    MsgBox "Ошибка при создании сводных таблиц:" & vbCrLf & Err.Description, _
           vbCritical, "CreatePivotTables"
End Sub


' ═══════════════════════════════════════════════════════════════
'  КАК ИСПОЛЬЗОВАТЬ:
'  1) Открыть ФИНАНСОВЫЙ_КОНТРОЛЬ.xlsx в Excel
'  2) Alt+F11 → File → Import File → выбрать ФИНАНСОВЫЙ_КОНТРОЛЬ_VBA.bas
'  3) Alt+F8 → выбрать "SetupAll" → нажать "Выполнить"
'     Кнопки появятся на листах, макросы назначатся автоматически.
'  4) Файл → Сохранить как → Книга Excel с поддержкой макросов (.xlsm)
' ═══════════════════════════════════════════════════════════════
'''


def write_vba_file(path="ФИНАНСОВЫЙ_КОНТРОЛЬ_VBA.bas"):
    code = VBA_CODE
    # Replace characters not in cp1251 (box-drawing, etc.)
    # VBA IDE imports .bas files using system ANSI (cp1251 on Russian Windows)
    code = code.replace("\u2550", "=")   # ═
    code = code.replace("\u2551", "|")   # ║
    code = code.replace("\u2500", "-")   # ─
    code = code.replace("\u2502", "|")   # │
    code = code.replace("\u2192", "->")  # →
    code = code.replace("\u00B7", ".")   # ·
    # ₽ is handled via ChrW(8381) in the VBA source itself
    with open(path, "w", encoding="cp1251", errors="replace") as f:
        f.write(code)
    return path


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

SHEET_NAMES = [
    "Пульт", "Ввод_Касса", "Ввод_Расходы", "БАЗА_ДДС",
    "Запись_Выплат", "Календарь_Выплат", "Дашборд", "Настройки",
    "Отчёт_Рук", "Сводные",
]
TAB_COLORS = {
    "Пульт":            "0B4F54",
    "Ввод_Касса":       "059669",
    "Ввод_Расходы":     "DC2626",
    "БАЗА_ДДС":         "1D4ED8",
    "Запись_Выплат":    "7C3AED",
    "Календарь_Выплат": "D97706",
    "Дашборд":          "B45309",
    "Настройки":        "6B7280",
    "Отчёт_Рук":        "065F46",
    "Сводные":          "0E7490",
}


def build_otchet_rukovoditelya(ws):
    """ОТЧЁТ РУКОВОДИТЕЛЯ — ежемесячный управленческий отчёт."""
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation  = "landscape"
    ws.page_setup.paperSize    = 9      # A4
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.page_setup.fitToPage    = True
    ws.print_options.gridLines = False

    NC = 7   # number of columns used
    set_widths(ws, [
        ("A", 4), ("B", 30), ("C", 18), ("D", 12),
        ("E", 18), ("F", 14), ("G", 12),
    ])

    months_arr = ('{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";'
                  '"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"}')

    # ── Internal helpers ──────────────────────────────────────
    def _mb(color="111827"):
        return Border(
            left=Side(style="medium", color=color),
            right=Side(style="medium", color=color),
            top=Side(style="medium", color=color),
            bottom=Side(style="medium", color=color),
        )

    def _row_hdr(row, labels, bg, height=22):
        ws.row_dimensions[row].height = height
        for col, lbl in enumerate(labels, 1):
            c = ws.cell(row, col, lbl)
            c.fill = mkfill(bg)
            c.font = mkfont(WHITE, 9, True)
            c.alignment = mkalign("center" if col != 2 else "left", "center")
            c.border = mkborder()

    def _fill_row(row, alt=False, total=False, height=20):
        ws.row_dimensions[row].height = height
        bg = ("FFD1FAE5" if total else (GRAY_L if alt else WHITE))
        for col in range(1, NC + 1):
            ws.cell(row, col).fill = mkfill(bg)

    def _cell(row, col, val, fmt=None, aln="center", bold=False, bg=None, color=NAVY):
        c = ws.cell(row, col, val)
        if bg:
            c.fill = mkfill(bg)
        c.font = mkfont(color, 10, bold)
        c.alignment = mkalign(aln, "center")
        if fmt:
            c.number_format = fmt
        c.border = mkborder()
        return c

    # ── ROW 1: Title ──────────────────────────────────────────
    ws.row_dimensions[1].height = 44
    ws.merge_cells("A1:G1")
    c = ws.cell(1, 1, "  ЕЖЕМЕСЯЧНЫЙ ОТЧЁТ  —  ФИНАНСОВЫЙ КОНТРОЛЬ")
    c.fill = mkfill(NAVY); c.font = mkfont(WHITE, 17, True)
    c.alignment = mkalign("center", "center")
    for ci in range(2, NC + 1):
        ws.cell(1, ci).fill = mkfill(NAVY)

    # ── ROW 2: Subtitle ───────────────────────────────────────
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:G2")
    c = ws.cell(2, 1, "  Управленческий отчёт по движению денежных средств магазина")
    c.fill = mkfill(TEAL); c.font = mkfont(WHITE, 11)
    c.alignment = mkalign("left", "center")

    # ── ROW 3: Period controls ────────────────────────────────
    ws.row_dimensions[3].height = 36
    ctrl_bg = "FFFFFFFF"
    for col, (lbl, val, merge_end) in enumerate([
        ("Месяц:", MONTHS_RU[0], None),
        (None,     None,          None),     # B3 = month value
        ("Год:",   YEAR,         None),
        (None,     None,          None),     # D3 = year value
        ("Дата отчёта:", "=TEXT(TODAY(),\"DD MMMM YYYY\"&\" г.\")", "G3"),
    ], 1):
        pass  # done manually below

    # A3: label "Месяц:"
    c = ws.cell(3, 1, "Месяц:")
    c.fill = mkfill(NAVY); c.font = mkfont(WHITE, 10, True)
    c.alignment = mkalign("right", "center")
    # B3: month dropdown
    c = ws.cell(3, 2, MONTHS_RU[0])
    c.fill = mkfill(ctrl_bg); c.font = mkfont(NAVY, 12, True)
    c.alignment = mkalign("center", "center"); c.border = _mb()
    _add_dv(ws, '"' + ",".join(MONTHS_RU) + '"', "B3")
    # C3: label "Год:"
    c = ws.cell(3, 3, "Год:")
    c.fill = mkfill(NAVY); c.font = mkfont(WHITE, 10, True)
    c.alignment = mkalign("right", "center")
    # D3: year input
    c = ws.cell(3, 4, YEAR)
    c.fill = mkfill(ctrl_bg); c.font = mkfont(NAVY, 12, True)
    c.alignment = mkalign("center", "center"); c.border = _mb()
    # E3: label "Сформирован:"
    c = ws.cell(3, 5, "Сформирован:")
    c.fill = mkfill(NAVY); c.font = mkfont(WHITE, 10, True)
    c.alignment = mkalign("right", "center")
    # F3-G3: today's date
    ws.merge_cells("F3:G3")
    c = ws.cell(3, 6, '=TEXT(TODAY(),"DD MMMM YYYY")&" г."')
    c.fill = mkfill(ctrl_bg); c.font = mkfont(NAVY, 11, True)
    c.alignment = mkalign("center", "center"); c.border = _mb()
    ws.cell(3, 7).fill = mkfill(ctrl_bg)

    # ── ROW 4: Hidden date helpers ────────────────────────────
    ws.row_dimensions[4].height = 0
    ws.cell(4, 1, f'=DATE(D3,MATCH(B3,{months_arr},0),1)').number_format = FMT_DATE
    ws.cell(4, 2, '=EOMONTH($A$4,0)').number_format = FMT_DATE
    ws.cell(4, 3, '=DATE(YEAR($A$4),MONTH($A$4)-1,1)').number_format = FMT_DATE
    ws.cell(4, 4, '=$A$4-1').number_format = FMT_DATE
    for ci in range(1, NC + 1):
        ws.cell(4, ci).font = mkfont(WHITE, 1)

    P  = '$A$4'   # period start (current month)
    Q  = '$B$4'   # period end
    PP = '$C$4'   # previous month start
    QP = '$D$4'   # previous month end

    # Formula shorthands (return strings WITHOUT leading =)
    def rev(p, q):
        return (f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
                f'tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')

    def exp_(p, q, cat=''):
        base = (f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
                f'tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')
        if cat:
            base = (f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Расход",'
                    f'tblБаза[Категория],"{cat}",'
                    f'tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')
        return base

    def rev_pay(p, q, method):
        return (f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Приход",'
                f'tblБаза[Способ оплаты],"{method}",'
                f'tblБаза[Дата],">="&{p},tblБаза[Дата],"<="&{q})')

    r = 5
    ws.row_dimensions[r].height = 10
    r += 1

    # ── Generic row writer ────────────────────────────────────
    def data_row(num, label, cf, pf, pct_ref=None,
                 fmt=FMT_RUB, alt=False, bold=False, total=False, inv=False):
        """Write one data row. cf/pf = formula WITHOUT leading '='. """
        nonlocal r
        bg = "FFD1FAE5" if total else (GRAY_L if alt else WHITE)
        h  = 22 if total else 20
        ws.row_dimensions[r].height = h
        fnt_sz = 11 if total else 10

        def wr(col, val, n_fmt=None, aln="center"):
            c = ws.cell(r, col, val)
            c.fill = mkfill(bg)
            c.font = mkfont(NAVY, fnt_sz, bold or total)
            c.alignment = mkalign(aln, "center")
            if n_fmt:
                c.number_format = n_fmt
            c.border = mkborder()

        wr(1, str(num) if num else "")
        wr(2, label, aln="left")
        if cf:
            wr(3, f'={cf}', fmt)
        else:
            wr(3, "")

        if pct_ref and cf:
            wr(4, f'=IFERROR(C{r}/{pct_ref},0)', "0.0%")
        elif cf and total:
            wr(4, '=IFERROR(C{0}/C{0},0)'.format(r), "0.0%")
        else:
            wr(4, "")

        if pf:
            wr(5, f'={pf}', fmt)
        else:
            wr(5, "")

        if cf and pf:
            if inv:
                wr(6, f'=C{r}-E{r}',
                   '[Red]"▲ "#,##0;[Green]"▼ "#,##0;"-"')
                wr(7, f'=IFERROR((C{r}-E{r})/ABS(E{r}),0)',
                   '[Red]"+0.0%";[Green]"-0.0%";"—"')
            else:
                wr(6, f'=C{r}-E{r}',
                   '[Green]"▲ "#,##0;[Red]"▼ "#,##0;"-"')
                wr(7, f'=IFERROR((C{r}-E{r})/ABS(E{r}),0)',
                   '[Green]"+0.0%";[Red]"-0.0%";"—"')
        else:
            wr(6, ""); wr(7, "")

        row_ref = r
        r += 1
        return row_ref

    def spacer(h=8):
        nonlocal r
        ws.row_dimensions[r].height = h
        for ci in range(1, NC + 1):
            ws.cell(r, ci).fill = mkfill(WHITE)
        r += 1

    def sec_title(text, bg=NAVY):
        nonlocal r
        sec_hdr(ws, r, text, bg=bg, ncols=NC, height=24)
        r += 1

    def tbl_hdr(labels, bg=TEAL_M):
        nonlocal r
        _row_hdr(r, labels, bg)
        r += 1

    # ════════════════════════════════════════════════════════
    #  TABLE 1 — ФИНАНСОВАЯ СВОДКА
    # ════════════════════════════════════════════════════════
    sec_title("  1. ФИНАНСОВАЯ СВОДКА", GREEN)
    tbl_hdr(["№", "Показатель", "Текущий месяц", "% выручки",
             "Прошлый месяц", "Δ (Откл.)", "Δ %"])

    r_rev = data_row(1, "Выручка (все поступления)",
                     rev(P, Q), rev(PP, QP), pct_ref=None)
    rev_ref = f'$C${r_rev}'

    data_row(2, "Расходы всего",
             exp_(P, Q), exp_(PP, QP), pct_ref=rev_ref, inv=True)
    data_row(3, "в т.ч. Закуп товара",
             exp_(P, Q, 'Закуп товара'), exp_(PP, QP, 'Закуп товара'),
             pct_ref=rev_ref, inv=True)
    data_row(4, "в т.ч. ЗП сотрудников",
             exp_(P, Q, 'ЗП'), exp_(PP, QP, 'ЗП'),
             pct_ref=rev_ref, inv=True)
    data_row(5, "в т.ч. Аренда",
             exp_(P, Q, 'Аренда'), exp_(PP, QP, 'Аренда'),
             pct_ref=rev_ref, inv=True)
    data_row(6, "Прочие расходы",
             f'{exp_(P,Q)}-{exp_(P,Q,"Закуп товара")}-{exp_(P,Q,"ЗП")}-{exp_(P,Q,"Аренда")}',
             f'{exp_(PP,QP)}-{exp_(PP,QP,"Закуп товара")}-{exp_(PP,QP,"ЗП")}-{exp_(PP,QP,"Аренда")}',
             pct_ref=rev_ref, inv=True)

    r_profit = data_row(7, "Чистая прибыль",
                        f'{rev(P,Q)}-{exp_(P,Q)}',
                        f'{rev(PP,QP)}-{exp_(PP,QP)}',
                        pct_ref=rev_ref)
    data_row(8, "Рентабельность %",
             f'IFERROR(({rev(P,Q)}-{exp_(P,Q)})/MAX(1,{rev(P,Q)}),0)',
             f'IFERROR(({rev(PP,QP)}-{exp_(PP,QP)})/MAX(1,{rev(PP,QP)}),0)',
             pct_ref=None, fmt="0.0%")
    data_row(9, "Маржа (валовая) %",
             f'IFERROR(({rev(P,Q)}-{exp_(P,Q,"Закуп товара")})/MAX(1,{rev(P,Q)}),0)',
             f'IFERROR(({rev(PP,QP)}-{exp_(PP,QP,"Закуп товара")})/MAX(1,{rev(PP,QP)}),0)',
             pct_ref=None, fmt="0.0%")

    spacer()

    # ════════════════════════════════════════════════════════
    #  TABLE 2 — ВЫРУЧКА ПО СПОСОБАМ ОПЛАТЫ
    # ════════════════════════════════════════════════════════
    sec_title("  2. ВЫРУЧКА ПО СПОСОБАМ ОПЛАТЫ", BLUE)
    tbl_hdr(["№", "Способ оплаты", "Текущий месяц", "% выручки",
             "Прошлый месяц", "Δ (Откл.)", "Δ %"])

    for i, method in enumerate(["Наличные", "Карта", "Перевод"], 1):
        data_row(i, method, rev_pay(P, Q, method), rev_pay(PP, QP, method),
                 pct_ref=rev_ref, alt=(i % 2 == 0))

    # Total row
    tot_r = r
    data_row("", "ИТОГО", rev(P, Q), rev(PP, QP), pct_ref=None,
             bold=True, total=True)
    # Fix % for total: should be 100%
    ws.cell(tot_r, 4, "=100.0%").number_format = "0.0%"
    ws.cell(tot_r, 4).font = mkfont(NAVY, 11, True)
    ws.cell(tot_r, 4).fill = mkfill("FFD1FAE5")
    ws.cell(tot_r, 4).alignment = mkalign("center", "center")
    ws.cell(tot_r, 4).border = mkborder()

    spacer()

    # ════════════════════════════════════════════════════════
    #  TABLE 3 — РАСХОДЫ ПО КАТЕГОРИЯМ
    # ════════════════════════════════════════════════════════
    sec_title("  3. РАСХОДЫ ПО КАТЕГОРИЯМ", RED)
    tbl_hdr(["№", "Категория расходов", "Текущий месяц", "% расходов",
             "Прошлый месяц", "Δ (Откл.)", "Δ %"])

    # Need a reference to total expenses cell for this table's % column
    exp_ref_row = r   # we'll write total at the end; need to forward-reference
    # Write a hidden total cell for % reference — use col 8 (off-screen)
    ws.cell(r, 8, f'={exp_(P, Q)}')
    ws.cell(r, 8).font = mkfont(WHITE, 1)
    exp_hidden_ref = f'$H${r}'

    for i, cat in enumerate(CATS_EXPENSE, 1):
        data_row(i, cat, exp_(P, Q, cat), exp_(PP, QP, cat),
                 pct_ref=exp_hidden_ref, alt=(i % 2 == 0), inv=True)

    r_exp_tot = r
    data_row("", "ИТОГО РАСХОДОВ", exp_(P, Q), exp_(PP, QP),
             pct_ref=None, bold=True, total=True, inv=True)
    ws.cell(r_exp_tot, 4, "=100.0%").number_format = "0.0%"
    ws.cell(r_exp_tot, 4).font = mkfont(NAVY, 11, True)
    ws.cell(r_exp_tot, 4).fill = mkfill("FFD1FAE5")
    ws.cell(r_exp_tot, 4).alignment = mkalign("center", "center")
    ws.cell(r_exp_tot, 4).border = mkborder()

    spacer()

    # ════════════════════════════════════════════════════════
    #  TABLE 4 — ОПЕРАЦИОННАЯ СТАТИСТИКА
    # ════════════════════════════════════════════════════════
    sec_title("  4. ОПЕРАЦИОННАЯ СТАТИСТИКА", TEAL)
    tbl_hdr(["№", "Показатель", "Текущий месяц", "—",
             "Прошлый месяц", "Δ (Откл.)", "—"],
            bg=TEAL_M)

    data_row(1, "Дней в периоде (кал.)", f'{Q}-{P}+1', f'{QP}-{PP}+1',
             fmt='0 "дн."')
    data_row(2, "Дней с данными",
             f'IFERROR(SUMPRODUCT((tblБаза[Дата]>={P})*(tblБаза[Дата]<={Q})*(tblБаза[Тип]="Приход"))/3,0)',
             f'IFERROR(SUMPRODUCT((tblБаза[Дата]>={PP})*(tblБаза[Дата]<={QP})*(tblБаза[Тип]="Приход"))/3,0)',
             fmt='0 "дн."')
    data_row(3, "Смен проведено",
             f'IFERROR(COUNTIFS(tblБаза[Тип],"Приход",tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})/3,0)',
             f'IFERROR(COUNTIFS(tblБаза[Тип],"Приход",tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})/3,0)',
             fmt='0 "см."')
    data_row(4, "Средняя выручка в день",
             f'IFERROR({rev(P,Q)}/MAX(1,{Q}-{P}+1),0)',
             f'IFERROR({rev(PP,QP)}/MAX(1,{QP}-{PP}+1),0)')
    data_row(5, "Средняя выручка за смену",
             f'IFERROR({rev(P,Q)}/MAX(1,COUNTIFS(tblБаза[Тип],"Приход",tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})/3),0)',
             f'IFERROR({rev(PP,QP)}/MAX(1,COUNTIFS(tblБаза[Тип],"Приход",tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})/3),0)')
    data_row(6, "Расхождение касс (сумма)",
             f'SUMIFS(tblБаза[Расхождение],tblБаза[Тип],"Приход",tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
             f'SUMIFS(tblБаза[Расхождение],tblБаза[Тип],"Приход",tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})',
             inv=True)
    data_row(7, "Кол-во расхождений",
             f'COUNTIFS(tblБаза[Расхождение],"<>0",tblБаза[Тип],"Приход",tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
             f'COUNTIFS(tblБаза[Расхождение],"<>0",tblБаза[Тип],"Приход",tblБаза[Дата],">="&{PP},tblБаза[Дата],"<="&{QP})',
             fmt="0", inv=True)

    spacer()

    # ════════════════════════════════════════════════════════
    #  TABLE 5 — ОБЯЗАТЕЛЬСТВА И ВЫПЛАТЫ
    # ════════════════════════════════════════════════════════
    sec_title("  5. ОБЯЗАТЕЛЬСТВА И ВЫПЛАТЫ ПОСТАВЩИКАМ", PURPLE)
    tbl_hdr(["№", "Статус / Показатель", "Сумма", "% итого",
             "—", "Кол-во", "—"])

    # Total for % reference — hidden in col 8
    r_vt = r
    ws.cell(r_vt, 8, '=SUMIFS(tblВыплаты[Сумма],tblВыплаты[Сумма],">0")')
    ws.cell(r_vt, 8).font = mkfont(WHITE, 1)
    vy_ref = f'$H${r_vt}'

    data_row(1, "Запланировано (в периоде)",
             f'SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
             f'tblВыплаты[Дата плановой оплаты],">="&{P},'
             f'tblВыплаты[Дата плановой оплаты],"<="&{Q})',
             None, pct_ref=vy_ref)
    ws.cell(r - 1, 6,
            f'=COUNTIFS(tblВыплаты[Статус],"Запланировано",'
            f'tblВыплаты[Дата плановой оплаты],">="&{P},'
            f'tblВыплаты[Дата плановой оплаты],"<="&{Q})').number_format = '0 "шт."'
    ws.cell(r - 1, 6).fill = mkfill(WHITE)
    ws.cell(r - 1, 6).font = mkfont(NAVY, 10)
    ws.cell(r - 1, 6).alignment = mkalign("center", "center")
    ws.cell(r - 1, 6).border = mkborder()

    data_row(2, "Просрочено (не оплачено, дата прошла)",
             'SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Запланировано",'
             'tblВыплаты[Дата плановой оплаты],"<"&TODAY())',
             None, pct_ref=vy_ref, inv=True, alt=True)
    ws.cell(r - 1, 6,
            '=COUNTIFS(tblВыплаты[Статус],"Запланировано",'
            'tblВыплаты[Дата плановой оплаты],"<"&TODAY())').number_format = '0 "шт."'
    ws.cell(r - 1, 6).fill = mkfill(GRAY_L)
    ws.cell(r - 1, 6).font = mkfont(RED[2:], 10, True)
    ws.cell(r - 1, 6).alignment = mkalign("center", "center")
    ws.cell(r - 1, 6).border = mkborder()

    data_row(3, "Оплачено (всего за период)",
             f'SUMIFS(tblВыплаты[Сумма],tblВыплаты[Статус],"Оплачено")',
             None, pct_ref=vy_ref)
    ws.cell(r - 1, 6,
            '=COUNTIFS(tblВыплаты[Статус],"Оплачено")').number_format = '0 "шт."'
    ws.cell(r - 1, 6).fill = mkfill(WHITE)
    ws.cell(r - 1, 6).font = mkfont(GREEN[2:], 10, True)
    ws.cell(r - 1, 6).alignment = mkalign("center", "center")
    ws.cell(r - 1, 6).border = mkborder()

    r_vy_tot = r
    data_row("", "ВСЕГО ОБЯЗАТЕЛЬСТВ",
             'SUMIFS(tblВыплаты[Сумма],tblВыплаты[Сумма],">0")',
             None, bold=True, total=True)
    ws.cell(r_vy_tot, 4, "=100.0%").number_format = "0.0%"
    ws.cell(r_vy_tot, 4).font = mkfont(NAVY, 11, True)
    ws.cell(r_vy_tot, 4).fill = mkfill("FFD1FAE5")
    ws.cell(r_vy_tot, 4).alignment = mkalign("center", "center")
    ws.cell(r_vy_tot, 4).border = mkborder()

    spacer()

    # ════════════════════════════════════════════════════════
    #  TABLE 6 — ДВИЖЕНИЕ ДОЛГА
    # ════════════════════════════════════════════════════════
    sec_title("  6. ДВИЖЕНИЕ ДОЛГА В ПЕРИОДЕ", AMBER)
    tbl_hdr(["№", "Показатель", "Сумма", "—",
             "Справка: всего", "—", "—"])

    debt_all = ('SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг")'
                '-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП")')
    data_row(1, "Долг на начало периода (остаток)",
             f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
             f'tblБаза[Дата],"<"&{P})'
             f'-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП",'
             f'tblБаза[Дата],"<"&{P})',
             None)
    data_row(2, "Взято в долг за период",
             f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
             f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
             None, inv=True, alt=True)
    data_row(3, "Выплачено долгов за период",
             f'SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП",'
             f'tblБаза[Дата],">="&{P},tblБаза[Дата],"<="&{Q})',
             None)
    data_row(4, "Долг на конец периода (остаток)",
             f'SUMIFS(tblБаза[Сумма],tblБаза[Тип],"Долг",'
             f'tblБаза[Дата],"<="&{Q})'
             f'-SUMIFS(tblБаза[Сумма],tblБаза[Категория],"Оплата ТП",'
             f'tblБаза[Дата],"<="&{Q})',
             None, inv=True, alt=True, bold=True)
    data_row("", "Итого долг (все время)",
             debt_all, None, total=True, bold=True)

    spacer(h=14)

    # ════════════════════════════════════════════════════════
    #  SIGNATURE BLOCK
    # ════════════════════════════════════════════════════════
    def sig_row(col_l, lbl_l, col_r=None, lbl_r=None):
        nonlocal r
        ws.row_dimensions[r].height = 28
        # Left signature
        ws.merge_cells(start_row=r, start_column=col_l,
                       end_row=r, end_column=col_l + 2)
        c = ws.cell(r, col_l, lbl_l)
        c.font = mkfont(NAVY, 10, True)
        c.alignment = mkalign("left", "center")
        # underline for signature
        ws.merge_cells(start_row=r, start_column=col_l + 3,
                       end_row=r, end_column=col_l + 3)
        c2 = ws.cell(r, col_l + 3, "________________________")
        c2.font = mkfont(GRAY_D, 9)
        c2.alignment = mkalign("center", "bottom")
        if col_r and lbl_r:
            ws.merge_cells(start_row=r, start_column=col_r,
                           end_row=r, end_column=col_r + 2)
            c3 = ws.cell(r, col_r, lbl_r)
            c3.font = mkfont(NAVY, 10, True)
            c3.alignment = mkalign("left", "center")
            ws.merge_cells(start_row=r, start_column=col_r + 3,
                           end_row=r, end_column=col_r + 3)
            c4 = ws.cell(r, col_r + 3, "________________________")
            c4.font = mkfont(GRAY_D, 9)
            c4.alignment = mkalign("center", "bottom")
        r += 1

    # Separator line before signatures
    ws.merge_cells(f"A{r}:G{r}")
    ws.row_dimensions[r].height = 2
    ws.cell(r, 1).fill = mkfill(GRAY_M)
    r += 1
    ws.row_dimensions[r].height = 8
    r += 1

    sig_row(1, "Составил:")
    sig_row(1, "Принял:")

    ws.row_dimensions[r].height = 6
    r += 1

    # Note
    ws.merge_cells(f"A{r}:G{r}")
    c = ws.cell(r, 1, "  * Отчёт сформирован автоматически на основе данных БАЗА_ДДС")
    c.font = mkfont(GRAY_D, 8)
    c.alignment = mkalign("left", "center")

    ws.freeze_panes = "A5"


def build_svodnye(ws):
    """Сводные — landing page; VBA CreatePivotTables builds real Excel PivotTables."""
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation  = "landscape"
    ws.page_setup.paperSize    = 9
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.page_setup.fitToPage    = True
    ws.print_options.gridLines = False

    NC = 12
    set_widths(ws, [
        ("A",  4), ("B", 24), ("C", 18), ("D", 18), ("E", 18),
        ("F", 18), ("G", 18), ("H", 18), ("I", 18), ("J", 18),
        ("K", 18), ("L", 18),
    ])

    sheet_title(ws, "  СВОДНЫЕ ТАБЛИЦЫ",
                "  Нажмите кнопку ниже — VBA построит 5 реальных Excel PivotTables",
                ncols=NC)

    # Row 3: big button (replaced by DrawingML shape via inject_button_shapes)
    ws.row_dimensions[3].height = 48
    _form_btn(ws, 3, 1, "СОЗДАТЬ СВОДНЫЕ ТАБЛИЦЫ", bg=TEAL, ncols=NC, size=16)

    ws.row_dimensions[4].height = 10

    # Instructions block
    sec_hdr(ws, 5, "  КАК СОЗДАТЬ СВОДНЫЕ ТАБЛИЦЫ", bg=NAVY, ncols=NC, height=26)
    steps = [
        "1.  Нажмите кнопку «СОЗДАТЬ СВОДНЫЕ ТАБЛИЦЫ» выше (или Alt+F8 → CreatePivotTables)",
        "2.  VBA автоматически создаст 5 настоящих Excel PivotTables из таблицы БАЗА_ДДС",
        "3.  Для обновления данных: нажмите кнопку снова (старые таблицы будут заменены)",
        "4.  После создания — перетаскивайте поля прямо в PivotTable для своего анализа",
        "5.  Для фильтрации по году/месяцу — используйте встроенные фильтры PivotTable",
    ]
    for i, text in enumerate(steps):
        rn = 6 + i
        ws.row_dimensions[rn].height = 22
        alt = (i % 2 == 0)
        c0 = ws.cell(rn, 1)
        c0.fill = mkfill(GRAY_L if alt else WHITE); c0.border = mkborder()
        ws.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=NC)
        c = ws.cell(rn, 2, text)
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10)
        c.alignment = mkalign("left", "center")
        c.border = mkborder()

    ws.row_dimensions[11].height = 14

    # Table of 5 PivotTables to be created
    sec_hdr(ws, 12, "  ЧТО БУДЕТ СОЗДАНО", bg=TEAL_M, ncols=NC, height=24)
    ws.row_dimensions[13].height = 26
    for ci, h in enumerate(["", "№", "Название", "", "Описание"], 1):
        c = ws.cell(13, ci, h)
        c.fill = mkfill(TEAL); c.font = mkfont(WHITE, 10, True)
        c.alignment = mkalign("center" if ci != 4 else "left", "center")
        c.border = mkborder()
    ws.merge_cells("D13:L13")

    pt_list = [
        ("ПТ1", "Выручка по месяцам",
         "Фильтр: Тип=Приход | Строки: Дата (по месяцам) | Значения: Сумма"),
        ("ПТ2", "Расходы по категориям",
         "Фильтр: Тип=Расход | Строки: Категория | Значения: Сумма"),
        ("ПТ3", "Выручка: кассиры × смены",
         "Фильтр: Тип=Приход | Строки: Кассир | Колонки: Смена | Значения: Сумма"),
        ("ПТ4", "Долги по поставщикам",
         "Фильтр: Тип=Долг | Строки: Поставщик (Комментарий) | Значения: Сумма"),
        ("ПТ5", "Итоговая сводная",
         "Строки: Тип + Категория | Колонки: Способ оплаты | Значения: Сумма"),
    ]
    for i, (num, name, desc) in enumerate(pt_list):
        rn = 14 + i
        ws.row_dimensions[rn].height = 22
        alt = (i % 2 == 1)
        d_cell(ws, rn, 1, "", alt)
        d_cell(ws, rn, 2, num, alt, "center", bold=True)
        d_cell(ws, rn, 3, name, alt, "left", bold=True)
        ws.merge_cells(start_row=rn, start_column=4, end_row=rn, end_column=NC)
        d_cell(ws, rn, 4, desc, alt, "left")

    ws.row_dimensions[19].height = 10

    # Note about import
    sec_hdr(ws, 20, "  ТРЕБОВАНИЯ", bg=AMBER, ncols=NC, height=22)
    notes = [
        "Файл должен быть сохранён как .xlsm (Книга Excel с поддержкой макросов)",
        "VBA макрос FinKontrolMacros.bas должен быть импортирован (Alt+F11 → Import File)",
    ]
    for i, note in enumerate(notes):
        rn = 21 + i
        ws.row_dimensions[rn].height = 22
        alt = (i % 2 == 0)
        c0 = ws.cell(rn, 1)
        c0.fill = mkfill(GRAY_L if alt else WHITE); c0.border = mkborder()
        ws.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=NC)
        c = ws.cell(rn, 2, f"  {note}")
        c.fill = mkfill(GRAY_L if alt else WHITE)
        c.font = mkfont(NAVY, 10)
        c.alignment = mkalign("left", "center")
        c.border = mkborder()

    ws.freeze_panes = "B4"



def inject_button_shapes(xlsx_path):
    """
    Post-process the saved XLSX: embed DrawingML shapes as real clickable buttons.

    Shapes carry macro= attribute. After the user imports ФИНАНСОВЫЙ_КОНТРОЛЬ_VBA.bas
    and saves as .xlsm, clicking a shape triggers the assigned macro directly —
    no SetupAll needed.

    Positions (0-indexed col/row, exclusive to-boundary):
      Ввод_Касса  : СОХРАНИТЬ A17:G17 → (0,16)→(7,17)
                    СЕГОДНЯ   E4:F4   → (4,3)→(6,4)
      Ввод_Расходы: СОХРАНИТЬ A16:D16 → (0,15)→(4,16)
                    СЕГОДНЯ   C3:D3   → (2,2)→(4,3)
      Календарь   : СЕГОДНЯ   I3:J3   → (8,2)→(10,3)
      Дашборд     : ОБНОВИТЬ  K3:L3   → (10,2)→(12,3)
      Сводные     : СОЗДАТЬ   A3:L3   → (0,2)→(12,3)
    """
    import zipfile, re, os

    BUTTONS = {
        "Ввод_Касса": [
            {"id": 1, "name": "Btn_SaveKassa",
             "caption": "СОХРАНИТЬ КАССУ",
             "macro": "FinKontrolMacros.SaveKassa",
             "fc": 0, "fr": 16, "tc": 7, "tr": 17, "color": "059669", "sz": 1400},
            {"id": 2, "name": "Btn_TodayKassa",
             "caption": "СЕГОДНЯ",
             "macro": "FinKontrolMacros.InsertToday_Kassa",
             "fc": 4, "fr": 3, "tc": 6, "tr": 4, "color": "1D4ED8", "sz": 1100},
        ],
        "Ввод_Расходы": [
            {"id": 1, "name": "Btn_SaveRashod",
             "caption": "СОХРАНИТЬ",
             "macro": "FinKontrolMacros.SaveRashod",
             "fc": 0, "fr": 15, "tc": 4, "tr": 16, "color": "059669", "sz": 1400},
            {"id": 2, "name": "Btn_TodayRashod",
             "caption": "СЕГОДНЯ",
             "macro": "FinKontrolMacros.InsertToday_Rashod",
             "fc": 2, "fr": 2, "tc": 4, "tr": 3, "color": "1D4ED8", "sz": 1100},
        ],
        "Календарь_Выплат": [
            {"id": 1, "name": "Btn_TodayCal",
             "caption": "СЕГОДНЯ",
             "macro": "FinKontrolMacros.InsertToday_Calendar",
             "fc": 8, "fr": 2, "tc": 10, "tr": 3, "color": "1D4ED8", "sz": 1100},
        ],
        "Дашборд": [
            {"id": 1, "name": "Btn_Refresh",
             "caption": "ОБНОВИТЬ",
             "macro": "FinKontrolMacros.RefreshDashboard",
             "fc": 10, "fr": 2, "tc": 12, "tr": 3, "color": "D97706", "sz": 1100},
        ],
        "Сводные": [
            {"id": 1, "name": "Btn_CreatePivot",
             "caption": "СОЗДАТЬ СВОДНЫЕ ТАБЛИЦЫ",
             "macro": "FinKontrolMacros.CreatePivotTables",
             "fc": 0, "fr": 2, "tc": 12, "tr": 3, "color": "0E7490", "sz": 1400},
        ],
    }

    def make_anchor(b):
        return (
            f'<xdr:twoCellAnchor editAs="oneCell">'
            f'<xdr:from><xdr:col>{b["fc"]}</xdr:col><xdr:colOff>0</xdr:colOff>'
            f'<xdr:row>{b["fr"]}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
            f'<xdr:to><xdr:col>{b["tc"]}</xdr:col><xdr:colOff>0</xdr:colOff>'
            f'<xdr:row>{b["tr"]}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
            f'<xdr:sp macro="{b["macro"]}" textlink="">'
            f'<xdr:nvSpPr>'
            f'<xdr:cNvPr id="{b["id"]}" name="{b["name"]}"/>'
            f'<xdr:cNvSpPr><a:spLocks noTextEdit="0"/></xdr:cNvSpPr>'
            f'</xdr:nvSpPr>'
            f'<xdr:spPr>'
            f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
            f'<a:prstGeom prst="roundRect">'
            f'<a:avLst><a:gd name="adj" fmla="val 10000"/></a:avLst>'
            f'</a:prstGeom>'
            f'<a:solidFill><a:srgbClr val="{b["color"]}"/></a:solidFill>'
            f'<a:ln><a:noFill/></a:ln>'
            f'</xdr:spPr>'
            f'<xdr:txBody>'
            f'<a:bodyPr anchor="ctr"/><a:lstStyle/>'
            f'<a:p><a:pPr algn="ctr"/>'
            f'<a:r>'
            f'<a:rPr lang="ru-RU" sz="{b["sz"]}" b="1" dirty="0">'
            f'<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill>'
            f'<a:latin typeface="Calibri"/>'
            f'</a:rPr>'
            f'<a:t>{b["caption"]}</a:t>'
            f'</a:r></a:p>'
            f'</xdr:txBody>'
            f'</xdr:sp>'
            f'<xdr:clientData/>'
            f'</xdr:twoCellAnchor>'
        )

    def make_drawing_xml(btn_list):
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<xdr:wsDr '
            'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
            'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            + ''.join(make_anchor(b) for b in btn_list)
            + '</xdr:wsDr>'
        )

    DRAWING_NS  = ("http://schemas.openxmlformats.org/officeDocument/"
                   "2006/relationships/drawing")
    DRAWING_CT  = ("application/vnd.openxmlformats-officedocument.drawing+xml")

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        files = {n: zin.read(n) for n in zin.namelist()}

    wb_xml  = files['xl/workbook.xml'].decode('utf-8')
    wb_rels = files['xl/_rels/workbook.xml.rels'].decode('utf-8')

    # Parse each <Relationship> element individually (attribute order varies)
    rid2tgt = {}
    for m in re.finditer(r'<Relationship\b([^>]+)>', wb_rels):
        attrs = m.group(1)
        id_m  = re.search(r'Id="([^"]+)"', attrs)
        tgt_m = re.search(r'Target="([^"]+)"', attrs)
        if id_m and tgt_m:
            tgt = tgt_m.group(1).lstrip('/')   # strip leading / on absolute paths
            rid2tgt[id_m.group(1)] = tgt

    sheet2file = {}
    for m in re.finditer(r'<sheet\b[^>]+name="([^"]+)"[^>]+r:id="([^"]+)"', wb_xml):
        tgt = rid2tgt.get(m.group(2), '')
        if tgt:
            sheet2file[m.group(1)] = tgt

    existing_drw = [n for n in files
                    if re.match(r'xl/drawings/drawing\d+\.xml$', n)]
    drw_idx = max((int(re.search(r'\d+', os.path.basename(n)).group())
                   for n in existing_drw), default=0) + 1

    ct_xml = files['[Content_Types].xml'].decode('utf-8')

    XDR_NS = ('xmlns:xdr="http://schemas.openxmlformats.org/'
               'drawingml/2006/spreadsheetDrawing"')
    R_NS = ('xmlns:r="http://schemas.openxmlformats.org/'
             'officeDocument/2006/relationships"')

    for sheet_name, btn_list in BUTTONS.items():
        if sheet_name not in sheet2file:
            continue
        sheet_file = sheet2file[sheet_name]
        sheet_base = os.path.basename(sheet_file).replace('.xml', '')
        rels_path  = f'xl/worksheets/_rels/{sheet_base}.xml.rels'

        ws_xml = files[sheet_file].decode('utf-8')
        if R_NS not in ws_xml:
            ws_xml = ws_xml.replace('<worksheet ', '<worksheet ' + R_NS + ' ', 1)

        # If the worksheet already has a <drawing> element, merge button shapes
        # into that existing drawing file (a worksheet can only have one <drawing>).
        existing_m = re.search(r'<drawing\b[^>]+r:id="([^"]+)"', ws_xml)
        if existing_m and rels_path in files:
            existing_rid = existing_m.group(1)
            rels_xml = files[rels_path].decode('utf-8')
            # Find the <Relationship> element with the matching Id (attribute order varies)
            tgt_m = None
            for elem_m in re.finditer(r'<Relationship\b[^>]*/>', rels_xml):
                elem_str = elem_m.group(0)
                if f'Id="{existing_rid}"' in elem_str:
                    tgt_m = re.search(r'Target="([^"]+)"', elem_str)
                    break
            if tgt_m:
                tgt_rel = tgt_m.group(1)
                drw_merge = ('xl/drawings/' + os.path.basename(tgt_rel)
                             if not tgt_rel.startswith('/') else tgt_rel.lstrip('/'))
                if drw_merge in files:
                    drw_xml = files[drw_merge].decode('utf-8')
                    # Ensure xdr: prefix is declared so our shapes parse correctly
                    if XDR_NS not in drw_xml:
                        drw_xml = re.sub(r'(<\w*:?wsDr\b)', r'\1 ' + XDR_NS, drw_xml, count=1)
                    shapes = ''.join(make_anchor(b) for b in btn_list)
                    drw_xml = re.sub(r'</(\w*:?wsDr)>', shapes + r'</\1>', drw_xml, count=1)
                    files[drw_merge] = drw_xml.encode('utf-8')
                    files[sheet_file] = ws_xml.encode('utf-8')
                    continue  # no new drawing file needed

        # No existing <drawing> — create a new drawing file and wire it up
        drw_file   = f'xl/drawings/drawing{drw_idx}.xml'
        drw_rel_id = f'rId_drw{drw_idx}'
        drw_target = f'../drawings/drawing{drw_idx}.xml'

        files[drw_file] = make_drawing_xml(btn_list).encode('utf-8')

        part = '/' + drw_file
        if part not in ct_xml:
            ct_xml = ct_xml.replace(
                '</Types>',
                f'<Override PartName="{part}" ContentType="{DRAWING_CT}"/></Types>')

        rel_entry  = (f'<Relationship Id="{drw_rel_id}" '
                      f'Type="{DRAWING_NS}" Target="{drw_target}"/>')
        if rels_path in files:
            rels_xml = files[rels_path].decode('utf-8')
            rels_xml = rels_xml.replace('</Relationships>',
                                        rel_entry + '</Relationships>')
        else:
            rels_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/'
                'package/2006/relationships">' + rel_entry + '</Relationships>')
        files[rels_path] = rels_xml.encode('utf-8')

        drw_elem = f'<drawing r:id="{drw_rel_id}"/>'
        if '<tableParts' in ws_xml:
            ws_xml = ws_xml.replace('<tableParts', drw_elem + '<tableParts', 1)
        else:
            ws_xml = ws_xml.replace('</worksheet>', drw_elem + '</worksheet>', 1)
        files[sheet_file] = ws_xml.encode('utf-8')

        drw_idx += 1

    files['[Content_Types].xml'] = ct_xml.encode('utf-8')

    tmp = xlsx_path + '.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    os.replace(tmp, xlsx_path)


def main():
    print("Генерируем демо-данные...")
    baza = gen_baza()
    vyplaty = gen_vyplaty()
    print(f"  БАЗА_ДДС:       {len(baza)} строк")
    print(f"  ЗАПИСЬ_ВЫПЛАТ:  {len(vyplaty)} строк")

    print("\nСоздаём книгу...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheets = {}
    for name in SHEET_NAMES:
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = TAB_COLORS[name]
        sheets[name] = ws

    print("Строим БАЗА_ДДС...")
    build_baza(sheets["БАЗА_ДДС"], baza)

    print("Строим ВВОД_КАССА...")
    build_vvod_kassa(sheets["Ввод_Касса"])

    print("Строим ВВОД_РАСХОДЫ...")
    build_vvod_rashody(sheets["Ввод_Расходы"])

    print("Строим ЗАПИСЬ_ВЫПЛАТ...")
    build_zapis_vyplat(sheets["Запись_Выплат"], vyplaty)

    print("Строим НАСТРОЙКИ...")
    build_nastroyki(sheets["Настройки"])

    print("Строим ПУЛЬТ...")
    build_pult(sheets["Пульт"])

    print("Строим КАЛЕНДАРЬ_ВЫПЛАТ...")
    build_calendar(sheets["Календарь_Выплат"])

    print("Строим ДАШБОРД...")
    build_dashboard(sheets["Дашборд"])
    build_dashboard_charts(sheets["Дашборд"])

    print("Строим ОТЧЁТ РУКОВОДИТЕЛЯ...")
    build_otchet_rukovoditelya(sheets["Отчёт_Рук"])

    print("Строим СВОДНЫЕ ТАБЛИЦЫ...")
    build_svodnye(sheets["Сводные"])

    # Stubs for remaining sheets
    for name in SHEET_NAMES:
        if name in ("БАЗА_ДДС", "Ввод_Касса", "Ввод_Расходы",
                    "Запись_Выплат", "Настройки",
                    "Пульт", "Календарь_Выплат", "Дашборд", "Отчёт_Рук", "Сводные"):
            continue
        ws = sheets[name]
        ws.merge_cells("A1:F1")
        c = ws.cell(1, 1, f"[ {name} ]  —  будет добавлен в следующих блоках")
        c.font = mkfont(GRAY_D, 12, True)
        c.alignment = mkalign("center", "center")
        ws.row_dimensions[1].height = 36

    # Force full recalculation on file open — avoids Excel's
    # "removed and restored" dialog caused by missing calcChain
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"

    out = "ФИНАНСОВЫЙ_КОНТРОЛЬ.xlsx"
    wb.save(out)
    inject_button_shapes(out)
    sz = os.path.getsize(out) // 1024
    print(f"\n✓ Сохранено: {out}  ({sz} KB)")
    print(f"  Строк данных в БАЗА_ДДС: {len(baza)}")

    vba_path = write_vba_file("ФИНАНСОВЫЙ_КОНТРОЛЬ_VBA.bas")
    vsz = os.path.getsize(vba_path) // 1024
    print(f"✓ Сохранено: {vba_path}  ({vsz} KB)")

    print("""
==============================================================
  ИНСТРУКЦИЯ
==============================================================
  1. Откройте ФИНАНСОВЫЙ_КОНТРОЛЬ.xlsx в Excel
     (кнопки уже видны как зелёные/синие/жёлтые фигуры)
  2. Alt+F11 -> File -> Import File -> ФИНАНСОВЫЙ_КОНТРОЛЬ_VBA.bas
  3. Файл -> Сохранить как -> .xlsm
     (кнопки станут кликабельными)
==============================================================""")


if __name__ == "__main__":
    main()
