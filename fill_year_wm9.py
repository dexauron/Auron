#!/usr/bin/env python3
"""
Fill WAY_MARKET_v9.xlsx with realistic full-year data for 2025.
Measures write time and file size. Reports any capacity overflows.
"""
import time, sys, os, subprocess, random
from datetime import date, timedelta
from openpyxl import load_workbook

random.seed(42)

XLSX = "/home/user/Auron/WAY_MARKET_v9.xlsx"

# ── Step 1: rebuild fresh template ────────────────────────────────────────────
print("Строю чистый шаблон...")
r = subprocess.run(["python3", "build_wm9.py"], capture_output=True, text=True, cwd="/home/user/Auron")
if r.returncode != 0:
    print("FAIL build_wm9.py:\n", r.stderr[:500])
    sys.exit(1)
print("✓ Шаблон построен")

t_total = time.time()
wb = load_workbook(XLSX)
ws_baza = wb["БАЗА_ДДС"]
ws_vipl = wb["ЗАПИСЬ_НА_ВЫПЛАТУ"]

# ── Справочники ────────────────────────────────────────────────────────────────
KASSIRY   = ["Анна Петрова", "Мария Сидорова", "Сергей Иванов", "Татьяна Козлова"]
SUPPLIERS = ["ТД Метро / Metro Cash&Carry", "ООО Лента Оптторг", "Вкусвилл",
             "Мираторг", "Красное&Белое", "Агрофирма Дмитровские"]
MISC_KATS = ["Хозтовары", "Реклама", "Обслуживание", "Банковские комиссии", "Другое"]
MONTH_EXP = [
    ("Аренда",       350_000),
    ("ЗП",           480_000),
    ("Коммунальные",  48_000),
    ("Охрана",        26_000),
    ("Налоги",       130_000),
    ("Интернет+связь",  5_000),
]

# Коэф выручки по дням недели (пн=0..вс=6)
DAY_W = [0.88, 0.85, 0.87, 0.90, 1.05, 1.22, 1.15]

# ── Буферы для записи ─────────────────────────────────────────────────────────
baza_buf: list = []   # [(dt, smena, kassir, typ, kat, sposob, summa, komm)]
vipl_buf: list = []   # [(dt, postavshik, summa, status, nakl, sposob, komm)]

BAZA_MAX = 2999  # rows 4..3002 (last usable)
VIPL_MAX = 499   # rows 4..502

# ── Генерация 365 дней 2025 ────────────────────────────────────────────────────
print("Генерирую данные за 2025 год...")
t_gen = time.time()

dt = date(2025, 1, 1)
end = date(2025, 12, 31)
nakl_cnt = 1000

while dt <= end:
    dow = dt.weekday()
    w   = DAY_W[dow]
    k_d = random.choice(KASSIRY)
    k_v = random.choice(KASSIRY)
    k_n = random.choice(KASSIRY)

    # ── ДОХОДЫ (3 смены) ──────────────────────────────────────────────────────
    # Структура: День=Нал+Экв+Пер, Вечер=Нал+Экв, Ночь=Нал
    # Итого ~2190 строк за год (хорошо укладывается в 3000)
    for smena, kassir, sfactor, add_ekv, add_per in [
        ("День",  k_d, 1.00, True,  True),
        ("Вечер", k_v, 0.55, True,  False),
        ("Ночь",  k_n, 0.28, False, False),
    ]:
        base = 90_000 * w * sfactor * random.uniform(0.88, 1.12)
        ekv  = base * random.uniform(0.28, 0.42)
        per  = base * random.uniform(0.10, 0.20)
        nal  = base - ekv - per

        baza_buf.append((dt, smena, kassir, "Доход", "", "Наличка",   round(nal, -2),  "Z-отчёт"))
        if add_ekv:
            baza_buf.append((dt, smena, kassir, "Доход", "", "Эквайринг", round(ekv, -2), "Z-Эквайринг"))
        if add_per:
            baza_buf.append((dt, smena, kassir, "Доход", "", "Перевод",   round(per, -2), "Z-Перевод"))

        # Расхождение: ~10% смен
        if random.random() < 0.10:
            rash = round(random.uniform(-1800, 1800), -2)
            baza_buf.append((dt, smena, kassir, "Расхождение", "", "Наличка", rash, "Расхождение по смене"))

    # ── ИМАН (хозяин добавляет наличку): ~раз в месяц ────────────────────────
    if dt.day in (12, 13, 14) and random.random() < 0.45:
        summa = round(random.uniform(50_000, 180_000), -3)
        baza_buf.append((dt, "-", "", "Иман", "", "Наличка", summa, "Пополнение от хозяина"))

    # ── ВЫПЛАТА С КАССЫ: ~3 раза в месяц ─────────────────────────────────────
    if dt.day in (5, 15, 25) and random.random() < 0.7:
        summa = round(random.uniform(3_000, 15_000), -2)
        baza_buf.append((dt, "День", k_d, "Расход", "Выплата с кассы", "Наличка", summa, "Выдано из кассы"))

    # ── ПОСТОЯННЫЕ РАСХОДЫ: 1-го числа каждого месяца ────────────────────────
    if dt.day == 1:
        for kat, base_s in MONTH_EXP:
            s = round(base_s * random.uniform(0.97, 1.03), -2)
            baza_buf.append((dt, "-", "", "Расход", kat, "Перевод", s, f"{kat} {dt.strftime('%m.%Y')}"))

    # ── ЗАКУП ТОВАРА: пн и чт ─────────────────────────────────────────────────
    if dow in (0, 3) and random.random() < 0.88:
        sup = random.choice(SUPPLIERS)
        s   = round(random.uniform(120_000, 380_000), -3)
        baza_buf.append((dt, "-", "", "Расход", "Закуп товара", "Перевод", s, f"Закуп: {sup}"))

    # ── МЕЛКИЕ РАСХОДЫ: ~3 раза в неделю ─────────────────────────────────────
    if random.random() < 3/7:
        kat  = random.choice(MISC_KATS)
        spay = random.choice(["Наличка", "Наличка", "Перевод"])
        s    = round(random.uniform(1_500, 22_000), -2)
        baza_buf.append((dt, "День", k_d, "Расход", kat, spay, s, kat))

    # ── ОПЛАТА ПОСТАВЩИКАМ: вт и пт ──────────────────────────────────────────
    if dow in (1, 4) and len(vipl_buf) < VIPL_MAX:
        sup    = random.choice(SUPPLIERS)
        s      = round(random.uniform(85_000, 320_000), -3)
        nakl   = f"НК-{nakl_cnt}"; nakl_cnt += 1
        status = "Выплачено"  # весь 2025 уже в прошлом
        vipl_buf.append((dt, sup, s, status, nakl, "Перевод", f"Оплата {nakl}"))
        # отражаем в БАЗЕ
        baza_buf.append((dt, "-", "", "Оплата долга", "Выплата поставщику", "Перевод", s, f"Оплата {nakl}"))

    dt += timedelta(days=1)

t_gen_done = time.time() - t_gen
print(f"  Сгенерировано: {len(baza_buf):,} строк БАЗА_ДДС, {len(vipl_buf):,} выплат ({t_gen_done:.1f}с)")

# ── Обрезаем до лимитов ────────────────────────────────────────────────────────
overflow_baza = 0
overflow_vipl = 0
if len(baza_buf) > BAZA_MAX:
    overflow_baza = len(baza_buf) - BAZA_MAX
    baza_buf = baza_buf[:BAZA_MAX]
if len(vipl_buf) > VIPL_MAX:
    overflow_vipl = len(vipl_buf) - VIPL_MAX
    vipl_buf = vipl_buf[:VIPL_MAX]

# ── Запись в БАЗА_ДДС ─────────────────────────────────────────────────────────
print(f"Пишу {len(baza_buf):,} строк в БАЗА_ДДС...")
t_w1 = time.time()
for i, (dt, smena, kassir, typ, kat, sposob, summa, komm) in enumerate(baza_buf):
    r = 4 + i
    ws_baza.cell(r, 1).value = dt;     ws_baza.cell(r, 1).number_format = "DD.MM.YYYY"
    ws_baza.cell(r, 2).value = smena
    ws_baza.cell(r, 3).value = kassir
    ws_baza.cell(r, 4).value = typ
    ws_baza.cell(r, 5).value = kat
    ws_baza.cell(r, 6).value = sposob
    ws_baza.cell(r, 7).value = summa;  ws_baza.cell(r, 7).number_format = "#,##0;[Red]-#,##0"
    ws_baza.cell(r, 8).value = komm
t_w1_done = time.time() - t_w1
print(f"  ✓ БАЗА_ДДС записана за {t_w1_done:.1f}с")

# ── Запись в ЗАПИСЬ_НА_ВЫПЛАТУ ────────────────────────────────────────────────
print(f"Пишу {len(vipl_buf):,} строк в ЗАПИСЬ_НА_ВЫПЛАТУ...")
t_w2 = time.time()
for i, (dt, sup, summa, status, nakl, sposob, komm) in enumerate(vipl_buf):
    r = 4 + i
    ws_vipl.cell(r, 2).value = dt;    ws_vipl.cell(r, 2).number_format = "DD.MM.YYYY"
    ws_vipl.cell(r, 3).value = sup
    ws_vipl.cell(r, 4).value = summa; ws_vipl.cell(r, 4).number_format = "#,##0;[Red]-#,##0"
    ws_vipl.cell(r, 5).value = status
    ws_vipl.cell(r, 6).value = nakl
    ws_vipl.cell(r, 7).value = sposob
    ws_vipl.cell(r, 8).value = komm
t_w2_done = time.time() - t_w2
print(f"  ✓ ЗАПИСЬ_НА_ВЫПЛАТУ записана за {t_w2_done:.1f}с")

# ── Сохранение ────────────────────────────────────────────────────────────────
print("Сохраняю файл...")
t_save = time.time()
wb.save(XLSX)
t_save_done = time.time() - t_save

size_bytes = os.path.getsize(XLSX)
size_mb = size_bytes / 1_048_576
t_all = time.time() - t_total

# ── Итоговый отчёт ────────────────────────────────────────────────────────────
print()
print("═" * 60)
print("  ИТОГИ ЗАПОЛНЕНИЯ — WAY MARKET v9 (год: 2025)")
print("═" * 60)
print(f"  БАЗА_ДДС строк записано  : {len(baza_buf):>6,}  (макс 2999)")
print(f"  ЗАПИСЬ_НА_ВЫПЛАТУ строк  : {len(vipl_buf):>6,}  (макс 499)")
print()
print(f"  Генерация данных         : {t_gen_done:>5.1f} с")
print(f"  Запись БАЗА_ДДС          : {t_w1_done:>5.1f} с")
print(f"  Запись ЗАПИСЬ_НА_ВЫПЛАТУ : {t_w2_done:>5.1f} с")
print(f"  Сохранение xlsx          : {t_save_done:>5.1f} с")
print(f"  Полное время             : {t_all:>5.1f} с")
print()
print(f"  Размер файла             : {size_mb:>5.1f} МБ  ({size_bytes:,} байт)")
print()

# ── Подсчёт по типам транзакций ───────────────────────────────────────────────
from collections import Counter
typ_counts = Counter(row[3] for row in baza_buf)
print("  Структура БАЗА_ДДС по типам:")
for typ, cnt in sorted(typ_counts.items(), key=lambda x: -x[1]):
    print(f"    {typ:<20} {cnt:>5,} строк")
print()

# ── Предупреждения ────────────────────────────────────────────────────────────
if overflow_baza:
    print(f"  ⚠️  БАЗА_ДДС: обрезано {overflow_baza} строк (превысили 3000-строчный лимит таблицы)")
    print(f"      РЕКОМЕНДАЦИЯ: увеличить tblБаза до A3:H5003 (доп. 2000 строк)")
if overflow_vipl:
    print(f"  ⚠️  ЗАПИСЬ_НА_ВЫПЛАТУ: обрезано {overflow_vipl} строк")
if not overflow_baza and not overflow_vipl:
    print("  ✅  Все данные уместились в таблицах без обрезки")

if size_mb > 10:
    print(f"  ⚠️  Файл > 10 МБ — возможны тормоза при открытии в Excel")
elif size_mb > 5:
    print(f"  ⚠️  Файл > 5 МБ — рекомендуем отключить автопересчёт при открытии")
else:
    print(f"  ✅  Размер файла в норме (до 5 МБ)")

if t_save_done > 30:
    print(f"  ⚠️  Сохранение > 30с — много формул или условного форматирования")
else:
    print(f"  ✅  Скорость сохранения приемлема")

print("═" * 60)
