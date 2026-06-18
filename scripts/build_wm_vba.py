# -*- coding: utf-8 -*-
"""
Вай Маркет — монолитная учётная система (Storage + UI), скелет .xlsx + VBA .bas.
Профессиональный дизайн на базе дизайн-системы Auron (индиго-бренд, светлый фон, карточки).
Слои:
  Storage     : БАЗА_ДДС (скрытая, защищённая, плоский реестр)
  Controllers : VBA-модуль Модуль_ВайМаркет.bas
  UI/State    : Настройки · Ввод_Касса · Ввод_Расходы · Пульт · Календарь_Выплат · Сводные
Функциональные адреса ячеек сохранены — макросы не меняются.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ---- Палитра / типографика ----
FONT = "Segoe UI"
BRAND = "5E5CE6"      # индиго — заголовки, бренд
INK = "1C1C1E"        # основной текст
INK2 = "8E8E93"       # вторичный текст
BG = "F2F2F7"         # фон страницы
CARD = "FFFFFF"       # карточки
LINE = "E5E5EA"       # разделители
GREEN = "1DA851"      # доход / доступно (текст)
RED = "E0352B"        # расход / долг (текст)
ORANGE = "D98200"     # план / предупреждение (текст)
GREEN_BAR = "30D158"
RED_BAR = "FF453A"
ORANGE_BAR = "FF9F0A"
INPUT_BG = "FFF7E6"   # ячейки ручного ввода
RED_FILL = PatternFill("solid", fgColor="FFE3E0")

RUB = '#,##0\\ "₽"'
DATEFMT = "DD.MM.YYYY"
PWD = "wm"
R = 5001
CR = 1001

thinL = Side(style="thin", color=LINE)
THIN = Border(left=thinL, right=thinL, top=thinL, bottom=thinL)


def F(size=11, bold=False, color=INK, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def fill_bg(ws, r1, c1, r2, c2, color):
    pf = PatternFill("solid", fgColor=color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = pf


def outline(ws, r1, c1, r2, c2, color=LINE, top_accent=None):
    """Тонкая рамка вокруг диапазона; опционально толстая цветная линия сверху (акцент карточки)."""
    s = Side(style="thin", color=color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            b = cell.border
            left = s if c == c1 else b.left
            right = s if c == c2 else b.right
            top = s if r == r1 else b.top
            bottom = s if r == r2 else b.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    if top_accent:
        acc = Side(style="thick", color=top_accent)
        for c in range(c1, c2 + 1):
            cell = ws.cell(r1, c)
            b = cell.border
            cell.border = Border(left=b.left, right=b.right, top=acc, bottom=b.bottom)


def titlebar(ws, last_col, title, subtitle=None, color=BRAND):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    t = ws.cell(1, 1, title)
    t.font = F(18, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=color)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 38
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        s = ws.cell(2, 1, subtitle)
        s.font = F(10, color=INK2)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20


def head_cell(ws, r, c, text, w=None):
    cell = ws.cell(r, c, text)
    cell.font = F(11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BRAND)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN
    if w:
        ws.column_dimensions[get_column_letter(c)].width = w
    return cell


def field_label(ws, r, c, text):
    cell = ws.cell(r, c, text)
    cell.font = F(11, color=INK2)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    return cell


def input_cell(ws, r, c, fmt=None):
    cell = ws.cell(r, c)
    cell.protection = Protection(locked=False)
    cell.fill = PatternFill("solid", fgColor=INPUT_BG)
    cell.font = F(12, color=INK)
    cell.alignment = Alignment(vertical="center", indent=1)
    cell.border = THIN
    if fmt:
        cell.number_format = fmt
    ws.row_dimensions[r].height = 26
    return cell


def _dv(ws, formula1, cells):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cells)


def _protect(ws):
    ws.protection.sheet = True
    ws.protection.password = PWD


def _clean(ws, tab=None):
    ws.sheet_view.showGridLines = False
    if tab:
        ws.sheet_properties.tabColor = tab


# ============================================================ БАЗА_ДДС
def build_baza(wb):
    ws = wb.create_sheet("БАЗА_ДДС")
    headers = ["ID", "Дата", "Тип Операции", "Статья", "Счет",
               "Сумма Дохода", "Сумма Расхода", "Описание", "Месяц", "Год"]
    widths = [8, 13, 17, 20, 14, 15, 15, 30, 8, 8]
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        head_cell(ws, 1, c, h, w)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    for r in range(2, R + 1):
        ws.cell(r, 2).number_format = DATEFMT
        ws.cell(r, 6).number_format = RUB
        ws.cell(r, 7).number_format = RUB
        ws.cell(r, 9, f'=IF(B{r}="","",MONTH(B{r}))')
        ws.cell(r, 10, f'=IF(B{r}="","",YEAR(B{r}))')
    ws.sheet_state = "hidden"
    _protect(ws)
    _clean(ws)
    return ws


# ============================================================ Настройки
def build_nastroyki(wb):
    ws = wb.create_sheet("Настройки")
    fill_bg(ws, 1, 1, 26, 12, BG)
    titlebar(ws, 11, "Настройки системы", "Заполняется один раз. Списки подставляются в карточки и формулы Пульта.")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    for col in ("C", "E", "G", "I", "K"):
        ws.column_dimensions[col].width = 3

    # карточка «Начальный остаток долга» (значение в B4 — на него ссылается формула Пульта)
    fill_bg(ws, 4, 1, 5, 2, CARD)
    outline(ws, 4, 1, 5, 2, top_accent=ORANGE_BAR)
    ws.cell(4, 1, "Начальный остаток долга").font = F(10, color=INK2)
    ws.cell(4, 1).alignment = Alignment(indent=1, vertical="center")
    ws.cell(5, 1, "введите сумму →").font = F(9, italic=True, color=INK2)
    ws.cell(5, 1).alignment = Alignment(indent=1, vertical="center")
    b4 = ws.cell(4, 2, 0); b4.protection = Protection(locked=False)
    b4.number_format = RUB; b4.font = F(15, bold=True, color=ORANGE)
    b4.fill = PatternFill("solid", fgColor=INPUT_BG)
    b4.alignment = Alignment(horizontal="right", indent=1, vertical="center")
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 20

    # Списки. Адреса фиксированы (валидации ссылаются): Статьи D7:D16, Счета F7:F9, Тип H7:H8, Статусы J7:J10.
    head_cell(ws, 6, 4, "Статьи", 22)
    for i, v in enumerate(["Оплата поставщику", "Аренда", "Зарплата", "Налоги", "Коммуналка",
                           "ГСМ", "Расходники", "Хозрасходы", "Выручка", "Прочее"]):
        cell = ws.cell(7 + i, 4, v); cell.border = THIN; cell.font = F(11, color=INK)
        cell.fill = PatternFill("solid", fgColor=CARD); cell.alignment = Alignment(indent=1)
    head_cell(ws, 6, 6, "Счета", 16)
    for i, v in enumerate(["Наличные", "Карта", "Банк"]):
        cell = ws.cell(7 + i, 6, v); cell.border = THIN; cell.font = F(11, color=INK)
        cell.fill = PatternFill("solid", fgColor=CARD); cell.alignment = Alignment(indent=1)
    head_cell(ws, 6, 8, "Тип операции", 20)
    for i, v in enumerate(["Расход", "Увеличение долга"]):
        cell = ws.cell(7 + i, 8, v); cell.border = THIN; cell.font = F(11, color=INK)
        cell.fill = PatternFill("solid", fgColor=CARD); cell.alignment = Alignment(indent=1)
    head_cell(ws, 6, 10, "Статусы выплат", 18)
    for i, v in enumerate(["Запланировано", "Оплачено", "Просрочено", "Отменено"]):
        cell = ws.cell(7 + i, 10, v); cell.border = THIN; cell.font = F(11, color=INK)
        cell.fill = PatternFill("solid", fgColor=CARD); cell.alignment = Alignment(indent=1)
    _clean(ws, tab=INK2)
    return ws


# ============================================================ форма ввода (общее)
def _form_sheet(wb, name, title, subtitle, fields, dvs, btn_row, btn_hint, tab):
    ws = wb.create_sheet(name)
    fill_bg(ws, 1, 1, btn_row + 3, 4, BG)
    titlebar(ws, 3, title, subtitle, color=tab)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 3
    # карточка формы
    last = fields[-1][0]
    fill_bg(ws, 3, 1, last, 2, CARD)
    outline(ws, 3, 1, last, 2, top_accent=tab)
    for (r, lab, fmt) in fields:
        field_label(ws, r, 1, lab)
        input_cell(ws, r, 2, fmt)
    for formula1, cell in dvs:
        _dv(ws, formula1, cell)
    h = ws.cell(btn_row, 1, btn_hint); h.font = F(9, italic=True, color=INK2)
    ws.merge_cells(start_row=btn_row, start_column=1, end_row=btn_row, end_column=2)
    _protect(ws)
    _clean(ws, tab=tab)
    return ws


def build_vvod_kassa(wb):
    return _form_sheet(
        wb, "Ввод_Касса", "Ввод кассы (выручка)",
        "Тип операции жёстко «Доход». Заполни поля и нажми кнопку.",
        [(3, "Дата", DATEFMT), (4, "Счёт", None), (5, "Сумма", RUB), (6, "Комментарий", None)],
        [("=Настройки!$F$7:$F$9", "B4")],
        8, "↓ кнопка «Записать кассу» — после запуска УстановитьКнопки", GREEN_BAR)


def build_vvod_rashody(wb):
    return _form_sheet(
        wb, "Ввод_Расходы", "Ввод расходов и долга",
        "«Увеличение долга» долг растит, но кассу НЕ уменьшает.",
        [(3, "Дата", DATEFMT), (4, "Тип операции", None), (5, "Статья", None),
         (6, "Счёт", None), (7, "Сумма", RUB), (8, "Комментарий", None)],
        [("=Настройки!$H$7:$H$8", "B4"), ("=Настройки!$D$7:$D$16", "B5"),
         ("=Настройки!$F$7:$F$9", "B6")],
        10, "↓ кнопка «Записать расход» — после запуска УстановитьКнопки", RED_BAR)


# ============================================================ Пульт (дашборд)
def build_pult(wb):
    ws = wb.create_sheet("Пульт")
    fill_bg(ws, 1, 1, 16, 12, BG)
    titlebar(ws, 11, "Пульт — сводка магазина", "Все цифры считаются автоматически. Период — в фильтре ниже.")
    for col, w in [("A", 2), ("B", 17), ("C", 9), ("D", 9), ("E", 2),
                   ("F", 17), ("G", 9), ("H", 9), ("I", 2), ("J", 17), ("K", 9), ("L", 9)]:
        ws.column_dimensions[col].width = w

    # фильтр периода (month=B3, year=B4 — формулы используют $B$3/$B$4)
    field_label(ws, 3, 1, None)
    lab = ws.cell(3, 1, "Период:"); lab.font = F(11, bold=True, color=INK); lab.alignment = Alignment(indent=1, vertical="center")
    m = ws.cell(3, 2); m.value = 6; m.protection = Protection(locked=False)
    m.fill = PatternFill("solid", fgColor=INPUT_BG); m.border = THIN; m.font = F(12, color=INK)
    m.alignment = Alignment(horizontal="center"); ws.cell(3, 3, "месяц").font = F(9, color=INK2)
    y = ws.cell(4, 2); y.value = 2026; y.protection = Protection(locked=False)
    y.fill = PatternFill("solid", fgColor=INPUT_BG); y.border = THIN; y.font = F(12, color=INK)
    y.alignment = Alignment(horizontal="center"); ws.cell(4, 3, "год").font = F(9, color=INK2)
    ws.cell(4, 1, "").font = F(11)

    B = "БАЗА_ДДС!"

    def card(r_label, c1, c2, label_text, value_cell, formula, accent, value_color, big=True):
        # карточка: строка label (серый, капс) + строка value (крупно, цвет)
        rv = r_label + 1
        fill_bg(ws, r_label, c1, rv, c2, CARD)
        outline(ws, r_label, c1, rv, c2, top_accent=accent)
        ws.merge_cells(start_row=r_label, start_column=c1, end_row=r_label, end_column=c2)
        L = ws.cell(r_label, c1, label_text); L.font = F(10, bold=True, color=INK2)
        L.alignment = Alignment(indent=1, vertical="center")
        ws.merge_cells(start_row=rv, start_column=c1, end_row=rv, end_column=c2)
        V = ws.cell(rv, c1, formula); V.number_format = RUB
        V.font = F(22 if big else 15, bold=True, color=value_color)
        V.alignment = Alignment(horizontal="right", indent=1, vertical="center")
        ws.row_dimensions[r_label].height = 20
        ws.row_dimensions[rv].height = 34

    # верхний ряд — три главные метрики
    card(6, 2, 4, "ДОСТУПНЫЕ СРЕДСТВА", "B7", None, GREEN_BAR, GREEN)
    ws["B7"] = f'=SUM({B}F2:F{R})-SUMIF({B}C2:C{R},"Расход",{B}G2:G{R})'
    card(6, 6, 8, "ОБЩИЙ ДОЛГ ПОСТАВЩИКАМ", "F7", None, RED_BAR, RED)
    ws["F7"] = (f'=Настройки!B4+SUMIF({B}C2:C{R},"Увеличение долга",{B}G2:G{R})'
                f'-SUMIF({B}D2:D{R},"Оплата поставщику",{B}G2:G{R})')
    card(6, 10, 12, "ПЛАН ВЫПЛАТ ЗА МЕСЯЦ", "J7", None, ORANGE_BAR, ORANGE)
    ws["J7"] = (f'=SUMIFS(Календарь_Выплат!C3:C{CR},Календарь_Выплат!D3:D{CR},"Запланировано",'
                f'Календарь_Выплат!F3:F{CR},$B$3,Календарь_Выплат!G3:G{CR},$B$4)')

    # нижний ряд — месячные метрики
    card(10, 2, 4, "ДОХОДЫ ЗА МЕСЯЦ", "B11", None, GREEN_BAR, GREEN, big=False)
    ws["B11"] = f'=SUMIFS({B}F2:F{R},{B}I2:I{R},$B$3,{B}J2:J{R},$B$4)'
    card(10, 6, 8, "РАСХОДЫ ЗА МЕСЯЦ", "F11", None, RED_BAR, RED, big=False)
    ws["F11"] = f'=SUMIFS({B}G2:G{R},{B}I2:I{R},$B$3,{B}J2:J{R},$B$4,{B}C2:C{R},"Расход")'

    h = ws.cell(13, 2, "↓ кнопка «Обновить сводную» — после запуска УстановитьКнопки")
    h.font = F(9, italic=True, color=INK2)
    ws.merge_cells("B13:D13")
    _protect(ws)
    _clean(ws, tab=BRAND)
    return ws


# ============================================================ Календарь_Выплат
def build_calendar(wb):
    ws = wb.create_sheet("Календарь_Выплат")
    fill_bg(ws, 1, 1, 30, 8, BG)
    titlebar(ws, 7, "Календарь выплат — планирование (просрочка подсветится красным)")
    heads = ["Дата", "Назначение", "План сумма", "Статус", "Факт сумма", "Месяц", "Год"]
    widths = [13, 28, 15, 17, 15, 8, 8]
    for c, (h, w) in enumerate(zip(heads, widths), start=1):
        head_cell(ws, 2, c, h, w)
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"
    for r in range(3, CR + 1):
        for c in (1, 2, 3, 4, 5):
            cell = ws.cell(r, c); cell.protection = Protection(locked=False)
            cell.border = THIN; cell.font = F(11, color=INK); cell.fill = PatternFill("solid", fgColor=CARD)
        ws.cell(r, 1).number_format = DATEFMT
        ws.cell(r, 3).number_format = RUB
        ws.cell(r, 5).number_format = RUB
        ws.cell(r, 6, f'=IF(A{r}="","",MONTH(A{r}))').font = F(10, color=INK2)
        ws.cell(r, 7, f'=IF(A{r}="","",YEAR(A{r}))').font = F(10, color=INK2)
    _dv(ws, "=Настройки!$J$7:$J$10", f"D3:D{CR}")
    ws.conditional_formatting.add(
        f"A3:E{CR}",
        FormulaRule(formula=['AND($A3<>"",$A3<TODAY(),$D3="Запланировано")'], fill=RED_FILL))
    _protect(ws)
    _clean(ws, tab=ORANGE_BAR)
    return ws


# ============================================================ Сводные
def build_svodnye(wb):
    ws = wb.create_sheet("Сводные")
    fill_bg(ws, 1, 1, 20, 12, BG)
    titlebar(ws, 11, "Сводные таблицы", "Нажми «Обновить сводную» на Пульте.")
    fill_bg(ws, 4, 1, 6, 8, CARD)
    outline(ws, 4, 1, 6, 8, top_accent=BRAND)
    ws.merge_cells("A4:H6")
    msg = ws.cell(4, 1, "Макрос построит сводную из БАЗА_ДДС с Временной шкалой (фильтр по дате) "
                        "и срезом по счёту. Период листай Временной шкалой сверху.")
    msg.font = F(11, color=INK); msg.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws.column_dimensions["A"].width = 16
    _clean(ws, tab=INK2)
    return ws


# ============================================================ VBA (.bas, cp1251) — без изменений
VBA = r'''Attribute VB_Name = "Модуль_ВайМаркет"
' ===============================================================
' ВАЙ МАРКЕТ — контроллеры (Storage <- UI), сводная и кнопки.
' Установка: Alt+F11 -> File -> Import File -> этот .bas
'            Alt+F8 -> УстановитьКнопки -> Run
' ===============================================================
Option Explicit

Private Const ПАРОЛЬ As String = "wm"

Function ПоследняяСтрока(ws As Worksheet) As Long
    ПоследняяСтрока = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
End Function

Function НовыйID(ws As Worksheet) As Long
    Dim лс As Long
    лс = ПоследняяСтрока(ws)
    If лс < 2 Then
        НовыйID = 1
    Else
        НовыйID = Val(ws.Cells(лс, 1).Value) + 1
    End If
End Function

Sub ОбеспечитьМесяцГод(ws As Worksheet, r As Long)
    If ws.Cells(r, 9).Formula = "" Then
        ws.Cells(r, 9).Formula = "=IF(B" & r & "="""","""",MONTH(B" & r & "))"
        ws.Cells(r, 10).Formula = "=IF(B" & r & "="""","""",YEAR(B" & r & "))"
    End If
End Sub

Sub ЗаписатьКассу()
    Dim wsФ As Worksheet, wsБ As Worksheet
    Set wsФ = ThisWorkbook.Sheets("Ввод_Касса")
    Set wsБ = ThisWorkbook.Sheets("БАЗА_ДДС")

    Dim дата As Variant, счет As String, сумма As Variant, ком As String
    дата = wsФ.Range("B3").Value
    счет = Trim(CStr(wsФ.Range("B4").Value))
    сумма = wsФ.Range("B5").Value
    ком = Trim(CStr(wsФ.Range("B6").Value))

    If Not IsDate(дата) Then MsgBox "Укажите корректную дату.", vbExclamation: Exit Sub
    If счет = "" Then MsgBox "Выберите счёт.", vbExclamation: Exit Sub
    If Not IsNumeric(сумма) Then MsgBox "Введите сумму числом.", vbExclamation: Exit Sub
    If CDbl(сумма) <= 0 Then MsgBox "Сумма должна быть больше нуля.", vbExclamation: Exit Sub

    Dim r As Long
    wsБ.Unprotect ПАРОЛЬ
    r = ПоследняяСтрока(wsБ) + 1
    wsБ.Cells(r, 1).Value = НовыйID(wsБ)
    wsБ.Cells(r, 2).Value = CDate(дата)
    wsБ.Cells(r, 3).Value = "Доход"
    wsБ.Cells(r, 4).Value = "Выручка"
    wsБ.Cells(r, 5).Value = счет
    wsБ.Cells(r, 6).Value = CDbl(сумма)
    wsБ.Cells(r, 7).Value = 0
    wsБ.Cells(r, 8).Value = ком
    ОбеспечитьМесяцГод wsБ, r
    wsБ.Protect ПАРОЛЬ

    wsФ.Range("B3:B6").ClearContents
    MsgBox "Касса записана: " & Format(CDbl(сумма), "#,##0") & " руб.", vbInformation
End Sub

Sub ЗаписатьРасход()
    Dim wsФ As Worksheet, wsБ As Worksheet
    Set wsФ = ThisWorkbook.Sheets("Ввод_Расходы")
    Set wsБ = ThisWorkbook.Sheets("БАЗА_ДДС")

    Dim дата As Variant, тип As String, статья As String
    Dim счет As String, сумма As Variant, ком As String
    дата = wsФ.Range("B3").Value
    тип = Trim(CStr(wsФ.Range("B4").Value))
    статья = Trim(CStr(wsФ.Range("B5").Value))
    счет = Trim(CStr(wsФ.Range("B6").Value))
    сумма = wsФ.Range("B7").Value
    ком = Trim(CStr(wsФ.Range("B8").Value))

    If Not IsDate(дата) Then MsgBox "Укажите корректную дату.", vbExclamation: Exit Sub
    If тип <> "Расход" And тип <> "Увеличение долга" Then _
        MsgBox "Выберите тип операции.", vbExclamation: Exit Sub
    If статья = "" Then MsgBox "Выберите статью.", vbExclamation: Exit Sub
    If счет = "" Then MsgBox "Выберите счёт.", vbExclamation: Exit Sub
    If Not IsNumeric(сумма) Then MsgBox "Введите сумму числом.", vbExclamation: Exit Sub
    If CDbl(сумма) <= 0 Then MsgBox "Сумма должна быть больше нуля.", vbExclamation: Exit Sub

    Dim r As Long
    wsБ.Unprotect ПАРОЛЬ
    r = ПоследняяСтрока(wsБ) + 1
    wsБ.Cells(r, 1).Value = НовыйID(wsБ)
    wsБ.Cells(r, 2).Value = CDate(дата)
    wsБ.Cells(r, 3).Value = тип
    wsБ.Cells(r, 4).Value = статья
    wsБ.Cells(r, 5).Value = счет
    wsБ.Cells(r, 6).Value = 0
    wsБ.Cells(r, 7).Value = CDbl(сумма)
    wsБ.Cells(r, 8).Value = ком
    ОбеспечитьМесяцГод wsБ, r
    wsБ.Protect ПАРОЛЬ

    wsФ.Range("B3:B8").ClearContents
    MsgBox "Записано: " & тип & " — " & Format(CDbl(сумма), "#,##0") & " руб.", vbInformation
End Sub

Sub ПостроитьСводные()
    Dim wsБ As Worksheet, wsС As Worksheet
    Dim pc As PivotCache, pt As PivotTable
    Dim лс As Long, рнг As Range
    Set wsБ = ThisWorkbook.Sheets("БАЗА_ДДС")
    лс = ПоследняяСтрока(wsБ)
    If лс < 2 Then MsgBox "Нет данных для сводной.", vbExclamation: Exit Sub

    Application.DisplayAlerts = False
    On Error Resume Next
    ThisWorkbook.Sheets("Сводные").Delete
    On Error GoTo 0
    Application.DisplayAlerts = True
    Set wsС = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
    wsС.Name = "Сводные"

    Set рнг = wsБ.Range("A1:J" & лс)
    Set pc = ThisWorkbook.PivotCaches.Create(xlDatabase, рнг)
    Set pt = pc.CreatePivotTable(wsС.Range("B8"), "СводнаяДДС")
    With pt
        .PivotFields("Статья").Orientation = xlRowField
        .PivotFields("Тип Операции").Orientation = xlColumnField
        .AddDataField .PivotFields("Сумма Дохода"), "Доход", xlSum
        .AddDataField .PivotFields("Сумма Расхода"), "Расход", xlSum
        .ShowTableStyleRowStripes = True
    End With

    On Error Resume Next
    Dim scT As SlicerCache
    Set scT = ThisWorkbook.SlicerCaches.Add2(pt, "Дата", , xlTimeline)
    scT.Slicers.Add wsС, , "ШкалаДата", "Период", 6, 350, 280, 100
    Dim scS As SlicerCache
    Set scS = ThisWorkbook.SlicerCaches.Add2(pt, "Счет")
    scS.Slicers.Add wsС, , "СрезСчет", "Счёт", 130, 350, 160, 120
    On Error GoTo 0

    wsС.Activate
    MsgBox "Сводная построена. Период фильтруй Временной шкалой сверху.", vbInformation
End Sub

Sub ДобавитьКнопку(имяЛиста As String, макрос As String, подпись As String, адрес As String)
    Dim ws As Worksheet, b As Button, shp As Shape
    Set ws = ThisWorkbook.Sheets(имяЛиста)
    ws.Unprotect ПАРОЛЬ
    For Each shp In ws.Shapes
        If shp.Type = msoFormControl Then shp.Delete
    Next shp
    Set b = ws.Buttons.Add(ws.Range(адрес).Left, ws.Range(адрес).Top, 150, 34)
    b.OnAction = макрос
    b.Caption = подпись
    ws.Protect ПАРОЛЬ
End Sub

Sub УстановитьКнопки()
    ДобавитьКнопку "Ввод_Касса", "ЗаписатьКассу", "Записать кассу", "B8"
    ДобавитьКнопку "Ввод_Расходы", "ЗаписатьРасход", "Записать расход", "B10"
    ДобавитьКнопку "Пульт", "ПостроитьСводные", "Обновить сводную", "B13"
    MsgBox "Кнопки установлены.", vbInformation
End Sub
'''


def write_bas():
    name = "Модуль_ВайМаркет.bas"
    with open(name, "wb") as f:
        f.write(VBA.replace("\n", "\r\n").encode("cp1251"))
    return name


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_baza(wb)
    build_nastroyki(wb)
    build_vvod_kassa(wb)
    build_vvod_rashody(wb)
    build_pult(wb)
    build_calendar(wb)
    build_svodnye(wb)
    # порядок вкладок: Пульт первым
    wb.move_sheet("Пульт", -(wb.sheetnames.index("Пульт")))
    wb.active = wb.sheetnames.index("Пульт")
    out = "Вай_Маркет.xlsx"
    wb.save(out)
    bas = write_bas()
    print("saved", out, "| листы:", wb.sheetnames)
    print("saved", bas)


if __name__ == "__main__":
    main()
