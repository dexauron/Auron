# -*- coding: utf-8 -*-
"""
Генератор Excel-шаблона Auron Finance для ежедневного заполнения (Way Market).
Логика и формулы соответствуют docs/08-BUILD-PLAN.md.
Суммы — в рублях (это инструмент для человека; правило «копейки» относится к коду приложения).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# --- Палитра дизайн-системы (04-DESIGN-SYSTEM) ---
BRAND   = "5E5CE6"  # индиго
INCOME  = "30D158"  # зелёный
EXPENSE = "FF453A"  # красный
TRANSFER= "FF9F0A"  # оранжевый
HEAD_BG = "5E5CE6"
SUB_BG  = "EEEEF7"
GREY    = "8E8E93"
LINE    = "E5E5EA"

RUB = '#,##0\\ "₽"'
PCT = '0.0%'

thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def title(ws, text, span=6):
    c = ws.cell(1, 1, text)
    c.font = Font(name="Calibri", size=16, bold=True, color=BRAND)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.row_dimensions[1].height = 28

def hint(ws, row, text, span=6):
    c = ws.cell(row, 1, text)
    c.font = Font(size=10, italic=True, color=GREY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def money_col(ws, col, r1, r2):
    for r in range(r1, r2 + 1):
        ws.cell(r, col).number_format = RUB

def grid(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = BORDER

wb = openpyxl.Workbook()

# ============================================================= ИНСТРУКЦИЯ
ws = wb.active
ws.title = "Инструкция"
title(ws, "Auron Finance — ежедневный учёт (Way Market)", 2)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 95
lines = [
    ("Что это", "Рабочая тетрадь магазина. Заполняется каждый день. Считает всё само по формулам приложения."),
    ("Сначала — один раз", "Лист «Справочники»: впиши поставщиков с начальными долгами, сотрудников со ставками, статьи накоплений, отделы, целевую рентабельность."),
    ("Каждое утро", "Лист «Касса»: вводи Z-отчёт по сменам. Расхождение посчитается само."),
    ("В течение дня", "Лист «Выплаты» — оплаты и накладные поставщикам. Лист «Заказы» — что заказали и что пришло."),
    ("Каждый день", "Лист «Табель» — кто работал. Лист «Накопления» — сколько отложил на аренду/зарплату/налоги."),
    ("Долги", "Лист «Долги» считает текущий долг каждого поставщика сам из листа «Выплаты»."),
    ("Зарплата", "Лист «Зарплата» считает к выплате: ставка × отработано − авансы."),
    ("Итог месяца", "Лист «Отчёт» собирает выручку, расходы, прибыль и рентабельность за месяц."),
    ("", ""),
    ("ФОРМУЛЫ (как в приложении)", ""),
    ("Расхождение кассы", "(Фактическая наличка + Выплаты с кассы) − Z-отчёт по наличке.  < 0 недостача · > 0 излишек · = 0 идеально."),
    ("Долг поставщику", "Начальный долг + Накладные (взяли в долг) − Оплачено − Возвраты."),
    ("Зарплата к выплате", "Ставка × Отработано (дней/часов) − Авансы."),
    ("Откладывать в день", "(План статьи − Уже отложено) / Дней до конца месяца."),
    ("Рентабельность", "(Выручка − Себестоимость) / Выручка × 100%."),
    ("Свободно", "Наличка в кассе − плановые выплаты − резервы накоплений."),
]
r = 3
for k, v in lines:
    a = ws.cell(r, 1, k); a.font = Font(bold=True, color=BRAND if k and not v else "1C1C1E", size=11)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b = ws.cell(r, 2, v); b.font = Font(size=11); b.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 1

# ============================================================= СПРАВОЧНИКИ
ws = wb.create_sheet("Справочники")
title(ws, "Справочники — заполнить один раз", 9)
hint(ws, 2, "Имена отсюда подставляются в выпадающих списках на других листах. Можно добавлять строки.")

# Поставщики A:C (данные 5..54)
header_row(ws, 4, ["Поставщик", "Телефон", "Начальный долг"], [24, 16, 16])
money_col(ws, 3, 5, 54); grid(ws, 4, 1, 54, 3)
ex_sup = [("ООО Альфа", "+7 900 000-00-01", 120000), ("Иванов П.", "+7 900 000-00-02", 45000), ("Нурмат ИП", "+7 900 000-00-03", 0)]
for i, (n, p, d) in enumerate(ex_sup):
    ws.cell(5 + i, 1, n); ws.cell(5 + i, 2, p); ws.cell(5 + i, 3, d)

# Сотрудники E:H (данные 5..34)
header_row(ws, 4, [None]*4)  # placeholder to keep style; will overwrite below
for col, h, w in [(5, "Сотрудник", 22), (6, "Должность", 18), (7, "Оплата (день/час)", 14), (8, "Ставка", 12)]:
    c = ws.cell(4, col, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=HEAD_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = w
money_col(ws, 8, 5, 34); grid(ws, 4, 5, 34, 8)
ex_emp = [("Марина", "Бухгалтер", "день", 2500), ("Олег", "Администратор", "день", 2000), ("Гуля", "Продавец", "день", 1800)]
for i, (n, role, t, rate) in enumerate(ex_emp):
    ws.cell(5 + i, 5, n); ws.cell(5 + i, 6, role); ws.cell(5 + i, 7, t); ws.cell(5 + i, 8, rate)

# Статьи накоплений J:K (данные 5..24)
for col, h, w in [(10, "Статья накоплений", 22), (11, "План в месяц", 14)]:
    c = ws.cell(4, col, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=HEAD_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
ws.column_dimensions['J'].width = 22; ws.column_dimensions['K'].width = 14
money_col(ws, 11, 5, 24); grid(ws, 4, 10, 24, 11)
for i, (n, pl) in enumerate([("Аренда", 60000), ("Зарплата", 150000), ("Налоги", 40000), ("Коммуналка", 15000)]):
    ws.cell(5 + i, 10, n); ws.cell(5 + i, 11, pl)

# Отделы M (данные 5..24) и Настройки O
c = ws.cell(4, 13, "Отделы"); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=HEAD_BG)
c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws.column_dimensions['M'].width = 18; grid(ws, 4, 13, 24, 13)
for i, n in enumerate(["Молочка", "Бакалея", "Заморозка", "Химия", "Овощи"]):
    ws.cell(5 + i, 13, n)
c = ws.cell(4, 15, "Настройки"); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=HEAD_BG)
c.alignment = Alignment(horizontal="center"); c.border = BORDER
ws.merge_cells("O4:P4")
ws.column_dimensions['O'].width = 24; ws.column_dimensions['P'].width = 14
sett = [("Целевая рентабельность", 0.22, PCT), ("Метод накоплений", "виртуальный резерв", None), ("Дней в месяце", 30, None)]
for i, (k, v, fmt) in enumerate(sett):
    ws.cell(5 + i, 15, k).font = Font(bold=True)
    cell = ws.cell(5 + i, 16, v)
    if fmt: cell.number_format = fmt
    grid(ws, 5 + i, 15, 5 + i, 16)

SUP_RANGE = "Справочники!$A$5:$A$54"
EMP_RANGE = "Справочники!$E$5:$E$34"
DEP_RANGE = "Справочники!$M$5:$M$24"

# ============================================================= КАССА (Z-отчёт)
ws = wb.create_sheet("Касса")
title(ws, "Касса — Z-отчёт по сменам", 12)
hint(ws, 2, "Расхождение = (Факт наличка + Выплаты с кассы) − Z наличка. Карта и СБП идут в банк, в расхождение по налу не входят.")
heads = ["Дата", "Смена", "Кассир", "Z наличка", "Z карта", "Z СБП", "Выручка смены",
         "Факт наличка", "Выплаты с кассы", "Расхождение", "Статус", "Комментарий"]
header_row(ws, 3, heads, [12, 10, 16, 12, 12, 12, 13, 12, 14, 13, 13, 24])
R1, R2 = 4, 203
for r in range(R1, R2 + 1):
    ws.cell(r, 7, f"=D{r}+E{r}+F{r}")                       # выручка смены
    ws.cell(r, 10, f"=H{r}+I{r}-D{r}")                      # расхождение
    ws.cell(r, 11, f'=IF(J{r}=0,"OK",IF(J{r}<0,"Недостача","Излишек"))')
for col in (4, 5, 6, 7, 8, 9, 10):
    money_col(ws, col, R1, R2)
grid(ws, 3, 1, R2, 12)
dv_shift = DataValidation(type="list", formula1='"Дневная,Ночная"', allow_blank=True)
ws.add_data_validation(dv_shift); dv_shift.add(f"B{R1}:B{R2}")
dv_cash = DataValidation(type="list", formula1=EMP_RANGE, allow_blank=True)
ws.add_data_validation(dv_cash); dv_cash.add(f"C{R1}:C{R2}")
ws.freeze_panes = "A4"

# ============================================================= ВЫПЛАТЫ поставщикам
ws = wb.create_sheet("Выплаты")
title(ws, "Выплаты и накладные поставщикам", 6)
hint(ws, 2, "Вводи каждую операцию строкой. Лист «Долги» сам пересчитает текущий долг каждого поставщика.")
header_row(ws, 3, ["Дата", "Поставщик", "Оплачено наличными", "Новый долг (накладная)", "Возврат", "Комментарий"],
           [12, 24, 18, 20, 14, 26])
money_col(ws, 3, 4, 203); money_col(ws, 4, 4, 203); money_col(ws, 5, 4, 203)
grid(ws, 3, 1, 203, 6)
dv = DataValidation(type="list", formula1=SUP_RANGE, allow_blank=True)
ws.add_data_validation(dv); dv.add("B4:B203")
ws.freeze_panes = "A4"

# ============================================================= ДОЛГИ (сводка)
ws = wb.create_sheet("Долги")
title(ws, "Долги поставщикам — считается автоматически", 5)
hint(ws, 2, "Текущий долг = Начальный + Накладные − Оплачено − Возвраты (из листа «Выплаты»).")
header_row(ws, 3, ["Поставщик", "Начальный долг", "Накладные", "Оплачено", "Текущий долг"],
           [24, 16, 16, 16, 16])
for i in range(50):
    r = 4 + i; src = 5 + i
    ws.cell(r, 1, f"=IF(Справочники!A{src}=\"\",\"\",Справочники!A{src})")
    ws.cell(r, 2, f"=IF(A{r}=\"\",\"\",Справочники!C{src})")
    ws.cell(r, 3, f'=IF(A{r}="","",SUMIF(Выплаты!$B:$B,A{r},Выплаты!$D:$D))')
    ws.cell(r, 4, f'=IF(A{r}="","",SUMIF(Выплаты!$B:$B,A{r},Выплаты!$C:$C)+SUMIF(Выплаты!$B:$B,A{r},Выплаты!$E:$E))')
    ws.cell(r, 5, f'=IF(A{r}="","",B{r}+C{r}-D{r})')
for col in (2, 3, 4, 5):
    money_col(ws, col, 4, 53)
grid(ws, 3, 1, 53, 5)
tr = ws.cell(54, 1, "ИТОГО долг"); tr.font = Font(bold=True)
ws.cell(54, 5, "=SUM(E4:E53)").font = Font(bold=True); ws.cell(54, 5).number_format = RUB

# ============================================================= ЗАКАЗЫ
ws = wb.create_sheet("Заказы")
title(ws, "Заказы поставщикам", 8)
hint(ws, 2, "Создал заказ — статус «Ожидается». Пришёл товар — поставь «Получен» и впиши фактическую сумму.")
header_row(ws, 3, ["Дата заказа", "Поставщик", "Отдел", "Сумма заказа", "Ожид. дата", "Статус", "Факт. сумма", "Дата прихода"],
           [12, 22, 16, 14, 12, 14, 14, 12])
money_col(ws, 4, 4, 203); money_col(ws, 7, 4, 203)
grid(ws, 3, 1, 203, 8)
dv = DataValidation(type="list", formula1=SUP_RANGE, allow_blank=True); ws.add_data_validation(dv); dv.add("B4:B203")
dv2 = DataValidation(type="list", formula1=DEP_RANGE, allow_blank=True); ws.add_data_validation(dv2); dv2.add("C4:C203")
dv3 = DataValidation(type="list", formula1='"Ожидается,В пути,Получен,Отменён"', allow_blank=True)
ws.add_data_validation(dv3); dv3.add("F4:F203")
ws.freeze_panes = "A4"

# ============================================================= ТАБЕЛЬ
ws = wb.create_sheet("Табель")
title(ws, "Табель — кто работал", 36)
hint(ws, 2, "Коды: Р работал · П полдня · В выходной · Б больничный · О отпуск · Н прогул. Отработано считается само.", 12)
heads = ["Сотрудник"] + [str(d) for d in range(1, 32)] + ["Отработано"]
header_row(ws, 3, heads)
ws.column_dimensions['A'].width = 20
for d in range(1, 32):
    ws.column_dimensions[get_column_letter(1 + d)].width = 4
ws.column_dimensions[get_column_letter(33)].width = 12
for i in range(30):
    r = 4 + i; src = 5 + i
    ws.cell(r, 1, f"=IF(Справочники!E{src}=\"\",\"\",Справочники!E{src})")
    first = get_column_letter(2); last = get_column_letter(32)
    ws.cell(r, 33, f'=IF(A{r}="","",COUNTIF({first}{r}:{last}{r},"Р")+0.5*COUNTIF({first}{r}:{last}{r},"П"))')
grid(ws, 3, 1, 33, 33)
dv = DataValidation(type="list", formula1='"Р,П,В,Б,О,Н"', allow_blank=True)
ws.add_data_validation(dv); dv.add("B4:AF33")
ws.freeze_panes = "B4"

# ============================================================= ЗАРПЛАТА
ws = wb.create_sheet("Зарплата")
title(ws, "Зарплата за месяц", 6)
hint(ws, 2, "К выплате = Ставка × Отработано − Авансы. Отработано берётся из листа «Табель».")
header_row(ws, 3, ["Сотрудник", "Ставка", "Отработано", "Начислено", "Авансы", "К выплате"],
           [22, 12, 12, 14, 12, 14])
for i in range(30):
    r = 4 + i; src = 5 + i
    ws.cell(r, 1, f"=IF(Справочники!E{src}=\"\",\"\",Справочники!E{src})")
    ws.cell(r, 2, f"=IF(A{r}=\"\",\"\",Справочники!H{src})")
    ws.cell(r, 3, f"=IF(A{r}=\"\",\"\",Табель!AG{4 + i})")
    ws.cell(r, 4, f'=IF(A{r}="","",B{r}*C{r})')
    ws.cell(r, 6, f'=IF(A{r}="","",MAX(0,D{r}-E{r}))')
for col in (2, 4, 5, 6):
    money_col(ws, col, 4, 33)
grid(ws, 3, 1, 34, 6)
ws.cell(34, 1, "ИТОГО к выплате").font = Font(bold=True)
ws.cell(34, 6, "=SUM(F4:F33)").font = Font(bold=True); ws.cell(34, 6).number_format = RUB

# ============================================================= НАКОПЛЕНИЯ
ws = wb.create_sheet("Накопления")
title(ws, "Накопления на постоянные расходы", 6)
hint(ws, 2, "Откладывать в день = (План − Отложено) / Дней до конца месяца. План берётся из «Справочников».")
header_row(ws, 3, ["Статья", "План в месяц", "Отложено", "Осталось", "Дней до конца", "В день"],
           [22, 14, 14, 14, 14, 12])
for i in range(20):
    r = 4 + i; src = 5 + i
    ws.cell(r, 1, f"=IF(Справочники!J{src}=\"\",\"\",Справочники!J{src})")
    ws.cell(r, 2, f"=IF(A{r}=\"\",\"\",Справочники!K{src})")
    ws.cell(r, 4, f'=IF(A{r}="","",B{r}-C{r})')
    ws.cell(r, 6, f'=IF(OR(A{r}="",E{r}=0),"",MAX(0,D{r})/E{r})')
for col in (2, 3, 4, 6):
    money_col(ws, col, 4, 23)
grid(ws, 3, 1, 23, 6)

# ============================================================= ЦЕЛИ
ws = wb.create_sheet("Цели")
title(ws, "Цели владельца", 6)
header_row(ws, 3, ["Цель", "Тип", "План", "Факт", "Прогресс", "Осталось"], [26, 14, 14, 14, 12, 14])
for r in range(4, 24):
    ws.cell(r, 5, f'=IF(C{r}=0,"",D{r}/C{r})').number_format = PCT
    ws.cell(r, 6, f'=IF(C{r}="","",C{r}-D{r})')
money_col(ws, 3, 4, 23); money_col(ws, 4, 4, 23); money_col(ws, 6, 4, 23)
grid(ws, 3, 1, 23, 6)
dv = DataValidation(type="list", formula1='"Выручка,Расходы,Прибыль,Долг,Накопления"', allow_blank=True)
ws.add_data_validation(dv); dv.add("B4:B23")

# ============================================================= ПРОЧИЕ РАСХОДЫ
ws = wb.create_sheet("Расходы")
title(ws, "Прочие расходы и списания", 4)
header_row(ws, 3, ["Дата", "Категория", "Сумма", "Комментарий"], [12, 20, 14, 34])
money_col(ws, 3, 4, 203); grid(ws, 3, 1, 203, 4)
dv = DataValidation(type="list", formula1='"ГСМ,Расходники,Списание товара,Хозрасходы,Прочее"', allow_blank=True)
ws.add_data_validation(dv); dv.add("B4:B203")
ws.freeze_panes = "A4"

# ============================================================= ОТЧЁТ (месяц)
ws = wb.create_sheet("Отчёт")
title(ws, "Итог месяца", 3)
hint(ws, 2, "Собирается автоматически из всех листов. Себестоимость впиши вручную для расчёта рентабельности.")
ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 18; ws.column_dimensions['C'].width = 40
rows = [
    ("ВЫРУЧКА", "head", ""),
    ("Наличные (Z)", "=SUM(Касса!D4:D203)", ""),
    ("Карта (Z)", "=SUM(Касса!E4:E203)", ""),
    ("Переводы СБП (Z)", "=SUM(Касса!F4:F203)", ""),
    ("Выручка всего", "=SUM(Касса!G4:G203)", "наличка + карта + СБП"),
    ("", "", ""),
    ("РАСХОДЫ", "head", ""),
    ("Оплачено поставщикам", "=SUM(Выплаты!C4:C203)", "наличными"),
    ("Зарплата к выплате", "=Зарплата!F34", "из табеля"),
    ("Накопления отложено", "=SUM(Накопления!C4:C23)", "аренда/зарплата/налоги"),
    ("Прочие расходы", "=SUM(Расходы!C4:C203)", ""),
    ("Расходы всего", "=B10+B11+B12+B13", ""),
    ("", "", ""),
    ("ИТОГИ", "head", ""),
    ("Себестоимость проданного", 0, "впиши вручную или из 1С"),
    ("Валовая прибыль", "=B7-B17", "выручка − себестоимость"),
    ("Рентабельность", "=IF(B7=0,\"\",B18/B7)", "цель: см. Справочники"),
    ("Долг поставщикам (конец)", "=Долги!E54", ""),
    ("Расхождения за месяц", "=SUM(Касса!J4:J203)", "сумма всех расхождений кассы"),
]
r = 3
for name, val, note in rows:
    a = ws.cell(r, 1, name)
    if val == "head":
        a.font = Font(bold=True, color="FFFFFF"); a.fill = PatternFill("solid", fgColor=HEAD_BG)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    else:
        a.font = Font(bold=(name in ("Выручка всего", "Расходы всего", "Валовая прибыль", "Рентабельность")))
        if val != "":
            cell = ws.cell(r, 2, val)
            cell.number_format = PCT if name == "Рентабельность" else RUB
            cell.font = Font(bold=(name in ("Выручка всего", "Расходы всего", "Валовая прибыль")))
        ws.cell(r, 3, note).font = Font(size=10, italic=True, color=GREY)
    r += 1

# ширины по умолчанию для листов без явных
for sh in wb.worksheets:
    if sh.sheet_view:
        sh.sheet_view.showGridLines = False

OUT = "Auron_шаблон_ежедневный.xlsx"
wb.save(OUT)
print("saved", OUT, "sheets:", [s.title for s in wb.worksheets])
