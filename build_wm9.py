#!/usr/bin/env python3
"""WAY MARKET v9 — Full Build (11 sheets). Professional management accounting template."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from datetime import date, datetime

today = date.today()

# ── Color Palette ──────────────────────────────────────────────────────────────
NAVY="FF111827"; INDIGO="FF4F46E5"
GREEN="FF10B981"; GREEN_L="FFD1FAE5"
BLUE="FF3B82F6";  BLUE_L="FFEFF6FF"
RED="FFEF4444";   RED_L="FFFEE2E2"
AMBER="FFF59E0B"; AMBER_L="FFFEF3C7"
PURPLE="FF8B5CF6";PURP_L="FFEDE9FE"
TEAL="FF14B8A6";  TEAL_L="FFCCFBF1"
GRAY="FF6B7280";  LGRAY="FFF7F8FA"
BORDER="FFE5E7EB";WHITE="FFFFFFFF"
INP="FFEFF6FF";   INP_BD="FF3B82F6"
DISABLED="FFF3F4F6"
MONEY="#,##0;[Red]-#,##0"; MONEY2="#,##0.0;[Red]-#,##0.0"; DATE_F="DD.MM.YYYY"
MONTHS_RU=["Январь","Февраль","Март","Апрель","Май","Июнь",
           "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

# ── Style Helpers ──────────────────────────────────────────────────────────────
def F(c): return PatternFill("solid",start_color=c,fgColor=c)
def fnt(sz=10,bold=False,col="FF000000",it=False):
    return Font(name="Calibri",size=sz,bold=bold,color=col,italic=it)
def brd(c=BORDER,s="thin"):
    sd=Side(style=s,color=c); return Border(left=sd,right=sd,top=sd,bottom=sd)
def brd_med(c=INP_BD): sd=Side(style="medium",color=c); return Border(left=sd,right=sd,top=sd,bottom=sd)
def CA(wrap=True): return Alignment(horizontal="center",vertical="center",wrap_text=wrap)
def LA(): return Alignment(horizontal="left",vertical="center",wrap_text=True)
def RA(): return Alignment(horizontal="right",vertical="center")
def prot(locked=True): return Protection(locked=locked)

def banner(ws,txt,merge,bg=NAVY,sz=14,col="FFFFFFFF"):
    ws.merge_cells(merge)
    c=ws[merge.split(":")[0]]; c.value=txt
    c.font=fnt(sz,True,col); c.fill=F(bg); c.alignment=CA()
    r=int(''.join(x for x in merge.split(":")[0] if x.isdigit()))
    ws.row_dimensions[r].height=38

def sec_hdr(ws,row,txt,ncols,bg,h=26):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row,1); c.value=txt
    c.font=fnt(10,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=LA()
    ws.row_dimensions[row].height=h

def hrow(ws,row,headers,bg=NAVY,h=24):
    for ci,h_ in enumerate(headers,1):
        c=ws.cell(row,ci); c.value=h_
        c.font=fnt(9,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=CA(); c.border=brd()
    ws.row_dimensions[row].height=h

def cw(ws,widths):
    for col,w in widths.items(): ws.column_dimensions[col].width=w

def lbl_cell(ws,row,c1,c2,text,bg=LGRAY,h=24):
    if c1!=c2: ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=text
    c.font=fnt(10); c.fill=F(bg); c.border=brd(); c.alignment=LA()
    ws.row_dimensions[row].height=h

def inp_cell(ws,row,c1,c2,money=True,val=None,fmt=None):
    if c1!=c2: ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1)
    if val is not None: c.value=val
    c.font=fnt(12,True,INDIGO); c.fill=F(INP); c.border=brd_med(); c.alignment=CA()
    c.protection=prot(False)
    if money: c.number_format=MONEY
    if fmt: c.number_format=fmt
    return c

def calc_cell(ws,row,c1,c2,formula,money=True,col=GREEN,bg=GREEN_L,sz=11):
    if c1!=c2: ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=formula
    c.font=fnt(sz,True,col); c.fill=F(bg); c.border=brd()
    c.alignment=RA() if money else CA()
    if money: c.number_format=MONEY
    return c

def hint_cell(ws,row,c1,c2,text):
    if c1!=c2: ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=text
    c.font=fnt(9,it=True,col=GRAY); c.fill=F(LGRAY); c.border=brd(); c.alignment=LA()

# ── Dashboard KPI Card Builder ─────────────────────────────────────────────────
def kpi_card(ws, row, col_start, col_end, label, value_formula, prev_formula,
             bg_label=LGRAY, bg_val=WHITE, val_color=NAVY, money_fmt=True):
    """Creates a 2-row KPI card: row N=label, row N+1=value+trend"""
    # Label row
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row, col_start)
    c.value = label
    c.font = fnt(9, False, GRAY); c.fill = F(bg_label)
    c.border = brd(); c.alignment = CA()
    ws.row_dimensions[row].height = 18

    # Value row
    ws.merge_cells(start_row=row+1, start_column=col_start, end_row=row+1, end_column=col_end-1)
    cv = ws.cell(row+1, col_start)
    cv.value = value_formula
    cv.font = fnt(14, True, val_color); cv.fill = F(bg_val)
    cv.border = brd(); cv.alignment = CA(wrap=False)
    if money_fmt: cv.number_format = MONEY

    # Trend arrow cell
    ct = ws.cell(row+1, col_end)
    if prev_formula:
        ct.value = f'=IF({value_formula[1:]}>{prev_formula[1:]},"▲","▼")'
        ct.font = fnt(11, True, GREEN)
    else:
        ct.value = ""
    ct.fill = F(bg_val); ct.border = brd(); ct.alignment = CA()
    ws.row_dimensions[row+1].height = 32

    # Traffic light conditional formatting
    addr_val = f"{get_column_letter(col_start)}{row+1}"
    return cv, ct

wb = Workbook()

# ════════════════════════════════════════════════════════════
# 1. НАСТРОЙКИ — центр управления шаблоном (9 разделов)
# ════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "НАСТРОЙКИ"; ws.sheet_view.showGridLines = False
cw(ws,{"A":28,"B":3,"C":20,"D":3,"E":18,"F":3,"G":26,"H":16})

banner(ws, "⚙  НАСТРОЙКИ — ЦЕНТР УПРАВЛЕНИЯ ШАБЛОНОМ", "A1:H1", INDIGO)
ws.merge_cells("A2:H2")
ws.cell(2,1).value="Все настройки магазина — только здесь. Остальные листы подстраиваются автоматически."
ws.cell(2,1).font=fnt(10,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22

# Helpers scoped to НАСТРОЙКИ
def n_sec(r_,title_,clr_=INDIGO):
    ws.merge_cells(f"A{r_}:H{r_}")
    ws.cell(r_,1).value=title_; ws.cell(r_,1).font=fnt(11,True,"FFFFFFFF")
    ws.cell(r_,1).fill=F(clr_); ws.cell(r_,1).alignment=LA()
    ws.row_dimensions[r_].height=28

def n_colhdr(r_,cols_,clr_):
    for ci_,tx_ in cols_:
        ws.cell(r_,ci_).value=tx_; ws.cell(r_,ci_).font=fnt(9,True,"FFFFFFFF")
        ws.cell(r_,ci_).fill=F(clr_); ws.cell(r_,ci_).border=brd(); ws.cell(r_,ci_).alignment=CA()
    ws.row_dimensions[r_].height=22

def n_param(r_,label_,val_,typ_="text",hint_=""):
    ws.cell(r_,1).value=label_; ws.cell(r_,1).font=fnt(10); ws.cell(r_,1).fill=F(LGRAY)
    ws.cell(r_,1).border=brd(); ws.cell(r_,1).alignment=LA()
    ws.cell(r_,5).value=val_; ws.cell(r_,5).font=fnt(11,True,INDIGO); ws.cell(r_,5).fill=F(INP)
    ws.cell(r_,5).border=brd_med(); ws.cell(r_,5).alignment=CA(); ws.cell(r_,5).protection=prot(False)
    if typ_=="date": ws.cell(r_,5).number_format=DATE_F
    elif typ_=="pct": ws.cell(r_,5).number_format="0%"
    elif typ_=="money": ws.cell(r_,5).number_format=MONEY
    if hint_:
        ws.cell(r_,7).value=hint_; ws.cell(r_,7).font=fnt(8,it=True,col=GRAY)
        ws.cell(r_,7).fill=F(LGRAY); ws.cell(r_,7).alignment=LA()
    ws.row_dimensions[r_].height=24

def n_toggle(r_,label_,default_="Вкл",hint_=""):
    ws.cell(r_,1).value=label_; ws.cell(r_,1).font=fnt(10); ws.cell(r_,1).fill=F(LGRAY)
    ws.cell(r_,1).border=brd(); ws.cell(r_,1).alignment=LA()
    ws.cell(r_,5).value=default_; ws.cell(r_,5).font=fnt(11,True)
    ws.cell(r_,5).fill=F(GREEN_L if default_=="Вкл" else RED_L)
    ws.cell(r_,5).border=brd_med(); ws.cell(r_,5).alignment=CA(); ws.cell(r_,5).protection=prot(False)
    if hint_:
        ws.cell(r_,7).value=hint_; ws.cell(r_,7).font=fnt(8,it=True,col=GRAY)
        ws.cell(r_,7).fill=F(LGRAY); ws.cell(r_,7).alignment=LA()
    ws.row_dimensions[r_].height=24

# ── РАЗДЕЛ 1: ПАРАМЕТРЫ МАГАЗИНА (E5–E11) ───────────────────
n_sec(4, "  РАЗДЕЛ 1 — ПАРАМЕТРЫ МАГАЗИНА")
n_colhdr(4, [], INDIGO)  # header already set above
n_param(5,  "Название магазина",           "WAY MARKET №2", "text",  "Отображается в заголовках отчётов")
n_param(6,  "Дата начала учёта",            today,          "date",  "Точка отсчёта для дашборда")
n_param(7,  "Доля в фонд (маржа %)",        0.25,           "pct",   "Выручка × Доля = маржинальная прибыль")
n_param(8,  "Лимит на закуп (%)",           0.75,           "pct",   "Максимум на закуп от выручки")
n_param(9,  "Начальный долг поставщикам",   500000,         "money", "Долг поставщикам на старте учёта")
n_param(10, "Округление сумм",              "до 100 ₽",     "text",  "Подсказка при вводе данных")
n_param(11, "Период сравнения",             "Прошлый месяц","text",  "Используется в отчёте руководителю")

# ── РАЗДЕЛ 2: АКТИВНЫЕ СМЕНЫ (E15–E17) ──────────────────────
n_sec(12, "  РАЗДЕЛ 2 — АКТИВНЫЕ СМЕНЫ", TEAL)
n_colhdr(13, [(1,"Смена"),(5,"Вкл / Выкл"),(7,"Эффект")], TEAL)
ws.merge_cells("A14:H14")
ws.cell(14,1).value=" Включите только те смены, которые работают в вашем магазине"
ws.cell(14,1).font=fnt(9,it=True,col=TEAL); ws.cell(14,1).fill=F(TEAL_L); ws.row_dimensions[14].height=20
n_toggle(15, "Смена ДЕНЬ",  "Вкл", "Показывает/скрывает блок ДЕНЬ на форме ввода")
n_toggle(16, "Смена ВЕЧЕР", "Вкл", "Показывает/скрывает блок ВЕЧЕР на форме ввода")
n_toggle(17, "Смена НОЧЬ",  "Вкл", "Показывает/скрывает блок НОЧЬ на форме ввода")

# ── РАЗДЕЛ 3: Z-ОТЧЁТ (E20–E24) ─────────────────────────────
n_sec(18, "  РАЗДЕЛ 3 — Z-ОТЧЁТ (источники выручки)", BLUE)
n_colhdr(19, [(1,"Источник"),(5,"Вкл / Выкл"),(7,"Описание")], BLUE)
n_toggle(20, "Эквайринг",              "Вкл",  "Терминал безналичной оплаты")
n_toggle(21, "Перевод (СБП / банк)",   "Вкл",  "Оплата по QR-коду или банковским переводом")
n_toggle(22, "Онлайн",                 "Выкл", "Онлайн-заказы, доставка")
n_toggle(23, "Иман (хозяин)",          "Вкл",  "Личные средства владельца, внесённые в кассу")
n_toggle(24, "Выплата с кассы",        "Вкл",  "Расходы, выданные напрямую из кассы")

# ── РАЗДЕЛ 4: КОНТРОЛЬ КАССЫ (E27–E29) ──────────────────────
n_sec(25, "  РАЗДЕЛ 4 — КОНТРОЛЬ КАССЫ (сверка Z vs факт)", RED)
n_colhdr(26, [(1,"Вид сверки"),(5,"Вкл / Выкл"),(7,"Формула расхождения")], RED)
n_toggle(27, "Сверка по наличке",      "Вкл",  "Z-наличка − Выплаты − Факт наличка")
n_toggle(28, "Сверка по эквайрингу",   "Вкл",  "Z-эквайринг − Факт эквайринг")
n_toggle(29, "Сверка по переводу",     "Вкл",  "Z-перевод − Факт перевод")

# ── РАЗДЕЛ 5: ИНВЕНТАРЬ (E32–E34) ───────────────────────────
n_sec(30, "  РАЗДЕЛ 5 — ИНВЕНТАРЬ", AMBER)
n_colhdr(31, [(1,"Функция"),(5,"Вкл / Выкл"),(7,"Где отображается")], AMBER)
n_toggle(32, "Списание товара",         "Выкл", "Поле «Списание» на форме ввода расходов")
n_toggle(33, "Возврат поставщику",      "Выкл", "Поле «Возврат поставщику» на форме ввода")
n_toggle(34, "Касса утром / вечером",   "Выкл", "Поля остатков наличных в кассе")

# ── РАЗДЕЛ 6: ПОРОГИ УВЕДОМЛЕНИЙ (E37–E39) ──────────────────
n_sec(35, "  РАЗДЕЛ 6 — ПОРОГИ УВЕДОМЛЕНИЙ", GRAY)
n_colhdr(36, [(1,"Параметр"),(5,"Значение"),(7,"Где используется")], GRAY)
n_param(37, "Расхождение кассы > (₽)",   5000,    "money", "Подсветка красным на дашборде")
n_param(38, "Общий долг > (₽)",          1000000, "money", "Подсветка долга на дашборде")
n_param(39, "Просрочка > (дней)",        7,       "text",  "Выплаты в КАЛЕНДАРЬ_ВЫПЛАТ")

# ── РАЗДЕЛ 7: СПРАВОЧНИКИ (A44–G79) ─────────────────────────
n_sec(40, "  РАЗДЕЛ 7 — СПРАВОЧНИКИ (выпадающие списки)", INDIGO)
ws.merge_cells("A41:H41")
ws.cell(41,1).value="  Добавляйте значения — они сразу появятся в выпадающих меню на всех листах"
ws.cell(41,1).font=fnt(9,it=True,col=INDIGO); ws.cell(41,1).fill=F(BLUE_L); ws.row_dimensions[41].height=20
ws.row_dimensions[42].height=6

for ci_,nm_,clr_ in [(1,"Кассиры",PURPLE),(3,"Категории расходов",AMBER),(5,"Способы оплаты",BLUE),(7,"Типы операций",NAVY)]:
    ws.cell(43,ci_).value=nm_; ws.cell(43,ci_).font=fnt(10,True,"FFFFFFFF")
    ws.cell(43,ci_).fill=F(clr_); ws.cell(43,ci_).border=brd(); ws.cell(43,ci_).alignment=CA()
ws.row_dimensions[43].height=26

kass_seed=["Сотрудник 1","Сотрудник 2"]
kat_seed =["Закуп","ЗП","Маркетинг","Логистика","Хоз.нужды","Аренда","Коммунальные","Охрана","Реклама","Ремонт"]
spos_seed=["Наличка","Эквайринг","Перевод","Иман","Долг"]
tip_seed =["Доход","Расход","Долг","Оплата долга","Расхождение","Иман","Списание","Возврат","Касса"]

for ri_ in range(44,80):
    alt_=ri_%2==0
    for ci_,seed_,clr_ in [(1,kass_seed,PURP_L),(3,kat_seed,AMBER_L),(5,spos_seed,BLUE_L),(7,tip_seed,BLUE_L)]:
        idx_=ri_-44
        c_=ws.cell(ri_,ci_)
        c_.value=seed_[idx_] if idx_<len(seed_) else None
        c_.font=fnt(10); c_.fill=F(clr_ if alt_ else WHITE)
        c_.border=brd(); c_.alignment=LA(); c_.protection=prot(False)
    ws.row_dimensions[ri_].height=22

tbl_kass=Table(displayName="tblКассиры",ref="A43:A79")
tbl_kass.tableStyleInfo=TableStyleInfo(name="TableStyleLight4",showRowStripes=True); ws.add_table(tbl_kass)
tbl_kat=Table(displayName="tblКатегории",ref="C43:C79")
tbl_kat.tableStyleInfo=TableStyleInfo(name="TableStyleLight3",showRowStripes=True); ws.add_table(tbl_kat)
tbl_spos=Table(displayName="tblСпособыОплаты",ref="E43:E79")
tbl_spos.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True); ws.add_table(tbl_spos)
tbl_tip=Table(displayName="tblТипыОпераций",ref="G43:G79")
tbl_tip.tableStyleInfo=TableStyleInfo(name="TableStyleLight2",showRowStripes=True); ws.add_table(tbl_tip)

# ── РАЗДЕЛ 8: ПОСТОЯННЫЕ РАСХОДЫ ПО МЕСЯЦАМ (rows 81–94) ───
n_sec(80, "  РАЗДЕЛ 8 — ПОСТОЯННЫЕ РАСХОДЫ ПО МЕСЯЦАМ (справочно)", TEAL)
hrow(ws,81,["Месяц","ЗП","Аренда","Налоги","Интернет","Охрана","Другое","ИТОГО"],TEAL,26)
for ri_,mon_ in enumerate(MONTHS_RU,82):
    ws.cell(ri_,1).value=mon_; ws.cell(ri_,1).font=fnt(10); ws.cell(ri_,1).fill=F(LGRAY)
    ws.cell(ri_,1).border=brd(); ws.cell(ri_,1).alignment=LA()
    for ci_ in range(2,8):
        c_=ws.cell(ri_,ci_); c_.font=fnt(10); c_.fill=F(INP); c_.border=brd()
        c_.alignment=RA(); c_.number_format=MONEY; c_.protection=prot(False)
    ws.cell(ri_,8).value=f"=IFERROR(SUM(B{ri_}:G{ri_}),0)"
    ws.cell(ri_,8).font=fnt(10,True,TEAL); ws.cell(ri_,8).fill=F(TEAL_L)
    ws.cell(ri_,8).border=brd(); ws.cell(ri_,8).alignment=RA(); ws.cell(ri_,8).number_format=MONEY
    ws.row_dimensions[ri_].height=22
ws.cell(94,1).value="ИТОГО (год)"; ws.cell(94,1).font=fnt(10,True,TEAL); ws.cell(94,1).fill=F(TEAL_L)
ws.cell(94,1).border=brd(); ws.cell(94,1).alignment=LA(); ws.row_dimensions[94].height=24
for ci_,cl_ in enumerate("BCDEFGH",2):
    ws.cell(94,ci_).value=f"=IFERROR(SUM({cl_}82:{cl_}93),0)"
    ws.cell(94,ci_).font=fnt(10,True,TEAL); ws.cell(94,ci_).fill=F(TEAL_L)
    ws.cell(94,ci_).border=brd(); ws.cell(94,ci_).alignment=RA(); ws.cell(94,ci_).number_format=MONEY
tbl_post=Table(displayName="tblПостоянные",ref="A81:H94")
tbl_post.tableStyleInfo=TableStyleInfo(name="TableStyleLight6",showRowStripes=True); ws.add_table(tbl_post)

# ── РАЗДЕЛ 9: СПРАВОЧНИК ПОСТАВЩИКОВ (ТП) (rows 98–1098) ────
ws.row_dimensions[95].height=8
n_sec(96, "  РАЗДЕЛ 9 — СПРАВОЧНИК ПОСТАВЩИКОВ / ТОРГОВЫХ ПРЕДСТАВИТЕЛЕЙ", PURPLE)
ws.merge_cells("A97:H97")
ws.cell(97,1).value="  Добавляйте ТП — они появятся в выпадающем списке на листе ЗАПИСЬ_НА_ВЫПЛАТУ"
ws.cell(97,1).font=fnt(9,it=True,col=PURPLE); ws.cell(97,1).fill=F(PURP_L); ws.row_dimensions[97].height=20
hrow(ws,98,["№","Название ТП / Поставщика","Телефон / Контакт","","","","",""],PURPLE,26)
_sup_ex={99:"ТД Метро / Metro Cash&Carry", 100:"ООО Лента Оптторг", 101:"Вкусвилл"}
for ri_ in range(99,1099):
    alt_=ri_%2==0; bg_=PURP_L if alt_ else WHITE
    ws.cell(ri_,1).value=f'=IF(B{ri_}="","",ROW()-98)'
    ws.cell(ri_,1).font=fnt(9,col=GRAY); ws.cell(ri_,1).fill=F(bg_)
    ws.cell(ri_,1).border=brd(); ws.cell(ri_,1).alignment=CA()
    ws.cell(ri_,2).value=_sup_ex.get(ri_,"")
    ws.cell(ri_,2).font=fnt(10); ws.cell(ri_,2).fill=F(INP if ri_<=101 else bg_)
    ws.cell(ri_,2).border=brd(); ws.cell(ri_,2).alignment=LA(); ws.cell(ri_,2).protection=prot(False)
    ws.cell(ri_,3).font=fnt(10); ws.cell(ri_,3).fill=F(bg_)
    ws.cell(ri_,3).border=brd(); ws.cell(ri_,3).alignment=LA(); ws.cell(ri_,3).protection=prot(False)
    ws.row_dimensions[ri_].height=20
tbl_tp=Table(displayName="tblПоставщики",ref="A98:C1098")
tbl_tp.tableStyleInfo=TableStyleInfo(name="TableStyleLight5",showRowStripes=True); ws.add_table(tbl_tp)

# ── DataValidation: Вкл/Выкл на всех переключателях ─────────
dv_vv=DataValidation(type="list",formula1='"Вкл,Выкл"'); ws.add_data_validation(dv_vv)
for rng_ in ["E15:E17","E20:E24","E27:E29","E32:E34"]:
    dv_vv.add(rng_)

# ── CF: Вкл → зелёный, Выкл → красный ───────────────────────
for rng_,anch_ in [("E15:E17","$E15"),("E20:E24","$E20"),("E27:E29","$E27"),("E32:E34","$E32")]:
    ws.conditional_formatting.add(rng_,FormulaRule(formula=[f'{anch_}="Вкл"'],  fill=F(GREEN_L),font=fnt(11,True,GREEN)))
    ws.conditional_formatting.add(rng_,FormulaRule(formula=[f'{anch_}="Выкл"'], fill=F(RED_L),  font=fnt(11,True,RED)))

ws.freeze_panes="A3"
ws.sheet_properties.tabColor="FF374151"
print("✓ НАСТРОЙКИ (9 разделов)")

# ════════════════════════════════════════════════════════════
# 2. БАЗА_ДДС (умная таблица tblБаза)
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("БАЗА_ДДС"); ws.sheet_view.showGridLines = False
banner(ws, "БАЗА ДДС — ТРАНЗАКЦИОННЫЙ ЛОГ (умная таблица tblБаза)", "A1:H1", NAVY)
ws.merge_cells("A2:H2")
ws.cell(2,1).value = "Умная таблица tblБаза — фильтрация по любому столбцу. Данные вносятся через ВВОД_КАССА и ВВОД_РАСХОДЫ."
ws.cell(2,1).font=fnt(10,it=True,col=BLUE); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22

headers_b=["Дата","Смена","Кассир","Тип операции","Категория","Способ оплаты","Сумма","Комментарий"]
hrow(ws,3,headers_b,NAVY,32)
for r in range(4,3004):
    alt=r%2==0
    for ci in range(1,9):
        c=ws.cell(r,ci); c.border=brd(); c.fill=F(LGRAY if alt else WHITE); c.font=fnt(10)
        c.alignment=CA() if ci in [2,4,6] else LA() if ci in [3,5,8] else RA() if ci==7 else CA()
        c.protection=prot(False)
    ws.cell(r,1).number_format=DATE_F; ws.cell(r,7).number_format=MONEY
    ws.row_dimensions[r].height=20

for tipo,fill_,font_ in [("Доход",BLUE_L,BLUE),("Расход",RED_L,RED),("Долг",AMBER_L,AMBER),
                          ("Оплата долга",GREEN_L,GREEN),("Расхождение",RED_L,RED),
                          ("Иман",PURP_L,PURPLE),("Списание",RED_L,RED),("Возврат",GREEN_L,GREEN),("Касса",TEAL_L,TEAL)]:
    ws.conditional_formatting.add("D4:D3003",FormulaRule(formula=[f'$D4="{tipo}"'],fill=F(fill_),font=fnt(10,True,font_)))

dv_tp=DataValidation(type="list",formula1='"Доход,Расход,Долг,Оплата долга,Расхождение,Иман,Списание,Возврат,Касса"'); ws.add_data_validation(dv_tp); dv_tp.add("D4:D3003")
dv_pay=DataValidation(type="list",formula1="=НАСТРОЙКИ!$E$44:$E$79"); ws.add_data_validation(dv_pay); dv_pay.add("F4:F3003")
dv_cat=DataValidation(type="list",formula1="=НАСТРОЙКИ!$C$44:$C$79"); ws.add_data_validation(dv_cat); dv_cat.add("E4:E3003")
dv_ksr=DataValidation(type="list",formula1="=НАСТРОЙКИ!$A$44:$A$79"); ws.add_data_validation(dv_ksr); dv_ksr.add("C4:C3003")
dv_sm=DataValidation(type="list",formula1='"День,Вечер,Ночь,-"'); ws.add_data_validation(dv_sm); dv_sm.add("B4:B3003")

tbl_b=Table(displayName="tblБаза",ref="A3:H3003")
tbl_b.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showFirstColumn=False)
ws.add_table(tbl_b)
cw(ws,{"A":12,"B":10,"C":18,"D":16,"E":18,"F":14,"G":14,"H":30})
ws.freeze_panes="A4"; ws.sheet_properties.tabColor="FF6B7280"

print("✓ БАЗА_ДДС")

# ════════════════════════════════════════════════════════════
# 3. ВВОД_КАССА  (10 колонок, Stage 4)
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("ВВОД_КАССА"); ws.sheet_view.showGridLines = False
banner(ws, "ВВОД ДАННЫХ КАССЫ  |  ► кнопки: СОХРАНИТЬ / СЕГОДНЯ / ВЧЕРА", "A1:J1", BLUE)
ws.merge_cells("A2:J2")
ws.cell(2,1).value="Введите Z-отчёты и фактику по трём каналам. Расхождение рассчитывается автоматически."
ws.cell(2,1).font=fnt(10,it=True,col=BLUE); ws.cell(2,1).fill=F(BLUE_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22
# Smart table tblВводКасса — 10 колонок
hdrs_k=["Дата","Смена","Кассир","Выручка (Z-отчёт)","Эквайринг (Z)","Перевод (Z)","Факт.наличка","Факт.эквайринг","Факт.перевод","Расхождение"]
hrow(ws,3,hdrs_k,BLUE,30)
ws.row_dimensions[3].height=30
dv_sm_k=DataValidation(type="list",formula1='"День,Вечер,Ночь,-"'             ); ws.add_data_validation(dv_sm_k)
dv_ks_k=DataValidation(type="list",formula1="=НАСТРОЙКИ!$A$44:$A$79"        ); ws.add_data_validation(dv_ks_k)
for r in range(4,504):
    alt=r%2==0; bg=LGRAY if alt else WHITE; seed=r<=6
    ws.cell(r,1).font=fnt(10); ws.cell(r,1).fill=F(INP if seed else bg); ws.cell(r,1).border=brd()
    ws.cell(r,1).alignment=CA(); ws.cell(r,1).number_format=DATE_F; ws.cell(r,1).protection=prot(False)
    if seed: ws.cell(r,1).value=today
    if r==4: ws.cell(r,2).value="День"
    elif r==5: ws.cell(r,2).value="Вечер"
    elif r==6: ws.cell(r,2).value="Ночь"
    ws.cell(r,2).font=fnt(10); ws.cell(r,2).fill=F(INP if seed else bg); ws.cell(r,2).border=brd()
    ws.cell(r,2).alignment=CA(); ws.cell(r,2).protection=prot(False)
    ws.cell(r,3).font=fnt(10); ws.cell(r,3).fill=F(INP if seed else bg); ws.cell(r,3).border=brd()
    ws.cell(r,3).alignment=LA(); ws.cell(r,3).protection=prot(False)
    # cols 4-9: Z-отчёты (D-F) + фактика (G-I)
    for ci in range(4,10):
        c=ws.cell(r,ci); c.font=fnt(10,True,INDIGO); c.fill=F(INP if seed else bg)
        c.border=brd(); c.alignment=RA(); c.number_format=MONEY; c.protection=prot(False)
    # col 10: Расхождение = (G+H+I) - (D+E+F)
    ws.cell(r,10).value=f"=IFERROR((G{r}+H{r}+I{r})-(D{r}+E{r}+F{r}),0)"
    ws.cell(r,10).font=fnt(10,True,RED); ws.cell(r,10).fill=F(RED_L if seed else bg)
    ws.cell(r,10).border=brd(); ws.cell(r,10).alignment=RA(); ws.cell(r,10).number_format=MONEY
    ws.row_dimensions[r].height=22
dv_sm_k.add("B4:B503"); dv_ks_k.add("C4:C503")
tbl_vk=Table(displayName="tblВводКасса",ref="A3:J503")
tbl_vk.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showFirstColumn=False)
ws.add_table(tbl_vk)
cw(ws,{"A":12,"B":10,"C":16,"D":17,"E":14,"F":13,"G":14,"H":15,"I":14,"J":14})
# CF: dim columns whose payment method is disabled in НАСТРОЙКИ
_DIM_=PatternFill("solid",fgColor="FFD1D5DB",start_color="FFD1D5DB")
_DIMF_=fnt(10,col="FF9CA3AF")
for _rng_,_tgl_ in [("E4:E503","НАСТРОЙКИ!$E$20"),("F4:F503","НАСТРОЙКИ!$E$21"),
                    ("H4:H503","НАСТРОЙКИ!$E$28"),("I4:I503","НАСТРОЙКИ!$E$29")]:
    ws.conditional_formatting.add(_rng_,FormulaRule(formula=[f'{_tgl_}="Выкл"'],fill=_DIM_,font=_DIMF_))
ws.freeze_panes="A4"; ws.sheet_properties.tabColor="FF3B82F6"
print("✓ ВВОД_КАССА")

# ════════════════════════════════════════════════════════════
# 4. ВВОД_РАСХОДЫ  (5 колонок без Кассира, Stage 4)
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("ВВОД_РАСХОДЫ"); ws.sheet_view.showGridLines = False
banner(ws, "ВВОД РАСХОДОВ  |  ► кнопка СОХРАНИТЬ РАСХОДЫ", "A1:E1", RED)
ws.merge_cells("A2:E2")
ws.cell(2,1).value="Заполните таблицу расходов. Нажмите СОХРАНИТЬ РАСХОДЫ."
ws.cell(2,1).font=fnt(10,it=True,col=RED); ws.cell(2,1).fill=F(RED_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22
# Smart table tblВводРасходы — 5 колонок (без Кассира)
hdrs_r=["Дата","Категория","Способ оплаты","Сумма","Комментарий"]
hrow(ws,3,hdrs_r,RED,30)
ws.row_dimensions[3].height=30
dv_cat_r=DataValidation(type="list",formula1="=НАСТРОЙКИ!$C$44:$C$79"); ws.add_data_validation(dv_cat_r)
dv_pay_r=DataValidation(type="list",formula1="=НАСТРОЙКИ!$E$44:$E$79"); ws.add_data_validation(dv_pay_r)
for r in range(4,504):
    alt=r%2==0; bg=LGRAY if alt else WHITE; seed=r==4
    if seed: ws.cell(r,1).value=today
    ws.cell(r,1).font=fnt(10); ws.cell(r,1).fill=F(INP if seed else bg)
    ws.cell(r,1).border=brd(); ws.cell(r,1).alignment=CA()
    ws.cell(r,1).number_format=DATE_F; ws.cell(r,1).protection=prot(False)
    ws.cell(r,2).font=fnt(10); ws.cell(r,2).fill=F(INP if seed else bg)
    ws.cell(r,2).border=brd(); ws.cell(r,2).alignment=LA(); ws.cell(r,2).protection=prot(False)
    ws.cell(r,3).font=fnt(10); ws.cell(r,3).fill=F(INP if seed else bg)
    ws.cell(r,3).border=brd(); ws.cell(r,3).alignment=CA(); ws.cell(r,3).protection=prot(False)
    ws.cell(r,4).font=fnt(11,True,INDIGO); ws.cell(r,4).fill=F(INP if seed else bg)
    ws.cell(r,4).border=brd(); ws.cell(r,4).alignment=RA()
    ws.cell(r,4).number_format=MONEY; ws.cell(r,4).protection=prot(False)
    ws.cell(r,5).font=fnt(10); ws.cell(r,5).fill=F(bg)
    ws.cell(r,5).border=brd(); ws.cell(r,5).alignment=LA(); ws.cell(r,5).protection=prot(False)
    ws.row_dimensions[r].height=22
dv_cat_r.add("B4:B503"); dv_pay_r.add("C4:C503")
tbl_vr=Table(displayName="tblВводРасходы",ref="A3:E503")
tbl_vr.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True,showFirstColumn=False)
ws.add_table(tbl_vr)
cw(ws,{"A":12,"B":22,"C":16,"D":16,"E":34})
ws.freeze_panes="A4"; ws.sheet_properties.tabColor="FFEF4444"
print("✓ ВВОД_РАСХОДЫ")

# ════════════════════════════════════════════════════════════
# 5. ЗАПИСЬ_НА_ВЫПЛАТУ (умная таблица tblВыплаты)
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("ЗАПИСЬ_НА_ВЫПЛАТУ"); ws.sheet_view.showGridLines = False
banner(ws, "ЗАПИСЬ НА ВЫПЛАТУ ПОСТАВЩИКАМ (умная таблица tblВыплаты)", "A1:K1", PURPLE)
ws.merge_cells("A2:K2")
ws.cell(2,1).value="Вносите выплаты вручную. Календарь выплат обновится автоматически."
ws.cell(2,1).font=fnt(10,it=True,col=PURPLE); ws.cell(2,1).fill=F(PURP_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22

hdrs_z=["№","Дата выплаты","Поставщик (ТП)","Сумма (₽)","Статус","Накладная №","Способ оплаты","Комментарий","","","Idx"]
hrow(ws,3,hdrs_z,PURPLE,30)
ws.column_dimensions["K"].hidden=True

dv_st_z=DataValidation(type="list",formula1='"Запланировано,Выплачено,Просрочено,Отменено"')
ws.add_data_validation(dv_st_z); dv_st_z.add("E4:E503")
dv_tp_z=DataValidation(type="list",formula1="=НАСТРОЙКИ!$B$99:$B$1098",allow_blank=True)
ws.add_data_validation(dv_tp_z); dv_tp_z.add("C4:C503")
dv_sp_z=DataValidation(type="list",formula1="=НАСТРОЙКИ!$E$44:$E$79",allow_blank=True)
ws.add_data_validation(dv_sp_z); dv_sp_z.add("G4:G503")

for r in range(4,504):
    alt=r%2==0
    bg=PURP_L if alt else WHITE
    ws.cell(r,1).value=f'=IF(C{r}="","",ROW()-3)'
    for ci in range(1,9):
        c=ws.cell(r,ci); c.border=brd(); c.fill=F(bg); c.font=fnt(10,bold=(ci==3))
        c.alignment=CA() if ci in [1,2,5,7] else LA() if ci in [3,6,8] else RA() if ci==4 else CA()
    ws.cell(r,2).number_format=DATE_F; ws.cell(r,4).number_format=MONEY
    ws.cell(r,11).value=f'=IF($B{r}="",99999,COUNTIFS($B$4:$B{r},$B{r}))'
    ws.cell(r,11).font=fnt(8,col=GRAY); ws.row_dimensions[r].height=22

for tipo_z,fill_z,fn_z in [("Запланировано",BLUE_L,BLUE),("Выплачено",GREEN_L,GREEN),
                             ("Просрочено",RED_L,RED),("Отменено",LGRAY,GRAY)]:
    ws.conditional_formatting.add("E4:E503",FormulaRule(formula=[f'$E4="{tipo_z}"'],fill=F(fill_z),font=fnt(10,True,fn_z)))
ws.conditional_formatting.add("B4:B503",FormulaRule(
    formula=['AND($E4="Запланировано",$B4<TODAY())'],fill=F(RED_L),font=fnt(10,True,RED)))

tbl_z=Table(displayName="tblВыплаты",ref="A3:K503")
tbl_z.tableStyleInfo=TableStyleInfo(name="TableStyleMedium5",showRowStripes=True)
ws.add_table(tbl_z)
cw(ws,{"A":5,"B":13,"C":24,"D":14,"E":15,"F":14,"G":13,"H":24,"I":4,"J":4,"K":4})
ws.freeze_panes="A4"; ws.sheet_properties.tabColor="FF"+PURPLE[2:]
print("✓ ЗАПИСЬ_НА_ВЫПЛАТУ (tblВыплаты)")

# ════════════════════════════════════════════════════════════
# 6. КАЛЕНДАРЬ_ВЫПЛАТ — интерактивный календарь + боковая панель
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("КАЛЕНДАРЬ_ВЫПЛАТ"); ws.sheet_view.showGridLines = False
banner(ws, "КАЛЕНДАРЬ ВЫПЛАТ ПОСТАВЩИКАМ", "A1:N1", PURPLE)
ws.merge_cells("A2:N2")
ws.cell(2,1).value="Выберите месяц и год. Кликните на день — справа появится список выплат за этот день."
ws.cell(2,1).font=fnt(10,it=True,col=PURPLE); ws.cell(2,1).fill=F(PURP_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=22

# Filter bar row 4
ws.cell(4,1).value="Месяц:"; ws.cell(4,1).font=fnt(11,True); ws.cell(4,1).alignment=RA(); ws.cell(4,1).fill=F(LGRAY)
ws.merge_cells("B4:D4")
ws.cell(4,2).value=MONTHS_RU[today.month-1]
ws.cell(4,2).font=fnt(14,True,INDIGO); ws.cell(4,2).fill=F(INP); ws.cell(4,2).border=brd_med(); ws.cell(4,2).alignment=CA()
dv_mon_k=DataValidation(type="list",formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"')
ws.add_data_validation(dv_mon_k); dv_mon_k.add("B4")
ws.cell(4,6).value="Год:"; ws.cell(4,6).font=fnt(11,True); ws.cell(4,6).alignment=RA(); ws.cell(4,6).fill=F(LGRAY)
ws.merge_cells("G4:H4")
ws.cell(4,7).value=today.year
ws.cell(4,7).font=fnt(14,True,INDIGO); ws.cell(4,7).fill=F(INP); ws.cell(4,7).border=brd_med(); ws.cell(4,7).alignment=CA()
dv_yr_k=DataValidation(type="list",formula1='"2024,2025,2026,2027,2028"')
ws.add_data_validation(dv_yr_k); dv_yr_k.add("G4")
ws.row_dimensions[4].height=36

# Hidden helper cells (column P)
ws.cell(4,16).value='=MATCH(B4,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0)'
ws.cell(5,16).value='=DATE(G4,P4,1)'; ws.cell(5,16).number_format=DATE_F
ws.cell(6,16).value='=EOMONTH(P5,0)'; ws.cell(6,16).number_format=DATE_F
ws.cell(7,16).value='=WEEKDAY(P5,2)'
ws.column_dimensions['P'].hidden=True

# KPI summary row 6-8
sec_hdr(ws, 6, "  СВОДКА ПО ВЫБРАННОМУ МЕСЯЦУ", 14, INDIGO)
def kpi_cal(ws, row, c1, c2, lbl_, f_, bg):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c=ws.cell(row,c1); c.value=lbl_; c.font=fnt(10,True,"FFBBBBBB"); c.fill=F("FF1F2937"); c.alignment=CA()
    ws.merge_cells(start_row=row+1, start_column=c1, end_row=row+1, end_column=c2)
    c=ws.cell(row+1,c1); c.value=f_; c.font=fnt(18,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=CA(); c.number_format=MONEY
    ws.row_dimensions[row].height=20; ws.row_dimensions[row+1].height=40

kpi_cal(ws,7,1,3,"К ВЫПЛАТЕ ВСЕГО",
    '=IFERROR(SUMPRODUCT((ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503>=$P$5)*(ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503<=$P$6)*ЗАПИСЬ_НА_ВЫПЛАТУ!$D$4:$D$503),0)',PURPLE)
kpi_cal(ws,7,5,7,"ВЫПЛАЧЕНО",
    '=IFERROR(SUMPRODUCT((ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503>=$P$5)*(ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503<=$P$6)*(ЗАПИСЬ_НА_ВЫПЛАТУ!$E$4:$E$503="Выплачено")*ЗАПИСЬ_НА_ВЫПЛАТУ!$D$4:$D$503),0)',GREEN)
kpi_cal(ws,7,9,11,"ЗАПЛАНИРОВАНО",
    '=IFERROR(SUMPRODUCT((ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503>=$P$5)*(ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503<=$P$6)*(ЗАПИСЬ_НА_ВЫПЛАТУ!$E$4:$E$503="Запланировано")*ЗАПИСЬ_НА_ВЫПЛАТУ!$D$4:$D$503),0)',AMBER)
kpi_cal(ws,7,12,14,"ПРОСРОЧЕНО",
    '=IFERROR(SUMPRODUCT((ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503>=$P$5)*(ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503<=$P$6)*(ЗАПИСЬ_НА_ВЫПЛАТУ!$E$4:$E$503="Просрочено")*ЗАПИСЬ_НА_ВЫПЛАТУ!$D$4:$D$503),0)',RED)

# Day-of-week header row 10
ws.row_dimensions[10].height=28
for i,d in enumerate(['ПН','ВТ','СР','ЧТ','ПТ','СБ','ВС']):
    cs=1+i*2
    ws.merge_cells(start_row=10,start_column=cs,end_row=10,end_column=cs+1)
    c=ws.cell(10,cs); c.value=d; c.font=fnt(10,True,"FFFFFFFF")
    c.fill=F(RED if i>=5 else NAVY); c.alignment=CA(); c.border=brd()

# 6-week calendar grid (rows 11-34, 4 rows per week)
CELL_ROWS=4; SR=11
for week in range(6):
    br=SR+week*CELL_ROWS
    for dp in range(7):
        cs=1+dp*2; ce=cs+1; di=week*7+dp
        dr=f'DATE($G$4,$P$4,1)+{di}-($P$7-1)'
        check=f"${get_column_letter(cs)}${br}"
        # Day number + total amount
        ws.cell(br,cs).value=f'=IFERROR(IF(AND({dr}>=$P$5,{dr}<=$P$6),DAY({dr}),""),"")'
        ws.cell(br,cs).font=fnt(13,True,NAVY); ws.cell(br,cs).alignment=LA()
        ws.cell(br,ce).value=f'=IFERROR(IF({check}="","",SUMPRODUCT((ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503={dr})*ЗАПИСЬ_НА_ВЫПЛАТУ!$D$4:$D$503)),"")'
        ws.cell(br,ce).font=fnt(11,True,RED); ws.cell(br,ce).alignment=RA(); ws.cell(br,ce).number_format='#,##0;;;'
        # Up to 2 suppliers
        for tp_i in range(2):
            rt=br+1+tp_i; n=tp_i+1
            ws.cell(rt,cs).value=f'=IFERROR(IF({check}="","",INDEX(ЗАПИСЬ_НА_ВЫПЛАТУ!$C:$C,SUMPRODUCT((ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503={dr})*(ЗАПИСЬ_НА_ВЫПЛАТУ!$K$4:$K$503={n})*ROW(ЗАПИСЬ_НА_ВЫПЛАТУ!$B$4:$B$503)))),"")'
            ws.cell(rt,cs).font=fnt(8,col=BLUE); ws.cell(rt,cs).alignment=LA()
            ws.merge_cells(start_row=rt,start_column=cs,end_row=rt,end_column=ce)
        # "+ N more" hint
        ws.cell(br+3,cs).value=f'=IFERROR(IF(OR({check}="",COUNTIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!$B:$B,{dr})<=2),"","+ ещё "&(COUNTIFS(ЗАПИСЬ_НА_ВЫПЛАТУ!$B:$B,{dr})-2)),"")'
        ws.cell(br+3,cs).font=fnt(8,it=True,col=GRAY); ws.cell(br+3,cs).alignment=CA()
        ws.merge_cells(start_row=br+3,start_column=cs,end_row=br+3,end_column=ce)
        # Cell background
        bg_c="FFFFF5F5" if dp>=5 else "FFFAFAFA"
        for ri in range(br,br+4):
            for ci_ in [cs,ce]:
                ws.cell(ri,ci_).fill=F(bg_c); ws.cell(ri,ci_).border=brd("FF999999" if ri==br else BORDER)
        for ri in [br,br+1,br+2,br+3]:
            ws.row_dimensions[ri].height=22 if ri==br else 16

# ── Sidebar (columns O-R): populated by VBA SelectionChange ──────────────────
SB=15  # column O
ws.merge_cells(start_row=1,start_column=SB,end_row=1,end_column=SB+3)
ws.cell(1,SB).value="ДЕТАЛИ ДНЯ"; ws.cell(1,SB).font=fnt(12,True,"FFFFFFFF")
ws.cell(1,SB).fill=F(INDIGO); ws.cell(1,SB).alignment=CA()

sidebar_hdrs=["Поставщик","Сумма","Статус","Накладная"]
for si,h in enumerate(sidebar_hdrs):
    ws.cell(2,SB+si).value=h; ws.cell(2,SB+si).font=fnt(9,True,"FFFFFFFF")
    ws.cell(2,SB+si).fill=F(NAVY); ws.cell(2,SB+si).alignment=CA(); ws.cell(2,SB+si).border=brd()

ws.merge_cells(start_row=3,start_column=SB,end_row=3,end_column=SB+3)
ws.cell(3,SB).value="← Кликните на день в календаре"
ws.cell(3,SB).font=fnt(10,it=True,col=GRAY); ws.cell(3,SB).fill=F(LGRAY); ws.cell(3,SB).alignment=CA()

for r in range(4,30):
    for si in range(4):
        c=ws.cell(r,SB+si); c.fill=F(LGRAY if r%2==0 else WHITE); c.border=brd(); c.font=fnt(10)
        c.alignment=LA() if si in [0,3] else RA() if si==1 else CA()
    ws.cell(r,SB+1).number_format=MONEY; ws.row_dimensions[r].height=20

for col_i in range(1,19):
    ws.column_dimensions[get_column_letter(col_i)].width=13
# Wider sidebar columns
ws.column_dimensions[get_column_letter(SB)].width=22
ws.column_dimensions[get_column_letter(SB+1)].width=14
ws.column_dimensions[get_column_letter(SB+2)].width=14
ws.column_dimensions[get_column_letter(SB+3)].width=16

ws.freeze_panes="A11"; ws.sheet_properties.tabColor="FF"+PURPLE[2:]
print("✓ КАЛЕНДАРЬ_ВЫПЛАТ (интерактивный)")

# ════════════════════════════════════════════════════════════
# 7. ДАННЫЕ (hidden sheet for chart data)
# ════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("ДАННЫЕ"); ws_d.sheet_view.showGridLines = False
ws_d.sheet_state = "hidden"

# Monthly aggregates table (rows 2-14): header + 12 months
hrow(ws_d, 1, ["Месяц","Выручка","Расходы","Закуп","Прибыль","Долг"], NAVY, 20)
month_names_short = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]
for mi,mon in enumerate(month_names_short,1):
    r=mi+1; yr=today.year
    start_d=f"DATE({yr},{mi},1)"
    end_d=f"EOMONTH(DATE({yr},{mi},1),0)"
    ws_d.cell(r,1).value=mon
    ws_d.cell(r,2).value=f'=SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=({start_d}))*(БАЗА_ДДС!$A$4:$A$3003<=({end_d}))*(БАЗА_ДДС!$D$4:$D$3003="Доход")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,3).value=f'=SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=({start_d}))*(БАЗА_ДДС!$A$4:$A$3003<=({end_d}))*(БАЗА_ДДС!$D$4:$D$3003="Расход")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,4).value=f'=SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=({start_d}))*(БАЗА_ДДС!$A$4:$A$3003<=({end_d}))*(БАЗА_ДДС!$E$4:$E$3003="Закуп товара")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,5).value=f"=B{r}-C{r}"
    ws_d.cell(r,6).value=f'=SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=({start_d}))*(БАЗА_ДДС!$A$4:$A$3003<=({end_d}))*(БАЗА_ДДС!$D$4:$D$3003="Долг")*БАЗА_ДДС!$G$4:$G$3003)'
    for ci in range(2,7): ws_d.cell(r,ci).number_format=MONEY
    ws_d.row_dimensions[r].height=18

# Shift data (rows 17-20): label + 3 shifts
hrow(ws_d, 16, ["Смена","Выручка","Записей"], NAVY, 20)
for si,(sname) in enumerate(["День","Вечер","Ночь"],1):
    r=16+si
    ws_d.cell(r,1).value=sname
    ws_d.cell(r,2).value=f'=SUMPRODUCT((БАЗА_ДДС!$B$4:$B$3003="{sname}")*(БАЗА_ДДС!$D$4:$D$3003="Доход")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,3).value=f'=SUMPRODUCT((БАЗА_ДДС!$B$4:$B$3003="{sname}")*(БАЗА_ДДС!$D$4:$D$3003="Доход")*(БАЗА_ДДС!$G$4:$G$3003<>""))'
    ws_d.cell(r,2).number_format=MONEY; ws_d.row_dimensions[r].height=18

# Top expense categories (rows 24-35)
hrow(ws_d, 23, ["Категория","Сумма"], NAVY, 20)
exp_cats=["Закуп товара","Зарплата","Аренда","Коммунальные","Налог","ГСМ","Расходный материал","Маркетинг","Охрана","Прочие расходы"]
for ei,cat in enumerate(exp_cats,1):
    r=23+ei
    ws_d.cell(r,1).value=cat
    ws_d.cell(r,2).value=f'=SUMPRODUCT((БАЗА_ДДС!$E$4:$E$3003="{cat}")*(БАЗА_ДДС!$D$4:$D$3003="Расход")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,2).number_format=MONEY; ws_d.row_dimensions[r].height=18

# Last 30 days (rows 38-68)
hrow(ws_d, 37, ["День","Выручка","Расходы"], NAVY, 20)
ws_d.cell(38,1).value="=TODAY()-29"; ws_d.cell(38,1).number_format=DATE_F
for i in range(1,30):
    ws_d.cell(38+i,1).value=f"=A{37+i}+1"; ws_d.cell(38+i,1).number_format=DATE_F
for r in range(38,68):
    ws_d.cell(r,2).value=f'=SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003=A{r})*(БАЗА_ДДС!$D$4:$D$3003="Доход")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,3).value=f'=SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003=A{r})*(БАЗА_ДДС!$D$4:$D$3003="Расход")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,2).number_format=MONEY; ws_d.cell(r,3).number_format=MONEY; ws_d.row_dimensions[r].height=16

# Debt dynamics (rows 71-82): monthly debt totals
hrow(ws_d, 70, ["Месяц","Долг взят","Долг погашен","Баланс"], NAVY, 20)
for mi,mon in enumerate(month_names_short,1):
    r=70+mi; yr=today.year
    start_d=f"DATE({yr},{mi},1)"; end_d=f"EOMONTH(DATE({yr},{mi},1),0)"
    ws_d.cell(r,1).value=mon
    ws_d.cell(r,2).value=f'=SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=({start_d}))*(БАЗА_ДДС!$A$4:$A$3003<=({end_d}))*(БАЗА_ДДС!$D$4:$D$3003="Долг")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,3).value=f'=SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=({start_d}))*(БАЗА_ДДС!$A$4:$A$3003<=({end_d}))*(БАЗА_ДДС!$D$4:$D$3003="Оплата долга")*БАЗА_ДДС!$G$4:$G$3003)'
    ws_d.cell(r,4).value=f"=B{r}-C{r}"
    for ci in range(2,5): ws_d.cell(r,ci).number_format=MONEY
    ws_d.row_dimensions[r].height=16

cw(ws_d,{"A":14,"B":16,"C":16,"D":16,"E":16,"F":16})
print("✓ ДАННЫЕ (hidden)")

# ════════════════════════════════════════════════════════════
# 8. ДАШБОРД — main dashboard with 26 KPIs + 6 charts
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("ДАШБОРД"); ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 90

# Row 1: Banner
banner(ws, "WAY MARKET — ДАШБОРД УПРАВЛЕНЧЕСКОГО УЧЁТА", "A1:L1", NAVY, 16)

# Row 2: Store info
ws.merge_cells("A2:E2")
ws.cell(2,1).value='=НАСТРОЙКИ!E5&"  |  Данные с: "&TEXT(НАСТРОЙКИ!E6,"DD.MM.YYYY")'
ws.cell(2,1).font=fnt(11,False,GRAY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=LA()
ws.merge_cells("F2:L2")
ws.cell(2,6).value='="Всего записей: "&COUNTA(БАЗА_ДДС!$A$4:$A$3003)&"  |  Последнее обновление: "&TEXT(NOW(),"DD.MM.YYYY HH:MM")'
ws.cell(2,6).font=fnt(10,False,GRAY); ws.cell(2,6).fill=F(LGRAY); ws.cell(2,6).alignment=RA()
ws.row_dimensions[2].height=24

# Row 3: spacer
ws.row_dimensions[3].height=6

# Row 4: Filter bar
ws.merge_cells("A4:A4")
ws.cell(4,1).value="ПЕРИОД:"; ws.cell(4,1).font=fnt(10,True,"FFFFFFFF"); ws.cell(4,1).fill=F(INDIGO); ws.cell(4,1).alignment=CA()

# B4: Month dropdown
ws.merge_cells("B4:C4")
c_month = ws.cell(4,2); c_month.value = MONTHS_RU[today.month-1]
c_month.font=fnt(12,True,INDIGO); c_month.fill=F(INP); c_month.border=brd_med(); c_month.alignment=CA()
c_month.protection=prot(False)
dv_mon_dash=DataValidation(type="list",formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"')
ws.add_data_validation(dv_mon_dash); dv_mon_dash.add("B4:C4")

ws.cell(4,4).value="ГОД:"; ws.cell(4,4).font=fnt(10,True,"FFFFFFFF"); ws.cell(4,4).fill=F(INDIGO); ws.cell(4,4).alignment=CA()

# E4: Year dropdown
ws.merge_cells("E4:F4")
c_year = ws.cell(4,5); c_year.value = today.year
c_year.font=fnt(12,True,INDIGO); c_year.fill=F(INP); c_year.border=brd_med(); c_year.alignment=CA()
c_year.protection=prot(False)
dv_yr_dash=DataValidation(type="list",formula1='"2024,2025,2026,2027,2028"')
ws.add_data_validation(dv_yr_dash); dv_yr_dash.add("E4:F4")

# Period info
ws.merge_cells("G4:I4")
ws.cell(4,7).value='=TEXT(A5,"DD.MM.YYYY")&" — "&TEXT(B5,"DD.MM.YYYY")'
ws.cell(4,7).font=fnt(10,False,GRAY); ws.cell(4,7).fill=F(LGRAY); ws.cell(4,7).alignment=CA()

# Tip
ws.merge_cells("J4:L4")
ws.cell(4,10).value='[ОБНОВИТЬ] — кнопка VBA  |  Ctrl+Shift+D'
ws.cell(4,10).font=fnt(9,it=True,col=GRAY); ws.cell(4,10).fill=F(LGRAY); ws.cell(4,10).alignment=CA()
ws.row_dimensions[4].height=32

# Row 5: Hidden helper date cells
ws.row_dimensions[5].height=0  # hide by setting height to 0

# A5: period start
ws.cell(5,1).value='=DATE($E$4,MATCH($B$4,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0),1)'
ws.cell(5,1).number_format=DATE_F
# B5: period end
ws.cell(5,2).value='=EOMONTH($A$5,0)'
ws.cell(5,2).number_format=DATE_F
# C5: prev period start
ws.cell(5,3).value='=EOMONTH($A$5,-2)+1'
ws.cell(5,3).number_format=DATE_F
# D5: prev period end
ws.cell(5,4).value='=EOMONTH($A$5,-1)'
ws.cell(5,4).number_format=DATE_F

# Row 6: spacer
ws.merge_cells("A6:L6")
ws.cell(6,1).value="  ОСНОВНЫЕ ПОКАЗАТЕЛИ"
ws.cell(6,1).font=fnt(10,True,"FFFFFFFF"); ws.cell(6,1).fill=F(NAVY); ws.cell(6,1).alignment=LA()
ws.row_dimensions[6].height=22

# ── KPI Layout: 12 columns (A-L), 3 cols per card = 4 cards per row ──────────
# Each block: row N=section header, rows N+1,N+2 = KPI cards (label+value)
# Strategy: store curr/prev formulas in hidden ДАННЫЕ sheet to avoid deep nesting.

KPI_DATA_ROW = [99]  # next free row in ДАННЫЕ for KPI values
KPI_REGISTRY = {}    # label -> (curr_ref, prev_ref)

def store_kpi(label, curr_formula, prev_formula, mfmt=True):
    """Allocate row in ДАННЫЕ, write formulas, return (curr_ref, prev_ref)"""
    KPI_DATA_ROW[0] += 1
    r = KPI_DATA_ROW[0]
    ws_d.cell(r, 1).value = label
    ws_d.cell(r, 1).font = fnt(9, False, GRAY)
    if curr_formula:
        ws_d.cell(r, 2).value = curr_formula
        if mfmt: ws_d.cell(r, 2).number_format = MONEY
    if prev_formula:
        ws_d.cell(r, 3).value = prev_formula
        if mfmt: ws_d.cell(r, 3).number_format = MONEY
    curr_ref = f"ДАННЫЕ!$B${r}" if curr_formula else None
    prev_ref = f"ДАННЫЕ!$C${r}" if prev_formula else None
    if label:
        KPI_REGISTRY[label] = (curr_ref, prev_ref)
    return curr_ref, prev_ref

def dash_section(ws, row, title, bg):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=12)
    c=ws.cell(row,1); c.value=title
    c.font=fnt(10,True,"FFFFFFFF"); c.fill=F(bg); c.alignment=LA()
    ws.row_dimensions[row].height=20

def kpi_block(ws, label_row, val_row, cards, bg_hdr=LGRAY, bg_val=WHITE, val_col=NAVY):
    """cards = [(label, val_formula, prev_formula, money_fmt), ...] up to 4"""
    ncards=len(cards); cols_per=12//max(ncards,1)
    for i,(label,val_f,prev_f,mfmt) in enumerate(cards):
        c1=i*cols_per+1; c2=c1+cols_per-1
        if i==ncards-1: c2=12  # last card takes remaining space
        val_c2=c2-1 if ncards>1 else c2

        # Label row
        ws.merge_cells(start_row=label_row,start_column=c1,end_row=label_row,end_column=c2)
        lc=ws.cell(label_row,c1); lc.value=label
        lc.font=fnt(9,False,GRAY); lc.fill=F(bg_hdr); lc.border=brd(); lc.alignment=CA()
        ws.row_dimensions[label_row].height=20

        # Store formulas in ДАННЫЕ, get short cell references
        curr_ref, prev_ref = store_kpi(label, val_f if val_f else None, prev_f if prev_f else None, mfmt)

        # Value row — reference ДАННЫЕ cell
        ws.merge_cells(start_row=val_row,start_column=c1,end_row=val_row,end_column=val_c2)
        vc=ws.cell(val_row,c1)
        if curr_ref:
            vc.value = f"={curr_ref}"
        else:
            vc.value = ""
        vc.font=fnt(16,True,val_col); vc.fill=F(bg_val); vc.border=brd(); vc.alignment=CA(wrap=False)
        if mfmt: vc.number_format=MONEY

        # Trend cell — short comparison via cell refs
        tc=ws.cell(val_row,c2)
        if curr_ref and prev_ref:
            tc.value = f'=IF({curr_ref}>{prev_ref},"▲","▼")'
            tc.font=fnt(12,True,GREEN)
        else:
            tc.value=""
        tc.fill=F(bg_val); tc.border=brd(); tc.alignment=CA()

        # Conditional formatting on trend cell
        addr_t=f"{get_column_letter(c2)}{val_row}"
        ws.conditional_formatting.add(addr_t,FormulaRule(formula=[f'{addr_t}="▲"'],font=fnt(12,True,GREEN)))
        ws.conditional_formatting.add(addr_t,FormulaRule(formula=[f'{addr_t}="▼"'],font=fnt(12,True,RED)))

        ws.row_dimensions[val_row].height=38

# Helper SUMIFS — clean construction, no trailing-paren bug
def sumifs_periodo(tipo=None, cat=None, pay=None, fld="$G$4:$G$3003"):
    conds = ["(БАЗА_ДДС!$A$4:$A$3003>=$A$5)", "(БАЗА_ДДС!$A$4:$A$3003<=$B$5)"]
    if tipo: conds.append(f'(БАЗА_ДДС!$D$4:$D$3003="{tipo}")')
    if cat:  conds.append(f'(БАЗА_ДДС!$E$4:$E$3003="{cat}")')
    if pay:  conds.append(f'(БАЗА_ДДС!$F$4:$F$3003="{pay}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(conds)}*БАЗА_ДДС!{fld}),0)'

def sumifs_prev(tipo=None, cat=None, pay=None, fld="$G$4:$G$3003"):
    conds = ["(БАЗА_ДДС!$A$4:$A$3003>=$C$5)", "(БАЗА_ДДС!$A$4:$A$3003<=$D$5)"]
    if tipo: conds.append(f'(БАЗА_ДДС!$D$4:$D$3003="{tipo}")')
    if cat:  conds.append(f'(БАЗА_ДДС!$E$4:$E$3003="{cat}")')
    if pay:  conds.append(f'(БАЗА_ДДС!$F$4:$F$3003="{pay}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(conds)}*БАЗА_ДДС!{fld}),0)'

def countifs_periodo(tipo=None):
    conds = ["(БАЗА_ДДС!$A$4:$A$3003>=$A$5)", "(БАЗА_ДДС!$A$4:$A$3003<=$B$5)"]
    if tipo: conds.append(f'(БАЗА_ДДС!$D$4:$D$3003="{tipo}")')
    conds.append('(БАЗА_ДДС!$G$4:$G$3003<>"")')
    return f'=IFERROR(SUMPRODUCT({"*".join(conds)}),0)'

# Revenues
v_выручка = sumifs_periodo("Доход")
p_выручка = sumifs_prev("Доход")

v_days = '=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Доход")*(БАЗА_ДДС!$G$4:$G$3003<>"")*(1/COUNTIFS(БАЗА_ДДС!$A$4:$A$3003,БАЗА_ДДС!$A$4:$A$3003,БАЗА_ДДС!$D$4:$D$3003,"Доход"))),0)'
p_days = '=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$C$5)*(БАЗА_ДДС!$A$4:$A$3003<=$D$5)*(БАЗА_ДДС!$D$4:$D$3003="Доход")*(БАЗА_ДДС!$G$4:$G$3003<>"")*(1/COUNTIFS(БАЗА_ДДС!$A$4:$A$3003,БАЗА_ДДС!$A$4:$A$3003,БАЗА_ДДС!$D$4:$D$3003,"Доход"))),0)'
v_выр_день = f'=IFERROR(({v_выручка[1:]})/MAX(1,{v_days[1:]}),0)'
p_выр_день = f'=IFERROR(({p_выручка[1:]})/MAX(1,{p_days[1:]}),0)'

cnt_smeny = countifs_periodo("Доход")
v_выр_смену = f'=IFERROR(({v_выручка[1:]})/MAX(1,{cnt_smeny[1:]}),0)'
p_cnt_smeny = '=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$C$5)*(БАЗА_ДДС!$A$4:$A$3003<=$D$5)*(БАЗА_ДДС!$D$4:$D$3003="Доход")*(БАЗА_ДДС!$G$4:$G$3003<>"")*(1)),0)'
p_выр_смену = f'=IFERROR(({p_выручка[1:]})/MAX(1,{p_cnt_smeny[1:]}),0)'

v_макс_день = '=IFERROR(MAXIFS(БАЗА_ДДС!$G$4:$G$3003,БАЗА_ДДС!$A$4:$A$3003,">="&$A$5,БАЗА_ДДС!$A$4:$A$3003,"<="&$B$5,БАЗА_ДДС!$D$4:$D$3003,"Доход"),0)'
p_макс_день = '=IFERROR(MAXIFS(БАЗА_ДДС!$G$4:$G$3003,БАЗА_ДДС!$A$4:$A$3003,">="&$C$5,БАЗА_ДДС!$A$4:$A$3003,"<="&$D$5,БАЗА_ДДС!$D$4:$D$3003,"Доход"),0)'

# Cash control
v_выплаты = sumifs_periodo("Расход", "Выплата с кассы")
p_выплаты = sumifs_prev("Расход", "Выплата с кассы")
v_расх = sumifs_periodo("Расхождение")
p_расх = sumifs_prev("Расхождение")
v_расх_кол = countifs_periodo("Расхождение")
p_расх_кол = '=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$C$5)*(БАЗА_ДДС!$A$4:$A$3003<=$D$5)*(БАЗА_ДДС!$D$4:$D$3003="Расхождение")*(БАЗА_ДДС!$G$4:$G$3003<>"")*(1)),0)'
v_kasса_ост = '=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Касса")*БАЗА_ДДС!$G$4:$G$3003),0)'
p_kasса_ост = '=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$C$5)*(БАЗА_ДДС!$A$4:$A$3003<=$D$5)*(БАЗА_ДДС!$D$4:$D$3003="Касса")*БАЗА_ДДС!$G$4:$G$3003),0)'

# Debt
v_долг_тек = '=IFERROR(НАСТРОЙКИ!E9+SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003<=$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Долг")*БАЗА_ДДС!$G$4:$G$3003)-SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003<=$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Оплата долга")*БАЗА_ДДС!$G$4:$G$3003),0)'
p_долг_тек = '=IFERROR(НАСТРОЙКИ!E9+SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003<=$D$5)*(БАЗА_ДДС!$D$4:$D$3003="Долг")*БАЗА_ДДС!$G$4:$G$3003)-SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003<=$D$5)*(БАЗА_ДДС!$D$4:$D$3003="Оплата долга")*БАЗА_ДДС!$G$4:$G$3003),0)'
v_долг_взят = sumifs_periodo("Долг")
p_долг_взят = sumifs_prev("Долг")
v_долг_выпл = sumifs_periodo("Оплата долга")
k_oplate = f'=IFERROR(({v_долг_тек[1:]}),0)'

# Profit
v_закуп = sumifs_periodo("Расход", "Закуп товара")
p_закуп = sumifs_prev("Расход", "Закуп товара")
v_расходы = sumifs_periodo("Расход")
p_расходы = sumifs_prev("Расход")
v_прибыль = f'=IFERROR(({v_выручка[1:]})-({v_расходы[1:]}),0)'
p_прибыль = f'=IFERROR(({p_выручка[1:]})-({p_расходы[1:]}),0)'
v_рент = f'=IFERROR(({v_прибыль[1:]})/MAX(1,({v_выручка[1:]}))*100,0)'
p_рент = f'=IFERROR(({p_прибыль[1:]})/MAX(1,({p_выручка[1:]}))*100,0)'

# Efficiency
v_маржа = f'=IFERROR((({v_выручка[1:]})-({v_закуп[1:]}))/MAX(1,({v_выручка[1:]}))*100,0)'
p_маржа = f'=IFERROR((({p_выручка[1:]})-({p_закуп[1:]}))/MAX(1,({p_выручка[1:]}))*100,0)'
v_эфф_закупа = f'=IFERROR(({v_выручка[1:]})/MAX(1,({v_закуп[1:]})),0)'
p_эфф_закупа = f'=IFERROR(({p_выручка[1:]})/MAX(1,({p_закуп[1:]})),0)'
v_нагр_долга = f'=IFERROR(({v_долг_тек[1:]})/MAX(1,({v_выручка[1:]}))*100,0)'
p_нагр_долга = f'=IFERROR(({p_долг_тек[1:]})/MAX(1,({p_выручка[1:]}))*100,0)'
v_ср_расх_день = f'=IFERROR(({v_расходы[1:]})/MAX(1,({v_days[1:]})),0)'
p_ср_расх_день = f'=IFERROR(({p_расходы[1:]})/MAX(1,{p_days[1:]}),0)'

# Operations
v_просроч = '=IFERROR(SUMPRODUCT((КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003>=$A$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003<=$B$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$G$4:$G$1003="Просрочено")*КАЛЕНДАРЬ_ВЫПЛАТ!$D$4:$D$1003),0)'
p_просроч = '=IFERROR(SUMPRODUCT((КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003>=$C$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003<=$D$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$G$4:$G$1003="Просрочено")*КАЛЕНДАРЬ_ВЫПЛАТ!$D$4:$D$1003),0)'
v_просроч_кол = '=IFERROR(SUMPRODUCT((КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003>=$A$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003<=$B$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$G$4:$G$1003="Просрочено")*(КАЛЕНДАРЬ_ВЫПЛАТ!$D$4:$D$1003<>"")*(1)),0)'
p_просроч_кол = '=IFERROR(SUMPRODUCT((КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003>=$C$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003<=$D$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$G$4:$G$1003="Просрочено")*(КАЛЕНДАРЬ_ВЫПЛАТ!$D$4:$D$1003<>"")*(1)),0)'
v_vypl_plan = '=IFERROR(SUMPRODUCT((КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003>=$A$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003<=$B$5)*КАЛЕНДАРЬ_ВЫПЛАТ!$D$4:$D$1003),0)'
v_vypl_fakt = '=IFERROR(SUMPRODUCT((КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003>=$A$5)*(КАЛЕНДАРЬ_ВЫПЛАТ!$A$4:$A$1003<=$B$5)*КАЛЕНДАРЬ_ВЫПЛАТ!$F$4:$F$1003),0)'
v_vypl_pct = f'=IFERROR(({v_vypl_fakt[1:]})/MAX(1,({v_vypl_plan[1:]}))*100,0)'
p_vypl_pct = '=50'  # placeholder for previous
v_zakup_dolg = '=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=$B$5)*(БАЗА_ДДС!$E$4:$E$3003="Закуп товара")*(БАЗА_ДДС!$F$4:$F$3003="Долг")*БАЗА_ДДС!$G$4:$G$3003),0)'
p_zakup_dolg = '=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$C$5)*(БАЗА_ДДС!$A$4:$A$3003<=$D$5)*(БАЗА_ДДС!$E$4:$E$3003="Закуп товара")*(БАЗА_ДДС!$F$4:$F$3003="Долг")*БАЗА_ДДС!$G$4:$G$3003),0)'
v_zakup_dolg_pct = f'=IFERROR(({v_zakup_dolg[1:]})/MAX(1,({v_закуп[1:]}))*100,0)'

# Statistics
v_дней = v_days
v_записей = countifs_periodo("Доход")
v_макс = v_макс_день
v_мин = '=IFERROR(MINIFS(БАЗА_ДДС!$G$4:$G$3003,БАЗА_ДДС!$A$4:$A$3003,">="&$A$5,БАЗА_ДДС!$A$4:$A$3003,"<="&$B$5,БАЗА_ДДС!$D$4:$D$3003,"Доход"),0)'
v_иман = sumifs_periodo("Иман")
p_иман = sumifs_prev("Иман")
v_списания = sumifs_periodo("Списание")
p_списания = sumifs_prev("Списание")

# ── KPI Blocks ────────────────────────────────────────────────────────────────
row = 7

dash_section(ws, row, "  ВЫРУЧКА", BLUE); row+=1
kpi_block(ws, row, row+1, [
    ("Общая выручка", v_выручка, p_выручка, True),
    ("Среднее в день",v_выр_день, p_выр_день, True),
    ("Среднее за смену", v_выр_смену, p_выр_смену, True),
    ("Лучший день", v_макс_день, p_макс_день, True),
], BLUE_L, WHITE, BLUE); row+=2

dash_section(ws, row, "  КОНТРОЛЬ КАССЫ", TEAL); row+=1
kpi_block(ws, row, row+1, [
    ("Выплаты из кассы", v_выплаты, p_выплаты, True),
    ("Расхождений сумма", v_расх, p_расх, True),
    ("Кол-во расхождений", v_расх_кол, p_расх_кол, False),
    ("Остаток кассы", v_kasса_ост, p_kasса_ост, True),
], TEAL_L, WHITE, TEAL); row+=2

dash_section(ws, row, "  ДОЛГИ И ОБЯЗАТЕЛЬСТВА", RED); row+=1
kpi_block(ws, row, row+1, [
    ("Текущий долг", v_долг_тек, p_долг_тек, True),
    ("Взято в долг", v_долг_взят, p_долг_взят, True),
    ("Выплачено долгов", v_долг_выпл, p_долг_тек, True),
    ("К оплате (=долг)", k_oplate, p_долг_тек, True),
], RED_L, WHITE, RED); row+=2

dash_section(ws, row, "  ПРИБЫЛЬ", GREEN); row+=1
kpi_block(ws, row, row+1, [
    ("Закуп товара", v_закуп, p_закуп, True),
    ("Все расходы", v_расходы, p_расходы, True),
    ("Чистая прибыль", v_прибыль, p_прибыль, True),
    ("Рентабельность %", v_рент, p_рент, False),
], GREEN_L, WHITE, GREEN); row+=2

dash_section(ws, row, "  ЭФФЕКТИВНОСТЬ", PURPLE); row+=1
kpi_block(ws, row, row+1, [
    ("Маржа %", v_маржа, p_маржа, False),
    ("Эффект. закупа (x)", v_эфф_закупа, p_эфф_закупа, False),
    ("Нагрузка долга %", v_нагр_долга, p_нагр_долга, False),
    ("Ср. расход/день", v_ср_расх_день, p_ср_расх_день, True),
], PURP_L, WHITE, PURPLE); row+=2

dash_section(ws, row, "  ОПЕРАЦИИ И ВЫПЛАТЫ", AMBER); row+=1
kpi_block(ws, row, row+1, [
    ("Просроч. выплаты", v_просроч, p_просроч, True),
    ("Просроч. кол-во", v_просроч_кол, p_просроч_кол, False),
    ("Выплачено %", v_vypl_pct, p_vypl_pct, False),
    ("Закуп в долг %", v_zakup_dolg_pct, p_vypl_pct, False),
], AMBER_L, WHITE, AMBER); row+=2

dash_section(ws, row, "  СТАТИСТИКА ПЕРИОДА", GRAY); row+=1
kpi_block(ws, row, row+1, [
    ("Дней с данными", v_дней, None, False),
    ("Всего операций", v_записей, None, False),
    ("Макс. выручка/день", v_макс, v_мин, True),
    ("Мин. выручка/день", v_мин, None, True),
], LGRAY, WHITE, NAVY); row+=2

# Second statistics row
kpi_block(ws, row, row+1, [
    ("Иман (хозяин)", v_иман, p_иман, True),
    ("Списания+Возвраты", v_списания, p_списания, True),
    ("", "=0", None, False),
    ("", "=0", None, False),
], LGRAY, WHITE, NAVY); row+=2

# Row spacer
ws.merge_cells(f"A{row}:L{row}")
ws.cell(row,1).fill=F(LGRAY); ws.row_dimensions[row].height=10; row+=1

# ── Expense Detail Section ─────────────────────────────────────────────────────
dash_section(ws, row, "  ДЕТАЛИЗАЦИЯ РАСХОДОВ ЗА ПЕРИОД", INDIGO); row+=1
hrow(ws, row, ["Категория","Сумма","Доля %","Гистограмма","—","—","—","—","—","—","—","—"], INDIGO, 22); row+=1

exp_cats_list=["Закуп товара","Зарплата","Аренда","Коммунальные","Налог","ГСМ","Расходный материал","Маркетинг","Охрана","Прочие расходы"]
exp_start_row=row
for ei,cat in enumerate(exp_cats_list):
    bg=LGRAY if ei%2==0 else WHITE
    ws.cell(row,1).value=cat; ws.cell(row,1).font=fnt(10); ws.cell(row,1).fill=F(bg); ws.cell(row,1).border=brd(); ws.cell(row,1).alignment=LA()
    sf=f'=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=$B$5)*(БАЗА_ДДС!$E$4:$E$3003="{cat}")*БАЗА_ДДС!$G$4:$G$3003),0)'
    ws.cell(row,2).value=sf; ws.cell(row,2).font=fnt(10,True,RED if ei==0 else NAVY); ws.cell(row,2).fill=F(bg); ws.cell(row,2).border=brd(); ws.cell(row,2).alignment=RA(); ws.cell(row,2).number_format=MONEY
    pct_f=f'=IFERROR(B{row}/{v_расходы[1:]}*100,0)'
    ws.cell(row,3).value=pct_f; ws.cell(row,3).font=fnt(9,False,GRAY); ws.cell(row,3).fill=F(bg); ws.cell(row,3).border=brd(); ws.cell(row,3).alignment=CA(); ws.cell(row,3).number_format="0.0%"
    ws.merge_cells(start_row=row,start_column=4,end_row=row,end_column=12)
    bar_f=f'=IFERROR(REPT("|",MAX(1,INT(B{row}/MAX($B${exp_start_row}:$B${exp_start_row+9})*20))),"")'
    ws.cell(row,4).value=bar_f; ws.cell(row,4).font=Font(name="Calibri",size=9,color=INDIGO); ws.cell(row,4).fill=F(BLUE_L if ei%2==0 else WHITE); ws.cell(row,4).border=brd(); ws.cell(row,4).alignment=LA()
    ws.row_dimensions[row].height=22; row+=1

ws.merge_cells(f"A{row}:L{row}")
ws.cell(row,1).fill=F(LGRAY); ws.row_dimensions[row].height=8; row+=1

# ── Charts Section ─────────────────────────────────────────────────────────────
charts_start_row = row

dash_section(ws, row, "  ГРАФИКИ И ДИАГРАММЫ", NAVY); row+=1

# Chart 1: Выручка по месяцам (Line)
chart1 = LineChart()
chart1.title = "Выручка по месяцам"
chart1.style = 10; chart1.height = 10; chart1.width = 16
data1 = Reference(ws_d, min_col=2, max_col=2, min_row=1, max_row=13)
cats1 = Reference(ws_d, min_col=1, min_row=2, max_row=13)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.series[0].graphicalProperties.line.solidFill = "3B82F6"
chart1.series[0].graphicalProperties.line.width = 20000
ws.add_chart(chart1, f"A{row}")

# Chart 2: Выручка по сменам (Bar)
chart2 = BarChart()
chart2.title = "Выручка по сменам"
chart2.style = 10; chart2.height = 10; chart2.width = 14
chart2.type = "col"
data2 = Reference(ws_d, min_col=2, max_col=2, min_row=16, max_row=19)
cats2 = Reference(ws_d, min_col=1, min_row=17, max_row=19)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.solidFill = "10B981"
ws.add_chart(chart2, f"G{row}")

row += 20

# Chart 3: Структура расходов (Pie)
chart3 = PieChart()
chart3.title = "Структура расходов"
chart3.style = 10; chart3.height = 10; chart3.width = 16
data3 = Reference(ws_d, min_col=2, min_row=23, max_row=33)
cats3 = Reference(ws_d, min_col=1, min_row=24, max_row=33)
chart3.add_data(data3, titles_from_data=False)
chart3.set_categories(cats3)
ws.add_chart(chart3, f"A{row}")

# Chart 4: ТОП расходов (Bar horizontal)
chart4 = BarChart()
chart4.title = "ТОП категорий расходов"
chart4.style = 10; chart4.height = 10; chart4.width = 14
chart4.type = "bar"  # horizontal
data4 = Reference(ws_d, min_col=2, min_row=23, max_row=33)
cats4 = Reference(ws_d, min_col=1, min_row=24, max_row=33)
chart4.add_data(data4, titles_from_data=False)
chart4.set_categories(cats4)
chart4.series[0].graphicalProperties.solidFill = "EF4444"
ws.add_chart(chart4, f"G{row}")

row += 20

# Chart 5: Последние 30 дней (Line)
chart5 = LineChart()
chart5.title = "Выручка за последние 30 дней"
chart5.style = 10; chart5.height = 10; chart5.width = 16
data5 = Reference(ws_d, min_col=2, max_col=3, min_row=37, max_row=67)
chart5.add_data(data5, titles_from_data=True)
chart5.series[0].graphicalProperties.line.solidFill = "3B82F6"
chart5.series[1].graphicalProperties.line.solidFill = "EF4444"
ws.add_chart(chart5, f"A{row}")

# Chart 6: Динамика долга (Line 2 series)
chart6 = LineChart()
chart6.title = "Динамика долга по месяцам"
chart6.style = 10; chart6.height = 10; chart6.width = 14
data6 = Reference(ws_d, min_col=2, max_col=3, min_row=70, max_row=82)
cats6 = Reference(ws_d, min_col=1, min_row=71, max_row=82)
chart6.add_data(data6, titles_from_data=True)
chart6.set_categories(cats6)
chart6.series[0].graphicalProperties.line.solidFill = "F59E0B"
chart6.series[1].graphicalProperties.line.solidFill = "10B981"
ws.add_chart(chart6, f"G{row}")

cw(ws,{"A":10,"B":10,"C":10,"D":10,"E":10,"F":10,"G":10,"H":10,"I":10,"J":10,"K":10,"L":10})
ws.freeze_panes="A7"
ws.sheet_properties.tabColor="FF4F46E5"

print("✓ ДАШБОРД")

# ════════════════════════════════════════════════════════════
# 9. ОТЧЁТ_РУКОВОДИТЕЛЮ
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("ОТЧЁТ_РУКОВОДИТЕЛЮ"); ws.sheet_view.showGridLines = False

# ── Period date references (from ДАШБОРД hidden row 5) ────────
_rP="БАЗА_ДДС!$A$4:$A$3003"; _rD="БАЗА_ДДС!$D$4:$D$3003"
_rE="БАЗА_ДДС!$E$4:$E$3003"; _rG="БАЗА_ДДС!$G$4:$G$3003"
DS="ДАШБОРД!$A$5"; DE="ДАШБОРД!$B$5"; PS="ДАШБОРД!$C$5"; PE="ДАШБОРД!$D$5"

def rsp(typ=None,cat=None,s=None,e=None):
    """SUMPRODUCT filtered by current period."""
    s_=s or DS; e_=e or DE
    c=[f"({_rP}>={s_})",f"({_rP}<={e_})"]
    if typ: c.append(f'({_rD}="{typ}")')
    if cat: c.append(f'({_rE}="{cat}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(c)}*{_rG}),0)'

def rsa(typ=None):
    """SUMPRODUCT all-time (no period filter)."""
    filt=f'({_rD}="{typ}")*' if typ else ""
    return f'=IFERROR(SUMPRODUCT({filt}{_rG}),0)'

# ── Banner ────────────────────────────────────────────────────
banner(ws,"ОТЧЁТ РУКОВОДИТЕЛЮ — финансовая модель 25/75","A1:F1",NAVY,15)
ws.row_dimensions[1].height=40

# Row 2: info bar — individual cells, no merge (formula constraint)
ws.cell(2,1).value='=НАСТРОЙКИ!E5&"  |  Период: "&TEXT(ДАШБОРД!$A$5,"DD.MM.YYYY")&" — "&TEXT(ДАШБОРД!$B$5,"DD.MM.YYYY")'
ws.cell(2,1).font=fnt(11,False,NAVY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).border=brd(); ws.cell(2,1).alignment=LA()
ws.cell(2,6).value='="Сформирован: "&TEXT(NOW(),"DD.MM.YYYY HH:MM")'
ws.cell(2,6).font=fnt(9,it=True,col=GRAY); ws.cell(2,6).fill=F(LGRAY); ws.cell(2,6).border=brd(); ws.cell(2,6).alignment=RA()
for ci_ in range(2,6): ws.cell(2,ci_).fill=F(LGRAY); ws.cell(2,ci_).border=brd()
ws.row_dimensions[2].height=28

# Row 3: column headers
for ci_,(h_,w_) in enumerate([("Показатель",32),("Сумма (₽)",18),("% от выручки",13),
                                ("Пред. период (₽)",15),("▲▼",5),("Лимит / Справка",16)],1):
    c=ws.cell(3,ci_); c.value=h_; c.font=fnt(9,True,"FFFFFFFF"); c.fill=F(NAVY); c.border=brd(); c.alignment=CA()
    ws.column_dimensions[get_column_letter(ci_)].width=w_
ws.row_dimensions[3].height=24

# ── БЛОК 1: СВОДНЫЕ ПОКАЗАТЕЛИ ────────────────────────────────
sec_hdr(ws,4,"  БЛОК 1: СВОДНЫЕ ПОКАЗАТЕЛИ ЗА ПЕРИОД",6,NAVY)

_f_vyr=rsp("Доход"); _f_rsh=rsp("Расход")
_p_vyr=rsp("Доход",s=PS,e=PE); _p_rsh=rsp("Расход",s=PS,e=PE)
_f_ost=f'=IFERROR({rsa("Доход")[1:]}-{rsa("Расход")[1:]},0)'

b1_rows=[
    (5,"Общая выручка (Доход)",_f_vyr,_p_vyr,"=1","100% базы",True),
    (6,"Общие расходы (Расход)",_f_rsh,_p_rsh,"=IFERROR(B6/MAX(1,B5),0)","% выручки",False),
    (7,"Чистая прибыль (Выручка − Расходы)","=IFERROR(B5-B6,0)","=IFERROR(D5-D6,0)","=IFERROR(B7/MAX(1,B5),0)","% рентабельн.",True),
    (8,"Фактический остаток в системе",_f_ost,None,"всё время","все периоды",None),
]
for (ri_,lbl_,cur_,prv_,pct_,note_,up_) in b1_rows:
    ws.cell(ri_,1).value=lbl_; ws.cell(ri_,1).font=fnt(10); ws.cell(ri_,1).fill=F(LGRAY); ws.cell(ri_,1).border=brd(); ws.cell(ri_,1).alignment=LA()
    ws.cell(ri_,2).value=cur_; ws.cell(ri_,2).font=fnt(12,True,NAVY); ws.cell(ri_,2).fill=F(BLUE_L); ws.cell(ri_,2).border=brd(); ws.cell(ri_,2).alignment=RA(); ws.cell(ri_,2).number_format=MONEY
    if isinstance(pct_,str) and pct_.startswith('='):
        ws.cell(ri_,3).value=pct_; ws.cell(ri_,3).number_format="0.0%"
    else:
        ws.cell(ri_,3).value=pct_
    ws.cell(ri_,3).font=fnt(9,col=GRAY); ws.cell(ri_,3).fill=F(LGRAY); ws.cell(ri_,3).border=brd(); ws.cell(ri_,3).alignment=CA()
    if prv_ is not None:
        ws.cell(ri_,4).value=prv_; ws.cell(ri_,4).font=fnt(10,col=GRAY); ws.cell(ri_,4).fill=F(LGRAY); ws.cell(ri_,4).border=brd(); ws.cell(ri_,4).alignment=RA(); ws.cell(ri_,4).number_format=MONEY
        tr_=f'=IF(B{ri_}>D{ri_},"▲","▼")' if up_ else f'=IF(B{ri_}<D{ri_},"▲","▼")'
        ws.cell(ri_,5).value=tr_; ws.cell(ri_,5).fill=F(LGRAY); ws.cell(ri_,5).border=brd(); ws.cell(ri_,5).alignment=CA(); ws.cell(ri_,5).font=fnt(12,True,GREEN)
        ws.conditional_formatting.add(f"E{ri_}",FormulaRule(formula=[f'E{ri_}="▲"'],font=fnt(12,True,GREEN)))
        ws.conditional_formatting.add(f"E{ri_}",FormulaRule(formula=[f'E{ri_}="▼"'],font=fnt(12,True,RED)))
    else:
        ws.cell(ri_,4).fill=F(LGRAY); ws.cell(ri_,4).border=brd()
        ws.cell(ri_,5).fill=F(LGRAY); ws.cell(ri_,5).border=brd()
    ws.cell(ri_,6).value=note_; ws.cell(ri_,6).font=fnt(9,it=True,col=GRAY); ws.cell(ri_,6).fill=F(LGRAY); ws.cell(ri_,6).border=brd(); ws.cell(ri_,6).alignment=LA()
    ws.row_dimensions[ri_].height=28

# Traffic light: прибыль (B7) и остаток (B8)
for ri_ in [7,8]:
    ws.conditional_formatting.add(f"B{ri_}",FormulaRule(formula=[f"B{ri_}>0"],fill=F(GREEN_L),font=fnt(12,True,GREEN)))
    ws.conditional_formatting.add(f"B{ri_}",FormulaRule(formula=[f"B{ri_}<=0"],fill=F(RED_L),font=fnt(12,True,RED)))
ws.row_dimensions[9].height=6  # spacer

# ── БЛОК 2: ФИНАНСОВАЯ МОДЕЛЬ 25/75 ────────────────────────────
sec_hdr(ws,10,"  БЛОК 2: ФИНАНСОВАЯ МОДЕЛЬ — РАСПРЕДЕЛЕНИЕ 25/75",6,INDIGO)

b2_rows=[
    (11,"Маржинальная прибыль (25%)","=IFERROR(B5*НАСТРОЙКИ!$E$7,0)","=НАСТРОЙКИ!$E$7",'=TEXT(НАСТРОЙКИ!$E$7,"0%")&" × выручка"'),
    (12,"Лимит на закуп товаров (75%)","=IFERROR(B5*НАСТРОЙКИ!$E$8,0)","=НАСТРОЙКИ!$E$8",'=TEXT(НАСТРОЙКИ!$E$8,"0%")&" × выручка"'),
    (13,"Фактический закуп за период",rsp("Расход","Закуп"),"=IFERROR(B13/MAX(1,B5),0)","Категория: Закуп"),
    (14,"Остаток лимита на закуп","=IFERROR(B12-B13,0)","=IFERROR(B14/MAX(1,B5),0)","= Лимит − Факт.закуп"),
]
for (ri_,lbl_,cur_,pct_,note_) in b2_rows:
    ws.cell(ri_,1).value=lbl_; ws.cell(ri_,1).font=fnt(10); ws.cell(ri_,1).fill=F(LGRAY); ws.cell(ri_,1).border=brd(); ws.cell(ri_,1).alignment=LA()
    ws.cell(ri_,2).value=cur_; ws.cell(ri_,2).font=fnt(12,True,INDIGO); ws.cell(ri_,2).fill=F(PURP_L); ws.cell(ri_,2).border=brd(); ws.cell(ri_,2).alignment=RA(); ws.cell(ri_,2).number_format=MONEY
    ws.cell(ri_,3).value=pct_
    if isinstance(pct_,str) and pct_.startswith('='):
        ws.cell(ri_,3).number_format="0.0%" if "НАСТРОЙКИ" not in pct_ else "0%"
    ws.cell(ri_,3).font=fnt(9,col=INDIGO); ws.cell(ri_,3).fill=F(LGRAY); ws.cell(ri_,3).border=brd(); ws.cell(ri_,3).alignment=CA()
    ws.cell(ri_,4).fill=F(LGRAY); ws.cell(ri_,4).border=brd()
    ws.cell(ri_,5).fill=F(LGRAY); ws.cell(ri_,5).border=brd()
    ws.cell(ri_,6).value=note_; ws.cell(ri_,6).font=fnt(9,it=True,col=GRAY); ws.cell(ri_,6).fill=F(LGRAY); ws.cell(ri_,6).border=brd(); ws.cell(ri_,6).alignment=LA()
    ws.row_dimensions[ri_].height=28

# Traffic light: остаток лимита (B14) — green=в норме, red=перерасход
ws.conditional_formatting.add("B14",FormulaRule(formula=["B14>=0"],fill=F(GREEN_L),font=fnt(12,True,GREEN)))
ws.conditional_formatting.add("B14",FormulaRule(formula=["B14<0"],fill=F(RED_L),font=fnt(12,True,RED)))
ws.row_dimensions[15].height=6  # spacer

# ── БЛОК 3: ДОЛГОВАЯ НАГРУЗКА ──────────────────────────────────
sec_hdr(ws,16,"  БЛОК 3: ДОЛГОВАЯ НАГРУЗКА И ОБЯЗАТЕЛЬСТВА",6,RED)

_f_dolg_tek="=IFERROR(B17+B18-B19,0)"
b3_rows=[
    (17,"Начальный долг поставщикам (баланс)","=НАСТРОЙКИ!$E$9","НАСТРОЙКИ!E9"),
    (18,"Взято в долг — всё время",rsa("Долг"),"тип транзакции: Долг"),
    (19,"Выплачено по долгам — всё время",rsa("Оплата долга"),"тип: Оплата долга"),
    (20,"Текущий долг поставщикам (итого)",_f_dolg_tek,"= Нач. + Взято − Выплачено"),
    (21,"Кассовый разрыв (прогноз)","=IFERROR(B8-B20,0)","= Фактич.остаток − Долг"),
]
for (ri_,lbl_,cur_,note_) in b3_rows:
    ws.cell(ri_,1).value=lbl_; ws.cell(ri_,1).font=fnt(10); ws.cell(ri_,1).fill=F(LGRAY); ws.cell(ri_,1).border=brd(); ws.cell(ri_,1).alignment=LA()
    ws.cell(ri_,2).value=cur_; ws.cell(ri_,2).font=fnt(12,True,RED); ws.cell(ri_,2).fill=F(RED_L); ws.cell(ri_,2).border=brd(); ws.cell(ri_,2).alignment=RA(); ws.cell(ri_,2).number_format=MONEY
    for ci_ in [3,4,5]: ws.cell(ri_,ci_).fill=F(LGRAY); ws.cell(ri_,ci_).border=brd()
    ws.cell(ri_,6).value=note_; ws.cell(ri_,6).font=fnt(9,it=True,col=GRAY); ws.cell(ri_,6).fill=F(LGRAY); ws.cell(ri_,6).border=brd(); ws.cell(ri_,6).alignment=LA()
    ws.row_dimensions[ri_].height=28

# Traffic light: текущий долг (B20) — red=долг есть
ws.conditional_formatting.add("B20",FormulaRule(formula=["B20>0"],fill=F(RED_L),font=fnt(12,True,RED)))
ws.conditional_formatting.add("B20",FormulaRule(formula=["B20<=0"],fill=F(GREEN_L),font=fnt(12,True,GREEN)))
# Кассовый разрыв (B21) — green=положительный, red=отрицательный (разрыв)
ws.conditional_formatting.add("B21",FormulaRule(formula=["B21>0"],fill=F(GREEN_L),font=fnt(12,True,GREEN)))
ws.conditional_formatting.add("B21",FormulaRule(formula=["B21<=0"],fill=F(RED_L),font=fnt(12,True,RED)))
ws.row_dimensions[22].height=6  # spacer

# Enhanced alert CF on B20: deep red when debt exceeds threshold from НАСТРОЙКИ!$E$38
ws.conditional_formatting.add("B20",FormulaRule(formula=["AND(B20>0,B20>НАСТРОЙКИ!$E$38)"],fill=F("FFEF4444"),font=fnt(12,True,"FFDC2626")))

# ── БЛОК 4: ВЫПЛАТЫ ПОСТАВЩИКАМ (мини-отчёт по ТОП-5) ─────────
sec_hdr(ws,23,"  БЛОК 4: ВЫПЛАТЫ ПОСТАВЩИКАМ — ТОП-5 ПО НАСТРОЙКАМ",6,PURPLE)
for ci_,h_ in enumerate(["Поставщик / ТП","Запланировано (₽)","Выплачено (₽)","Остаток (₽)","","Источник"],1):
    c=ws.cell(24,ci_); c.value=h_; c.font=fnt(9,True,"FFFFFFFF"); c.fill=F(PURPLE); c.border=brd(); c.alignment=CA()
ws.row_dimensions[24].height=24

_ZP_="ЗАПИСЬ_НА_ВЫПЛАТУ!$C$4:$C$503"
_ZS_="ЗАПИСЬ_НА_ВЫПЛАТУ!$D$4:$D$503"
_ZT_="ЗАПИСЬ_НА_ВЫПЛАТУ!$E$4:$E$503"
for i_,ri_ in enumerate(range(25,30)):
    sr_=f"НАСТРОЙКИ!$B${99+i_}"
    lbl_=f'=IF({sr_}="","(Поставщик {i_+1})",{sr_})'
    f_plan=f'=IFERROR(SUMPRODUCT(({_ZP_}={sr_})*({_ZT_}="Запланировано")*{_ZS_}),0)'
    f_vip =f'=IFERROR(SUMPRODUCT(({_ZP_}={sr_})*({_ZT_}="Выплачено")*{_ZS_}),0)'
    f_ost =f'=IFERROR(C{ri_}-D{ri_},0)'
    ws.cell(ri_,1).value=lbl_; ws.cell(ri_,1).font=fnt(10); ws.cell(ri_,1).fill=F(LGRAY); ws.cell(ri_,1).border=brd(); ws.cell(ri_,1).alignment=LA()
    ws.cell(ri_,2).value=f_plan; ws.cell(ri_,2).font=fnt(11,True,PURPLE); ws.cell(ri_,2).fill=F(PURP_L); ws.cell(ri_,2).border=brd(); ws.cell(ri_,2).alignment=RA(); ws.cell(ri_,2).number_format=MONEY
    ws.cell(ri_,3).value=f_vip;  ws.cell(ri_,3).font=fnt(11,True,GREEN);  ws.cell(ri_,3).fill=F(GREEN_L); ws.cell(ri_,3).border=brd(); ws.cell(ri_,3).alignment=RA(); ws.cell(ri_,3).number_format=MONEY
    ws.cell(ri_,4).value=f_ost;  ws.cell(ri_,4).font=fnt(11,True,AMBER);  ws.cell(ri_,4).fill=F(AMBER_L); ws.cell(ri_,4).border=brd(); ws.cell(ri_,4).alignment=RA(); ws.cell(ri_,4).number_format=MONEY
    ws.cell(ri_,5).fill=F(LGRAY); ws.cell(ri_,5).border=brd()
    ws.cell(ri_,6).value=f"НАСТРОЙКИ!B{99+i_}"; ws.cell(ri_,6).font=fnt(9,it=True,col=GRAY); ws.cell(ri_,6).fill=F(LGRAY); ws.cell(ri_,6).border=brd(); ws.cell(ri_,6).alignment=LA()
    ws.row_dimensions[ri_].height=26
    ws.conditional_formatting.add(f"D{ri_}",FormulaRule(formula=[f"D{ri_}>0"],fill=F(RED_L),font=fnt(11,True,RED)))
    ws.conditional_formatting.add(f"D{ri_}",FormulaRule(formula=[f"D{ri_}<=0"],fill=F(GREEN_L),font=fnt(11,True,GREEN)))
ws.row_dimensions[30].height=6  # spacer

ws.freeze_panes="A4"
ws.sheet_properties.tabColor="FF111827"
print("✓ ОТЧЁТ_РУКОВОДИТЕЛЮ")

# ════════════════════════════════════════════════════════════
# 10. ИНСТРУКЦИЯ
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("ИНСТРУКЦИЯ"); ws.sheet_view.showGridLines = False
banner(ws, "WAY MARKET v9 — ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ", "A1:J1", INDIGO, 14)

steps=[
    ("1. НАСТРОЙКИ — ЦЕНТР УПРАВЛЕНИЯ", INDIGO,
     ["Раздел 1 (строки 5–11): Название магазина, дата, маржа, долг, лимит, период",
      "Раздел 2 (строки 15–17): Вкл/Выкл смен ДЕНЬ / ВЕЧЕР / НОЧЬ",
      "Раздел 3 (строки 20–24): Вкл/Выкл способов оплаты (Эквайринг, Перевод и др.)",
      "Раздел 6 (строки 37–39): Пороги уведомлений — расхождение, долг, просрочка",
      "Раздел 7 (строки 44–79): Справочники кассиров, категорий, способов, типов операций",
      "Раздел 8 (строки 82–93): Постоянные расходы — нажмите ЗАЧИСЛИТЬ РАСХОДЫ",
      "Раздел 9 (строки 99–1098): Справочник поставщиков для выплат и мини-отчёта"]),
    ("2. ЕЖЕДНЕВНЫЙ ВВОД ДАННЫХ", BLUE,
     ["ВВОД_КАССА: дата, смена, кассир, Z-отчёты по активным способам оплаты",
      "Колонки затемнены, если способ оплаты выключен в НАСТРОЙКАХ",
      "Нажмите СОХРАНИТЬ КАССУ — данные запишутся в БАЗА_ДДС",
      "ВВОД_РАСХОДЫ: дата, категория (из справочника), способ, сумма",
      "Нажмите СОХРАНИТЬ РАСХОДЫ — данные запишутся в БАЗА_ДДС"]),
    ("3. ВЫПЛАТЫ ПОСТАВЩИКАМ", AMBER,
     ["ЗАПИСЬ_НА_ВЫПЛАТУ: выберите поставщика из списка (Раздел 9 НАСТРОЙКИ)",
      "Укажите дату, сумму, способ оплаты, накладную и статус",
      "Нажмите СОХРАНИТЬ ВЫПЛАТУ",
      "Просроченные выплаты подсвечиваются красным в КАЛЕНДАРЬ_ВЫПЛАТ"]),
    ("4. ДАШБОРД", GREEN,
     ["ДАШБОРД: выберите месяц и год в фильтре (строка 4) → нажмите ОБНОВИТЬ",
      "Все показатели пересчитываются автоматически по БАЗА_ДДС",
      "Стрелки ▲▼ — динамика vs предыдущий период",
      "Цветовая индикация: зелёный = норма, красный = превышение"]),
    ("5. ОТЧЁТ РУКОВОДИТЕЛЮ", PURPLE,
     ["ОТЧЁТ_РУКОВОДИТЕЛЮ: 4 блока финансовых показателей",
      "Блок 1: Выручка, Расходы, Прибыль, Остаток в системе",
      "Блок 2: Финансовая модель 25/75 — маржа и лимит закупа",
      "Блок 3: Долговая нагрузка и кассовый разрыв",
      "Блок 4: Мини-отчёт по ТОП-5 поставщикам из НАСТРОЙКИ"]),
]

r=3
for title,bg,items in steps:
    sec_hdr(ws, r, f"  {title}", 10, bg); r+=1
    for item in items:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=10)
        c=ws.cell(r,1); c.value=f"• {item}"
        c.font=fnt(10); c.fill=F(LGRAY if r%2==0 else WHITE); c.border=brd(); c.alignment=LA()
        ws.row_dimensions[r].height=22; r+=1
    ws.row_dimensions[r].height=8; r+=1

ws.merge_cells(f"A{r}:J{r}")
ws.cell(r,1).value="Горячие клавиши: Ctrl+Shift+S — Сохранить кассу | Ctrl+Shift+R — Сохранить расходы | Ctrl+Shift+D — Обновить дашборд"
ws.cell(r,1).font=fnt(10,True,INDIGO); ws.cell(r,1).fill=F(BLUE_L); ws.cell(r,1).alignment=CA(); ws.row_dimensions[r].height=30

cw(ws,{"A":30,"B":16,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14})
ws.sheet_properties.tabColor="FF4F46E5"
print("✓ ИНСТРУКЦИЯ")

# ════════════════════════════════════════════════════════════
# 11. КАК_СДЕЛАТЬ_ДАШБОРД
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("КАК_СДЕЛАТЬ_ДАШБОРД"); ws.sheet_view.showGridLines = False
banner(ws, "КАК ПОДКЛЮЧИТЬ МАКРОСЫ VBA", "A1:J1", PURPLE, 14)

steps2=[
    "1. Откройте Excel и нажмите Alt+F11 (Редактор VBA)",
    "2. В меню: Файл → Импорт файла → выберите Модуль_WM9.bas",
    "3. Закройте редактор VBA (Alt+Q)",
    "4. Перейдите на лист ВВОД_КАССА → Alt+F8 → УстановитьВсеКнопки → Run",
    "5. На листах появятся кнопки для сохранения данных",
    "6. Если появится 'Предупреждение безопасности' — нажмите 'Включить содержимое'",
    "",
    "ВАЖНО: Файл нужно сохранить в формате .xlsm (Книга Excel с поддержкой макросов)",
    "ВАЖНО: Защита листов установлена. Редактировать можно только синие ячейки.",
    "",
    "Для LibreOffice Calc: импорт Basic-модуля через Сервис → Макросы → Редактор Basic",
]
for i,txt in enumerate(steps2,3):
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=10)
    c=ws.cell(i,1); c.value=txt
    if txt.startswith("ВАЖНО"):
        c.font=fnt(10,True,RED); c.fill=F(RED_L)
    elif txt=="":
        c.fill=F(LGRAY)
    else:
        c.font=fnt(10); c.fill=F(LGRAY if i%2==0 else WHITE)
    c.border=brd(); c.alignment=LA(); ws.row_dimensions[i].height=24

cw(ws,{"A":60,"B":14,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14})
ws.sheet_properties.tabColor="FF8B5CF6"
print("✓ КАК_СДЕЛАТЬ_ДАШБОРД")

# ════════════════════════════════════════════════════════════
# Reorder sheets
# ════════════════════════════════════════════════════════════
desired_order=["ИНСТРУКЦИЯ","КАК_СДЕЛАТЬ_ДАШБОРД","ВВОД_КАССА","ВВОД_РАСХОДЫ",
               "ЗАПИСЬ_НА_ВЫПЛАТУ","КАЛЕНДАРЬ_ВЫПЛАТ","БАЗА_ДДС","ДАННЫЕ","ДАШБОРД",
               "ОТЧЁТ_РУКОВОДИТЕЛЮ","НАСТРОЙКИ"]
sheet_dict={ws_.title: ws_ for ws_ in wb.worksheets}
for i,title in enumerate(desired_order):
    if title in sheet_dict:
        wb._sheets.remove(sheet_dict[title])
        wb._sheets.insert(i, sheet_dict[title])

# ════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════
out_path="/home/user/Auron/WAY_MARKET_v9.xlsx"
wb.save(out_path)
import os
size=os.path.getsize(out_path)
print(f"\n✅ WAY_MARKET_v9.xlsx saved — {size//1024} KB, {len(wb.worksheets)} sheets")
