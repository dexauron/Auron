#!/usr/bin/env python3
"""WAY MARKET — Ideal Edition v2. Teal+Gold palette. 12 sheets."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import date
import os

today = date.today()

# ── Palette: Deep Teal + Gold ──────────────────────────────────
TEAL   = "FF0B4F54"; TEAL_M = "FF0D9488"; TEAL_L = "FFE0F2F1"
GOLD   = "FFB45309"; GOLD_M = "FFF59E0B"; GOLD_L = "FFFEF3C7"
NAVY   = "FF111827"; DGRAY  = "FF374151"
GREEN  = "FF059669"; GREEN_L= "FFD1FAE5"
RED    = "FFDC2626"; RED_L  = "FFFEE2E2"
AMBER  = "FFD97706"; AMBER_L= "FFFEF3C7"
PURPLE = "FF7C3AED"; PURP_L = "FFEDE9FE"
GRAY   = "FF6B7280"; LGRAY  = "FFF9FAFB"
WHITE  = "FFFFFFFF"; INP    = "FFF0FDFA"; INP_BD = "FF0D9488"
BORDER = "FFE5E7EB"; DISABLED="FFF3F4F6"
MONEY="#,##0;[Red]-#,##0"; MONEY2="#,##0.0"; DATE_F="DD.MM.YYYY"
MONTHS_RU=["Январь","Февраль","Март","Апрель","Май","Июнь",
           "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

def F(c): return PatternFill("solid",fgColor=c)
def fnt(sz=10,bold=False,col=DGRAY,it=False):
    return Font(name="Calibri",size=sz,bold=bold,color=col,italic=it)
def brd(c=BORDER,s="thin"):
    sd=Side(style=s,color=c); return Border(left=sd,right=sd,top=sd,bottom=sd)
def brd_m(c=INP_BD):
    sd=Side(style="medium",color=c); return Border(left=sd,right=sd,top=sd,bottom=sd)
def CA(w=False): return Alignment(horizontal="center",vertical="center",wrap_text=w)
def LA(w=True):  return Alignment(horizontal="left",  vertical="center",wrap_text=w)
def RA():        return Alignment(horizontal="right", vertical="center")
def prot(locked=True): return Protection(locked=locked)
def cw(ws,d):
    for col,w in d.items(): ws.column_dimensions[col].width=w

def banner(ws,txt,merge,bg=TEAL,sz=14,fg=WHITE):
    ws.merge_cells(merge)
    c=ws[merge.split(":")[0]]; c.value=txt
    c.font=fnt(sz,True,fg); c.fill=F(bg); c.alignment=CA()
    r=int(''.join(x for x in merge.split(":")[0] if x.isdigit()))
    ws.row_dimensions[r].height=40
    # fill all cells
    from openpyxl.utils.cell import range_boundaries
    mc,mr,xc,xr = range_boundaries(merge)
    for row in range(mr,xr+1):
        for col in range(mc,xc+1):
            ws.cell(row,col).fill=F(bg)

def sec_hdr(ws,row,txt,ncols,bg=TEAL_M,h=24):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row,1); c.value=txt
    c.font=fnt(10,True,WHITE); c.fill=F(bg); c.alignment=LA(False)
    ws.row_dimensions[row].height=h

def hdr_row(ws,row,headers,bg=TEAL,h=24):
    for ci,(txt,w) in enumerate(headers,1):
        c=ws.cell(row,ci); c.value=txt
        c.font=fnt(9,True,WHITE); c.fill=F(bg); c.alignment=CA(True); c.border=brd()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[row].height=h

def inp(ws,row,col,val=None,fmt=MONEY,locked=False):
    c=ws.cell(row,col)
    if val is not None: c.value=val
    c.fill=F(INP); c.border=brd_m(); c.font=fnt(11,True,TEAL_M)
    c.alignment=CA(False); c.number_format=fmt; c.protection=prot(locked)
    return c

wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════
# 1. ПУЛЬТ — ежедневный пульт управления
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ПУЛЬТ"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = "FF0B4F54"
cw(ws,{"A":3,"B":24,"C":18,"D":3,"E":24,"F":18,"G":3,"H":24,"I":18})

banner(ws,'⚡  ПУЛЬТ УПРАВЛЕНИЯ — WAY MARKET',"A1:I1",TEAL,16)
ws.row_dimensions[1].height=44

# Row 2: shop name + date
ws.merge_cells("B2:C2")
ws.cell(2,2).value='=НАСТРОЙКИ!E5&" — "&TEXT(TODAY(),"DD MMMM YYYY")'
ws.cell(2,2).font=fnt(12,True,TEAL); ws.cell(2,2).fill=F(LGRAY); ws.cell(2,2).alignment=LA(False)
ws.merge_cells("E2:I2")
ws.cell(2,5).value='Данные обновляются при каждом открытии файла'
ws.cell(2,5).font=fnt(9,it=True,col=GRAY); ws.cell(2,5).fill=F(LGRAY); ws.cell(2,5).alignment=RA()
ws.row_dimensions[2].height=28

# ── Блок 1: KPI за текущий месяц (строки 4-11) ───────────────────
sec_hdr(ws,3,"  📊  МЕСЯЦ ДО СЕГОДНЯ (MTD) — ключевые показатели",9,TEAL,28)
ws.row_dimensions[3].height=28

_rA="БАЗА_ДДС!$A$4:$A$3003"; _rD="БАЗА_ДДС!$D$4:$D$3003"
_rE="БАЗА_ДДС!$E$4:$E$3003"; _rG="БАЗА_ДДС!$G$4:$G$3003"
mtd_s="DATE(YEAR(TODAY()),MONTH(TODAY()),1)"; mtd_e="TODAY()"
prev_s="DATE(YEAR(TODAY()),MONTH(TODAY())-1,1)"; prev_e="EOMONTH(DATE(YEAR(TODAY()),MONTH(TODAY())-1,1),0)"

def sp(typ=None,cat=None,s=None,e=None):
    s_=s or mtd_s; e_=e or mtd_e
    c=[f"({_rA}>={s_})",f"({_rA}<={e_})"]
    if typ: c.append(f'({_rD}="{typ}")')
    if cat: c.append(f'({_rE}="{cat}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(c)}*{_rG}),0)'

def sp_prev(typ=None):
    c=[f"({_rA}>={prev_s})",f"({_rA}<={prev_e})"]
    if typ: c.append(f'({_rD}="{typ}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(c)}*{_rG}),0)'

kpis = [
    ("B","C","💰 Выручка (MTD)",  sp("Доход"),    sp_prev("Доход"),    GREEN,   GREEN_L),
    ("E","F","📦 Расходы (MTD)",   sp("Расход"),   sp_prev("Расход"),   AMBER,   GOLD_L),
    ("H","I","📈 Прибыль (MTD)",   f'=IFERROR({sp("Доход")[1:]}-{sp("Расход")[1:]},0)', None, TEAL_M, TEAL_L),
]
for col_l,col_v,lbl,val_f,prev_f,val_clr,val_bg in kpis:
    cl = get_column_letter(ord(col_l)-64)
    cv_l = ord(col_l)-64; cv_v = ord(col_v)-64
    # Label
    ws.merge_cells(start_row=4,start_column=cv_l,end_row=4,end_column=cv_v)
    c=ws.cell(4,cv_l); c.value=lbl
    c.font=fnt(9,col=WHITE); c.fill=F(val_clr); c.alignment=CA(); c.border=brd()
    ws.row_dimensions[4].height=22
    # Value
    ws.merge_cells(start_row=5,start_column=cv_l,end_row=5,end_column=cv_v)
    c=ws.cell(5,cv_l); c.value=val_f
    c.font=fnt(18,True,val_clr); c.fill=F(val_bg); c.alignment=CA(False); c.border=brd(); c.number_format=MONEY
    ws.row_dimensions[5].height=38
    # Prev period comparison
    ws.merge_cells(start_row=6,start_column=cv_l,end_row=6,end_column=cv_v)
    if prev_f:
        pf_stripped = prev_f[1:]
        vf_stripped = val_f[1:] if val_f.startswith('=') else val_f
        c2=ws.cell(6,cv_l)
        c2.value=f'=IFERROR("Пред. мес.: "&TEXT({pf_stripped},"{MONEY}")&IF({vf_stripped}>{pf_stripped},"  ▲"," ▼"),"—")'
        c2.font=fnt(9,col=GRAY); c2.fill=F(LGRAY); c2.alignment=CA(); c2.border=brd()
    else:
        c2=ws.cell(6,cv_l); c2.value="Всё время"; c2.font=fnt(9,it=True,col=GRAY)
        c2.fill=F(LGRAY); c2.alignment=CA(); c2.border=brd()
    ws.row_dimensions[6].height=20

ws.row_dimensions[7].height=8  # spacer

# ── Блок 2: Показатели за сегодня (строки 8-11) ──────────────────
sec_hdr(ws,8,"  📅  СЕГОДНЯ",9,GOLD,24)
today_s="TODAY()"; today_e="TODAY()"
today_kpis = [
    ("B","C","Выручка сегодня",   sp("Доход",s=today_s,e=today_e),  GREEN,  GREEN_L),
    ("E","F","Расходы сегодня",   sp("Расход",s=today_s,e=today_e), AMBER,  GOLD_L),
    ("H","I","Прибыль сегодня",   f'=IFERROR({sp("Доход",s=today_s,e=today_e)[1:]}-{sp("Расход",s=today_s,e=today_e)[1:]},0)', TEAL_M, TEAL_L),
]
for col_l,col_v,lbl,val_f,val_clr,val_bg in today_kpis:
    cv_l=ord(col_l)-64; cv_v=ord(col_v)-64
    ws.merge_cells(start_row=9,start_column=cv_l,end_row=9,end_column=cv_v)
    c=ws.cell(9,cv_l); c.value=lbl; c.font=fnt(9,col=GRAY); c.fill=F(LGRAY); c.alignment=CA(); c.border=brd()
    ws.row_dimensions[9].height=20
    ws.merge_cells(start_row=10,start_column=cv_l,end_row=10,end_column=cv_v)
    c=ws.cell(10,cv_l); c.value=val_f
    c.font=fnt(15,True,val_clr); c.fill=F(val_bg); c.alignment=CA(False); c.border=brd(); c.number_format=MONEY
    ws.row_dimensions[10].height=32

ws.row_dimensions[11].height=8

# ── Блок 3: Сигналы / Алерты (строки 12-17) ─────────────────────
sec_hdr(ws,12,"  🚨  СИГНАЛЫ — требуют внимания",9,RED,24)
alerts = [
    (13,"Просроч. выплаты",
     f'=IFERROR(SUMPRODUCT((ЗАПИСЬ_ВЫПЛАТ!$B$4:$B$503<TODAY())*(ЗАПИСЬ_ВЫПЛАТ!$E$4:$E$503="Запланировано")*(ЗАПИСЬ_ВЫПЛАТ!$D$4:$D$503<>"")*(1)),0)',
     "шт.", RED, RED_L),
    (14,"Сумма просрочки",
     f'=IFERROR(SUMPRODUCT((ЗАПИСЬ_ВЫПЛАТ!$B$4:$B$503<TODAY())*(ЗАПИСЬ_ВЫПЛАТ!$E$4:$E$503="Запланировано")*ЗАПИСЬ_ВЫПЛАТ!$D$4:$D$503),0)',
     "₽", RED, RED_L),
    (15,"Иман хозяйки (месяц)",
     f'=IFERROR(SUMPRODUCT(({_rA}>={mtd_s})*({_rA}<={mtd_e})*({_rD}="Иман")*{_rG}),0)',
     "₽", AMBER, AMBER_L),
    (16,"Долг поставщикам",
     f'=IFERROR(SUMPRODUCT(({_rD}="Долг")*{_rG})-SUMPRODUCT(({_rD}="Оплата долга")*{_rG})+НАСТРОЙКИ!$E$9,0)',
     "₽", AMBER, AMBER_L),
]
for (row,lbl,formula,unit,clr,bg) in alerts:
    ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=3)
    c=ws.cell(row,2); c.value=lbl; c.font=fnt(9,col=DGRAY); c.fill=F(LGRAY); c.alignment=LA(False); c.border=brd()
    ws.merge_cells(start_row=row,start_column=5,end_row=row,end_column=6)
    cv=ws.cell(row,5); cv.value=formula
    cv.font=fnt(13,True,clr); cv.fill=F(bg); cv.alignment=CA(False); cv.border=brd()
    cv.number_format = MONEY if unit=="₽" else "#,##0"
    ws.merge_cells(start_row=row,start_column=8,end_row=row,end_column=9)
    cu=ws.cell(row,8); cu.value=unit; cu.font=fnt(9,it=True,col=GRAY); cu.fill=F(LGRAY); cu.alignment=LA(False); cu.border=brd()
    ws.row_dimensions[row].height=26
    ws.conditional_formatting.add(f"E{row}",FormulaRule(formula=[f"E{row}>0"],font=fnt(13,True,RED),fill=F(RED_L)))
    ws.conditional_formatting.add(f"E{row}",FormulaRule(formula=[f"E{row}<=0"],font=fnt(13,True,GREEN),fill=F(GREEN_L)))

ws.row_dimensions[17].height=8
ws.freeze_panes="B3"
print("✓ ПУЛЬТ")

# ════════════════════════════════════════════════════════════════════
# 2. НАСТРОЙКИ
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("НАСТРОЙКИ"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = TEAL
cw(ws,{"A":28,"B":3,"C":20,"D":3,"E":18,"F":3,"G":26,"H":16})
banner(ws,"⚙  НАСТРОЙКИ — ЦЕНТР УПРАВЛЕНИЯ","A1:H1",TEAL,14)

def n_sec(r_,title_,clr_=TEAL):
    ws.merge_cells(f"A{r_}:H{r_}")
    ws.cell(r_,1).value=title_; ws.cell(r_,1).font=fnt(11,True,WHITE)
    ws.cell(r_,1).fill=F(clr_); ws.cell(r_,1).alignment=LA(False)
    ws.row_dimensions[r_].height=28

def n_param(r_,label_,default_,typ_="money",hint_=""):
    ws.cell(r_,1).value=label_; ws.cell(r_,1).font=fnt(10)
    ws.cell(r_,1).fill=F(LGRAY); ws.cell(r_,1).border=brd(); ws.cell(r_,1).alignment=LA(False)
    c=ws.cell(r_,5); c.value=default_; c.font=fnt(11,True,TEAL_M)
    c.fill=F(INP); c.border=brd_m(); c.protection=prot(False)
    c.alignment=CA(False)
    if typ_=="money": c.number_format=MONEY
    elif typ_=="pct":  c.number_format="0%"
    elif typ_=="date": c.number_format=DATE_F
    elif typ_=="int":  c.number_format="#,##0"
    ws.cell(r_,7).value=hint_; ws.cell(r_,7).font=fnt(9,it=True,col=GRAY)
    ws.cell(r_,7).fill=F(LGRAY); ws.cell(r_,7).border=brd(); ws.cell(r_,7).alignment=LA(False)
    ws.row_dimensions[r_].height=24

# РАЗДЕЛ 1: Основное
n_sec(3,"  РАЗДЕЛ 1 — МАГАЗИН И ПЕРИОД")
n_param(4,"Название магазина","WAY MARKET","text","Отображается в отчётах и ПУЛЬТ")
n_param(5,"Год работы",today.year,"int","Используется в формулах ДАШБОРД")
n_param(6,"Финансовая модель (маржа)",0.25,"pct","25% = стандарт, изменить при необходимости")
n_param(7,"Лимит на закуп (% выручки)",0.75,"pct","75% = остаток после маржи")
n_param(8,"Начальный долг поставщикам",0,"money","Долг на момент начала ведения учёта")
n_param(9,"Порог расхождения кассы",500,"money","Сигнал на ПУЛЬТ при превышении")
n_param(10,"Порог просрочки Иман (дней)",30,"int","Через сколько дней Иман = просрочен")

# РАЗДЕЛ 2: Смены
n_sec(12,"  РАЗДЕЛ 2 — СМЕНЫ",TEAL_M)
for r_,sh_,def_ in [(13,"ДЕНЬ",True),(14,"ВЕЧЕР",True),(15,"НОЧЬ",False)]:
    ws.cell(r_,1).value=f"Смена {sh_} активна"; ws.cell(r_,1).font=fnt(10)
    ws.cell(r_,1).fill=F(LGRAY); ws.cell(r_,1).border=brd(); ws.cell(r_,1).alignment=LA(False)
    c=ws.cell(r_,5); c.value="Вкл" if def_ else "Выкл"
    c.font=fnt(11,True,GREEN if def_ else GRAY)
    c.fill=F(GREEN_L if def_ else LGRAY); c.border=brd_m(); c.protection=prot(False); c.alignment=CA()
    dv=DataValidation(type="list",formula1='"Вкл,Выкл"',allow_blank=False)
    ws.add_data_validation(dv); dv.add(c)
    ws.row_dimensions[r_].height=24

# РАЗДЕЛ 3: Способы оплаты
n_sec(17,"  РАЗДЕЛ 3 — СПОСОБЫ ОПЛАТЫ",TEAL_M)
for r_,pay_,def_ in [(18,"Наличные",True),(19,"Эквайринг",True),(20,"Перевод (СБП)",True),(21,"Иман (доверие)",True)]:
    ws.cell(r_,1).value=pay_; ws.cell(r_,1).font=fnt(10)
    ws.cell(r_,1).fill=F(LGRAY); ws.cell(r_,1).border=brd(); ws.cell(r_,1).alignment=LA(False)
    c=ws.cell(r_,5); c.value="Вкл" if def_ else "Выкл"
    c.font=fnt(11,True,GREEN if def_ else GRAY)
    c.fill=F(GREEN_L if def_ else LGRAY); c.border=brd_m(); c.protection=prot(False); c.alignment=CA()
    dv2=DataValidation(type="list",formula1='"Вкл,Выкл"',allow_blank=False)
    ws.add_data_validation(dv2); dv2.add(c)
    ws.row_dimensions[r_].height=24

# РАЗДЕЛ 4: Месячные планы (новый!)
n_sec(23,"  РАЗДЕЛ 4 — ПЛАН ВЫРУЧКИ ПО МЕСЯЦАМ (₽)",GOLD)
ws.merge_cells("A24:H24")
ws.cell(24,1).value="Укажите плановую выручку для каждого месяца — ДАШБОРД покажет выполнение плана"
ws.cell(24,1).font=fnt(9,it=True,col=GRAY); ws.cell(24,1).fill=F(GOLD_L)
ws.cell(24,1).alignment=LA(False); ws.row_dimensions[24].height=20
# 2 rows of months, 6 per row
month_defaults = [350000,320000,380000,400000,420000,450000,470000,450000,410000,380000,360000,340000]
for mi,mn in enumerate(MONTHS_RU):
    r_ = 25 if mi<6 else 26
    c_ = (mi%6)*2+1  # cols 1,3,5,7,9,11
    ws.cell(r_,c_).value=mn; ws.cell(r_,c_).font=fnt(9,col=GRAY)
    ws.cell(r_,c_).fill=F(LGRAY); ws.cell(r_,c_).border=brd(); ws.cell(r_,c_).alignment=CA()
    ws.column_dimensions[get_column_letter(c_)].width=10
    cv=ws.cell(r_,c_+1); cv.value=month_defaults[mi]; cv.font=fnt(10,True,GOLD)
    cv.fill=F(GOLD_L); cv.border=brd_m(); cv.number_format=MONEY
    cv.protection=prot(False); cv.alignment=CA()
    ws.column_dimensions[get_column_letter(c_+1)].width=13
for r_ in [25,26]: ws.row_dimensions[r_].height=26

# РАЗДЕЛ 5: Справочники
n_sec(28,"  РАЗДЕЛ 5 — СПРАВОЧНИКИ",TEAL_M)
ref_data = [
    (29,"Кассиры",     ["A","B","C","D","E","F","G"], "Ахмед,Заур,Лейла,Малика,Ибрагим,Хасан,Зара"),
    (35,"Категории",   ["C","B","A","D","E","F","G"], "Закуп товара,Аренда,ЗП,Коммунальные,Налоги,Охрана,Прочее"),
    (42,"Типы опер.",  ["E","F","G","H","I","J","K"], "Доход,Расход,Иман,Долг,Оплата долга,Расхождение,Списание"),
]
for start_r, name, _, items_str in ref_data:
    ws.cell(start_r,1).value=f"▸ {name}"; ws.cell(start_r,1).font=fnt(9,True,WHITE)
    ws.cell(start_r,1).fill=F(TEAL_M); ws.cell(start_r,1).alignment=LA(False); ws.row_dimensions[start_r].height=20
    for i,item in enumerate(items_str.split(",")):
        ws.cell(start_r+1+i,1).value=item.strip()
        ws.cell(start_r+1+i,1).font=fnt(10); ws.cell(start_r+1+i,1).fill=F(LGRAY)
        ws.cell(start_r+1+i,1).border=brd(); ws.cell(start_r+1+i,1).alignment=LA(False)
        ws.cell(start_r+1+i,1).protection=prot(False)
        ws.row_dimensions[start_r+1+i].height=22

# РАЗДЕЛ 6: Поставщики (строки 50-100)
n_sec(50,"  РАЗДЕЛ 6 — ПОСТАВЩИКИ (ТОП-5 для отчётов)",TEAL_M)
ws.merge_cells("A51:H51")
ws.cell(51,1).value="Введите основных поставщиков — они появятся в ОТЧЁТ и выпадающих списках ВЫПЛАТ"
ws.cell(51,1).font=fnt(9,it=True,col=GRAY); ws.cell(51,1).fill=F(LGRAY); ws.cell(51,1).alignment=LA(False); ws.row_dimensions[51].height=20
sup_hdrs=[("Поставщик",22),("Контакт",18),("Условия",14)]
for ci,(h,w) in enumerate(sup_hdrs,1):
    ws.cell(52,ci).value=h; ws.cell(52,ci).font=fnt(9,True,WHITE)
    ws.cell(52,ci).fill=F(TEAL); ws.cell(52,ci).alignment=CA(); ws.cell(52,ci).border=brd()
    ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[52].height=22
for i in range(5):
    for ci in range(1,4):
        c_=ws.cell(53+i,ci); c_.fill=F(INP if ci==1 else LGRAY)
        c_.border=brd_m() if ci==1 else brd()
        c_.protection=prot(False); c_.alignment=LA(False); c_.font=fnt(10)
    ws.row_dimensions[53+i].height=24

# РАЗДЕЛ 6: Пороги уведомлений
n_sec(60,"  РАЗДЕЛ 6 — ПОРОГИ УВЕДОМЛЕНИЙ",RED)
ws.merge_cells("A61:H61")
ws.cell(61,1).value="Превышение этих порогов подсвечивается красным на ПУЛЬТ и ДАШБОРД"
ws.cell(61,1).font=fnt(9,it=True,col=GRAY); ws.cell(61,1).fill=F(RED_L); ws.cell(61,1).alignment=LA(False); ws.row_dimensions[61].height=20
for r_,lbl_,def_,hint_ in [
    (62,"Расхождение кассы > (₽)",5000,"Подсветка расхождения красным на ПУЛЬТ"),
    (63,"Общий долг поставщикам > (₽)",1000000,"Подсветка долга на ДАШБОРД"),
    (64,"Просрочка выплат > (дней)",7,"Просрочена выплата если дата + X дней < сегодня"),
]:
    n_param(r_,lbl_,def_,"money" if r_<64 else "int",hint_)

ws.protection.sheet=True; ws.protection.password="wm2025"
ws.freeze_panes="A3"
print("✓ НАСТРОЙКИ")

# ════════════════════════════════════════════════════════════════════
# 3. ВВОД_КАССА
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ВВОД_КАССА"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = GREEN
cw(ws,{"A":13,"B":13,"C":13,"D":13,"E":13,"F":18,"G":16,"H":16,"I":12,"J":12,"K":12})
banner(ws,"📥  ВВОД КАССЫ — ежедневный Z-отчёт","A1:K1",GREEN,14)

ws.merge_cells("A2:K2")
ws.cell(2,1).value="Заполните дату, смену, кассира и суммы по способам оплаты → нажмите СОХРАНИТЬ КАССУ"
ws.cell(2,1).font=fnt(9,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY)
ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=20

# ── Строка 3: Дата + Смена + Кассир ──────────────────────────────
cw(ws,{"A":20,"B":14,"C":14,"D":14,"E":14,"F":14,"G":14})
for ci_,(lbl_,c1_,c2_) in enumerate(
        [("Дата",1,1),("Смена",3,3),("Кассир",5,6)],0):
    ws.cell(3,c1_*2-1).value=lbl_; ws.cell(3,c1_*2-1).font=fnt(10,True,DGRAY)
    ws.cell(3,c1_*2-1).fill=F(LGRAY); ws.cell(3,c1_*2-1).border=brd(); ws.cell(3,c1_*2-1).alignment=LA(False)
# direct placement is cleaner:
ws.cell(3,1).value="Дата"; ws.cell(3,1).font=fnt(10,True,DGRAY); ws.cell(3,1).fill=F(LGRAY); ws.cell(3,1).border=brd(); ws.cell(3,1).alignment=LA(False)
c_date=inp(ws,3,2,None,DATE_F,False); c_date.value="=TODAY()"
ws.cell(3,3).value="Смена"; ws.cell(3,3).font=fnt(10,True,DGRAY); ws.cell(3,3).fill=F(LGRAY); ws.cell(3,3).border=brd(); ws.cell(3,3).alignment=LA(False)
c_shift=inp(ws,3,4,"День","@",False)
dv_sh=DataValidation(type="list",formula1='"День,Вечер,Ночь"',allow_blank=False)
ws.add_data_validation(dv_sh); dv_sh.add(c_shift)
ws.cell(3,5).value="Кассир"; ws.cell(3,5).font=fnt(10,True,DGRAY); ws.cell(3,5).fill=F(LGRAY); ws.cell(3,5).border=brd(); ws.cell(3,5).alignment=LA(False)
ws.merge_cells("F3:G3"); c_cash=inp(ws,3,6,"","@",False)
dv_cas=DataValidation(type="list",formula1="НАСТРОЙКИ!$A$30:$A$36",allow_blank=True)
ws.add_data_validation(dv_cas); dv_cas.add(c_cash)
ws.row_dimensions[3].height=28

ws.row_dimensions[4].height=8

# ── Строки 5-10: Таблица ввода Z / Розетка / Факт ─────────────────
# Row 5: column headers
for ci_,(hdr_,bg_) in enumerate([("Способ оплаты",TEAL),
    ("Z-отчёт (касса, ₽)",TEAL_M),("Розетка/Терминал (₽)",TEAL_M),
    ("Факт (₽)",TEAL_M),("Расхождение",TEAL)],1):
    ws.cell(5,ci_).value=hdr_
    ws.cell(5,ci_).font=fnt(9,True,WHITE); ws.cell(5,ci_).fill=F(bg_)
    ws.cell(5,ci_).alignment=CA(True); ws.cell(5,ci_).border=brd()
ws.row_dimensions[5].height=30

# НАСТРОЙКИ rows for payment toggles: 18=Нал, 19=Эквайр, 20=Перевод, 21=Иман
_pay_rows = [
    (6,  "Наличные",        "НАСТРОЙКИ!$E$18", True),   # Z, Розетка(нет), Факт
    (7,  "Эквайринг",       "НАСТРОЙКИ!$E$19", False),  # Z, Розетка, Факт
    (8,  "Перевод (СБП)",   "НАСТРОЙКИ!$E$20", False),  # Z, Розетка, Факт
    (9,  "Иман хозяйки",    "НАСТРОЙКИ!$E$21", None),   # только Z (сумма взятого)
]
KASSA_CELLS={}  # row → (z_cell, roz_cell, fakt_cell)
for (r_,name_,setting_,cash_only_) in _pay_rows:
    # Label
    ws.cell(r_,1).value=name_; ws.cell(r_,1).font=fnt(10,bold=True,col=DGRAY)
    ws.cell(r_,1).fill=F(LGRAY); ws.cell(r_,1).border=brd(); ws.cell(r_,1).alignment=LA(False)
    # Z-отчёт (col B=2)
    c_z=inp(ws,r_,2,0,MONEY,False)
    # Розетка/Терминал (col C=3) — not for Иман
    if cash_only_ is None:  # Иман — no terminal
        ws.cell(r_,3).value="—"; ws.cell(r_,3).font=fnt(9,it=True,col=GRAY)
        ws.cell(r_,3).fill=F(LGRAY); ws.cell(r_,3).border=brd(); ws.cell(r_,3).alignment=CA()
        ws.cell(r_,4).value="—"; ws.cell(r_,4).font=fnt(9,it=True,col=GRAY)
        ws.cell(r_,4).fill=F(LGRAY); ws.cell(r_,4).border=brd(); ws.cell(r_,4).alignment=CA()
        ws.cell(r_,5).value="—"; ws.cell(r_,5).font=fnt(9,it=True,col=GRAY)
        ws.cell(r_,5).fill=F(LGRAY); ws.cell(r_,5).border=brd(); ws.cell(r_,5).alignment=CA()
        c_roz=None; c_fakt=None
    else:
        c_roz=inp(ws,r_,3,0,MONEY,False)
        c_fakt=inp(ws,r_,4,0,MONEY,False)
        # Расхождение = Факт − Z
        c_dif=ws.cell(r_,5)
        c_dif.value=f'=IFERROR(D{r_}-B{r_},0)'
        c_dif.font=fnt(11,True,GRAY); c_dif.fill=F(LGRAY); c_dif.border=brd()
        c_dif.alignment=CA(False); c_dif.number_format=MONEY
        ws.conditional_formatting.add(f"E{r_}",FormulaRule(
            formula=[f"AND(E{r_}<>0,{setting_}=\"Вкл\")"],fill=F(RED_L),font=fnt(11,True,RED)))
        ws.conditional_formatting.add(f"E{r_}",FormulaRule(
            formula=[f"E{r_}=0"],fill=F(GREEN_L),font=fnt(11,True,GREEN)))
    KASSA_CELLS[r_]=(c_z,c_roz,c_fakt)
    # Grey out if disabled in НАСТРОЙКИ
    for ci_ in range(1,6):
        ws.conditional_formatting.add(f"{get_column_letter(ci_)}{r_}",FormulaRule(
            formula=[f'{setting_}="Выкл"'],fill=F(DISABLED if ci_>1 else LGRAY),
            font=fnt(10,col=GRAY)))
    ws.row_dimensions[r_].height=26

ws.row_dimensions[10].height=8

# ── Строки 11-13: Итоги ───────────────────────────────────────────
for ci_,(hdr_,bg_) in enumerate([("ИТОГО",TEAL),("Z-отчёт",TEAL_M),
    ("Розетка",TEAL_M),("Факт",TEAL_M),("Расхождение",TEAL)],1):
    ws.cell(11,ci_).value=hdr_
    ws.cell(11,ci_).font=fnt(9,True,WHITE); ws.cell(11,ci_).fill=F(bg_)
    ws.cell(11,ci_).alignment=CA(True); ws.cell(11,ci_).border=brd()
ws.row_dimensions[11].height=22

ws.cell(12,1).value="Сумма смены"
ws.cell(12,1).font=fnt(10,True,TEAL); ws.cell(12,1).fill=F(TEAL_L); ws.cell(12,1).border=brd(); ws.cell(12,1).alignment=LA(False)
# Z total = Нал+Эквайр+Перевод+Иман
ws.cell(12,2).value="=IFERROR(B6+B7+B8+B9,0)"
ws.cell(12,2).font=fnt(13,True,GREEN); ws.cell(12,2).fill=F(GREEN_L); ws.cell(12,2).border=brd_m(); ws.cell(12,2).number_format=MONEY; ws.cell(12,2).alignment=CA()
# Розетка total
ws.cell(12,3).value="=IFERROR(C6+C7+C8,0)"
ws.cell(12,3).font=fnt(11,True,TEAL_M); ws.cell(12,3).fill=F(TEAL_L); ws.cell(12,3).border=brd(); ws.cell(12,3).number_format=MONEY; ws.cell(12,3).alignment=CA()
# Факт total
ws.cell(12,4).value="=IFERROR(D6+D7+D8,0)"
ws.cell(12,4).font=fnt(11,True,TEAL_M); ws.cell(12,4).fill=F(TEAL_L); ws.cell(12,4).border=brd(); ws.cell(12,4).number_format=MONEY; ws.cell(12,4).alignment=CA()
# Расхождение total = Факт+Иман − Z (Иман объясняет разницу)
ws.cell(12,5).value="=IFERROR(D12+B9-B12,0)"
ws.cell(12,5).font=fnt(11,True,GRAY); ws.cell(12,5).fill=F(LGRAY); ws.cell(12,5).border=brd(); ws.cell(12,5).number_format=MONEY; ws.cell(12,5).alignment=CA()
ws.conditional_formatting.add("E12",FormulaRule(formula=["E12<>0"],fill=F(RED_L),font=fnt(11,True,RED)))
ws.conditional_formatting.add("E12",FormulaRule(formula=["E12=0"],fill=F(GREEN_L),font=fnt(11,True,GREEN)))
ws.row_dimensions[12].height=30

ws.cell(13,1).value="Примечание"; ws.cell(13,1).font=fnt(9,col=GRAY); ws.cell(13,1).fill=F(LGRAY); ws.cell(13,1).border=brd(); ws.cell(13,1).alignment=LA(False)
ws.merge_cells("B13:G13"); c_note=inp(ws,13,2,"","@",False); c_note.font=fnt(10)
ws.row_dimensions[13].height=24
ws.row_dimensions[14].height=8

# СОХРАНИТЬ button
ws.merge_cells("A15:G15")
ws.cell(15,1).value="▶  СОХРАНИТЬ КАССУ  (VBA макрос)"
ws.cell(15,1).font=fnt(12,True,WHITE); ws.cell(15,1).fill=F(TEAL)
ws.cell(15,1).alignment=CA(); ws.row_dimensions[15].height=30
ws.row_dimensions[16].height=8

# History table tblВводКасса — расширенная с Z/Розетка/Факт
hdr_row(ws,17,[("Дата",12),("Смена",11),("Кассир",13),
               ("Z Нал",12),("Z Эквайр",12),("Z Перевод",12),("Z Иман",11),
               ("Роз.Нал",11),("Роз.Экв",11),("Роз.Пер",11),
               ("Факт Нал",12),("Факт Экв",12),("Факт Пер",12),
               ("Расхожд.",12),("Статус",12)],GREEN,24)
for r in range(18,518):
    for ci in range(1,16):
        c_=ws.cell(r,ci); c_.fill=F(WHITE if r%2==0 else LGRAY); c_.border=brd()
        c_.protection=prot(False)
    ws.cell(r,1).number_format=DATE_F
    for ci in range(4,15): ws.cell(r,ci).number_format=MONEY
    ws.row_dimensions[r].height=18

tbl_kassa=Table(displayName="tblВводКасса",ref="A17:O517")
tbl_kassa.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
ws.add_table(tbl_kassa)
ws.freeze_panes="A18"
print("✓ ВВОД_КАССА")

# ════════════════════════════════════════════════════════════════════
# 4. ВВОД_РАСХОДЫ
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ВВОД_РАСХОДЫ"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = AMBER
cw(ws,{"A":13,"B":18,"C":14,"D":16,"E":14,"F":18})
banner(ws,"📤  ВВОД РАСХОДОВ","A1:F1",AMBER,14,WHITE)

for r_,lbl_,fmt_ in [(3,"Дата",DATE_F),(4,"Категория","@"),(5,"Способ оплаты","@"),(6,"Сумма (₽)",MONEY),(7,"Комментарий","@")]:
    ws.cell(r_,1).value=lbl_; ws.cell(r_,1).font=fnt(10,True,DGRAY)
    ws.cell(r_,1).fill=F(LGRAY); ws.cell(r_,1).border=brd(); ws.cell(r_,1).alignment=LA(False)
    ws.merge_cells(start_row=r_,start_column=2,end_row=r_,end_column=4)
    c_=inp(ws,r_,2,None,fmt_,False)
    if lbl_=="Дата": c_.value="=TODAY()"; c_.number_format=DATE_F; c_.font=fnt(11,True,TEAL_M)
    if lbl_=="Категория":
        dv=DataValidation(type="list",formula1="НАСТРОЙКИ!$A$36:$A$42",allow_blank=False)
        ws.add_data_validation(dv); dv.add(c_)
    if lbl_=="Способ оплаты":
        dv=DataValidation(type="list",formula1='"Наличные,Эквайринг,Перевод,Безналичные"',allow_blank=False)
        ws.add_data_validation(dv); dv.add(c_)
    ws.row_dimensions[r_].height=26

ws.merge_cells("A9:D9")
ws.cell(9,1).value="▶  СОХРАНИТЬ РАСХОД  (VBA макрос)"
ws.cell(9,1).font=fnt(12,True,WHITE); ws.cell(9,1).fill=F(AMBER)
ws.cell(9,1).alignment=CA(); ws.row_dimensions[9].height=30

hdr_row(ws,11,[("Дата",12),("Категория",18),("Способ",14),("Сумма (₽)",16),("Комментарий",22),("Статус",12)],AMBER,24)
for r in range(12,512):
    for ci in range(1,7):
        c_=ws.cell(r,ci); c_.fill=F(WHITE if r%2==0 else LGRAY); c_.border=brd(); c_.protection=prot(False)
    ws.cell(r,1).number_format=DATE_F; ws.cell(r,4).number_format=MONEY
    ws.row_dimensions[r].height=20

tbl_rash=Table(displayName="tblВводРасходы",ref="A11:F511")
tbl_rash.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True)
ws.add_table(tbl_rash)
ws.freeze_panes="A12"
print("✓ ВВОД_РАСХОДЫ")

# ════════════════════════════════════════════════════════════════════
# 5. БАЗА_ДДС
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("БАЗА_ДДС"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = NAVY
cw(ws,{"A":13,"B":12,"C":14,"D":14,"E":16,"F":14,"G":16,"H":14})
banner(ws,"🗄  БАЗА ДДС — главная книга транзакций","A1:H1",NAVY,14)

ws.merge_cells("A2:H2")
ws.cell(2,1).value="Только для чтения. Данные вносятся через ВВОД_КАССА и ВВОД_РАСХОДЫ."
ws.cell(2,1).font=fnt(9,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=20

hdr_row(ws,3,[("Дата",13),("Смена",12),("Кассир",14),("Тип операций",14),
              ("Категория",16),("Способ оплаты",14),("Сумма (₽)",16),("Прим.",14)],NAVY,24)
for r in range(4,3004):
    for ci in range(1,9):
        c_=ws.cell(r,ci); c_.fill=F(WHITE if r%2==0 else LGRAY); c_.border=brd(); c_.protection=prot(False)
    ws.cell(r,1).number_format=DATE_F; ws.cell(r,7).number_format=MONEY
    ws.row_dimensions[r].height=18

# CF: rows with type=Доход → green, Расход → amber, Иман → purple
for typ_,clr_ in [("Доход",GREEN_L),("Расход",AMBER_L),("Иман",PURP_L),("Расхождение",RED_L)]:
    ws.conditional_formatting.add(f"A4:H3003",FormulaRule(
        formula=[f'$D4="{typ_}"'],fill=F(clr_)))

tbl_baza=Table(displayName="tblБаза",ref="A3:H3003")
tbl_baza.tableStyleInfo=TableStyleInfo(name="TableStyleMedium9",showRowStripes=False)
ws.add_table(tbl_baza)
ws.freeze_panes="A4"
print("✓ БАЗА_ДДС")


# ════════════════════════════════════════════════════════════════════
# 7. ВЫПЛАТЫ_ПЛАН — выплаты поставщикам
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ЗАПИСЬ_ВЫПЛАТ"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = AMBER
cw(ws,{"A":5,"B":13,"C":20,"D":16,"E":13,"F":14,"G":14,"H":15,"I":18})
banner(ws,"💳  ЗАПИСЬ ВЫПЛАТ ПОСТАВЩИКАМ","A1:H1",GOLD,14,WHITE)

ws.merge_cells("A2:H2")
ws.cell(2,1).value="Фиксируйте все запланированные и выполненные выплаты поставщикам и торговым представителям."
ws.cell(2,1).font=fnt(9,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=20

# Summary bar
for ci,(lbl,formula,clr,bg) in enumerate([
    ("Запланировано",
     '=IFERROR(SUMPRODUCT((ЗАПИСЬ_ВЫПЛАТ!$E$4:$E$503="Запланировано")*ЗАПИСЬ_ВЫПЛАТ!$D$4:$D$503),0)',
     AMBER, GOLD_L),
    ("Выплачено",
     '=IFERROR(SUMPRODUCT((ЗАПИСЬ_ВЫПЛАТ!$E$4:$E$503="Выплачено")*ЗАПИСЬ_ВЫПЛАТ!$D$4:$D$503),0)',
     GREEN, GREEN_L),
    ("Просрочено",
     '=IFERROR(SUMPRODUCT((ЗАПИСЬ_ВЫПЛАТ!$B$4:$B$503<TODAY())*(ЗАПИСЬ_ВЫПЛАТ!$E$4:$E$503="Запланировано")*ЗАПИСЬ_ВЫПЛАТ!$D$4:$D$503),0)',
     RED, RED_L),
],1):
    col2=(ci-1)*3+1
    ws.merge_cells(start_row=3,start_column=col2,end_row=3,end_column=col2+1)
    lc=ws.cell(3,col2); lc.value=lbl; lc.font=fnt(9,col=GRAY); lc.fill=F(LGRAY); lc.alignment=CA(); lc.border=brd()
    ws.merge_cells(start_row=4,start_column=col2,end_row=4,end_column=col2+1)
    vc=ws.cell(4,col2); vc.value=formula
    vc.font=fnt(13,True,clr); vc.fill=F(bg); vc.alignment=CA(False); vc.border=brd(); vc.number_format=MONEY
    ws.row_dimensions[3].height=20; ws.row_dimensions[4].height=30

ws.row_dimensions[5].height=8

hdr_row(ws,6,[("№",4),("Дата выплаты",13),("Поставщик (ТП)",20),("Сумма (₽)",16),
              ("Статус",13),("Накладная",14),("Способ",14),("Факт. оплата",15),("Примечание",18)],GOLD,24)
for r in range(7,507):
    ws.cell(r,1).value=f'=IF(B{r}<>"",ROW()-6,"")'
    ws.cell(r,1).font=fnt(9,col=GRAY); ws.cell(r,1).alignment=CA(); ws.cell(r,1).border=brd()
    ws.cell(r,1).fill=F(WHITE if r%2==0 else LGRAY)
    for ci in range(2,10):
        c_=ws.cell(r,ci); c_.fill=F(WHITE if r%2==0 else LGRAY); c_.border=brd(); c_.protection=prot(False)
    ws.cell(r,2).number_format=DATE_F; ws.cell(r,4).number_format=MONEY
    ws.cell(r,8).number_format=DATE_F   # Дата факт. оплаты
    ws.row_dimensions[r].height=20

dv_st=DataValidation(type="list",formula1='"Запланировано,Выплачено,Отменено"',allow_blank=False)
ws.add_data_validation(dv_st); dv_st.sqref="E7:E506"
dv_sup=DataValidation(type="list",formula1="НАСТРОЙКИ!$A$53:$A$57",allow_blank=True)
ws.add_data_validation(dv_sup); dv_sup.sqref="C7:C506"

ws.conditional_formatting.add("A7:J506",FormulaRule(formula=['$E7="Выплачено"'],fill=F(GREEN_L),font=fnt(10,col=GREEN)))
ws.conditional_formatting.add("A7:J506",FormulaRule(formula=['AND($E7="Запланировано",$B7<TODAY())'],fill=F(RED_L),font=fnt(10,True,RED)))

# Col J: Idx — порядковый номер выплаты на эту дату (для календаря)
ws.cell(6,10).value="Idx"
for r in range(7,507):
    ws.cell(r,10).value=f'=IF(B{r}<>"",COUNTIFS($B$7:$B{r},$B{r}),"")'
    ws.cell(r,10).font=fnt(9,col=GRAY); ws.cell(r,10).fill=F(LGRAY); ws.cell(r,10).border=brd()
    ws.cell(r,10).alignment=CA(); ws.cell(r,10).protection=prot(True)
ws.column_dimensions["J"].width=5

tbl_vip=Table(displayName="tblВыплаты",ref="A6:J506")
tbl_vip.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=False)
ws.add_table(tbl_vip)
ws.freeze_panes="B7"
print("✓ ЗАПИСЬ_ВЫПЛАТ")

# ════════════════════════════════════════════════════════════════════
# 8. КАЛЕНДАРЬ_ВЫПЛАТ
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("КАЛЕНДАРЬ_ВЫПЛАТ"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = GOLD
banner(ws,"📅  КАЛЕНДАРЬ ВЫПЛАТ ПОСТАВЩИКАМ","A1:N1",GOLD,14,WHITE)
ws.merge_cells("A2:N2")
ws.cell(2,1).value="Выберите месяц и год — календарь покажет все выплаты. Просроченные выделены красным."
ws.cell(2,1).font=fnt(9,it=True,col=GOLD); ws.cell(2,1).fill=F(GOLD_L); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=20

# Filter bar row 4
ws.cell(4,1).value="Месяц:"; ws.cell(4,1).font=fnt(10,True,DGRAY); ws.cell(4,1).alignment=RA(); ws.cell(4,1).fill=F(LGRAY)
ws.merge_cells("B4:D4")
ws.cell(4,2).value=MONTHS_RU[today.month-1]
ws.cell(4,2).font=fnt(13,True,TEAL); ws.cell(4,2).fill=F(INP); ws.cell(4,2).border=brd_m(); ws.cell(4,2).alignment=CA()
dv_mon_k=DataValidation(type="list",formula1='"Январь,Февраль,Март,Апрель,Май,Июнь,Июль,Август,Сентябрь,Октябрь,Ноябрь,Декабрь"')
ws.add_data_validation(dv_mon_k); dv_mon_k.add("B4")
ws.cell(4,6).value="Год:"; ws.cell(4,6).font=fnt(10,True,DGRAY); ws.cell(4,6).alignment=RA(); ws.cell(4,6).fill=F(LGRAY)
ws.merge_cells("G4:H4")
ws.cell(4,7).value=today.year
ws.cell(4,7).font=fnt(13,True,TEAL); ws.cell(4,7).fill=F(INP); ws.cell(4,7).border=brd_m(); ws.cell(4,7).alignment=CA()
ws.row_dimensions[4].height=34

# Hidden helper cells (column P)
ws.cell(4,16).value='=MATCH(B4,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0)'
ws.cell(5,16).value='=DATE(G4,P4,1)'; ws.cell(5,16).number_format=DATE_F
ws.cell(6,16).value='=EOMONTH(P5,0)'; ws.cell(6,16).number_format=DATE_F
ws.cell(7,16).value='=WEEKDAY(P5,2)'
ws.column_dimensions['P'].hidden=True

# KPI summary row 6-8
_ZV="ЗАПИСЬ_ВЫПЛАТ!$B$7:$B$506"; _ZD="ЗАПИСЬ_ВЫПЛАТ!$D$7:$D$506"; _ZS="ЗАПИСЬ_ВЫПЛАТ!$E$7:$E$506"
sec_hdr(ws,6,"  СВОДКА ПО ВЫБРАННОМУ МЕСЯЦУ",14,TEAL_M,20)
def kpi_kal(ws,row,c1,c2,lbl,f_,bg_):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row,c1); c.value=lbl; c.font=fnt(9,True,WHITE); c.fill=F(NAVY); c.alignment=CA()
    ws.merge_cells(start_row=row+1,start_column=c1,end_row=row+1,end_column=c2)
    c=ws.cell(row+1,c1); c.value=f_; c.font=fnt(16,True,WHITE); c.fill=F(bg_); c.alignment=CA(); c.number_format=MONEY
    ws.row_dimensions[row].height=20; ws.row_dimensions[row+1].height=38

kpi_kal(ws,7,1,3,"К ВЫПЛАТЕ ВСЕГО",
    f'=IFERROR(SUMPRODUCT(({_ZV}>=$P$5)*({_ZV}<=$P$6)*{_ZD}),0)',TEAL)
kpi_kal(ws,7,5,7,"ВЫПЛАЧЕНО",
    f'=IFERROR(SUMPRODUCT(({_ZV}>=$P$5)*({_ZV}<=$P$6)*({_ZS}="Выплачено")*{_ZD}),0)',GREEN)
kpi_kal(ws,7,9,11,"ЗАПЛАНИРОВАНО",
    f'=IFERROR(SUMPRODUCT(({_ZV}>=$P$5)*({_ZV}<=$P$6)*({_ZS}="Запланировано")*{_ZD}),0)',GOLD_M)
kpi_kal(ws,7,12,14,"ПРОСРОЧЕНО",
    f'=IFERROR(SUMPRODUCT(({_ZV}<TODAY())*({_ZS}="Запланировано")*{_ZD}),0)',RED)

# Day-of-week header row 10
ws.row_dimensions[10].height=26
for i_,d_ in enumerate(['ПН','ВТ','СР','ЧТ','ПТ','СБ','ВС']):
    cs_=1+i_*2
    ws.merge_cells(start_row=10,start_column=cs_,end_row=10,end_column=cs_+1)
    c=ws.cell(10,cs_); c.value=d_; c.font=fnt(10,True,WHITE)
    c.fill=F(RED if i_>=5 else TEAL); c.alignment=CA(); c.border=brd()

# 6-week grid
CELL_ROWS_=4; SR_=11
_ZC="ЗАПИСЬ_ВЫПЛАТ!$C$7:$C$506"; _ZI="ЗАПИСЬ_ВЫПЛАТ!$J$7:$J$506"
for week_ in range(6):
    br_=SR_+week_*CELL_ROWS_
    for dp_ in range(7):
        cs_=1+dp_*2; ce_=cs_+1; di_=week_*7+dp_
        dr_=f'DATE($G$4,$P$4,1)+{di_}-($P$7-1)'
        chk_=f"${get_column_letter(cs_)}${br_}"
        ws.cell(br_,cs_).value=f'=IFERROR(IF(AND({dr_}>=$P$5,{dr_}<=$P$6),DAY({dr_}),""),"")'
        ws.cell(br_,cs_).font=fnt(13,True,TEAL); ws.cell(br_,cs_).alignment=LA()
        ws.cell(br_,ce_).value=f'=IFERROR(IF({chk_}="","",SUMPRODUCT(({_ZV}={dr_})*{_ZD})),"")'
        ws.cell(br_,ce_).font=fnt(11,True,RED); ws.cell(br_,ce_).alignment=RA(); ws.cell(br_,ce_).number_format='#,##0;;;'
        for tp_i in range(2):
            rt_=br_+1+tp_i; n_=tp_i+1
            ws.cell(rt_,cs_).value=f'=IFERROR(IF({chk_}="","",INDEX(ЗАПИСЬ_ВЫПЛАТ!$C:$C,SUMPRODUCT(({_ZV}={dr_})*({_ZI}={n_})*ROW(ЗАПИСЬ_ВЫПЛАТ!$B$7:$B$506)))),"")'
            ws.cell(rt_,cs_).font=fnt(8,col=TEAL_M); ws.cell(rt_,cs_).alignment=LA()
            ws.merge_cells(start_row=rt_,start_column=cs_,end_row=rt_,end_column=ce_)
        ws.cell(br_+3,cs_).value=f'=IFERROR(IF(OR({chk_}="",COUNTIFS(ЗАПИСЬ_ВЫПЛАТ!$B:$B,{dr_})<=2),"","+ ещё "&(COUNTIFS(ЗАПИСЬ_ВЫПЛАТ!$B:$B,{dr_})-2)),"")'
        ws.cell(br_+3,cs_).font=fnt(8,it=True,col=GRAY); ws.cell(br_+3,cs_).alignment=CA()
        ws.merge_cells(start_row=br_+3,start_column=cs_,end_row=br_+3,end_column=ce_)
        bg_c=RED_L if dp_>=5 else LGRAY
        for ri_ in range(br_,br_+4):
            for ci__ in [cs_,ce_]:
                ws.cell(ri_,ci__).fill=F(bg_c); ws.cell(ri_,ci__).border=brd("FF999999" if ri_==br_ else BORDER)
        for ri_ in [br_,br_+1,br_+2,br_+3]:
            ws.row_dimensions[ri_].height=22 if ri_==br_ else 16

for col_i in range(1,15):
    ws.column_dimensions[get_column_letter(col_i)].width=12

ws.freeze_panes="A11"
print("✓ КАЛЕНДАРЬ_ВЫПЛАТ")

# ════════════════════════════════════════════════════════════════════
# 9. ДАННЫЕ — скрытый лист формул
# ════════════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("ДАННЫЕ"); ws_d.sheet_state="hidden"

# Row 1: headers
for ci,h in enumerate(["Месяц","Доход","Расход","Прибыль","Кол-во операций","Закуп"],1):
    ws_d.cell(1,ci).value=h

_rA="БАЗА_ДДС!$A$4:$A$3003"; _rD="БАЗА_ДДС!$D$4:$D$3003"
_rE="БАЗА_ДДС!$E$4:$E$3003"; _rG="БАЗА_ДДС!$G$4:$G$3003"

for mi in range(1,13):
    r=mi+1; yr_ref="ДАШБОРД!$E$3"
    s=f"DATE({yr_ref},{mi},1)"; e=f"EOMONTH(DATE({yr_ref},{mi},1),0)"
    ws_d.cell(r,1).value=MONTHS_RU[mi-1]
    ws_d.cell(r,2).value=f'=IFERROR(SUMPRODUCT(({_rA}>={s})*({_rA}<={e})*({_rD}="Доход")*{_rG}),0)'
    ws_d.cell(r,3).value=f'=IFERROR(SUMPRODUCT(({_rA}>={s})*({_rA}<={e})*({_rD}="Расход")*{_rG}),0)'
    ws_d.cell(r,4).value=f'=IFERROR(B{r}-C{r},0)'
    ws_d.cell(r,5).value=f'=IFERROR(SUMPRODUCT(({_rA}>={s})*({_rA}<={e})*({_rD}="Доход")*(1)),0)'
    ws_d.cell(r,6).value=f'=IFERROR(SUMPRODUCT(({_rA}>={s})*({_rA}<={e})*({_rD}="Расход")*({_rE}="Закуп товара")*{_rG}),0)'

# Budget plan references (НАСТРОЙКИ rows 25-26)
for mi in range(1,13):
    r=mi+1
    # Budget reference: months 1-6 → row 25, months 7-12 → row 26
    row_plan = 25 if mi <= 6 else 26
    col_plan = (mi-1)%6*2+2  # col 2,4,6,8,10,12
    ws_d.cell(r,7).value=f'=IFERROR(НАСТРОЙКИ!{get_column_letter(col_plan)}{row_plan},0)'

# Day of week analysis (rows 20-27: Mon-Sun)
ws_d.cell(19,1).value="День недели"
dow_names=["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"]
for di,dn in enumerate(dow_names,1):
    r=19+di
    ws_d.cell(r,1).value=dn
    ws_d.cell(r,2).value=(f'=IFERROR(SUMPRODUCT(({_rA}>=ДАШБОРД!$A$5)*({_rA}<=ДАШБОРД!$B$5)'
                         f'*({_rD}="Доход")*(WEEKDAY({_rA},2)={di})*{_rG}),0)')
    ws_d.cell(r,3).value=(f'=IFERROR(SUMPRODUCT(({_rA}>=ДАШБОРД!$A$5)*({_rA}<=ДАШБОРД!$B$5)'
                         f'*(WEEKDAY({_rA},2)={di})*(1)),0)')

# Shift analysis (rows 30-33)
ws_d.cell(29,1).value="Смена"
_rB="БАЗА_ДДС!$B$4:$B$3003"
for si,sn in enumerate(["День","Вечер","Ночь"],1):
    r=29+si
    ws_d.cell(r,1).value=sn
    ws_d.cell(r,2).value=(f'=IFERROR(SUMPRODUCT(({_rA}>=ДАШБОРД!$A$5)*({_rA}<=ДАШБОРД!$B$5)'
                         f'*({_rD}="Доход")*({_rB}="{sn}")*{_rG}),0)')

print("✓ ДАННЫЕ (hidden)")

# ════════════════════════════════════════════════════════════════════
# 9. ДАШБОРД
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ДАШБОРД"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = TEAL_M

# 12 columns: 4 KPIs × 3 cols each
cw(ws,{"A":17,"B":13,"C":4,"D":17,"E":13,"F":4,"G":17,"H":13,"I":4,"J":17,"K":13,"L":4})
banner(ws,"📊  WAY MARKET — ДАШБОРД УПРАВЛЕНЧЕСКОГО УЧЁТА","A1:L1",TEAL,15)

# Row 2: shop name + record count
ws.merge_cells("A2:F2")
ws.cell(2,1).value='=НАСТРОЙКИ!E4&"  |  Данные с: "&TEXT(НАСТРОЙКИ!E5,"DD.MM.YYYY")'
ws.cell(2,1).font=fnt(10,True,TEAL_M); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=LA(False)
ws.merge_cells("G2:L2")
ws.cell(2,7).value='=IFERROR("Всего записей: "&TEXT(COUNTA(БАЗА_ДДС!$A$4:$A$3003),"#,##0")&"  |  Последнее: "&TEXT(MAX(БАЗА_ДДС!$A$4:$A$3003),"DD.MM.YYYY"),"—")'
ws.cell(2,7).font=fnt(9,it=True,col=GRAY); ws.cell(2,7).fill=F(LGRAY); ws.cell(2,7).alignment=RA()
ws.row_dimensions[2].height=24

# Row 3: period filter
ws.cell(3,1).value="ПЕРИОД:"
ws.cell(3,1).font=fnt(9,True,DGRAY); ws.cell(3,1).fill=F(LGRAY); ws.cell(3,1).border=brd(); ws.cell(3,1).alignment=RA()
ws.merge_cells("B3:C3")
c_month=inp(ws,3,2,MONTHS_RU[today.month-1],"@",False)
dv_m=DataValidation(type="list",formula1='"'+",".join(MONTHS_RU)+'"',allow_blank=False)
ws.add_data_validation(dv_m); dv_m.add(c_month)
ws.cell(3,4).value="Год:"
ws.cell(3,4).font=fnt(9,True,DGRAY); ws.cell(3,4).fill=F(LGRAY); ws.cell(3,4).border=brd(); ws.cell(3,4).alignment=RA()
ws.merge_cells("E3:F3")
c_year=inp(ws,3,5,today.year,'"####"',False)
ws.merge_cells("G3:L3")
ws.cell(3,7).value="[ОБНОВИТЬ] — нажмите F9 для пересчёта формул  |  Ctrl+Shift+D"
ws.cell(3,7).font=fnt(9,it=True,col=GRAY); ws.cell(3,7).fill=F(LGRAY); ws.cell(3,7).alignment=CA()
ws.row_dimensions[3].height=30

# Row 4: spacer
ws.row_dimensions[4].height=6

# Row 5: hidden period calculations (A5=start, B5=end, C5=prev_start, D5=prev_end, E5=year)
ws.row_dimensions[5].height=0
ws.cell(5,1).value='=DATE($E$3,MATCH($B$3,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0),1)'
ws.cell(5,1).number_format=DATE_F
ws.cell(5,2).value='=EOMONTH($A$5,0)'; ws.cell(5,2).number_format=DATE_F
ws.cell(5,3).value='=DATE($E$3,MATCH($B$3,{"Январь";"Февраль";"Март";"Апрель";"Май";"Июнь";"Июль";"Август";"Сентябрь";"Октябрь";"Ноябрь";"Декабрь"},0)-1,1)'; ws.cell(5,3).number_format=DATE_F
ws.cell(5,4).value='=EOMONTH($C$5,0)'; ws.cell(5,4).number_format=DATE_F
ws.cell(5,5).value='=$E$3'

# ── Formula helpers ───────────────────────────────────────────────
def SP(typ=None,cat=None,pay=None):
    conds=["(БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)","(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)"]
    if typ: conds.append(f'(БАЗА_ДДС!$D$4:$D$3003="{typ}")')
    if cat: conds.append(f'(БАЗА_ДДС!$E$4:$E$3003="{cat}")')
    if pay: conds.append(f'(БАЗА_ДДС!$F$4:$F$3003="{pay}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(conds)}*БАЗА_ДДС!$G$4:$G$3003),0)'

def SP_cnt(typ=None,cat=None):
    conds=["(БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)","(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)"]
    if typ: conds.append(f'(БАЗА_ДДС!$D$4:$D$3003="{typ}")')
    if cat: conds.append(f'(БАЗА_ДДС!$E$4:$E$3003="{cat}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(conds)}*(1)),0)'

def SP_prev(typ=None,cat=None):
    conds=["(БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$C$5)","(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$D$5)"]
    if typ: conds.append(f'(БАЗА_ДДС!$D$4:$D$3003="{typ}")')
    if cat: conds.append(f'(БАЗА_ДДС!$E$4:$E$3003="{cat}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(conds)}*БАЗА_ДДС!$G$4:$G$3003),0)'

# ── 4-tile block builder ─────────────────────────────────────────
def tile_block(ws, start_row, title, tiles, bg, lbl_bg=None, val_bg=None):
    if lbl_bg is None: lbl_bg = bg
    if val_bg is None: val_bg = bg
    sec_hdr(ws, start_row, f"  {title}", 12, bg, 22)
    ws.row_dimensions[start_row+1].height=22
    ws.row_dimensions[start_row+2].height=38
    for i,(lbl,val_f,fmt,ind_f) in enumerate(tiles):
        c1=i*3+1; c3=c1+2
        ws.merge_cells(start_row=start_row+1,start_column=c1,end_row=start_row+1,end_column=c3)
        cl=ws.cell(start_row+1,c1); cl.value=lbl
        cl.font=fnt(9,col=WHITE); cl.fill=F(lbl_bg); cl.alignment=CA(True); cl.border=brd()
        ws.merge_cells(start_row=start_row+2,start_column=c1,end_row=start_row+2,end_column=c3-1)
        cv=ws.cell(start_row+2,c1); cv.value=val_f
        cv.font=fnt(15,True,WHITE); cv.fill=F(val_bg); cv.alignment=CA(False); cv.border=brd()
        cv.number_format=fmt if fmt else MONEY
        ci=ws.cell(start_row+2,c3)
        ci.value=ind_f if ind_f else "—"
        ci.font=fnt(11,True,WHITE); ci.fill=F(val_bg); ci.alignment=CA(); ci.border=brd()

# ── БЛОК 1: ВЫРУЧКА (rows 6-8) ───────────────────────────────────
_vyr=SP("Доход")
tile_block(ws,6,"  ВЫРУЧКА",[
    ("Общая выручка",    _vyr, MONEY,
     f'=IFERROR(IF({_vyr[1:]}>{SP_prev("Доход")[1:]},"▲","▼"),"▶")'),
    ("Среднее в день",
     f'=IFERROR({_vyr[1:]}/MAX(1,DAY(ДАШБОРД!$B$5)),0)', MONEY, None),
    ("Среднее за смену",
     f'=IFERROR({_vyr[1:]}/MAX(1,{SP_cnt("Доход")[1:]}),0)', MONEY, None),
    ("Выполнение плана",
     '=IFERROR('+_vyr[1:]+'/IFERROR(INDEX(НАСТРОЙКИ!$B$25:$L$26,IF(MATCH(ДАШБОРД!$B$3,{"Январь","Февраль","Март","Апрель","Май","Июнь"},0)>0,1,2),MOD(MATCH(ДАШБОРД!$B$3,{"Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"},0)-1,6)*2+1,1),1),0)',
     "0%", None),
],GREEN)

# ── БЛОК 2: КОНТРОЛЬ КАССЫ (rows 9-11) ──────────────────────────
_rash_sum='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Расхождение")*БАЗА_ДДС!$G$4:$G$3003),0)'
_rash_cnt='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Расхождение")*(1)),0)'
_kassa_f='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Касса")*БАЗА_ДДС!$G$4:$G$3003),0)'
_ost_f=f'=IFERROR({_vyr[1:]}-{SP("Расход")[1:]},0)'
tile_block(ws,9,"  КОНТРОЛЬ КАССЫ",[
    ("Выплаты с кассы",    _kassa_f,  MONEY, None),
    ("Расхождений сумма",  _rash_sum, MONEY,
     f'=IFERROR(IF({_rash_sum[1:]}>НАСТРОЙКИ!$E$62,"⚠","✓"),"—")'),
    ("Кол-во расхождений", _rash_cnt, "#,##0", None),
    ("Остаток кассы",      _ost_f,    MONEY,
     f'=IFERROR(IF({_ost_f[1:]}>=0,"▲","▼"),"—")'),
],TEAL_M)

# ── БЛОК 3: ДОЛГИ (rows 12-14) ───────────────────────────────────
_dolg_tek='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$D$4:$D$3003="Долг")*БАЗА_ДДС!$G$4:$G$3003)-SUMPRODUCT((БАЗА_ДДС!$D$4:$D$3003="Оплата долга")*БАЗА_ДДС!$G$4:$G$3003)+НАСТРОЙКИ!$E$8,0)'
_dolg_per=SP("Долг"); _opl_per=SP("Оплата долга")
tile_block(ws,12,"  ДОЛГИ И ОБЯЗАТЕЛЬСТВА",[
    ("Текущий долг", _dolg_tek, MONEY,
     f'=IFERROR(IF({_dolg_tek[1:]}>НАСТРОЙКИ!$E$63,"⚠","✓"),"—")'),
    ("Взято в долг",     _dolg_per, MONEY, None),
    ("Выплачено долгов", _opl_per,  MONEY, None),
    ("К оплате (=долг)", _dolg_tek, MONEY, None),
],RED)

# ── БЛОК 4: ПРИБЫЛЬ (rows 15-17) ────────────────────────────────
_zakup=SP("Расход","Закуп товара"); _rash_all=SP("Расход")
_pribyl=f'=IFERROR({_vyr[1:]}-{_rash_all[1:]},0)'
_rent=f'=IFERROR({_pribyl[1:]}/MAX(1,{_vyr[1:]}),0)'
tile_block(ws,15,"  ПРИБЫЛЬ",[
    ("Закуп товара",   _zakup,    MONEY, None),
    ("Все расходы",    _rash_all, MONEY,
     f'=IFERROR(IF({_rash_all[1:]}<{_vyr[1:]},"▲","▼"),"—")'),
    ("Чистая прибыль", _pribyl,   MONEY,
     f'=IFERROR(IF({_pribyl[1:]}>=0,"▲","▼"),"—")'),
    ("Рентабельность", _rent, "0.0%", None),
],GREEN)

# ── БЛОК 5: ЭФФЕКТИВНОСТЬ (rows 18-20) ──────────────────────────
_eff_zakup=f'=IFERROR({_vyr[1:]}/MAX(1,{_zakup[1:]}),0)'
_nagruzka=f'=IFERROR({_dolg_tek[1:]}/MAX(1,{_vyr[1:]}),0)'
_avg_rash=f'=IFERROR({_rash_all[1:]}/MAX(1,DAY(ДАШБОРД!$B$5)),0)'
tile_block(ws,18,"  ЭФФЕКТИВНОСТЬ",[
    ("Маржа %",            _rent,       "0.0%",  None),
    ("Эффект. закупа (x)", _eff_zakup,  "0.0x",  None),
    ("Нагрузка долга %",   _nagruzka,   "0.0%",
     f'=IFERROR(IF({_nagruzka[1:]}>1,"⚠","✓"),"—")'),
    ("Ср. расход / день",  _avg_rash,   MONEY,   None),
],TEAL_M)

# ── БЛОК 6: ОПЕРАЦИИ И ВЫПЛАТЫ (rows 21-23) ─────────────────────
_prosr_sum='=IFERROR(SUMPRODUCT((ЗАПИСЬ_ВЫПЛАТ!$B$7:$B$506<TODAY())*(ЗАПИСЬ_ВЫПЛАТ!$E$7:$E$506="Запланировано")*ЗАПИСЬ_ВЫПЛАТ!$D$7:$D$506),0)'
_prosr_cnt='=IFERROR(SUMPRODUCT((ЗАПИСЬ_ВЫПЛАТ!$B$7:$B$506<TODAY())*(ЗАПИСЬ_ВЫПЛАТ!$E$7:$E$506="Запланировано")*(ЗАПИСЬ_ВЫПЛАТ!$D$7:$D$506<>"")),0)'
_vipl_sum='=IFERROR(SUMPRODUCT((ЗАПИСЬ_ВЫПЛАТ!$E$7:$E$506="Выплачено")*ЗАПИСЬ_ВЫПЛАТ!$D$7:$D$506),0)'
_plan_sum='=IFERROR(SUMPRODUCT(ЗАПИСЬ_ВЫПЛАТ!$D$7:$D$506*(ЗАПИСЬ_ВЫПЛАТ!$D$7:$D$506<>"")),0)'
_vipl_pct=f'=IFERROR({_vipl_sum[1:]}/MAX(1,{_plan_sum[1:]}),0)'
_zakup_dolg_pct=f'=IFERROR({_dolg_per[1:]}/MAX(1,{_rash_all[1:]}),0)'
tile_block(ws,21,"  ОПЕРАЦИИ И ВЫПЛАТЫ",[
    ("Просроч. выплаты ₽", _prosr_sum, MONEY, f'=IFERROR(IF({_prosr_sum[1:]}>0,"⚠","✓"),"—")'),
    ("Просроч. кол-во",    _prosr_cnt, "#,##0", None),
    ("Выплачено %",        _vipl_pct,  "0%",   None),
    ("Закуп в долг %",     _zakup_dolg_pct, "0%", None),
],AMBER)

# ── БЛОК 7: СТАТИСТИКА ПЕРИОДА (rows 24-26) ─────────────────────
_total_ops=SP_cnt()
_max_day='=IFERROR(MAXIFS(БАЗА_ДДС!$G$4:$G$3003,БАЗА_ДДС!$D$4:$D$3003,"Доход",БАЗА_ДДС!$A$4:$A$3003,">="&ДАШБОРД!$A$5,БАЗА_ДДС!$A$4:$A$3003,"<="&ДАШБОРД!$B$5),0)'
_min_day='=IFERROR(MINIFS(БАЗА_ДДС!$G$4:$G$3003,БАЗА_ДДС!$D$4:$D$3003,"Доход",БАЗА_ДДС!$A$4:$A$3003,">="&ДАШБОРД!$A$5,БАЗА_ДДС!$A$4:$A$3003,"<="&ДАШБОРД!$B$5),0)'
_days_with='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Доход")/COUNTIFS(БАЗА_ДДС!$A$4:$A$3003,БАЗА_ДДС!$A$4:$A$3003,БАЗА_ДДС!$D$4:$D$3003,"Доход")),0)'
tile_block(ws,24,"  СТАТИСТИКА ПЕРИОДА",[
    ("Дней с данными",     _days_with, "#,##0", None),
    ("Всего операций",     _total_ops, "#,##0", None),
    ("Макс. выручка/день", _max_day,   MONEY,   None),
    ("Мин. выручка/день",  _min_day,   MONEY,   None),
],NAVY)

# ── БЛОК 8: ИМАН И ПРОЧЕЕ (rows 27-29) ──────────────────────────
_iman_per=SP("Иман")
_iman_all='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$D$4:$D$3003="Иман")*БАЗА_ДДС!$G$4:$G$3003),0)'
_spis='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*((БАЗА_ДДС!$D$4:$D$3003="Списание")+(БАЗА_ДДС!$D$4:$D$3003="Возврат"))*БАЗА_ДДС!$G$4:$G$3003),0)'
_vozvr='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Возврат")*БАЗА_ДДС!$G$4:$G$3003),0)'
tile_block(ws,27,"  ИМАН И ПРОЧЕЕ",[
    ("Иман хозяина (период)", _iman_per,  MONEY, f'=IFERROR(IF({_iman_per[1:]}>0,"▶","—"),"—")'),
    ("Иман (всё время)",       _iman_all,  MONEY, None),
    ("Списания+Возвраты",      _spis,      MONEY, None),
    ("Возвраты (период)",      _vozvr,     MONEY, None),
],PURPLE)

ws.row_dimensions[30].height=8

# ── ДЕТАЛИЗАЦИЯ РАСХОДОВ (rows 31-42) ────────────────────────────
sec_hdr(ws,31,"  ДЕТАЛИЗАЦИЯ РАСХОДОВ ЗА ПЕРИОД",12,NAVY,22)
ws.merge_cells("A32:D32"); ws.cell(32,1).value="Категория"
ws.cell(32,1).font=fnt(9,True,WHITE); ws.cell(32,1).fill=F(NAVY); ws.cell(32,1).alignment=LA(False); ws.cell(32,1).border=brd()
ws.cell(32,5).value="Сумма (₽)"; ws.cell(32,5).font=fnt(9,True,WHITE); ws.cell(32,5).fill=F(NAVY); ws.cell(32,5).alignment=CA(); ws.cell(32,5).border=brd()
ws.cell(32,6).value="Доля %"; ws.cell(32,6).font=fnt(9,True,WHITE); ws.cell(32,6).fill=F(NAVY); ws.cell(32,6).alignment=CA(); ws.cell(32,6).border=brd()
ws.merge_cells("G32:L32"); ws.cell(32,7).value="Гистограмма (длина = доля от общих расходов)"
ws.cell(32,7).font=fnt(9,True,WHITE); ws.cell(32,7).fill=F(NAVY); ws.cell(32,7).alignment=LA(False); ws.cell(32,7).border=brd()
ws.row_dimensions[32].height=22
_tot_rash_d=f'MAX(1,{_rash_all[1:]})'
for ci,cat in enumerate(["Закуп товара офисом","Закуп товара","Зарплата","Аренда","Коммунальные","Налоги","ГСМ","Расходный материал","Маркетинг","Прочие расходы"]):
    r=33+ci
    cat_f=f'=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Расход")*(БАЗА_ДДС!$E$4:$E$3003="{cat}")*БАЗА_ДДС!$G$4:$G$3003),0)'
    pct_f=f'=IFERROR({cat_f[1:]}/{_tot_rash_d},0)'
    bar_f=f'=IFERROR(REPT("█",ROUND({pct_f[1:]}*40,0)),"")' 
    bg_r=WHITE if ci%2==0 else LGRAY
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    ws.cell(r,1).value=cat; ws.cell(r,1).font=fnt(10); ws.cell(r,1).fill=F(bg_r); ws.cell(r,1).border=brd(); ws.cell(r,1).alignment=LA(False)
    ws.cell(r,5).value=cat_f; ws.cell(r,5).font=fnt(10,True,AMBER); ws.cell(r,5).fill=F(AMBER_L); ws.cell(r,5).border=brd(); ws.cell(r,5).alignment=RA(); ws.cell(r,5).number_format=MONEY
    ws.cell(r,6).value=pct_f; ws.cell(r,6).font=fnt(9); ws.cell(r,6).fill=F(bg_r); ws.cell(r,6).border=brd(); ws.cell(r,6).alignment=CA(); ws.cell(r,6).number_format="0.0%"
    ws.merge_cells(start_row=r,start_column=7,end_row=r,end_column=12)
    ws.cell(r,7).value=bar_f; ws.cell(r,7).font=fnt(9,col=AMBER); ws.cell(r,7).fill=F(bg_r); ws.cell(r,7).border=brd(); ws.cell(r,7).alignment=LA(False)
    ws.row_dimensions[r].height=20

ws.row_dimensions[43].height=8

# ── ВЫРУЧКА ПО ДНЯМ НЕДЕЛИ (rows 44-52) ─────────────────────────
sec_hdr(ws,44,"  📅  ВЫРУЧКА ПО ДНЯМ НЕДЕЛИ (текущий период)",12,TEAL_M,22)
for ci_,h_ in enumerate(["День недели","Выручка (₽)","Кол-во дней","Ср. в день","% от периода","Лучший?"],1):
    ws.cell(45,ci_).value=h_; ws.cell(45,ci_).font=fnt(9,True,WHITE); ws.cell(45,ci_).fill=F(TEAL_M); ws.cell(45,ci_).border=brd(); ws.cell(45,ci_).alignment=CA()
ws.row_dimensions[45].height=22
_tot_vyr_dow='IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Доход")*БАЗА_ДДС!$G$4:$G$3003),1)'
for di,dn in enumerate(["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"],1):
    r=45+di
    vyr_f=f'=ДАННЫЕ!B{19+di}'
    days_f=f'=ДАННЫЕ!C{19+di}'
    bg_r=RED_L if di>=6 else (WHITE if di%2==0 else LGRAY)
    for ci_ in range(1,7):
        ws.cell(r,ci_).fill=F(bg_r); ws.cell(r,ci_).border=brd()
    ws.cell(r,1).value=dn; ws.cell(r,1).font=fnt(10); ws.cell(r,1).alignment=LA(False)
    ws.cell(r,2).value=vyr_f; ws.cell(r,2).font=fnt(10,True,TEAL_M); ws.cell(r,2).fill=F(TEAL_L); ws.cell(r,2).border=brd(); ws.cell(r,2).alignment=RA(); ws.cell(r,2).number_format=MONEY
    ws.cell(r,3).value=days_f; ws.cell(r,3).font=fnt(10); ws.cell(r,3).alignment=CA()
    ws.cell(r,4).value=f'=IFERROR({vyr_f[1:]}/MAX(1,{days_f[1:]}),0)'; ws.cell(r,4).font=fnt(10); ws.cell(r,4).alignment=RA(); ws.cell(r,4).number_format=MONEY
    ws.cell(r,5).value=f'=IFERROR({vyr_f[1:]}/{_tot_vyr_dow},0)'; ws.cell(r,5).font=fnt(10); ws.cell(r,5).alignment=CA(); ws.cell(r,5).number_format="0.0%"
    ws.cell(r,6).value=f'=IF({vyr_f[1:]}=MAX(ДАННЫЕ!$B$20:$B$26),"⭐ ЛУЧ","")'; ws.cell(r,6).font=fnt(10,True,GOLD); ws.cell(r,6).alignment=CA()
    ws.row_dimensions[r].height=22

ws.row_dimensions[53].height=8

# ── ГОДОВОЙ ТРЕНД — chart (rows 54+) ─────────────────────────────
sec_hdr(ws,54,"  📉  ГОДОВОЙ ТРЕНД ВЫРУЧКИ И РАСХОДОВ",12,NAVY,22)
# Hidden data cols O-R (15-18)
ws.cell(55,15).value="Месяц"; ws.cell(55,16).value="Доход"; ws.cell(55,17).value="Расход"; ws.cell(55,18).value="План"
for mi in range(1,13):
    r=55+mi
    ws.cell(r,15).value=MONTHS_RU[mi-1]
    ws.cell(r,16).value=f'=ДАННЫЕ!B{mi+1}'
    ws.cell(r,17).value=f'=ДАННЫЕ!C{mi+1}'
    ws.cell(r,18).value=f'=ДАННЫЕ!G{mi+1}'
    ws.row_dimensions[r].height=0
# Hidden DOW data cols S-T (19-20)
ws.cell(55,19).value="День"; ws.cell(55,20).value="Выручка"
for di in range(7):
    ws.cell(56+di,19).value=["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"][di]
    ws.cell(56+di,20).value=f'=ДАННЫЕ!B{20+di}'
    ws.row_dimensions[56+di].height=0

chart=LineChart(); chart.title="Тренд выручки и расходов"
chart.style=10; chart.height=10; chart.width=22
chart.y_axis.title="Сумма (₽)"; chart.x_axis.title="Месяц"
data_ref=Reference(ws,min_col=16,max_col=18,min_row=55,max_row=67)
cats_ref=Reference(ws,min_col=15,min_row=56,max_row=67)
chart.add_data(data_ref,titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.line.solidFill=TEAL_M
chart.series[0].graphicalProperties.line.width=25000
chart.series[1].graphicalProperties.line.solidFill=AMBER
chart.series[1].graphicalProperties.line.width=18000
if len(chart.series)>2:
    chart.series[2].graphicalProperties.line.solidFill=GOLD_M
    chart.series[2].graphicalProperties.line.dashDot="dash"
    chart.series[2].graphicalProperties.line.width=15000
ws.add_chart(chart,"A55")

chart2=BarChart(); chart2.title="Выручка по дням недели"
chart2.style=10; chart2.height=8; chart2.width=14
dow_data=Reference(ws,min_col=20,max_col=20,min_row=55,max_row=62)
dow_cats=Reference(ws,min_col=19,min_row=56,max_row=62)
chart2.add_data(dow_data,titles_from_data=True)
chart2.set_categories(dow_cats)
chart2.series[0].graphicalProperties.solidFill=TEAL_M
ws.add_chart(chart2,"G55")

ws.freeze_panes="A4"
print("✓ ДАШБОРД")


# ════════════════════════════════════════════════════════════════════
# 10. АНАЛИТИКА — углублённый анализ
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("АНАЛИТИКА"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = TEAL_M
cw(ws,{"A":20,"B":14,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14})

banner(ws,"🔍  АНАЛИТИКА — ГЛУБОКИЙ АНАЛИЗ","A1:I1",TEAL_M,14)
ws.merge_cells("A2:I2")
ws.cell(2,1).value="Данные подтягиваются из БАЗА_ДДС за выбранный период (из ДАШБОРД)"
ws.cell(2,1).font=fnt(9,it=True,col=GRAY); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=CA(); ws.row_dimensions[2].height=20

# БЛОК 1: По дням недели
sec_hdr(ws,3,"  📅  ВЫРУЧКА ПО ДНЯМ НЕДЕЛИ (текущий период)",9,TEAL_M,24)
hdr_row(ws,4,[("День недели",20),("Выручка (₽)",14),("Кол-во дней",14),
              ("Ср. в день (₽)",14),("% от периода",14),("Лучший день?",14)],TEAL_M,22)
for di,dn in enumerate(dow_names,1):
    r=4+di
    vyr_f=f'=ДАННЫЕ!B{19+di}'
    days_f=f'=ДАННЫЕ!C{19+di}'
    avg_f=f'=IFERROR({vyr_f[1:]}/MAX(1,{days_f[1:]}),0)'
    tot_f='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Доход")*БАЗА_ДДС!$G$4:$G$3003),0)'
    ws.cell(r,1).value=dn; ws.cell(r,1).font=fnt(10); ws.cell(r,1).fill=F(LGRAY); ws.cell(r,1).border=brd(); ws.cell(r,1).alignment=LA(False)
    ws.cell(r,2).value=vyr_f; ws.cell(r,2).font=fnt(10,True,TEAL_M); ws.cell(r,2).fill=F(TEAL_L); ws.cell(r,2).border=brd(); ws.cell(r,2).alignment=RA(); ws.cell(r,2).number_format=MONEY
    ws.cell(r,3).value=days_f; ws.cell(r,3).font=fnt(10); ws.cell(r,3).fill=F(LGRAY); ws.cell(r,3).border=brd(); ws.cell(r,3).alignment=CA()
    ws.cell(r,4).value=f'={avg_f}'; ws.cell(r,4).font=fnt(10); ws.cell(r,4).fill=F(LGRAY); ws.cell(r,4).border=brd(); ws.cell(r,4).alignment=RA(); ws.cell(r,4).number_format=MONEY
    ws.cell(r,5).value=f'=IFERROR({vyr_f[1:]}/{tot_f[1:]},0)'; ws.cell(r,5).font=fnt(10); ws.cell(r,5).fill=F(LGRAY); ws.cell(r,5).border=brd(); ws.cell(r,5).alignment=CA(); ws.cell(r,5).number_format="0.0%"
    ws.cell(r,6).value=f'=IF({vyr_f[1:]}=MAX(ДАННЫЕ!$B$20:$B$26),"⭐ ЛУЧ","")'; ws.cell(r,6).font=fnt(10,True,GOLD); ws.cell(r,6).fill=F(LGRAY); ws.cell(r,6).border=brd(); ws.cell(r,6).alignment=CA()
    ws.row_dimensions[r].height=22
    # Color scale on revenue
    ws.conditional_formatting.add(f"B{r}",ColorScaleRule(start_type="min",start_color=LGRAY,end_type="max",end_color=GREEN_L))

ws.row_dimensions[12].height=10

# БЛОК 2: По сменам
sec_hdr(ws,13,"  🔄  ВЫРУЧКА ПО СМЕНАМ (текущий период)",9,TEAL_M,24)
hdr_row(ws,14,[("Смена",20),("Выручка (₽)",14),("Доля (%)",14),("Оценка",14)],TEAL_M,22)
for si,sn in enumerate(["День","Вечер","Ночь"],1):
    r=14+si
    vf=f'=ДАННЫЕ!B{29+si}'
    ws.cell(r,1).value=sn; ws.cell(r,1).font=fnt(10); ws.cell(r,1).fill=F(LGRAY); ws.cell(r,1).border=brd(); ws.cell(r,1).alignment=LA(False)
    ws.cell(r,2).value=vf; ws.cell(r,2).font=fnt(11,True,TEAL_M); ws.cell(r,2).fill=F(TEAL_L); ws.cell(r,2).border=brd(); ws.cell(r,2).alignment=RA(); ws.cell(r,2).number_format=MONEY
    tot_sh='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Доход")*БАЗА_ДДС!$G$4:$G$3003),0)'
    ws.cell(r,3).value=f'=IFERROR({vf[1:]}/MAX(1,{tot_sh[1:]}),0)'; ws.cell(r,3).font=fnt(10); ws.cell(r,3).fill=F(LGRAY); ws.cell(r,3).border=brd(); ws.cell(r,3).alignment=CA(); ws.cell(r,3).number_format="0.0%"
    ws.cell(r,4).value=f'=IF({vf[1:]}=MAX(ДАННЫЕ!$B$30:$B$32),"★ Лидер","")'; ws.cell(r,4).font=fnt(10,True,GOLD); ws.cell(r,4).fill=F(LGRAY); ws.cell(r,4).border=brd(); ws.cell(r,4).alignment=CA()
    ws.row_dimensions[r].height=24

ws.row_dimensions[18].height=10

# БЛОК 3: Топ категории расходов
sec_hdr(ws,19,"  💸  РАСХОДЫ ПО КАТЕГОРИЯМ (текущий период)",9,AMBER,24)
categories=["Закуп товара","Аренда","ЗП","Коммунальные","Налоги","Охрана","Прочее"]
hdr_row(ws,20,[("Категория",20),("Сумма (₽)",14),("% от расходов",14),("% от выручки",14)],AMBER,22)
tot_rash_f='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Расход")*БАЗА_ДДС!$G$4:$G$3003),0)'
tot_vyr_f='=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Доход")*БАЗА_ДДС!$G$4:$G$3003),0)'
for ci,cat in enumerate(categories,1):
    r=20+ci
    cat_f=f'=IFERROR(SUMPRODUCT((БАЗА_ДДС!$A$4:$A$3003>=ДАШБОРД!$A$5)*(БАЗА_ДДС!$A$4:$A$3003<=ДАШБОРД!$B$5)*(БАЗА_ДДС!$D$4:$D$3003="Расход")*(БАЗА_ДДС!$E$4:$E$3003="{cat}")*БАЗА_ДДС!$G$4:$G$3003),0)'
    ws.cell(r,1).value=cat; ws.cell(r,1).font=fnt(10); ws.cell(r,1).fill=F(LGRAY); ws.cell(r,1).border=brd(); ws.cell(r,1).alignment=LA(False)
    ws.cell(r,2).value=cat_f; ws.cell(r,2).font=fnt(10,True,AMBER); ws.cell(r,2).fill=F(AMBER_L); ws.cell(r,2).border=brd(); ws.cell(r,2).alignment=RA(); ws.cell(r,2).number_format=MONEY
    ws.cell(r,3).value=f'=IFERROR({cat_f[1:]}/MAX(1,{tot_rash_f[1:]}),0)'; ws.cell(r,3).font=fnt(10); ws.cell(r,3).fill=F(LGRAY); ws.cell(r,3).border=brd(); ws.cell(r,3).alignment=CA(); ws.cell(r,3).number_format="0.0%"
    ws.cell(r,4).value=f'=IFERROR({cat_f[1:]}/MAX(1,{tot_vyr_f[1:]}),0)'; ws.cell(r,4).font=fnt(10); ws.cell(r,4).fill=F(LGRAY); ws.cell(r,4).border=brd(); ws.cell(r,4).alignment=CA(); ws.cell(r,4).number_format="0.0%"
    ws.row_dimensions[r].height=22
    ws.conditional_formatting.add(f"C{r}",ColorScaleRule(start_type="min",start_color=GREEN_L,end_type="max",end_color=RED_L))

ws.freeze_panes="A3"
print("✓ АНАЛИТИКА")

# ════════════════════════════════════════════════════════════════════
# 11. ОТЧЁТ_РУКОВОДИТЕЛЮ
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ОТЧЁТ"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = NAVY
cw(ws,{"A":28,"B":18,"C":14,"D":18,"E":6,"F":16})

banner(ws,"📋  ОТЧЁТ РУКОВОДИТЕЛЮ — ФИНАНСОВАЯ МОДЕЛЬ 25/75","A1:F1",NAVY,14)
ws.merge_cells("A2:F2")
ws.cell(2,1).value='=НАСТРОЙКИ!E4&"  |  Период: "&TEXT(ДАШБОРД!$A$5,"DD.MM.YY")&" — "&TEXT(ДАШБОРД!$B$5,"DD.MM.YY")'
ws.cell(2,1).font=fnt(10,True,TEAL_M); ws.cell(2,1).fill=F(LGRAY); ws.cell(2,1).alignment=LA(False); ws.row_dimensions[2].height=26

def rsp(typ=None,cat=None,s=None,e=None):
    s_=s or "ДАШБОРД!$A$5"; e_=e or "ДАШБОРД!$B$5"
    c=[f"(БАЗА_ДДС!$A$4:$A$3003>={s_})",f"(БАЗА_ДДС!$A$4:$A$3003<={e_})"]
    if typ: c.append(f'(БАЗА_ДДС!$D$4:$D$3003="{typ}")')
    if cat: c.append(f'(БАЗА_ДДС!$E$4:$E$3003="{cat}")')
    return f'=IFERROR(SUMPRODUCT({"*".join(c)}*БАЗА_ДДС!$G$4:$G$3003),0)'

def rsa(typ): return f'=IFERROR(SUMPRODUCT((БАЗА_ДДС!$D$4:$D$3003="{typ}")*БАЗА_ДДС!$G$4:$G$3003),0)'

hdr_row(ws,3,[("Показатель",28),("Период (₽)",18),("% от выручки",14),
              ("Пред. период (₽)",18),("▲▼",6),("Лимит / Норма",16)],NAVY,24)

sec_hdr(ws,4,"  БЛОК 1: СВОДНЫЕ ПОКАЗАТЕЛИ",6,TEAL,24)
rep_rows=[
    (5,"Выручка (Доход)",rsp("Доход"),rsp("Доход",s="ДАШБОРД!$C$5",e="ДАШБОРД!$D$5"),"=1","100%",True,GREEN,GREEN_L),
    (6,"Расходы (Расход)",rsp("Расход"),rsp("Расход",s="ДАШБОРД!$C$5",e="ДАШБОРД!$D$5"),"=IFERROR(B6/MAX(1,B5),0)","% выручки",False,AMBER,AMBER_L),
    (7,"Прибыль (Доход − Расход)","=IFERROR(B5-B6,0)","=IFERROR(D5-D6,0)","=IFERROR(B7/MAX(1,B5),0)","рентабельность",True,GREEN,GREEN_L),
    (8,"Остаток кассы (всё время)",f'=IFERROR({rsa("Доход")[1:]}-{rsa("Расход")[1:]},0)',None,"всё время","—",None,TEAL_M,TEAL_L),
]

for row_data in rep_rows:
    ri,lbl,cur_,prv_,pct_,note_,up_,vc_,vb_ = row_data
    ws.cell(ri,1).value=lbl; ws.cell(ri,1).font=fnt(10); ws.cell(ri,1).fill=F(LGRAY); ws.cell(ri,1).border=brd(); ws.cell(ri,1).alignment=LA(False)
    ws.cell(ri,2).value=cur_; ws.cell(ri,2).font=fnt(12,True,vc_); ws.cell(ri,2).fill=F(vb_); ws.cell(ri,2).border=brd(); ws.cell(ri,2).alignment=RA(); ws.cell(ri,2).number_format=MONEY
    if isinstance(pct_,str) and pct_.startswith("="):
        ws.cell(ri,3).value=pct_; ws.cell(ri,3).number_format="0.0%"
    else:
        ws.cell(ri,3).value=pct_
    ws.cell(ri,3).font=fnt(9,col=GRAY); ws.cell(ri,3).fill=F(LGRAY); ws.cell(ri,3).border=brd(); ws.cell(ri,3).alignment=CA()
    if prv_:
        ws.cell(ri,4).value=prv_; ws.cell(ri,4).font=fnt(10,col=GRAY); ws.cell(ri,4).fill=F(LGRAY); ws.cell(ri,4).border=brd(); ws.cell(ri,4).alignment=RA(); ws.cell(ri,4).number_format=MONEY
        if up_ is not None:
            tr=f'=IF(B{ri}>D{ri},"▲","▼")' if up_ else f'=IF(B{ri}<D{ri},"▲","▼")'
            ws.cell(ri,5).value=tr; ws.cell(ri,5).font=fnt(12,True,GREEN); ws.cell(ri,5).fill=F(LGRAY); ws.cell(ri,5).border=brd(); ws.cell(ri,5).alignment=CA()
    for ci_ in [4,5]:
        if not ws.cell(ri,ci_).value: ws.cell(ri,ci_).fill=F(LGRAY); ws.cell(ri,ci_).border=brd()
    ws.cell(ri,6).value=note_; ws.cell(ri,6).font=fnt(9,it=True,col=GRAY); ws.cell(ri,6).fill=F(LGRAY); ws.cell(ri,6).border=brd(); ws.cell(ri,6).alignment=LA(False)
    ws.row_dimensions[ri].height=26

# БЛОК 2: Модель 25/75
sec_hdr(ws,10,"  БЛОК 2: ФИНАНСОВАЯ МОДЕЛЬ 25/75",6,GOLD,24)
b2=[
    (11,"Норма прибыли (25%)","=IFERROR(B5*НАСТРОЙКИ!$E$6,0)","=НАСТРОЙКИ!$E$6","25% норма"),
    (12,"Лимит на закуп (75%)","=IFERROR(B5*НАСТРОЙКИ!$E$7,0)","=НАСТРОЙКИ!$E$7","75% лимит"),
    (13,"Факт закупа за период",rsp("Расход","Закуп товара"),"=IFERROR(B13/MAX(1,B5),0)","кат. Закуп товара"),
    (14,"Остаток лимита","=IFERROR(B12-B13,0)","=IFERROR(B14/MAX(1,B5),0)","Лимит − Факт"),
]
for (ri,lbl,cur,pct,note) in b2:
    ws.cell(ri,1).value=lbl; ws.cell(ri,1).font=fnt(10); ws.cell(ri,1).fill=F(LGRAY); ws.cell(ri,1).border=brd(); ws.cell(ri,1).alignment=LA(False)
    ws.cell(ri,2).value=cur; ws.cell(ri,2).font=fnt(11,True,GOLD); ws.cell(ri,2).fill=F(GOLD_L); ws.cell(ri,2).border=brd(); ws.cell(ri,2).alignment=RA(); ws.cell(ri,2).number_format=MONEY
    ws.cell(ri,3).value=pct
    if isinstance(pct,str) and pct.startswith("="):
        ws.cell(ri,3).number_format="0%"
    ws.cell(ri,3).font=fnt(9,col=GOLD); ws.cell(ri,3).fill=F(LGRAY); ws.cell(ri,3).border=brd(); ws.cell(ri,3).alignment=CA()
    for ci in [4,5]: ws.cell(ri,ci).fill=F(LGRAY); ws.cell(ri,ci).border=brd()
    ws.cell(ri,6).value=note; ws.cell(ri,6).font=fnt(9,it=True,col=GRAY); ws.cell(ri,6).fill=F(LGRAY); ws.cell(ri,6).border=brd(); ws.cell(ri,6).alignment=LA(False)
    ws.row_dimensions[ri].height=26
ws.conditional_formatting.add("B14",FormulaRule(formula=["B14>=0"],fill=F(GREEN_L),font=fnt(11,True,GREEN)))
ws.conditional_formatting.add("B14",FormulaRule(formula=["B14<0"],fill=F(RED_L),font=fnt(11,True,RED)))

# БЛОК 3: Долги
sec_hdr(ws,16,"  БЛОК 3: ДОЛГИ И ОБЯЗАТЕЛЬСТВА",6,RED,24)
b3=[
    (17,"Нач. долг поставщикам","=НАСТРОЙКИ!$E$8","из НАСТРОЙКИ"),
    (18,"Взято в долг (всё время)",rsa("Долг"),"тип: Долг"),
    (19,"Выплачено по долгам",rsa("Оплата долга"),"тип: Оплата долга"),
    (20,"Текущий долг","=IFERROR(B17+B18-B19,0)","Нач + Взято − Выплачено"),
    (21,"Иман хозяйки (всё время)",rsa("Иман"),"тип: Иман в БАЗА_ДДС"),
]
for (ri,lbl,cur,note) in b3:
    ws.cell(ri,1).value=lbl; ws.cell(ri,1).font=fnt(10); ws.cell(ri,1).fill=F(LGRAY); ws.cell(ri,1).border=brd(); ws.cell(ri,1).alignment=LA(False)
    ws.cell(ri,2).value=cur; ws.cell(ri,2).font=fnt(11,True,RED); ws.cell(ri,2).fill=F(RED_L); ws.cell(ri,2).border=brd(); ws.cell(ri,2).alignment=RA(); ws.cell(ri,2).number_format=MONEY
    for ci in [3,4,5]: ws.cell(ri,ci).fill=F(LGRAY); ws.cell(ri,ci).border=brd()
    ws.cell(ri,6).value=note; ws.cell(ri,6).font=fnt(9,it=True,col=GRAY); ws.cell(ri,6).fill=F(LGRAY); ws.cell(ri,6).border=brd(); ws.cell(ri,6).alignment=LA(False)
    ws.row_dimensions[ri].height=26

ws.freeze_panes="A4"
print("✓ ОТЧЁТ")

# ════════════════════════════════════════════════════════════════════
# 12. ИНСТРУКЦИЯ
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ИНСТРУКЦИЯ"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = TEAL_M
cw(ws,{"A":4,"B":35,"C":55})
banner(ws,"📖  ИНСТРУКЦИЯ — WAY MARKET IDEAL EDITION","A1:C1",TEAL,14)
ws.row_dimensions[1].height=40

steps=[
    ("ПУЛЬТ",TEAL,"Главный экран. Открывайте каждое утро. Видите выручку за месяц, сегодня, сигналы — просроченные выплаты, Иман. Всё на одном экране."),
    ("НАСТРОЙКИ",TEAL_M,"Задайте название магазина, год, процент маржи (25%), долг поставщикам на старте. Заполните план выручки по месяцам — ДАШБОРД покажет выполнение."),
    ("ВВОД_КАССА",GREEN,"Каждую смену вводите Z-отчёты по каждому способу оплаты. Нажмите СОХРАНИТЬ — данные уйдут в БАЗА_ДДС."),
    ("ВВОД_РАСХОДЫ",AMBER,"Расход на закуп, аренду, ЗП — сразу пишите здесь. Выберите категорию из списка."),
    ("ЗАПИСЬ_ВЫПЛАТ",GOLD,"Планируете выплату поставщику — записывайте заранее. Поставьте статус 'Выплачено' когда перевели. Просрочки видны в ПУЛЬТ."),
    ("ДАШБОРД",TEAL,"Выберите месяц и год. Видите: выручку, расходы, прибыль, план, тренды и графики по дням недели."),
    ("АНАЛИТИКА",TEAL_M,"Самые прибыльные дни недели, лучшие смены, разбивка расходов по категориям — всё здесь."),
    ("ОТЧЁТ",NAVY,"Для руководителя: финансовая модель 25/75, долговая нагрузка, Иман в обращении."),
]
for i,(title,clr,desc) in enumerate(steps,1):
    r=i*3-1
    ws.cell(r,1).value=str(i); ws.cell(r,1).font=fnt(14,True,WHITE); ws.cell(r,1).fill=F(clr); ws.cell(r,1).alignment=CA(); ws.row_dimensions[r].height=30
    ws.cell(r,2).value=title; ws.cell(r,2).font=fnt(13,True,WHITE); ws.cell(r,2).fill=F(clr); ws.cell(r,2).alignment=LA(False); ws.cell(r,2).border=brd()
    ws.merge_cells(start_row=r,start_column=3,end_row=r+1,end_column=3)
    ws.cell(r,3).value=desc; ws.cell(r,3).font=fnt(10); ws.cell(r,3).fill=F(LGRAY); ws.cell(r,3).alignment=LA(True); ws.cell(r,3).border=brd()
    ws.row_dimensions[r+1].height=0

ws.freeze_panes="A2"
print("✓ ИНСТРУКЦИЯ")

# ════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════
fname="WAY_MARKET_ideal.xlsx"
wb.save(fname)
sz=os.path.getsize(fname)//1024
print(f"\n✅  {fname} saved — {sz} KB, {len(wb.sheetnames)} sheets")
print("Sheets:", ", ".join(wb.sheetnames))
