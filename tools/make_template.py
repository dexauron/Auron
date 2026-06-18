# -*- coding: utf-8 -*-
"""
Генератор Excel-шаблона Auron Finance.
Структура — по утверждённым документам docs/03-SCREENS.md и docs/04-DESIGN-SYSTEM.md.
Работает и на компьютере (Excel), и на телефоне (Google Таблицы / Excel mobile).
Все балансы, расхождения и итоги считаются формулами автоматически.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Цвета из дизайн-системы (04-DESIGN-SYSTEM.md) ──────────────────────────
BRAND   = "5E5CE6"   # индиго
INCOME  = "30D158"   # зелёный — доходы
EXPENSE = "FF453A"   # красный — расходы
WARN    = "FF9F0A"   # оранжевый — переводы/предупреждения
BG      = "F2F2F7"   # фон
CARD    = "FFFFFF"   # карточки
INK     = "1C1C1E"   # текст
INK2    = "8E8E93"   # вторичный текст
LINE    = "E5E5EA"   # разделители
HEADBG  = "2C2C42"   # тёмный заголовок секций

WHITE = "FFFFFF"

thin = Side(style="thin", color=LINE)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)

def title_cell(ws, cell, text, size=20):
    ws[cell] = text
    ws[cell].font = Font(name="Calibri", size=size, bold=True, color=INK)

def header_row(ws, row, headers, start_col=1, bg=BRAND, color=WHITE):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, color=color, size=11)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_all

def section_banner(ws, row, col_span, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=13, color=WHITE)
    c.fill = fill(HEADBG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 26

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 1: СПРАВОЧНИКИ  (источник выпадающих списков)
# ═══════════════════════════════════════════════════════════════════════════
ref = wb.active
ref.title = "Справочники"
ref.sheet_properties.tabColor = INK2

title_cell(ref, "A1", "⚙️  Справочники", 18)
ref["A2"] = "Заполните один раз. Отсюда берутся выпадающие списки на всех листах."
ref["A2"].font = Font(size=10, italic=True, color=INK2)

# Счета
section_banner(ref, 4, 3, "СЧЕТА (где лежат деньги)")
header_row(ref, 5, ["Счёт", "Начальный остаток, ₽", "Иконка"])
accounts = [
    ("Наличные", 0, "💵"),
    ("Карта/Банк", 0, "💳"),
    ("СБП", 0, "📱"),
    ("Сейф", 0, "🔒"),
]
for i, (name, bal, icon) in enumerate(accounts):
    r = 6 + i
    ref.cell(row=r, column=1, value=name).border = border_all
    cb = ref.cell(row=r, column=2, value=bal); cb.border = border_all; cb.number_format = '#,##0'
    ref.cell(row=r, column=3, value=icon).border = border_all

# Категории
section_banner(ref, 12, 3, "КАТЕГОРИИ ОПЕРАЦИЙ")
header_row(ref, 13, ["Категория", "Тип", "Иконка"])
categories = [
    ("Продажи", "Доход", "💰"), ("Z-отчёт", "Доход", "🧾"),
    ("Закупка", "Расход", "🛒"), ("ЗП", "Расход", "👥"),
    ("Аренда", "Расход", "🏠"), ("Коммуналка", "Расход", "💡"),
    ("Налоги", "Расход", "🏛"), ("Хозрасходы", "Расход", "🔧"),
    ("Реклама", "Расход", "📢"), ("Инкассация", "Расход", "🏦"),
    ("Оплата поставщику", "Расход", "📝"), ("Прочий расход", "Расход", "📋"),
    ("Перевод", "Перевод", "↔"),
]
for i, (name, typ, icon) in enumerate(categories):
    r = 14 + i
    ref.cell(row=r, column=1, value=name).border = border_all
    ref.cell(row=r, column=2, value=typ).border = border_all
    ref.cell(row=r, column=3, value=icon).border = border_all

# Сотрудники (имена для кассы/табеля) — справа
ref["E4"] = "СОТРУДНИКИ"
ref["E4"].font = Font(bold=True, size=13, color=WHITE)
ref["E4"].fill = fill(HEADBG)
ref.merge_cells("E4:G4")
ref.row_dimensions[4].height = 26
header_row(ref, 5, ["Имя", "Должность", "Ставка ₽/день"], start_col=5)
employees = [
    ("Иванова А.", "Продавец", 1500),
    ("Петров В.", "Продавец", 1500),
    ("", "", ""),
    ("", "", ""),
]
for i, (name, role, rate) in enumerate(employees):
    r = 6 + i
    ref.cell(row=r, column=5, value=name).border = border_all
    ref.cell(row=r, column=6, value=role).border = border_all
    rc = ref.cell(row=r, column=7, value=rate); rc.border = border_all
    if rate != "": rc.number_format = '#,##0'

# Типы операций список (для валидации)
ref["I1"] = "Доход"; ref["I2"] = "Расход"; ref["I3"] = "Перевод"
for c in ("I1", "I2", "I3"):
    ref[c].font = Font(color=INK2, size=9)

widths = {"A": 22, "B": 20, "C": 8, "D": 3, "E": 18, "F": 16, "G": 14, "H": 3, "I": 10}
for col, w in widths.items():
    ref.column_dimensions[col].width = w

# именованные диапазоны для выпадающих списков
wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
    "сч", attr_text="Справочники!$A$6:$A$9"))
wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
    "кат", attr_text="Справочники!$A$14:$A$26"))
wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
    "сотр", attr_text="Справочники!$E$6:$E$9"))
wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
    "типы", attr_text="Справочники!$I$1:$I$3"))

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 2: ОПЕРАЦИИ  (журнал всех доходов/расходов)
# ═══════════════════════════════════════════════════════════════════════════
ops = wb.create_sheet("Операции")
ops.sheet_properties.tabColor = BRAND
title_cell(ops, "A1", "📝  Операции", 18)
ops["A2"] = "Каждая строка — одна операция. Тип, категорию и счёт выбирайте из списка (стрелочка в ячейке)."
ops["A2"].font = Font(size=10, italic=True, color=INK2)

OPS_HEADERS = ["Дата", "Тип", "Категория", "Счёт", "Сумма, ₽", "Сотрудник", "Комментарий"]
header_row(ops, 4, OPS_HEADERS)
ops.freeze_panes = "A5"

OPS_ROWS = 500
for r in range(5, 5 + OPS_ROWS):
    for col in range(1, 8):
        ops.cell(row=r, column=col).border = border_all
    ops.cell(row=r, column=1).number_format = 'DD.MM.YYYY'
    amt = ops.cell(row=r, column=5)
    amt.number_format = '#,##0;[Red]-#,##0'

# Раскраска суммы по типу: условное форматирование
from openpyxl.formatting.rule import FormulaRule
ops.conditional_formatting.add(
    f"E5:E{4+OPS_ROWS}",
    FormulaRule(formula=['$B5="Доход"'], font=Font(color=INCOME, bold=True)))
ops.conditional_formatting.add(
    f"E5:E{4+OPS_ROWS}",
    FormulaRule(formula=['$B5="Расход"'], font=Font(color=EXPENSE, bold=True)))
ops.conditional_formatting.add(
    f"E5:E{4+OPS_ROWS}",
    FormulaRule(formula=['$B5="Перевод"'], font=Font(color=WARN, bold=True)))

# Валидации
dv_type = DataValidation(type="list", formula1="=типы", allow_blank=True)
dv_cat  = DataValidation(type="list", formula1="=кат",  allow_blank=True)
dv_acc  = DataValidation(type="list", formula1="=сч",   allow_blank=True)
dv_emp  = DataValidation(type="list", formula1="=сотр", allow_blank=True)
ops.add_data_validation(dv_type); dv_type.add(f"B5:B{4+OPS_ROWS}")
ops.add_data_validation(dv_cat);  dv_cat.add(f"C5:C{4+OPS_ROWS}")
ops.add_data_validation(dv_acc);  dv_acc.add(f"D5:D{4+OPS_ROWS}")
ops.add_data_validation(dv_emp);  dv_emp.add(f"F5:F{4+OPS_ROWS}")

ow = {"A": 13, "B": 11, "C": 20, "D": 14, "E": 14, "F": 16, "G": 34}
for col, w in ow.items():
    ops.column_dimensions[col].width = w

# пример первой строки
ops["A5"] = "=TODAY()"; ops["B5"] = "Расход"; ops["C5"] = "Закупка"
ops["D5"] = "Наличные"; ops["E5"] = 15000; ops["G5"] = "Пример: закупка молочки"

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 3: ГЛАВНАЯ  (дашборд — всё считается само)
# ═══════════════════════════════════════════════════════════════════════════
home = wb.create_sheet("Главная", 0)   # поставить первой
home.sheet_properties.tabColor = INCOME
home.sheet_view.showGridLines = False

title_cell(home, "B2", "🏪  Way Market — сводка", 22)
home["B3"] = "Цифры обновляются сами, когда вы вносите операции на листе «Операции»."
home["B3"].font = Font(size=10, italic=True, color=INK2)

def card(ws, top_left, bottom_right, color=CARD):
    from openpyxl.utils.cell import range_boundaries
    min_c, min_r, max_c, max_r = range_boundaries(f"{top_left}:{bottom_right}")
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            ws.cell(row=r, column=c).fill = fill(color)

# Карточка: Наличных в кассе (баланс счёта «Наличные»)
card(home, "B5", "E9", BRAND)
home["B5"] = "Наличных в кассе"
home["B5"].font = Font(size=12, color=WHITE, bold=True)
home["B6"] = ('=Справочники!B6 + SUMIFS(Операции!E:E,Операции!D:D,"Наличные",Операции!B:B,"Доход")'
              ' - SUMIFS(Операции!E:E,Операции!D:D,"Наличные",Операции!B:B,"Расход")')
home["B6"].font = Font(size=30, bold=True, color=WHITE)
home["B6"].number_format = '#,##0 ₽'

# Карточка: Выручка сегодня
card(home, "B11", "C14", CARD)
home["B11"] = "Выручка сегодня"
home["B11"].font = Font(size=11, color=INK2)
home["B12"] = '=SUMIFS(Операции!E:E,Операции!B:B,"Доход",Операции!A:A,TODAY())'
home["B12"].font = Font(size=20, bold=True, color=INCOME)
home["B12"].number_format = '#,##0 ₽'

# Карточка: Расходы сегодня
card(home, "D11", "E14", CARD)
home["D11"] = "Расходы сегодня"
home["D11"].font = Font(size=11, color=INK2)
home["D12"] = '=SUMIFS(Операции!E:E,Операции!B:B,"Расход",Операции!A:A,TODAY())'
home["D12"].font = Font(size=20, bold=True, color=EXPENSE)
home["D12"].number_format = '#,##0 ₽'

# Балансы по всем счетам
section_banner(home, 16, 5, "ОСТАТКИ ПО СЧЕТАМ")
header_row(home, 17, ["Счёт", "Остаток, ₽"], start_col=2)
for i in range(4):
    r = 18 + i
    acc_ref = f"Справочники!A{6+i}"
    start_ref = f"Справочники!B{6+i}"
    home.cell(row=r, column=2, value=f"={acc_ref}").border = border_all
    f = (f'=IF({acc_ref}="","",{start_ref}'
         f'+SUMIFS(Операции!E:E,Операции!D:D,{acc_ref},Операции!B:B,"Доход")'
         f'-SUMIFS(Операции!E:E,Операции!D:D,{acc_ref},Операции!B:B,"Расход"))')
    cc = home.cell(row=r, column=3, value=f); cc.border = border_all
    cc.number_format = '#,##0 ₽'; cc.font = Font(bold=True, color=INK)
# Итого
r = 22
tc = home.cell(row=r, column=2, value="ИТОГО"); tc.font = Font(bold=True); tc.border = border_all
ts = home.cell(row=r, column=3, value="=SUM(C18:C21)"); ts.border = border_all
ts.number_format = '#,##0 ₽'; ts.font = Font(bold=True, size=12, color=BRAND)

# Итоги за месяц
section_banner(home, 24, 5, "ЗА ТЕКУЩИЙ МЕСЯЦ")
home["B25"] = "Доходы"
home["C25"] = ('=SUMIFS(Операции!E:E,Операции!B:B,"Доход",'
               'Операции!A:A,">="&EOMONTH(TODAY(),-1)+1,Операции!A:A,"<="&TODAY())')
home["C25"].number_format = '#,##0 ₽'; home["C25"].font = Font(bold=True, color=INCOME)
home["B26"] = "Расходы"
home["C26"] = ('=SUMIFS(Операции!E:E,Операции!B:B,"Расход",'
               'Операции!A:A,">="&EOMONTH(TODAY(),-1)+1,Операции!A:A,"<="&TODAY())')
home["C26"].number_format = '#,##0 ₽'; home["C26"].font = Font(bold=True, color=EXPENSE)
home["B27"] = "Прибыль"
home["C27"] = "=C25-C26"
home["C27"].number_format = '#,##0 ₽'; home["C27"].font = Font(bold=True, size=13, color=BRAND)
for rr in (25, 26, 27):
    home.cell(row=rr, column=2).font = Font(color=INK2, size=11)

hw = {"A": 3, "B": 22, "C": 18, "D": 18, "E": 16}
for col, w in hw.items():
    home.column_dimensions[col].width = w
home.row_dimensions[6].height = 40

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 4: КАССА  (Z-отчёты по сменам, расхождение само)
# ═══════════════════════════════════════════════════════════════════════════
kassa = wb.create_sheet("Касса")
kassa.sheet_properties.tabColor = BRAND
title_cell(kassa, "A1", "🧾  Касса — Z-отчёты", 18)
kassa["A2"] = "Расхождение = (Факт наличные + Выплаты) − Z-наличные. Красное — недостача, оранжевое — излишек."
kassa["A2"].font = Font(size=10, italic=True, color=INK2)

KH = ["Дата", "Смена", "Кассир",
      "Z: Наличные", "Z: Карта", "Z: СБП",
      "Факт: Наличные", "Выплаты с кассы",
      "Z Итого", "Расхождение"]
header_row(kassa, 4, KH)
kassa.freeze_panes = "A5"
KROWS = 200
for r in range(5, 5 + KROWS):
    for col in range(1, 11):
        kassa.cell(row=r, column=col).border = border_all
    kassa.cell(row=r, column=1).number_format = 'DD.MM.YYYY'
    for col in (4, 5, 6, 7, 8, 9, 10):
        kassa.cell(row=r, column=col).number_format = '#,##0'
    # Z Итого = D+E+F
    kassa.cell(row=r, column=9, value=f"=D{r}+E{r}+F{r}")
    # Расхождение = (Факт нал + Выплаты) - Z наличные
    kassa.cell(row=r, column=10, value=f"=(G{r}+H{r})-D{r}")

kassa.conditional_formatting.add(
    f"J5:J{4+KROWS}",
    FormulaRule(formula=[f"J5<0"], font=Font(color=EXPENSE, bold=True)))
kassa.conditional_formatting.add(
    f"J5:J{4+KROWS}",
    FormulaRule(formula=[f"J5>0"], font=Font(color=WARN, bold=True)))

dv_shift = DataValidation(type="list", formula1='"Дневная,Ночная,Полный день"', allow_blank=True)
dv_kemp  = DataValidation(type="list", formula1="=сотр", allow_blank=True)
kassa.add_data_validation(dv_shift); dv_shift.add(f"B5:B{4+KROWS}")
kassa.add_data_validation(dv_kemp);  dv_kemp.add(f"C5:C{4+KROWS}")

kw = {"A": 13, "B": 13, "C": 16, "D": 12, "E": 11, "F": 11,
      "G": 14, "H": 16, "I": 12, "J": 14}
for col, w in kw.items():
    kassa.column_dimensions[col].width = w
kassa["A5"] = "=TODAY()"; kassa["B5"] = "Дневная"; kassa["C5"] = "Иванова А."
kassa["D5"] = 48000; kassa["E5"] = 30000; kassa["F5"] = 12000; kassa["G5"] = 47700; kassa["H5"] = 20000

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 5: ПОСТАВЩИКИ  (долги: взяли / оплатили / остаток)
# ═══════════════════════════════════════════════════════════════════════════
sup = wb.create_sheet("Поставщики")
sup.sheet_properties.tabColor = EXPENSE
title_cell(sup, "A1", "🤝  Поставщики и долги", 18)
sup["A2"] = "Каждая операция с поставщиком — отдельная строка. Остаток долга по каждому считается ниже."
sup["A2"].font = Font(size=10, italic=True, color=INK2)

# Журнал операций по долгам
section_banner(sup, 4, 6, "ЖУРНАЛ: ЗАКУПКИ В ДОЛГ И ОПЛАТЫ")
SH = ["Дата", "Поставщик", "Тип", "Сумма, ₽", "Счёт оплаты", "Комментарий"]
header_row(sup, 5, SH)
SROWS = 300
for r in range(6, 6 + SROWS):
    for col in range(1, 7):
        sup.cell(row=r, column=col).border = border_all
    sup.cell(row=r, column=1).number_format = 'DD.MM.YYYY'
    sup.cell(row=r, column=4).number_format = '#,##0'

dv_dtype = DataValidation(type="list", formula1='"Долг (закупка),Оплата"', allow_blank=True)
dv_sacc  = DataValidation(type="list", formula1="=сч", allow_blank=True)
sup.add_data_validation(dv_dtype); dv_dtype.add(f"C6:C{5+SROWS}")
sup.add_data_validation(dv_sacc);  dv_sacc.add(f"E6:E{5+SROWS}")
sup.conditional_formatting.add(
    f"D6:D{5+SROWS}",
    FormulaRule(formula=['$C6="Долг (закупка)"'], font=Font(color=EXPENSE, bold=True)))
sup.conditional_formatting.add(
    f"D6:D{5+SROWS}",
    FormulaRule(formula=['$C6="Оплата"'], font=Font(color=INCOME, bold=True)))

# Сводка по поставщикам (справа)
sup["H4"] = "ОСТАТОК ДОЛГА ПО ПОСТАВЩИКАМ"
sup["H4"].font = Font(bold=True, size=13, color=WHITE)
sup["H4"].fill = fill(HEADBG); sup.merge_cells("H4:J4")
sup.row_dimensions[4].height = 26
header_row(sup, 5, ["Поставщик", "Взяли в долг", "Остаток долга"], start_col=8)
supplier_names = ["Иванов П.", "ООО Альфа", "Петров К.", "", "", "", ""]
for i, nm in enumerate(supplier_names):
    r = 6 + i
    sup.cell(row=r, column=8, value=nm).border = border_all
    nref = f"H{r}"
    debt = (f'=IF({nref}="","",SUMIFS($D$6:$D${5+SROWS},$B$6:$B${5+SROWS},{nref},'
            f'$C$6:$C${5+SROWS},"Долг (закупка)"))')
    bal = (f'=IF({nref}="","",SUMIFS($D$6:$D${5+SROWS},$B$6:$B${5+SROWS},{nref},'
           f'$C$6:$C${5+SROWS},"Долг (закупка)")'
           f'-SUMIFS($D$6:$D${5+SROWS},$B$6:$B${5+SROWS},{nref},$C$6:$C${5+SROWS},"Оплата"))')
    dc = sup.cell(row=r, column=9, value=debt); dc.border = border_all; dc.number_format = '#,##0'
    bc = sup.cell(row=r, column=10, value=bal); bc.border = border_all
    bc.number_format = '#,##0'; bc.font = Font(bold=True, color=EXPENSE)
# общий долг
sup.cell(row=13, column=8, value="ИТОГО ДОЛГ").font = Font(bold=True)
tot = sup.cell(row=13, column=10, value="=SUM(J6:J12)")
tot.number_format = '#,##0 ₽'; tot.font = Font(bold=True, size=12, color=EXPENSE)

# выпадающий список поставщиков в журнале — из этого же списка
wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
    "пост", attr_text="Поставщики!$H$6:$H$12"))
dv_sname = DataValidation(type="list", formula1="=пост", allow_blank=True)
sup.add_data_validation(dv_sname); dv_sname.add(f"B6:B{5+SROWS}")

sw = {"A": 13, "B": 16, "C": 16, "D": 13, "E": 14, "F": 26,
      "G": 3, "H": 16, "I": 14, "J": 15}
for col, w in sw.items():
    sup.column_dimensions[col].width = w
sup["A6"] = "=TODAY()"; sup["B6"] = "Иванов П."; sup["C6"] = "Долг (закупка)"
sup["D6"] = 50000; sup["F6"] = "Пример: накладная молочки"

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 6: АНАЛИТИКА  (по категориям + кассиры)
# ═══════════════════════════════════════════════════════════════════════════
an = wb.create_sheet("Аналитика")
an.sheet_properties.tabColor = WARN
an.sheet_view.showGridLines = False
title_cell(an, "B2", "📈  Аналитика", 18)
an["B3"] = "Считается по всем операциям. Период — весь введённый."
an["B3"].font = Font(size=10, italic=True, color=INK2)

section_banner(an, 5, 4, "РАСХОДЫ ПО КАТЕГОРИЯМ")
header_row(an, 6, ["Категория", "Сумма, ₽", "Доля"], start_col=2)
expense_cats = ["Закупка", "ЗП", "Аренда", "Коммуналка", "Налоги",
                "Хозрасходы", "Реклама", "Оплата поставщику", "Прочий расход"]
start = 7
for i, cat in enumerate(expense_cats):
    r = start + i
    an.cell(row=r, column=2, value=cat).border = border_all
    sc = an.cell(row=r, column=3,
                 value=f'=SUMIFS(Операции!E:E,Операции!C:C,"{cat}",Операции!B:B,"Расход")')
    sc.border = border_all; sc.number_format = '#,##0'
    pc = an.cell(row=r, column=4, value=f"=IF($C${start+len(expense_cats)}=0,0,C{r}/$C${start+len(expense_cats)})")
    pc.border = border_all; pc.number_format = '0%'
rt = start + len(expense_cats)
an.cell(row=rt, column=2, value="ИТОГО РАСХОДЫ").font = Font(bold=True)
tc = an.cell(row=rt, column=3, value=f"=SUM(C{start}:C{rt-1})")
tc.number_format = '#,##0 ₽'; tc.font = Font(bold=True, color=EXPENSE)

# Доходы (справа)
an["F5"] = "ДОХОДЫ"
an["F5"].font = Font(bold=True, size=13, color=WHITE); an["F5"].fill = fill(HEADBG)
an.merge_cells("F5:H5");
header_row(an, 6, ["Категория", "Сумма, ₽"], start_col=6)
income_cats = ["Продажи", "Z-отчёт", "Прочий доход"]
for i, cat in enumerate(income_cats):
    r = 7 + i
    an.cell(row=r, column=6, value=cat).border = border_all
    sc = an.cell(row=r, column=7,
                 value=f'=SUMIFS(Операции!E:E,Операции!C:C,"{cat}",Операции!B:B,"Доход")')
    sc.border = border_all; sc.number_format = '#,##0'
an.cell(row=11, column=6, value="ИТОГО ДОХОДЫ").font = Font(bold=True)
ic = an.cell(row=11, column=7, value="=SUM(G7:G10)")
ic.number_format = '#,##0 ₽'; ic.font = Font(bold=True, color=INCOME)

# Кассиры — выручка и расхождения
section_banner(an, 18, 8, "КАССИРЫ: ВЫРУЧКА И РАСХОЖДЕНИЯ")
header_row(an, 19, ["Кассир", "Выручка (Z итого), ₽", "Смен", "Сумма расхождений, ₽"], start_col=2)
for i in range(4):
    r = 20 + i
    nref = f"Справочники!E{6+i}"
    an.cell(row=r, column=2, value=f"={nref}").border = border_all
    rev = an.cell(row=r, column=3,
                  value=f'=IF({nref}="","",SUMIFS(Касса!I:I,Касса!C:C,{nref}))')
    rev.border = border_all; rev.number_format = '#,##0'
    cnt = an.cell(row=r, column=4,
                  value=f'=IF({nref}="","",COUNTIFS(Касса!C:C,{nref}))')
    cnt.border = border_all
    dif = an.cell(row=r, column=5,
                  value=f'=IF({nref}="","",SUMIFS(Касса!J:J,Касса!C:C,{nref}))')
    dif.border = border_all; dif.number_format = '#,##0;[Red]-#,##0'

aw = {"A": 3, "B": 20, "C": 18, "D": 10, "E": 22, "F": 18, "G": 14, "H": 10}
for col, w in aw.items():
    an.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 7: ИНСТРУКЦИЯ
# ═══════════════════════════════════════════════════════════════════════════
info = wb.create_sheet("Инструкция", 0)
info.sheet_properties.tabColor = INK2
info.sheet_view.showGridLines = False
title_cell(info, "B2", "📖  Как пользоваться", 20)
lines = [
    ("", ""),
    ("1. Справочники", "Сначала откройте лист «Справочники». Впишите свои счета и начальные остатки, своих сотрудников. Категории уже готовы — можно добавить свои."),
    ("2. Операции", "Основной лист. Каждая трата или приход — новая строка. Дата, тип (Доход/Расход), категория, счёт, сумма. Всё выбирается из списка."),
    ("3. Главная", "Ничего не вводите — здесь всё считается само: остатки по счетам, выручка, прибыль за месяц."),
    ("4. Касса", "В конце смены вносите Z-отчёт. Расхождение посчитается автоматически: красное — недостача, оранжевое — излишек."),
    ("5. Поставщики", "Закупки в долг и оплаты. Остаток долга по каждому поставщику считается сам."),
    ("6. Аналитика", "Расходы по категориям, доходы, рейтинг кассиров — всё автоматически."),
    ("", ""),
    ("💻 На компьютере", "Откройте файл в Excel или загрузите в Google Таблицы (drive.google.com → Создать → Загрузить файл)."),
    ("📱 На телефоне", "Установите приложение «Google Таблицы» или «Excel». Откройте тот же файл — заполнять можно прямо с телефона, всё синхронизируется."),
    ("👥 Вместе", "Через Google Таблицы можно дать доступ бухгалтеру и владельцу — все видят одни и те же данные в реальном времени."),
    ("", ""),
    ("⚠️ Важно", "Суммы вводите числом без знака минус — тип «Расход» сам всё посчитает правильно. Не удаляйте лист «Справочники» — на нём держатся все списки."),
]
r = 4
for head, text in lines:
    if head:
        c = info.cell(row=r, column=2, value=head)
        c.font = Font(bold=True, size=12, color=BRAND)
        t = info.cell(row=r, column=3, value=text)
        t.font = Font(size=11, color=INK)
        t.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
info.column_dimensions["A"].width = 3
info.column_dimensions["B"].width = 20
info.column_dimensions["C"].width = 90
for rr in range(4, r):
    info.row_dimensions[rr].height = 34

# порядок листов: Инструкция, Главная, Операции, Касса, Поставщики, Аналитика, Справочники
order = ["Инструкция", "Главная", "Операции", "Касса", "Поставщики", "Аналитика", "Справочники"]
wb._sheets.sort(key=lambda s: order.index(s.title))

OUT = "/home/user/Auron/Auron_шаблон.xlsx"
wb.save(OUT)
print("Сохранено:", OUT)
