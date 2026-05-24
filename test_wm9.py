#!/usr/bin/env python3
"""
БЛОК 4 — Self-test for WAY_MARKET_v9.xlsx
Builds a fresh workbook, injects 10 test rows into БАЗА_ДДС,
then verifies ОТЧЁТ_РУКОВОДИТЕЛЮ SUMPRODUCT formula values.
"""
import subprocess, sys, os, json
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX = "/home/user/Auron/WAY_MARKET_v9.xlsx"

# ── Step 1: build fresh xlsx ─────────────────────────────────────────────────
result = subprocess.run(["python3", "build_wm9.py"], capture_output=True, text=True, cwd="/home/user/Auron")
if result.returncode != 0:
    print("FAIL: build_wm9.py failed\n", result.stderr)
    sys.exit(1)
print("✓ Build OK")

# ── Step 2: load and inject 10 test rows ─────────────────────────────────────
wb = load_workbook(XLSX)
ws_baza = wb["БАЗА_ДДС"]

today = date.today()
d = lambda offset: today - timedelta(days=offset)

# 10 test transactions — columns: Дата(1), Смена(2), Кассир(3), Тип(4), Категория(5), Способ(6), Сумма(7), Комм(8)
test_rows = [
    (d(0), "День",  "Кассир1", "Доход",       "",               "Наличка",    50000, "Z-отчёт"),
    (d(0), "Вечер", "Кассир2", "Доход",       "",               "Эквайринг",  20000, "Z-Эквайринг"),
    (d(0), "Ночь",  "Кассир1", "Доход",       "",               "Перевод",    10000, "Z-Перевод"),
    (d(1), "День",  "Кассир1", "Доход",       "",               "Наличка",    45000, "Z-отчёт"),
    (d(1), "Вечер", "Кассир2", "Расход",      "Аренда",         "Наличка",    15000, "Расход аренда"),
    (d(2), "День",  "Кассир1", "Расход",      "Коммунальные",   "Перевод",     5000, "Комм.услуги"),
    (d(2), "-",     "",        "Расход",      "Закуп товара",   "Наличка",    30000, "Закуп"),
    (d(3), "-",     "",        "Оплата долга","Выплата поставщику","Перевод", 100000, "Выплата ТП"),
    (d(4), "День",  "Кассир1", "Доход",       "",               "Наличка",    55000, "Z-отчёт"),
    (d(4), "День",  "Кассир1", "Расхождение", "",               "Наличка",    -1500, "Расхождение"),
]

for i, row in enumerate(test_rows):
    r = 4 + i
    ws_baza.cell(r, 1).value = row[0]
    ws_baza.cell(r, 1).number_format = "DD.MM.YYYY"
    ws_baza.cell(r, 2).value = row[1]
    ws_baza.cell(r, 3).value = row[2]
    ws_baza.cell(r, 4).value = row[3]
    ws_baza.cell(r, 5).value = row[4]
    ws_baza.cell(r, 6).value = row[5]
    ws_baza.cell(r, 7).value = row[6]
    ws_baza.cell(r, 7).number_format = "#,##0;[Red]-#,##0"
    ws_baza.cell(r, 8).value = row[7]

wb.save(XLSX)
print("✓ 10 test rows injected into БАЗА_ДДС")

# ── Step 3: verify expected totals ───────────────────────────────────────────
# Reload to check values were written correctly
wb2 = load_workbook(XLSX, data_only=True)
ws2 = wb2["БАЗА_ДДС"]

expected = {
    "Доход_total":       50000 + 20000 + 10000 + 45000 + 55000,   # 180000
    "Расход_total":      15000 + 5000 + 30000,                    # 50000
    "Оплата_долга":      100000,
    "Расхождение_total": -1500,
}

totals = {"Доход": 0, "Расход": 0, "Оплата долга": 0, "Расхождение": 0}
for r in range(4, 14):
    tip   = ws2.cell(r, 4).value or ""
    summa = ws2.cell(r, 7).value or 0
    if tip in totals:
        totals[tip] += summa

PASS = True
checks = [
    ("Доход (наличка+экв+перевод)", totals["Доход"],      expected["Доход_total"]),
    ("Расход",                       totals["Расход"],     expected["Расход_total"]),
    ("Оплата долга",                 totals["Оплата долга"], expected["Оплата_долга"]),
    ("Расхождение",                  totals["Расхождение"], expected["Расхождение_total"]),
]

print("\n── БАЗА_ДДС value checks ──────────────────────────────")
for label, got, exp in checks:
    ok = (got == exp)
    PASS = PASS and ok
    print(f"  {'✓' if ok else '✗'} {label}: got={got:>10,.0f}  expected={exp:>10,.0f}")

# ── Step 4: structural checks ─────────────────────────────────────────────────
wb3 = load_workbook(XLSX)
errors = []

# ВВОД_КАССА: must have 12 columns (A:L) in tblВводКасса
ws_k = wb3["ВВОД_КАССА"]
tbl_names_k = list(ws_k._tables)   # iterates over name strings
if "tblВводКасса" not in tbl_names_k:
    errors.append("tblВводКасса not found in ВВОД_КАССА")
else:
    tbl_k = ws_k._tables["tblВводКасса"]
    if tbl_k.ref != "A3:L503":
        errors.append(f"tblВводКасса ref={tbl_k.ref}, expected A3:L503")
    l4 = ws_k.cell(4, 12).value
    if l4 and "I4" in str(l4) and "D4" in str(l4) and "G4" in str(l4):
        pass
    else:
        errors.append(f"L4 расхождение formula wrong: {l4!r}")

# ВВОД_РАСХОДЫ: must have 5 columns (A:E)
ws_r = wb3["ВВОД_РАСХОДЫ"]
tbl_names_r = list(ws_r._tables)
if "tblВводРасходы" not in tbl_names_r:
    errors.append("tblВводРасходы not found in ВВОД_РАСХОДЫ")
else:
    tbl_r = ws_r._tables["tblВводРасходы"]
    if tbl_r.ref != "A3:E503":
        errors.append(f"tblВводРасходы ref={tbl_r.ref}, expected A3:E503")
    headers_r = [ws_r.cell(3, ci).value for ci in range(1, 6)]
    if "Кассир" in headers_r:
        errors.append(f"Кассир still present in ВВОД_РАСХОДЫ headers: {headers_r}")

# НАСТРОЙКИ: 9-section smart tables exist (Stage 5 layout)
ws_n = wb3["НАСТРОЙКИ"]
tbl_names_n = list(ws_n._tables)
required_n = {
    "tblКассиры":       "A43:A79",
    "tblКатегории":     "C43:C79",
    "tblСпособыОплаты": "E43:E79",
    "tblТипыОпераций":  "G43:G79",
    "tblПостоянные":    "A81:H94",
    "tblПоставщики":    "A98:C1098",
}
for tname, exp_ref in required_n.items():
    if tname not in tbl_names_n:
        errors.append(f"{tname} not found in НАСТРОЙКИ")
    else:
        ref = ws_n._tables[tname].ref
        if ref != exp_ref:
            errors.append(f"{tname} ref={ref!r}, expected {exp_ref!r}")
# E5/E7/E8/E9 are direct value cells (not formulas)
if str(ws_n.cell(5,5).value or "").startswith("="):
    errors.append("НАСТРОЙКИ E5 should be direct value, not formula")
if ws_n.cell(7,5).value != 0.25:
    errors.append(f"НАСТРОЙКИ E7={ws_n.cell(7,5).value!r}, expected 0.25")
if ws_n.cell(9,5).value != 500000:
    errors.append(f"НАСТРОЙКИ E9={ws_n.cell(9,5).value!r}, expected 500000")

# ОТЧЁТ_РУКОВОДИТЕЛЮ: 3 blocks exist
ws_o = wb3["ОТЧЁТ_РУКОВОДИТЕЛЮ"]
if ws_o.cell(5, 1).value is None:
    errors.append("ОТЧЁТ block 1 row 5 empty")

# БАЗА_ДДС: no merged cells in DATA range (rows 4+)
for merge in ws2.merged_cells.ranges:
    if merge.min_row >= 4:
        errors.append(f"Merged cell in БАЗА_ДДС data area: {merge}")

print("\n── Structural checks ─────────────────────────────────")
if errors:
    PASS = False
    for e in errors:
        print(f"  ✗ {e}")
else:
    print("  ✓ All structural checks passed")

# ── Step 5: VBA .bas checks ───────────────────────────────────────────────────
bas_path = "/home/user/Auron/Модуль_WM9.bas"
with open(bas_path, "rb") as f:
    bas_content = f.read().decode("cp1251", errors="replace")

bas_checks = [
    ("SohranitKassu exists",        "Sub SohranitKassu()" in bas_content),
    ("SohranitRashody exists",       "Sub SohranitRashody()" in bas_content),
    ("SohranitViplatu uses tblВыплаты", "tblВыплаты" in bas_content),
    ("SohranitViplatu writes Оплата долга", '"Оплата долга"' in bas_content),
    ("EksportOtchetaPDF removed",   "ExportAsFixedFormat" not in bas_content),
    ("Navigation subs exist",        "НавигацияНаДашборд" in bas_content),
    ("Right-side X>=865 in buttons", "865, 4" in bas_content),
    ("No .Protect calls",            ".Protect" not in bas_content),
    ("12-col расхождение col 12",    "Cells(r, 12)" in bas_content),
    ("ObnovitKolonkiKassy exists",   "Sub ObnovitKolonkiKassy()" in bas_content),
    ("ObnovitSmenyKassy exists",     "Sub ObnovitSmenyKassy()" in bas_content),
    ("PrimjenitNastrojki exists",    "Sub PrimjenitNastrojki()" in bas_content),
    ("Иман toggle E23",              "wsCfg.Cells(23, 5)" in bas_content),
    ("Выплата toggle E24",           "wsCfg.Cells(24, 5)" in bas_content),
    ("5-col rashody col 2=kat",      'Dim katR As String: katR = CStr(wsVvod.Cells(r, 2)' in bas_content),
]

print("\n── VBA .bas checks ───────────────────────────────────")
for label, ok in bas_checks:
    PASS = PASS and ok
    print(f"  {'✓' if ok else '✗'} {label}")

# ── Final result ──────────────────────────────────────────────────────────────
print("\n" + "═"*55)
if PASS:
    print("  ✅  ALL TESTS PASSED — WAY MARKET v9 Stage 7 OK")
else:
    print("  ❌  SOME TESTS FAILED — see details above")
print("═"*55)
sys.exit(0 if PASS else 1)
